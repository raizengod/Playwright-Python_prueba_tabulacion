import re # Importa el módulo de expresiones regulares
import time # Importa el módulo para funciones relacionadas con el tiempo
import random # Importa el módulo para generar números aleatorios
from playwright.sync_api import Page, expect, Error, TimeoutError, sync_playwright, Response, Dialog, Locator, BrowserContext # Importa clases y excepciones necesarias de Playwright
from datetime import datetime # Importa la clase datetime para trabajar con fechas y horas
import os # Importa el módulo os para interactuar con el sistema operativo (rutas de archivos, directorios)
from typing import List, Dict, Union, Callable, Tuple, Optional, Any # Importa tipos para mejorar la legibilidad y validación del código
from Simples.utils.config import LOGGER_DIR # Importa la ruta del directorio de logs desde config.py
from Simples.utils.logger import setup_logger # Importa la función setup_logger desde logger.py
import logging # Importa el módulo logging para configurar y usar loggers
import openpyxl # Librería para hacer uso del excel (para archivos .xlsx)
import csv # Importa la librería csv para manejar archivos CSV (para archivos .csv)
import json # Importa la librería json para manejar archivos JSON
import xml.etree.ElementTree as ET # Importa el módulo para trabajar con XML
import math

class Funciones_Globales:
    
    #1- Creamos una función incial 'Constructor'-----ES IMPORTANTE TENER ESTE INICIADOR-----
    def __init__(self, page):
        self.page = page
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""
        self._alerta_input_capturado = ""
        self._dialog_handler_registered = False # <--- ¡Esta línea es crucial!

        # --- Nuevas variables para el manejo de pestañas (popups) ---
        self._popup_detectado = False
        self._popup_page = None # Para almacenar el objeto Page de la nueva pestaña
        self._popup_url_capturado = ""
        self._popup_title_capturado = ""  
        
        # Nueva lista para almacenar todas las nuevas páginas abiertas durante una interacción
        self._all_new_pages_opened_by_click: List[Page] = []
        
        # Registramos el manejador de eventos para nuevas páginas
        # Limpiamos la lista al registrar para evitar resagos de pruebas anteriores
        self.page.context.on("page", self._on_new_page)
        # Esto es importante: Si se va a usar _all_new_pages_opened_by_click,
        # necesitamos una forma de reiniciarla o asegurarnos de que solo contenga
        # las páginas relevantes para la acción actual.
        # Una estrategia es limpiar la lista antes de la acción que abre la nueva ventana,
        # y luego recopilar las páginas.
        
        # Configurar el logger para esta clase
        self.logger = setup_logger(name='Funciones_Globales', console_level=logging.INFO, file_level=logging.DEBUG)
        
    #2- Función para generar el nombre de archivo con marca de tiempo
    def _generar_nombre_archivo_con_timestamp(self, prefijo):
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3] # Quita los últimos 3 dígitos para milisegundos más precisos
        return f"{timestamp}_{prefijo}"
    
    #3- Función para tomar captura de pantalla
    def tomar_captura(self, nombre_base, directorio):
        """
        Toma una captura de pantalla de la página y la guarda en el directorio especificado.
        Por defecto, usa SCREENSHOT_DIR de config.py.

        Args:
            nombre_base (str): El nombre base para el archivo de la captura de pantalla.
            directorio (str): El directorio donde se guardará la captura. Por defecto, SCREENSHOT_DIR.
        """
        try:
            if not os.path.exists(directorio):
                os.makedirs(directorio)
                self.logger.info(f"\n Directorio creado para capturas de pantalla: {directorio}") #

            nombre_archivo = self._generar_nombre_archivo_con_timestamp(nombre_base) #
            ruta_completa = os.path.join(directorio, f"{nombre_archivo}.png") # Cambiado a .png para mejor calidad
            self.page.screenshot(path=ruta_completa) #
            self.logger.info(f"\n 📸 Captura de pantalla guardada en: {ruta_completa}") #
        except Exception as e:
            self.logger.error(f"\n ❌ Error al tomar captura de pantalla '{nombre_base}': {e}") #
        
    #4- unción basica para tiempo de espera que espera recibir el parametro tiempo
    #En caso de no pasar el tiempo por parametro, el mismo tendra un valor de medio segundo
    def esperar_fijo(self, tiempo=0.5):
        """
        Espera un tiempo fijo en segundos.

        Args:
            tiempo (Union[int, float]): El tiempo en segundos a esperar. Por defecto, 0.5 segundos.
        """
        self.logger.debug(f"\n Esperando fijo por {tiempo} segundos...") #
        try:
            time.sleep(tiempo) #
            self.logger.info(f"Espera fija de {tiempo} segundos completada.") #
        except TypeError:
            self.logger.error(f"\n ❌ Error: El tiempo de espera debe ser un número. Se recibió: {tiempo}") #
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado durante la espera fija: {e}") #
        
    #5- Función para indicar el tiempo que se tardará en hacer el scroll
    def scroll_pagina(self, horz, vert, tiempo: Union[int, float] = 0.5):
        """
        Realiza un scroll en la página.

        Args:
            horz (int): Cantidad de scroll horizontal. Por defecto, 0.
            vert (int): Cantidad de scroll vertical. Por defecto, 0.
            tiempo (Union[int, float]): Tiempo de espera después del scroll en segundos. Por defecto, 0.5.
        """
        self.logger.debug(f"Realizando scroll - Horizontal: {horz}, Vertical: {vert}. Espera: {tiempo} segundos.") #
        try:
            # --- Medición de rendimiento: Inicio de la acción de scroll ---
            start_time_scroll_action = time.time()
            self.page.mouse.wheel(horz, vert)
            # --- Medición de rendimiento: Fin de la acción de scroll ---
            end_time_scroll_action = time.time()
            duration_scroll_action = end_time_scroll_action - start_time_scroll_action
            self.logger.info(f"PERFORMANCE: Duración de la acción de scroll (Playwright API): {duration_scroll_action:.4f} segundos.")
            
            self.esperar_fijo(tiempo) # Reutiliza la función esperar_fijo para el log y manejo de errores
            self.logger.info(f"Scroll completado (H: {horz}, V: {vert}).") #
        except Exception as e:
            self.logger.error(f"❌ Error al realizar scroll en la página: {e}") #
            
    # 6- Función para validar que un elemento es visible
    def validar_elemento_visible(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5, resaltar: bool = True) -> bool:
        """
        Valida que un elemento sea visible en la página dentro de un tiempo límite especificado.
        Esta función integra la medición del tiempo que tarda el elemento en volverse visible,
        lo que es útil para métricas de rendimiento de la interfaz de usuario.

        Args:
            selector: El selector del elemento. Puede ser una cadena (CSS, XPath, etc.) o
                      un objeto `Locator` de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para nombrar las capturas de pantalla
                               tomadas durante la ejecución de la validación.
            directorio (str): Ruta del directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el elemento
                                        sea considerado visible. Si el elemento no es visible
                                        dentro de este plazo, la validación fallará.
                                        Por defecto, 5.0 segundos.
            resaltar (bool): Si es `True`, el elemento visible será resaltado brevemente en la
                             página para una confirmación visual durante la ejecución. Por defecto, `True`.

        Returns:
            bool: `True` si el elemento es visible dentro del tiempo especificado; `False` en caso
                  de que no sea visible (por timeout) o si ocurre otro tipo de error.

        Raises:
            Error: Si ocurre un error específico de Playwright (ej., selector inválido,
                   elemento desprendido del DOM).
            Exception: Para cualquier otro error inesperado durante la ejecución.
        """
        self.logger.info(f"\nValidando visibilidad del elemento con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright.
        # Si 'selector' es una cadena, lo convierte a Locator; de lo contrario, usa el objeto directamente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector
        
        # --- Medición de rendimiento: Inicio de la espera por visibilidad ---
        # Registra el tiempo justo antes de iniciar la espera activa de Playwright.
        start_time_visible_check = time.time()

        try:
            # Espera explícita a que el elemento cumpla la condición de ser visible.
            # Playwright reintenta automáticamente hasta que la condición se cumple o
            # el 'timeout' (expresado en milisegundos) expira.
            expect(locator).to_be_visible() 

            # --- Medición de rendimiento: Fin de la espera por visibilidad ---
            # Registra el tiempo inmediatamente después de que el elemento se vuelve visible.
            end_time_visible_check = time.time()
            # Calcula la duración total que tardó el elemento en ser visible.
            duration_visible_check = end_time_visible_check - start_time_visible_check
            # Registra la métrica de rendimiento. Un tiempo elevado aquí puede indicar
            # problemas de carga o renderizado en la aplicación.
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en ser visible: {duration_visible_check:.4f} segundos.")

            if resaltar:
                # Resalta visualmente el elemento en la página para ayudar en el debugging o demostraciones.
                locator.highlight()
                self.logger.debug(f"Elemento '{selector}' resaltado.")

            # Toma una captura de pantalla para documentar que el elemento es visible.
            self.tomar_captura(f"{nombre_base}_visible", directorio)
            self.logger.info(f"\n✔ ÉXITO: El elemento '{selector}' es visible en la página.")
            
            # Realiza una espera fija adicional. Esto es útil para pausas visuales
            # o si el siguiente paso en la prueba requiere un breve momento después
            # de la aparición del elemento. Considera si esta espera es estrictamente
            # necesaria para la lógica de la prueba o si es solo para observación.
            self.esperar_fijo(tiempo) 

            return True

        except TimeoutError as e:
            # Manejo específico para cuando el elemento no se vuelve visible dentro del 'timeout'.
            # Se registra el tiempo transcurrido hasta el fallo.
            end_time_visible_check = time.time()
            duration_visible_check = end_time_visible_check - start_time_visible_check
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento con selector '{selector}' NO fue visible "
                f"después de {duration_visible_check:.4f} segundos (timeout configurado: {tiempo}s). Detalles: {e}"
            )
            self.logger.warning(error_msg)
            # Toma una captura de pantalla en caso de timeout para depuración.
            self.tomar_captura(f"{nombre_base}_NO_visible_timeout", directorio)
            return False

        except Error as e:
            # Manejo específico para errores generados por Playwright (ej. selector malformado,
            # elemento que se desprende del DOM antes de la verificación).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar la visibilidad de '{selector}'. "
                f"Posibles causas: Selector inválido, elemento desprendido del DOM. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # exc_info=True para incluir la traza completa.
            # Toma una captura de pantalla para el error de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise # Re-lanza la excepción para asegurar que la prueba falle.

        except Exception as e:
            # Manejo general para cualquier otra excepción inesperada que no sea de Playwright o Timeout.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al validar la visibilidad de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa critical para errores graves y exc_info.
            # Toma una captura para errores inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque finally se ejecuta siempre, independientemente de si hubo una excepción o no.
            # En este caso, no hay operaciones finales específicas necesarias aquí que no estén ya
            # manejadas en los bloques try/except.
            pass

    # 7- Función para validar que un elemento NO es visible
    def validar_elemento_no_visible(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Valida que un elemento NO es visible en la página dentro de un tiempo límite especificado.
        Esta función integra la medición del tiempo que tarda el elemento en ocultarse o desaparecer,
        lo que es útil para métricas de rendimiento de la interfaz de usuario.

        Args:
            selector (Union[str, Page.locator]): El selector del elemento (puede ser una cadena CSS/XPath)
                                                  o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera (en segundos) para que el elemento
                                        NO sea visible o se oculte. Por defecto, 5.0 segundos.

        Raises:
            AssertionError: Si el elemento permanece visible después del tiempo de espera.
            TimeoutError: Si la operación de espera se agota.
            Error: Si ocurre un error específico de Playwright.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nValidando que el elemento con selector '{selector}' NO es visible. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para su uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la espera por no visibilidad ---
        # Registra el tiempo justo antes de iniciar la espera activa de Playwright
        # para que el elemento se oculte.
        start_time_hidden_check = time.time()

        try:
            # Espera explícita a que el elemento cumpla la condición de estar oculto (no visible)
            # o de no existir en el DOM. Playwright reintenta automáticamente.
            # El 'timeout' se especifica en milisegundos.
            expect(locator).to_be_hidden()

            # --- Medición de rendimiento: Fin de la espera por no visibilidad ---
            # Registra el tiempo inmediatamente después de que el elemento se oculta.
            end_time_hidden_check = time.time()
            # Calcula la duración total que tardó el elemento en ocultarse.
            duration_hidden_check = end_time_hidden_check - start_time_hidden_check
            # Registra la métrica de rendimiento. Un tiempo elevado aquí podría indicar
            # que la aplicación tarda en ocultar elementos o en limpiar el DOM.
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en ocultarse/desaparecer: {duration_hidden_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: El elemento con selector '{selector}' NO es visible.")
            # La captura de éxito se maneja en el bloque `finally` para asegurar que se tome.

        except TimeoutError as e:
            # Captura específica para el error de tiempo de espera de Playwright.
            # Esto ocurre si el elemento sigue visible después del 'timeout' especificado.
            end_time_hidden_check = time.time() # Registra el tiempo al fallar el timeout.
            duration_hidden_check = end_time_hidden_check - start_time_hidden_check
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento con selector '{selector}' AÚN ES VISIBLE "
                f"después de {duration_hidden_check:.4f} segundos (timeout configurado: {tiempo}s). Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla en caso de fallo por timeout para depuración.
            self.tomar_captura(f"{nombre_base}_fallo_no_visible_timeout", directorio)
            raise # Re-lanza la excepción para que la prueba falle.

        except AssertionError as e:
            # Captura específica para AssertionErrors. Esto podría ocurrir si la aserción
            # es `to_be_hidden` y el elemento inesperadamente no se oculta.
            error_msg = (
                f"\n❌ FALLO (Assertion): El elemento con selector '{selector}' aún es visible o no se ocultó a tiempo. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_no_visible_assertion", directorio)
            raise # Re-lanza la excepción para que la prueba falle.
            
        except Error as e:
            # Captura específica para errores internos de Playwright (ej., selector inválido,
            # o problemas con el contexto de la página).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar que '{selector}' NO es visible. "
                f"Posibles causas: Selector inválido, problema de contexto de la página. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_no_visible", directorio)
            raise # Re-lanza la excepción para que la prueba falle.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al validar que '{selector}' NO es visible. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_no_visible", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, independientemente de si la validación fue exitosa o falló.
            # Es un buen lugar para tomar una captura de pantalla final que muestre el estado de la página.
            self.tomar_captura(f"{nombre_base}_estado_final_no_visible", directorio=directorio)

    # 8- Función para verificar que un elemento (o elementos) localizado en una página web contiene un texto específico
    def verificar_texto_contenido(self, selector: Union[str, Page.locator], texto_esperado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Verifica que un elemento localizado en una página web **contiene un texto específico**.
        Esta función está optimizada para **integrar métricas de rendimiento básicas**, midiendo
        el tiempo que tarda el elemento en volverse visible y en contener el texto esperado.

        Args:
            selector (Union[str, Page.locator]): El **selector del elemento** (puede ser una cadena CSS/XPath)
                                                  o un objeto `Locator` de Playwright preexistente.
            texto_esperado (str): El **texto exacto o parcial** que se espera encontrar dentro del elemento.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla** tomadas durante la validación,
                               facilitando la identificación en el directorio de salida.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el elemento
                                        sea visible Y contenga el texto esperado. Si alguna de estas
                                        condiciones no se cumple dentro de este plazo, la validación fallará.
                                        Por defecto, `5.0` segundos.

        Raises:
            TimeoutError: Si el elemento no se hace visible o no contiene el texto esperado
                          dentro del tiempo límite especificado.
            AssertionError: Si el elemento es visible, pero su contenido de texto no coincide
                            con el `texto_esperado`.
            Error: Si ocurre un error específico de Playwright durante la operación (ej.,
                   selector malformado, problema de comunicación con el navegador).
            Exception: Para cualquier otro error inesperado que no esté cubierto por las excepciones anteriores.
        """
        self.logger.info(f"Verificando que el elemento con selector '{selector}' contiene el texto: '{texto_esperado}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Playwright Locator.
        # Esto permite una interacción consistente, ya sea que se pase una cadena de selector
        # o un Locator ya definido.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la espera por visibilidad ---
        # Registra el tiempo en que comienza la operación de esperar a que el elemento sea visible.
        start_time_visible_check = time.time()
        try:
            # Playwright espera implícitamente a que el elemento cumpla la condición de visibilidad.
            # El `timeout` se convierte de segundos a milisegundos, como lo requiere Playwright.
            expect(locator).to_be_visible()
            
            # Registra el tiempo una vez que el elemento se ha vuelto visible.
            end_time_visible_check = time.time()
            # Calcula la duración de esta fase. Esta métrica es vital para entender
            # la latencia de renderizado de la UI.
            duration_visible_check = end_time_visible_check - start_time_visible_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en ser visible: {duration_visible_check:.4f} segundos.")
            self.logger.debug(f"Elemento con selector '{selector}' es visible.")

            # Opcional: **Resalta visualmente el elemento** en la página del navegador.
            # Esto es extremadamente útil para el debugging o para demos visuales de la prueba.
            locator.highlight()
            # Toma una captura de pantalla del estado actual de la página, antes de verificar el texto,
            # para documentar la visibilidad del elemento.
            self.tomar_captura(f"{nombre_base}_antes_verificacion_texto", directorio)

            # --- Medición de rendimiento: Inicio de la espera por el texto ---
            # Registra el tiempo en que comienza la operación de esperar a que el elemento contenga el texto.
            start_time_text_check = time.time()
            # Verifica que el elemento contiene el `texto_esperado`. Playwright también reintenta
            # esta aserción hasta que el texto coincide o el `timeout` se agota.
            expect(locator).to_contain_text(texto_esperado)
            
            # Registra el tiempo una vez que el texto esperado se ha encontrado.
            end_time_text_check = time.time()
            # Calcula la duración de esta fase. Esta métrica es importante si el texto se carga
            # dinámicamente o tarda en aparecer después de que el elemento base es visible.
            duration_text_check = end_time_text_check - start_time_text_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en contener el texto '{texto_esperado}': {duration_text_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            # Toma una captura de pantalla final para documentar la verificación exitosa del texto.
            self.tomar_captura(nombre_base=f"{nombre_base}_despues_verificacion_texto", directorio=directorio)
            
            # Realiza una **espera fija** después de la verificación. Esto puede ser útil para
            # propósitos de sincronización con el siguiente paso de la prueba o para permitir
            # una observación visual si la prueba se ejecuta en modo interactivo.
            self.esperar_fijo(tiempo)

        except TimeoutError as e:
            # Este bloque se ejecuta si el elemento no se hizo visible O no contenía el texto esperado
            # dentro del `tiempo` total especificado.
            end_time_fail = time.time() # Registra el tiempo final de la operación.
            # Calcula la duración total que tardó la operación completa hasta el fallo.
            duration_fail = end_time_fail - start_time_visible_check
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento con selector '{selector}' no se hizo visible o no contenía "
                f"el texto '{texto_esperado}' después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout para depuración.
            self.tomar_captura(f"{nombre_base}_fallo_verificacion_texto_timeout", directorio)
            raise # Re-lanza la excepción para asegurar que la prueba falle.

        except AssertionError as e:
            # Este bloque se ejecuta si el elemento era visible, pero el texto real no coincidía
            # con el `texto_esperado` después de las reintentos de `to_contain_text`.
            error_msg = (
                f"\n❌ FALLO (Aserción): El elemento con selector '{selector}' NO contiene el texto esperado: "
                f"'{texto_esperado}'. Contenido actual: '{locator.text_content() if locator else 'N/A'}' Detalle: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla en el momento del fallo de aserción.
            self.tomar_captura(f"{nombre_base}_fallo_verificacion_texto_contenido", directorio)
            raise # Re-lanza la excepción.

        except Error as e:
            # Este bloque maneja errores específicos de Playwright que no son timeouts ni aserciones fallidas,
            # como un selector malformado o un problema de comunicación con el navegador.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar texto para '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_verificacion_texto", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Este bloque captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el texto para el selector '{selector}'. "
                f"Detalles: {e}"
            )
            # Usa `critical` para errores graves e `exc_info=True` para incluir la traza completa.
            self.logger.critical(error_msg, exc_info=True)
            # Toma una captura para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificacion_texto", directorio)
            raise # Re-lanza la excepción.

    # 9- Función para rellenar campo de texto, tomar capturas y medir rendimiento
    def rellenar_campo_de_texto(self, selector: Union[str, Page.locator], texto, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Rellena un campo de texto con el valor especificado y toma capturas de pantalla
        en puntos clave de la operación. Esta función incluye una **medición de rendimiento**
        para registrar el tiempo que tarda la operación de rellenado (`.fill()`).

        Args:
            selector (Union[str, Page.locator]): El **selector del campo de texto**. Puede ser
                                                  una cadena (por ejemplo, un selector CSS o XPath)
                                                  o un objeto `Locator` de Playwright preexistente.
            texto: El **valor a introducir** en el campo de texto.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función. Esto ayuda a identificar
                               las imágenes en el directorio de salida (ej., "login_campo_usuario").
            directorio (str): **Ruta completa del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo de espera fijo** (en segundos) que se aplicará
                                        después de rellenar el campo. Es útil para pausas visuales
                                        o para permitir que la interfaz de usuario (UI) reaccione
                                        antes de la siguiente acción. Por defecto, `0.5` segundos.

        Raises:
            TimeoutError: Si la operación de `.fill()` excede el tiempo de espera, lo que indica
                          que el elemento no estaba visible, habilitado o editable a tiempo.
            Error: Si ocurre un problema específico de Playwright durante la interacción
                   (ej., el selector es inválido, el elemento se desprende del DOM).
            Exception: Para cualquier otro error inesperado que ocurra durante la ejecución de la función.
        """
        self.logger.info(f"\nRellenando campo con selector '{selector}' con el texto: '{texto}'.")

        # Asegura que 'selector' sea un objeto Locator de Playwright. Esto garantiza que
        # las interacciones (como 'highlight' y 'fill') se realicen de manera consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        try:
            # Resalta visualmente el campo de texto en el navegador. Esto es una ayuda visual
            # excelente durante la ejecución de la prueba o el debugging.
            locator.highlight()
            # Toma una captura de pantalla del estado del campo *antes* de introducir el texto.
            self.tomar_captura(f"{nombre_base}_antes_de_rellenar_texto", directorio)

            # --- Medición de rendimiento: Inicio de la operación de rellenado ---
            # Registra el momento exacto en que comenzamos la operación de 'fill'.
            start_time_fill = time.time()
            
            # Rellena el campo de texto con el valor proporcionado. El método `fill()` de Playwright
            # es robusto: espera automáticamente a que el elemento sea visible, habilitado y editable
            # antes de intentar escribir, lo que reduce la necesidad de esperas explícitas adicionales.
            locator.fill(texto) # Convertimos el 'texto' a cadena para asegurar compatibilidad.
            
            # --- Medición de rendimiento: Fin de la operación de rellenado ---
            # Registra el momento en que la operación de 'fill' ha finalizado.
            end_time_fill = time.time()
            # Calcula la duración total que tomó la operación de rellenado.
            # Esta métrica es fundamental para evaluar la **reactividad de los campos de entrada**
            # y el rendimiento percibido por el usuario.
            duration_fill = end_time_fill - start_time_fill
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en rellenar el campo '{selector}': {duration_fill:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Campo '{selector}' rellenado con éxito con el texto: '{texto}'.")

            # Toma una captura de pantalla del estado del campo *después* de introducir el texto.
            self.tomar_captura(f"{nombre_base}_despues_de_rellenar_texto", directorio)

        except TimeoutError as e:
            # Este bloque se ejecuta si la operación `locator.fill()` no pudo completarse
            # dentro del tiempo de espera implícito de Playwright (que se basa en el timeout
            # global de la página o el definido por el usuario para el locator).
            error_msg = (
                f"\n❌ ERROR (Timeout): El tiempo de espera se agotó al interactuar con el selector '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible, habilitado o editable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa para depuración.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_error_timeout_rellenar", directorio)
            # Re-lanza la excepción como un Error de Playwright para mantener la coherencia en el manejo de errores.
            raise Error(error_msg) from e

        except Error as e:
            # Captura errores específicos de Playwright que no son timeouts. Esto incluye problemas como
            # un selector malformado, un elemento que se desprende del DOM, o fallos de comunicación con el navegador.
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al interactuar con el selector '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_rellenar", directorio)
            raise # Re-lanza la excepción para que la prueba se marque como fallida.

        except Exception as e:
            # Este es un bloque de captura general para cualquier otra excepción inesperada
            # que no haya sido manejada por los tipos de errores anteriores.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al interactuar con el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel 'critical' para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_rellenar", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre, independientemente de si la operación fue exitosa
            # o si se produjo una excepción.
            # Aplica una espera fija después de la operación. Esta espera es útil para permitir
            # que la UI se actualice completamente o para propósitos de observación visual.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 10- Función para rellenar campo numérico positivo y hacer captura de la imagen con medición de rendimiento
    def rellenar_campo_numerico_positivo(self, selector: Union[str, Page.locator], valor_numerico: Union[int, float], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Rellena un campo de texto con un **valor numérico positivo** (entero o flotante)
        y toma capturas de pantalla en puntos clave. Esta función valida el tipo y el
        signo del número, e integra una **medición de rendimiento** para registrar el
        tiempo que tarda la operación de rellenado (`.fill()`).

        Args:
            selector (Union[str, Page.locator]): El **selector del campo de texto** donde se
                                                  introducirá el valor numérico. Puede ser una
                                                  cadena (CSS, XPath, etc.) o un objeto `Locator`.
            valor_numerico (Union[int, float]): El **valor numérico positivo** (entero o flotante)
                                                que se va a introducir en el campo.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas,
                               facilitando su identificación en el directorio de salida.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo de espera fijo** (en segundos) que se aplicará
                                        después de rellenar el campo. Útil para pausas visuales
                                        o para permitir que la UI reaccione. Por defecto, `0.5` segundos.

        Raises:
            ValueError: Si el `valor_numerico` no es un tipo numérico (int/float) o si es negativo.
            TimeoutError: Si la operación de `.fill()` se agota (el elemento no está listo).
            Error: Si ocurre un error específico de Playwright (selector inválido, etc.).
            TypeError: Si el `selector` proporcionado no es un tipo válido (`str` o `Locator`).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nRellenando campo con selector '{selector}' con el valor numérico POSITIVO: '{valor_numerico}'.")

        # --- Validaciones de entrada ---
        # 1. Valida que el 'valor_numerico' sea de tipo numérico (int o float).
        if not isinstance(valor_numerico, (int, float)):
            error_msg = f"\n❌ ERROR: El valor proporcionado '{valor_numerico}' no es un tipo numérico (int o float) válido."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_valor_no_numerico", directorio)
            raise ValueError(error_msg)

        # 2. Valida que el 'valor_numerico' sea positivo (mayor o igual a cero).
        if valor_numerico < 0:
            error_msg = f"\n❌ ERROR: El valor numérico '{valor_numerico}' no es positivo. Se esperaba un número mayor o igual a cero."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_valor_negativo", directorio)
            raise ValueError(error_msg)

        # Convierte el valor numérico a una cadena para poder rellenar el campo de texto.
        valor_a_rellenar_str = str(valor_numerico)

        # Asegura que 'selector' sea un objeto Locator de Playwright.
        # Esto permite una interacción consistente con Playwright.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        elif isinstance(selector, Page.locator): # Asegura que sea un objeto Locator válido
            locator = selector
        else:
            error_msg = f"\n❌ ERROR: El selector proporcionado '{type(selector)}' no es una cadena ni un objeto Locator válido."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_selector_numerico", directorio)
            raise TypeError(error_msg)

        try:
            # Resalta visualmente el campo de texto en el navegador.
            locator.highlight()
            # Toma una captura de pantalla del estado del campo *antes* de rellenarlo.
            self.tomar_captura(f"{nombre_base}_antes_de_rellenar_numerico", directorio)

            # --- Medición de rendimiento: Inicio de la operación de rellenado ---
            # Registra el tiempo justo antes de ejecutar la acción de 'fill'.
            start_time_fill = time.time()
            
            # Rellena el campo de texto con el valor numérico convertido a cadena.
            # El método `fill()` de Playwright esperará automáticamente a que el elemento
            # esté visible, habilitado y editable.
            locator.fill(valor_a_rellenar_str)
            
            # --- Medición de rendimiento: Fin de la operación de rellenado ---
            # Registra el tiempo inmediatamente después de que la operación de 'fill' se ha completado.
            end_time_fill = time.time()
            # Calcula la duración total de la operación de rellenado.
            # Esta métrica es crucial para evaluar la **reactividad de los campos de entrada**,
            # especialmente en formularios donde el rendimiento es crítico.
            duration_fill = end_time_fill - start_time_fill
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en rellenar el campo '{selector}' con '{valor_a_rellenar_str}': {duration_fill:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Campo '{selector}' rellenado con éxito con el valor: '{valor_a_rellenar_str}'.")

            # Toma una captura de pantalla del estado del campo *después* de rellenarlo.
            self.tomar_captura(f"{nombre_base}_despues_de_rellenar_numerico", directorio)

        except TimeoutError as e:
            # Captura específica para errores de tiempo de espera de Playwright.
            # Esto indica que el elemento no estaba listo (visible, habilitado, editable)
            # dentro del tiempo implícito de espera de Playwright para la operación `fill()`.
            error_msg = (
                f"\n❌ ERROR (Timeout): El tiempo de espera se agotó al interactuar con el selector '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/editable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_error_timeout_numerico", directorio)
            # Re-lanza la excepción como un Error de Playwright para mantener la coherencia.
            raise Error(error_msg) from e

        except Error as e:
            # Captura específica para errores de Playwright que no son timeouts (ej., selector malformado,
            # elemento desprendido del DOM, problemas con el contexto del navegador).
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al interactuar con el selector '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_numerico", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la operación.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al interactuar con el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_numerico", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esta espera es útil para
            # observar los cambios en la UI o para dar tiempo a la aplicación a procesar
            # la entrada antes de la siguiente acción de la prueba.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    # 11- Función para validar el título de una página con medición de rendimiento
    def validar_titulo_de_web(self, titulo_esperado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Valida el título de la página web actual. Esta función espera hasta que el título
        de la página coincida con el `titulo_esperado` dentro de un tiempo límite,
        e integra una **medición de rendimiento** para registrar cuánto tiempo tarda esta validación.

        Args:
            titulo_esperado (str): El **título exacto** que se espera que tenga la página web.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla** tomadas
                               durante la ejecución, facilitando su identificación.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        título de la página coincida. Si el título no coincide
                                        dentro de este plazo, la validación fallará.
                                        Por defecto, `5.0` segundos.

        Raises:
            TimeoutError: Si el título de la página no coincide con el `titulo_esperado`
                          dentro del `tiempo` límite.
            AssertionError: Si la aserción de título falla (aunque `TimeoutError` es más común
                            para esta aserción cuando se usa un timeout).
            Exception: Para cualquier otro error inesperado que ocurra durante la validación.
        """
        self.logger.info(f"\nValidando que el título de la página sea: '{titulo_esperado}'. Tiempo máximo de espera: {tiempo}s.")

        # --- Medición de rendimiento: Inicio de la espera por el título ---
        # Registra el tiempo justo antes de iniciar la espera activa de Playwright.
        start_time_title_check = time.time()

        try:
            # Playwright espera a que el título de la página coincida con el `titulo_esperado`.
            # El `timeout` se especifica en milisegundos.
            expect(self.page).to_have_title(titulo_esperado)
            
            # --- Medición de rendimiento: Fin de la espera por el título ---
            # Registra el tiempo una vez que el título ha sido validado con éxito.
            end_time_title_check = time.time()
            # Calcula la duración total que tardó la validación del título.
            # Esta métrica es importante para evaluar la **velocidad de carga y actualización**
            # del título de la página, un indicador clave del rendimiento de navegación.
            duration_title_check = end_time_title_check - start_time_title_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en validar el título de la página a '{titulo_esperado}': {duration_title_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Título de la página '{self.page.title()}' validado exitosamente.")
            # Toma una captura de pantalla al validar el título con éxito.
            self.tomar_captura(f"{nombre_base}_exito_titulo", directorio)

        except TimeoutError as e:
            # Captura específica para cuando el título no coincide dentro del tiempo especificado.
            # Se registra el tiempo transcurrido hasta el fallo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_title_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El título de la página no coincidió con '{titulo_esperado}' "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). Título actual: '{self.page.title()}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_titulo_timeout", directorio)
            raise # Re-lanza la excepción para que la prueba falle.

        except AssertionError as e:
            # Captura si la aserción de título falla por alguna otra razón (menos común con `to_have_title`
            # y timeout, ya que `TimeoutError` suele ser lo primero).
            error_msg = (
                f"\n❌ FALLO (Aserción): El título de la página NO coincide con '{titulo_esperado}'. "
                f"Título actual: '{self.page.title()}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla en el momento del fallo de aserción.
            self.tomar_captura(f"{nombre_base}_fallo_titulo_no_coincide", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la validación del título.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al validar el título de la página. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura para errores inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_titulo", directorio)
            raise # Re-lanza la excepción.
        
    # 12- Función para validar URL actual con medición de rendimiento
    def validar_url_actual(self, patron_url: str, tiempo: Union[int, float] = 0.5):
        """
        Valida la URL actual de la página usando un patrón de expresión regular.
        Esta función espera hasta que la URL de la página coincida con el `patron_url`
        dentro de un tiempo límite, e integra una **medición de rendimiento** para registrar
        cuánto tiempo tarda esta validación.

        Args:
            patron_url (str): El **patrón de expresión regular** (regex) que se espera
                              que coincida con la URL actual de la página. Por ejemplo,
                              `r".*\\/dashboard.*"` para URLs que contengan "/dashboard".
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que la
                                        URL de la página coincida con el patrón. Si la URL
                                        no coincide dentro de este plazo, la validación fallará.
                                        Por defecto, `5.0` segundos.

        Raises:
            TimeoutError: Si la URL actual de la página no coincide con el `patron_url`
                          dentro del `tiempo` límite especificado.
            AssertionError: Si la aserción de URL falla por alguna otra razón
                            (menos común con `to_have_url` y `timeout`, ya que `TimeoutError`
                            suele ser la excepción principal).
            Exception: Para cualquier otro error inesperado que ocurra durante la validación de la URL.
        """
        self.logger.info(f"\nValidando que la URL actual coincida con el patrón: '{patron_url}'. Tiempo máximo de espera: {tiempo}s.")

        # --- Medición de rendimiento: Inicio de la espera por la URL ---
        # Registra el tiempo justo antes de iniciar la espera activa de Playwright para la URL.
        start_time_url_check = time.time()

        try:
            # Playwright espera a que la URL de la página coincida con el patrón de expresión regular.
            # El `timeout` se especifica en milisegundos.
            # `re.compile(patron_url)` convierte la cadena del patrón en un objeto de expresión regular
            # que `to_have_url` puede utilizar.
            expect(self.page).to_have_url(re.compile(patron_url))
            
            # --- Medición de rendimiento: Fin de la espera por la URL ---
            # Registra el tiempo una vez que la URL ha sido validada con éxito.
            end_time_url_check = time.time()
            # Calcula la duración total que tardó la validación de la URL.
            # Esta métrica es crucial para evaluar la **velocidad de navegación y carga de la página**,
            # ya que la URL a menudo cambia una vez que la página está completamente cargada o enrutada.
            duration_url_check = end_time_url_check - start_time_url_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en validar la URL a '{patron_url}': {duration_url_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: URL '{self.page.url}' validada exitosamente con el patrón: '{patron_url}'.")
            # Nota sobre capturas de pantalla para URL:
            # Generalmente, una captura de pantalla no es visualmente útil para validar una URL,
            # ya que la URL se encuentra en la barra de direcciones del navegador.
            # Sin embargo, si deseas tener un registro visual del estado de la página
            # en el momento de la validación exitosa, podrías descomentar la siguiente línea
            # y asegurarte de pasar `nombre_base` y `directorio` como argumentos a esta función.
            # self.tomar_captura(f"{nombre_base}_exito_url", directorio)

        except TimeoutError as e:
            # Captura específica para cuando la URL no coincide con el patrón dentro del tiempo especificado.
            # Se registra el tiempo transcurrido hasta el fallo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_url_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): La URL actual '{self.page.url}' no coincidió con el patrón "
                f"'{patron_url}' después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Podrías añadir una captura de pantalla aquí en caso de fallo, si es necesario para depuración.
            # self.tomar_captura(f"{nombre_base}_fallo_url_timeout", directorio)
            raise # Re-lanza la excepción para asegurar que la prueba falle.

        except AssertionError as e:
            # Captura si la aserción de URL falla por alguna otra razón que no sea un timeout directo
            # (aunque con `to_have_url` y `timeout`, `TimeoutError` es más común).
            error_msg = (
                f"\n❌ FALLO (Aserción): La URL actual '{self.page.url}' NO coincide con el patrón: "
                f"'{patron_url}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Podrías añadir una captura de pantalla aquí en caso de fallo de aserción.
            # self.tomar_captura(f"{nombre_base}_fallo_url_no_coincide", directorio)
            raise # Re-lanza la excepción.
        
        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la validación de la URL.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al validar la URL. "
                f"URL actual: '{self.page.url}', Patrón esperado: '{patron_url}'. Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Podrías añadir una captura de pantalla aquí para errores inesperados.
            # self.tomar_captura(f"{nombre_base}_error_inesperado_url", directorio)
            raise # Re-lanza la excepción.
        
    # 13- Función para hacer click en un elemento, con capturas y medición de rendimiento
    def hacer_click_en_elemento(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, texto_esperado: str = None, tiempo: Union[int, float] = 0.5):
        """
        Realiza un click en un elemento de la página web. La función incluye
        validaciones opcionales del texto del elemento, toma capturas de pantalla
        antes y después del clic, e integra una **medición de rendimiento** para registrar
        el tiempo que tarda la operación de clic.

        Args:
            selector (Union[str, Page.locator]): El **selector del elemento** sobre el que
                                                  se desea hacer clic. Puede ser una cadena
                                                  (CSS, XPath, etc.) o un objeto `Locator`
                                                  de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            texto_esperado (str, optional): Texto que se espera que el elemento contenga
                                            **antes de hacer clic**. Si se proporciona,
                                            se realizará una aserción `to_have_text` antes del clic.
                                            Por defecto, `None` (no se verifica el texto).
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté clicable y para la aserción de texto (si aplica).
                                        También es el tiempo de espera fijo después del clic.
                                        Por defecto, `5.0` segundos.

        Raises:
            TimeoutError: Si el elemento no está visible, habilitado o clicable a tiempo,
                          o si no contiene el `texto_esperado` dentro del plazo.
            Error: Si ocurre un error específico de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando hacer click en el elemento con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración y visualización.
            locator.highlight()
            # Toma una captura de pantalla del estado de la página *antes* de realizar el clic.
            self.tomar_captura(f"{nombre_base}_antes_click", directorio)

            # Si se proporciona 'texto_esperado', valida que el elemento contenga ese texto.
            # Esta aserción también espera a que el texto esté presente.
            if texto_esperado:
                # Registra el tiempo antes de la aserción de texto.
                start_time_text_check = time.time()
                expect(locator).to_have_text(texto_esperado)
                # Registra el tiempo después de la aserción de texto y calcula la duración.
                end_time_text_check = time.time()
                duration_text_check = end_time_text_check - start_time_text_check
                self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en contener el texto '{texto_esperado}': {duration_text_check:.4f} segundos.")
                self.logger.info(f"\n✅ El elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            # --- Medición de rendimiento: Inicio de la operación de clic ---
            # Registra el tiempo justo antes de ejecutar la acción de 'click'.
            start_time_click = time.time()

            # Realiza el clic en el elemento. El método `click()` de Playwright
            # esperará automáticamente a que el elemento sea visible, habilitado y clicable.
            # El `timeout` aquí es para esta operación específica.
            locator.click()

            # --- Medición de rendimiento: Fin de la operación de clic ---
            # Registra el tiempo inmediatamente después de que la operación de clic se ha completado.
            end_time_click = time.time()
            # Calcula la duración total de la operación de clic.
            # Esta métrica es crucial para evaluar la **reactividad de los botones/enlaces**
            # y el rendimiento percibido por el usuario al interactuar.
            duration_click = end_time_click - start_time_click
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el clic en el elemento '{selector}': {duration_click:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Click realizado exitosamente en el elemento con selector '{selector}'.")
            # Toma una captura de pantalla del estado de la página *después* de realizar el clic.
            self.tomar_captura(f"{nombre_base}_despues_click", directorio)

        except TimeoutError as e:
            # Captura específica para errores de tiempo de espera de Playwright.
            # Esto indica que el elemento no estaba listo (visible, habilitado, clicable)
            # o que el texto esperado no apareció a tiempo.
            error_msg = (
                f"\n❌ ERROR (Timeout): El tiempo de espera se agotó al intentar hacer click en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/clicable a tiempo, "
                f"o no contenía el texto esperado (si se especificó).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_error_timeout_click", directorio)
            # Re-lanza la excepción como un Error de Playwright para mantener la coherencia.
            raise Error(error_msg) from e

        except Error as e:
            # Captura errores específicos de Playwright que no son timeouts (ej., selector malformado,
            # elemento desprendido del DOM, problemas con el contexto del navegador).
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer click en el selector '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_click", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la operación de clic.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer click en el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_click", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esta espera es útil para
            # observar los cambios en la UI que ocurran después del clic (ej., una navegación,
            # un modal apareciendo) o para dar tiempo a la aplicación a procesar la acción.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 14- Función para hacer doble click en un elemento, con capturas y medición de rendimiento
    def hacer_doble_click_en_elemento(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, texto_esperado: str = None, tiempo: Union[int, float] = 0.5):
        """
        Realiza un **doble click** en un elemento de la página web. La función incluye
        validaciones opcionales del texto del elemento, toma capturas de pantalla
        antes y después del doble clic, e integra una **medición de rendimiento** para
        registrar el tiempo que tarda la operación de doble clic.

        Args:
            selector (Union[str, Page.locator]): El **selector del elemento** sobre el que
                                                  se desea hacer doble clic. Puede ser una cadena
                                                  (CSS, XPath, etc.) o un objeto `Locator`
                                                  de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            texto_esperado (str, optional): Texto que se espera que el elemento contenga
                                            **antes de hacer doble clic**. Si se proporciona,
                                            se realizará una aserción `to_have_text` antes del doble clic.
                                            Por defecto, `None` (no se verifica el texto).
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté clicable y para la aserción de texto (si aplica).
                                        También es el tiempo de espera fijo después del doble clic.
                                        Por defecto, `5.0` segundos. (Se cambió de 1 a 5 para consistencia)

        Raises:
            TimeoutError: Si el elemento no está visible, habilitado o doble-clicable a tiempo,
                          o si no contiene el `texto_esperado` dentro del plazo.
            Error: Si ocurre un error específico de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando hacer doble click en el elemento con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración y visualización.
            locator.highlight()
            # Toma una captura de pantalla del estado de la página *antes* de realizar el doble clic.
            self.tomar_captura(f"{nombre_base}_antes_doble_click", directorio)

            # Si se proporciona 'texto_esperado', valida que el elemento contenga ese texto.
            # Esta aserción también espera a que el texto esté presente.
            if texto_esperado:
                # Registra el tiempo antes de la aserción de texto.
                start_time_text_check = time.time()
                expect(locator).to_have_text(texto_esperado)
                # Registra el tiempo después de la aserción de texto y calcula la duración.
                end_time_text_check = time.time()
                duration_text_check = end_time_text_check - start_time_text_check
                self.logger.info(f"PERFORMANCE: Tiempo que tardó el elemento '{selector}' en contener el texto '{texto_esperado}' antes del doble clic: {duration_text_check:.4f} segundos.")
                self.logger.info(f"\n✅ El elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            # --- Medición de rendimiento: Inicio de la operación de doble clic ---
            # Registra el tiempo justo antes de ejecutar la acción de 'dblclick'.
            start_time_dblclick = time.time()

            # Realiza el doble clic en el elemento. El método `dblclick()` de Playwright
            # esperará automáticamente a que el elemento sea visible, habilitado y doble-clicable.
            # El `timeout` aquí es para esta operación específica.
            locator.dblclick()

            # --- Medición de rendimiento: Fin de la operación de doble clic ---
            # Registra el tiempo inmediatamente después de que la operación de doble clic se ha completado.
            end_time_dblclick = time.time()
            # Calcula la duración total de la operación de doble clic.
            # Esta métrica es crucial para evaluar la **reactividad de la UI**
            # ante interacciones más complejas como el doble clic.
            duration_dblclick = end_time_dblclick - start_time_dblclick
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el doble clic en el elemento '{selector}': {duration_dblclick:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Doble click realizado exitosamente en el elemento con selector '{selector}'.")
            # Toma una captura de pantalla del estado de la página *después* de realizar el doble clic.
            self.tomar_captura(f"{nombre_base}_despues_doble_click", directorio)

        except TimeoutError as e:
            # Captura específica para errores de tiempo de espera de Playwright.
            # Esto indica que el elemento no estaba listo (visible, habilitado, doble-clicable)
            # o que el texto esperado no apareció a tiempo.
            error_msg = (
                f"\n❌ ERROR (Timeout): El tiempo de espera se agotó al intentar hacer doble click en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/doble-clicable a tiempo, "
                f"o no contenía el texto esperado (si se especificó).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_error_timeout_doble_click", directorio)
            # Re-lanza la excepción como un Error de Playwright para mantener la coherencia.
            raise Error(error_msg) from e

        except Error as e:
            # Captura errores específicos de Playwright que no son timeouts (ej., selector malformado,
            # elemento desprendido del DOM, problemas con el contexto del navegador).
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer doble click en el selector '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_doble_click", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la operación de doble clic.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer doble click en el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_doble_click", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esta espera es útil para
            # observar los cambios en la UI que ocurran después del doble clic.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    # 15- Función para hacer hover sobre un elemento, con capturas y medición de rendimiento
    def hacer_hover_en_elemento(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de **hover (pasar el ratón por encima)** sobre un elemento
        de la página web. La función toma capturas de pantalla antes y después del hover,
        e integra una **medición de rendimiento** para registrar el tiempo que tarda
        la operación de hover.

        Args:
            selector (Union[str, Page.locator]): El **selector del elemento** sobre el que
                                                  se desea realizar el hover. Puede ser una cadena
                                                  (CSS, XPath, etc.) o un objeto `Locator`
                                                  de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté visible y sea interactuable antes de realizar
                                        el hover. También es el tiempo de espera fijo después del hover.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para consistencia).

        Raises:
            TimeoutError: Si el elemento no está visible o interactuable a tiempo para el hover.
            Error: Si ocurre un error específico de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando hacer hover sobre el elemento con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración y visualización.
            locator.highlight()
            # Toma una captura de pantalla del estado de la página *antes* de realizar el hover.
            self.tomar_captura(f"{nombre_base}_antes_hover", directorio)

            # --- Medición de rendimiento: Inicio de la operación de hover ---
            # Registra el tiempo justo antes de ejecutar la acción de 'hover'.
            start_time_hover = time.time()

            # Realiza el hover sobre el elemento. El método `hover()` de Playwright
            # esperará automáticamente a que el elemento sea visible y esté listo para la interacción.
            # El `timeout` aquí es para esta operación específica.
            locator.hover()

            # --- Medición de rendimiento: Fin de la operación de hover ---
            # Registra el tiempo inmediatamente después de que la operación de hover se ha completado.
            end_time_hover = time.time()
            # Calcula la duración total de la operación de hover.
            # Esta métrica es importante para evaluar la **reactividad de la UI**
            # ante interacciones que revelan tooltips, menús desplegables, etc.
            duration_hover = end_time_hover - start_time_hover
            self.logger.info(f"PERFORMANCE: Tiempo que tardó el hover en el elemento '{selector}': {duration_hover:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Hover realizado exitosamente en el elemento con selector '{selector}'.")
            # Toma una captura de pantalla del estado de la página *después* de realizar el hover.
            # Esta captura es especialmente útil si el hover revela nuevos elementos (ej., un menú).
            self.tomar_captura(f"{nombre_base}_despues_hover", directorio)

        except TimeoutError as e:
            # Captura específica para errores de tiempo de espera de Playwright.
            # Esto indica que el elemento no estaba visible o interactuable a tiempo para el hover.
            error_msg = (
                f"\n❌ ERROR (Timeout): El tiempo de espera se agotó al intentar hacer hover en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció o no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_error_timeout_hover", directorio)
            # Re-lanza la excepción como un Error de Playwright para mantener la coherencia.
            raise Error(error_msg) from e

        except Error as e:
            # Captura errores específicos de Playwright que no son timeouts (ej., selector malformado,
            # elemento desprendido del DOM, problemas con el contexto del navegador).
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer hover en el selector '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_hover", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir durante la operación de hover.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer hover en el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_hover", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esto es útil para
            # observar los cambios en la UI que puedan activarse por el hover (ej., tooltips, menús).
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 16- Función para verificar si un elemento está habilitado (enabled) con medición de rendimiento
    def verificar_elemento_habilitado(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica si un elemento está **habilitado (enabled)** en la página web.
        Esta función espera hasta que el elemento esté habilitado dentro de un tiempo límite,
        y registra el tiempo que tarda esta verificación como una métrica de rendimiento.
        Toma capturas de pantalla tanto en caso de éxito como de fallo.

        Args:
            selector (Union[str, Page.locator]): El **selector del elemento** a verificar.
                                                  Puede ser una cadena (CSS, XPath, etc.)
                                                  o un objeto `Locator` de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento pase a estar habilitado. Si no lo está dentro
                                        de este plazo, la función devolverá `False`.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Returns:
            bool: `True` si el elemento está habilitado dentro del tiempo especificado;
                  `False` en caso contrario (timeout o aserción fallida).

        Raises:
            Error: Si ocurre un problema específico de Playwright que impida la verificación
                   (ej., selector inválido, problema con el navegador).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nVerificando si el elemento con selector '{selector}' está habilitado. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación de habilitación ---
        # Registra el tiempo justo antes de iniciar la aserción de habilitación.
        start_time_enabled_check = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()

            # Playwright espera a que el elemento esté habilitado.
            # El `timeout` se especifica en milisegundos.
            expect(locator).to_be_enabled()
            
            # --- Medición de rendimiento: Fin de la verificación ---
            # Registra el tiempo una vez que la aserción de habilitación ha sido exitosa.
            end_time_enabled_check = time.time()
            # Calcula la duración total de la verificación de habilitación.
            # Esta métrica es importante para evaluar la **velocidad con la que los elementos
            # interactivos de la UI se vuelven funcionales**. Un tiempo de habilitación
            # prolongado podría indicar problemas de carga de JavaScript o de renderizado.
            duration_enabled_check = end_time_enabled_check - start_time_enabled_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en verificar que el elemento '{selector}' está habilitado: {duration_enabled_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: El elemento '{selector}' está habilitado.")
            # Toma una captura de pantalla al verificar que el elemento está habilitado con éxito.
            self.tomar_captura(f"{nombre_base}_habilitado", directorio)
            return True

        except TimeoutError as e:
            # Captura específica para cuando el elemento no está habilitado dentro del tiempo especificado.
            # Se registra el tiempo transcurrido hasta el fallo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_enabled_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento con selector '{selector}' NO está habilitado "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). "
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' ya que la función devuelve False en lugar de fallar la prueba.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_NO_habilitado_timeout", directorio)
            return False

        except AssertionError as e:
            # Captura si la aserción de habilitación falla por alguna otra razón.
            # Con `to_be_enabled` y un timeout, `TimeoutError` es más común, pero `AssertionError`
            # podría ocurrir si el elemento existe pero la aserción de Playwright lo considera inhabilitado
            # sin agotar el timeout.
            error_msg = (
                f"\n❌ FALLO (Aserción): El elemento con selector '{selector}' NO está habilitado. "
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' aquí también.
            # Toma una captura de pantalla en el momento del fallo de aserción.
            self.tomar_captura(f"{nombre_base}_NO_habilitado_fallo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright que no son timeouts ni AssertionErrors (ej., selector malformado).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar habilitación de '{selector}'. "
                f"Esto indica un problema fundamental con el selector o el navegador. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_habilitado", directorio)
            raise # Re-lanza la excepción porque esto es un fallo de ejecución, no una verificación de estado.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar la habilitación de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_habilitado", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre.
            # Aplica una espera fija después de la operación. Puede ser útil para observar
            # el estado del elemento o esperar efectos secundarios en la UI.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 17- Función para mover el mouse a coordenadas X, Y y hacer clic, con medición de rendimiento
    def mouse_mueve_y_hace_clic_xy(self, x: int, y: int, nombre_base: str, directorio: str, tiempo: Union[int, float] = 1.0):
        """
        Mueve el cursor del mouse a las coordenadas de pantalla (X, Y) especificadas y luego
        realiza un clic en esas mismas coordenadas. Esta función es útil para interacciones
        precisas fuera del contexto de un elemento específico del DOM.
        Integra una **medición de rendimiento** para registrar el tiempo que tarda la secuencia
        completa (movimiento y clic).

        Args:
            x (int): La **coordenada X** (horizontal) en píxeles de la pantalla,
                     donde se moverá el mouse y se hará clic.
            y (int): La **coordenada Y** (vertical) en píxeles de la pantalla,
                     donde se moverá el mouse y se hará clic.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo de espera fijo** (en segundos) que se aplicará
                                        después de que el clic se haya completado. Útil para
                                        observar los efectos del clic o dar tiempo a la UI.
                                        Por defecto, `1.0` segundos (se ajustó de 0.5 para una espera más razonable).

        Raises:
            ValueError: Si las coordenadas X o Y no son números enteros.
            Exception: Para cualquier error inesperado que ocurra durante la operación del mouse.
        """
        self.logger.info(f"\nIntentando mover el mouse a X:{x}, Y:{y} y haciendo click.")

        # --- Validaciones de entrada ---
        # Asegura que las coordenadas sean de tipo entero para evitar errores inesperados con mouse.move/click.
        if not isinstance(x, int) or not isinstance(y, int):
            error_msg = f"\n❌ ERROR: Las coordenadas X ({x}) e Y ({y}) deben ser números enteros."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_coordenadas_invalidas", directorio)
            raise ValueError(error_msg)

        try:
            # Toma una captura de pantalla del estado de la página *antes* de mover y hacer clic.
            self.tomar_captura(f"{nombre_base}_antes_mouse_click_xy", directorio)
            
            # --- Medición de rendimiento: Inicio de la operación del mouse ---
            # Registra el tiempo justo antes de iniciar el movimiento y clic del mouse.
            start_time_mouse_action = time.time()

            # Mueve el cursor del mouse a las coordenadas especificadas.
            # `steps=5` hace que el movimiento sea más suave, simulando un usuario real.
            self.page.mouse.move(x, y, steps=5) 
            self.logger.debug(f"\nMouse movido a X:{x}, Y:{y}.")
            
            # Realiza un clic en las mismas coordenadas.
            self.page.mouse.click(x, y)

            # --- Medición de rendimiento: Fin de la operación del mouse ---
            # Registra el tiempo inmediatamente después de que el clic se ha completado.
            end_time_mouse_action = time.time()
            # Calcula la duración total de la secuencia de movimiento y clic.
            # Esta métrica es relevante para acciones de UI que dependen de interacciones
            # de ratón muy precisas y para evaluar la latencia percibida en estas acciones.
            duration_mouse_action = end_time_mouse_action - start_time_mouse_action
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en mover y hacer clic en X:{x}, Y:{y}: {duration_mouse_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Click realizado en X:{x}, Y:{y}.")
            # Toma una captura de pantalla del estado de la página *después* de la acción del mouse.
            self.tomar_captura(f"{nombre_base}_despues_mouse_click_xy", directorio)

        except Error as e:
            # Captura errores específicos de Playwright relacionados con la interacción del mouse.
            error_msg = (
                f"\n❌ ERROR (Playwright): Ocurrió un problema de Playwright al mover el mouse y hacer clic en X:{x}, Y:{y}.\n"
                f"Esto puede deberse a problemas con la ventana del navegador o el contexto de ejecución.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla en el momento del fallo.
            self.tomar_captura(f"{nombre_base}_error_playwright_mouse_click_xy", directorio)
            raise # Re-lanza la excepción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ ERROR (Inesperado): Se produjo un error desconocido al mover el mouse y hacer clic en X:{x}, Y:{y}.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_mouse_click_xy", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esto es útil para observar
            # los cambios visuales que el clic en las coordenadas pueda haber provocado.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 18- Función para marcar un checkbox, con verificación y medición de rendimiento
    def marcar_checkbox(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Marca un checkbox especificado por su selector y verifica que se haya marcado
        correctamente. Esta función toma capturas de pantalla antes y después de la
        acción, e integra una **medición de rendimiento** para registrar el tiempo
        que tarda la operación completa (marcar y verificar).

        Args:
            selector (Union[str, Page.locator]): El **selector del checkbox** que se desea marcar.
                                                  Puede ser una cadena (CSS, XPath, etc.) o un objeto
                                                  `Locator` de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        checkbox sea marcado y para que su estado sea verificado.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Raises:
            AssertionError: Si el checkbox no puede ser marcado o no se verifica como marcado
                            dentro del tiempo límite, o si ocurre un error de Playwright.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando marcar el checkbox con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de marcado y verificación ---
        # Registra el tiempo justo antes de iniciar la operación.
        start_time_checkbox_action = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado de la página *antes* de marcar el checkbox.
            self.tomar_captura(f"{nombre_base}_antes_marcar_checkbox", directorio)
            
            # Marca el checkbox. Playwright esperará automáticamente a que sea interactuable.
            locator.check()
            # Verifica que el checkbox esté marcado. Esta aserción también espera.
            expect(locator).to_be_checked() 
            
            # --- Medición de rendimiento: Fin de la operación ---
            # Registra el tiempo una vez que el checkbox ha sido marcado y verificado con éxito.
            end_time_checkbox_action = time.time()
            # Calcula la duración total de la operación.
            # Esta métrica es importante para evaluar la **capacidad de respuesta de los elementos
            # de formulario** y la velocidad de actualización de su estado en la UI.
            duration_checkbox_action = end_time_checkbox_action - start_time_checkbox_action
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en marcar y verificar el checkbox '{selector}': {duration_checkbox_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Checkbox con selector '{selector}' marcado y verificado exitosamente.")
            # Toma una captura de pantalla del estado de la página *después* de marcar el checkbox.
            self.tomar_captura(f"{nombre_base}_despues_marcar_checkbox", directorio)

        except TimeoutError as e:
            # Captura específica para cuando la operación de marcar o la verificación fallan por tiempo.
            # Registra el tiempo transcurrido hasta el fallo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_checkbox_action # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El checkbox con selector '{selector}' no pudo ser marcado "
                f"o verificado como marcado dentro de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_marcar", directorio)
            # Re-lanza la excepción como un AssertionError para que la prueba falle claramente.
            raise AssertionError(f"\nCheckbox no marcado/verificado (Timeout): {selector}") from e

        except Error as e: # Captura errores específicos de Playwright (ej., selector inválido)
            error_msg = (
                f"\n❌ FALLO (Playwright Error): Problema al interactuar con el checkbox '{selector}'.\n"
                f"Posibles causas: Selector inválido, elemento no interactuable, DOM no estable.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_marcar", directorio)
            raise AssertionError(f"\nError de Playwright con checkbox: {selector}") from e # Re-lanza.

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar marcar el checkbox '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_marcar", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esto puede ser útil para
            # observar cualquier cambio adicional en la UI provocado por el cambio de estado del checkbox.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 19- Función para desmarcar un checkbox, con verificación y medición de rendimiento
    def desmarcar_checkbox(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Desmarca un checkbox especificado por su selector y verifica que se haya desmarcado
        correctamente. Esta función toma capturas de pantalla antes y después de la acción,
        e integra una **medición de rendimiento** para registrar el tiempo que tarda la
        operación completa (desmarcar y verificar).

        Args:
            selector (Union[str, Page.locator]): El **selector del checkbox** que se desea desmarcar.
                                                  Puede ser una cadena (CSS, XPath, etc.) o un objeto
                                                  `Locator` de Playwright preexistente.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        checkbox sea desmarcado y para que su estado sea verificado.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Raises:
            AssertionError: Si el checkbox no puede ser desmarcado o no se verifica como desmarcado
                            dentro del tiempo límite, o si ocurre un error de Playwright.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando desmarcar el checkbox con selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de desmarcado y verificación ---
        # Registra el tiempo justo antes de iniciar la operación.
        start_time_checkbox_action = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado de la página *antes* de desmarcar el checkbox.
            self.tomar_captura(f"{nombre_base}_antes_desmarcar_checkbox", directorio)
            
            # Desmarca el checkbox. Playwright esperará automáticamente a que sea interactuable.
            locator.uncheck()
            # Verifica que el checkbox no esté marcado. Esta aserción también espera.
            expect(locator).not_to_be_checked() 
            
            # --- Medición de rendimiento: Fin de la operación ---
            # Registra el tiempo una vez que el checkbox ha sido desmarcado y verificado con éxito.
            end_time_checkbox_action = time.time()
            # Calcula la duración total de la operación.
            # Esta métrica es importante para evaluar la **capacidad de respuesta de los elementos
            # de formulario** y la velocidad de actualización de su estado en la UI.
            duration_checkbox_action = end_time_checkbox_action - start_time_checkbox_action
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en desmarcar y verificar el checkbox '{selector}': {duration_checkbox_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Checkbox con selector '{selector}' desmarcado y verificado exitosamente.")
            # Toma una captura de pantalla del estado de la página *después* de desmarcar el checkbox.
            self.tomar_captura(f"{nombre_base}_despues_desmarcar_checkbox", directorio)

        except TimeoutError as e:
            # Captura específica para cuando la operación de desmarcar o la verificación fallan por tiempo.
            # Registra el tiempo transcurrido hasta el fallo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_checkbox_action # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El checkbox con selector '{selector}' no pudo ser desmarcado "
                f"o verificado como desmarcado dentro de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_desmarcar", directorio)
            # Re-lanza la excepción como un AssertionError para que la prueba falle claramente.
            raise AssertionError(f"\nCheckbox no desmarcado/verificado (Timeout): {selector}") from e

        except Error as e: # Captura errores específicos de Playwright (ej., selector inválido)
            error_msg = (
                f"\n❌ FALLO (Playwright Error): Problema al interactuar con el checkbox '{selector}'.\n"
                f"Posibles causas: Selector inválido, elemento no interactuable, DOM no estable.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_desmarcar", directorio)
            raise AssertionError(f"\nError de Playwright con checkbox: {selector}") from e # Re-lanza.

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar desmarcar el checkbox '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_desmarcar", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija después de la operación. Esto puede ser útil para
            # observar cualquier cambio adicional en la UI provocado por el cambio de estado del checkbox.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    # 20- Función para verificar el valor de un campo de texto con medición de rendimiento
    def verificar_valor_campo(self, selector: Union[str, Page.locator], valor_esperado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica que el **valor de un campo de texto** coincida con el `valor_esperado`.
        Esta función espera hasta que el campo de texto contenga el valor deseado dentro
        de un tiempo límite, y registra el tiempo que tarda esta verificación como una
        métrica de rendimiento. Toma capturas de pantalla tanto en caso de éxito como de fallo.

        Args:
            selector (Union[str, Page.locator]): El **selector del campo de texto** a verificar.
                                                  Puede ser una cadena (CSS, XPath, etc.)
                                                  o un objeto `Locator` de Playwright preexistente.
            valor_esperado (str): El **valor de texto exacto** que se espera encontrar en el campo.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        campo contenga el `valor_esperado`. Si no lo contiene
                                        dentro de este plazo, la función devolverá `False`.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Returns:
            bool: `True` si el valor del campo coincide con `valor_esperado` dentro del tiempo especificado;
                  `False` en caso contrario (timeout o aserción fallida).

        Raises:
            Error: Si ocurre un problema específico de Playwright que impida la verificación
                   (ej., selector inválido, elemento no es un campo de texto).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nVerificando que el campo '{selector}' contiene el valor esperado: '{valor_esperado}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación del valor del campo ---
        # Registra el tiempo justo antes de iniciar la aserción del valor.
        start_time_value_check = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado del campo *antes* de la verificación.
            # Esto puede ser útil para ver el valor inicial si es diferente al esperado.
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_campo", directorio)

            # Playwright espera a que el campo contenga el valor especificado.
            # El `timeout` se especifica en milisegundos.
            expect(locator).to_have_value(valor_esperado)
            
            # --- Medición de rendimiento: Fin de la verificación ---
            # Registra el tiempo una vez que la aserción del valor ha sido exitosa.
            end_time_value_check = time.time()
            # Calcula la duración total de la verificación del valor.
            # Esta métrica es importante para evaluar la **velocidad con la que los campos
            # de texto se pueblan o actualizan** en la UI, lo cual puede depender de la carga
            # de datos o de la lógica de la aplicación.
            duration_value_check = end_time_value_check - start_time_value_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en verificar que el campo '{selector}' contiene el valor '{valor_esperado}': {duration_value_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: El campo '{selector}' contiene el valor esperado: '{valor_esperado}'.")
            # Toma una captura de pantalla al verificar que el campo tiene el valor esperado.
            self.tomar_captura(f"{nombre_base}_despues_verificar_valor_campo", directorio)
            return True

        except TimeoutError as e:
            # Captura específica para cuando el campo no contiene el valor esperado dentro del tiempo.
            # Se intenta obtener el valor actual del campo para incluirlo en el mensaje de error.
            actual_value = "No se pudo obtener el valor"
            try:
                actual_value = locator.input_value() # Intenta obtener el valor actual
            except Exception:
                pass # Ignora si no se puede obtener el valor (ej., elemento no existe o no es input)

            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_value_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El campo '{selector}' no contiene el valor esperado '{valor_esperado}' "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). "
                f"Valor actual: '{actual_value}'. Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' ya que la función devuelve False.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_campo", directorio)
            return False

        except AssertionError as e:
            # Captura si la aserción de valor falla por alguna otra razón, aunque TimeoutError es más común.
            actual_value = "No se pudo obtener el valor"
            try:
                actual_value = locator.input_value()
            except Exception:
                pass

            error_msg = (
                f"\n❌ FALLO (Aserción): El campo '{selector}' NO contiene el valor esperado '{valor_esperado}'. "
                f"Valor actual: '{actual_value}'. Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' aquí también.
            # Toma una captura de pantalla en el momento del fallo de aserción.
            self.tomar_captura(f"{nombre_base}_fallo_verificar_valor_campo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, elemento no es un campo de entrada).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar el valor del campo '{selector}'. "
                f"Esto indica un problema fundamental con el selector o el tipo de elemento. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_campo", directorio)
            raise # Re-lanza la excepción porque esto es un fallo de ejecución, no una verificación de estado.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar el valor del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_campo", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre.
            # Aplica una espera fija después de la operación. Puede ser útil para observar
            # el estado del elemento o esperar efectos secundarios en la UI.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 21- Función para verificar el valor de un campo numérico (entero) con medición de rendimiento
    def verificar_valor_campo_numerico_int(self, selector: Union[str, Page.locator], valor_numerico_esperado: int, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica que el **valor de un campo de texto**, interpretado como un **número entero**,
        coincida con el `valor_numerico_esperado`. Esta función espera hasta que el campo
        contenga el valor deseado (como cadena) dentro de un tiempo límite, y registra el
        tiempo que tarda esta verificación como una métrica de rendimiento.
        Toma capturas de pantalla tanto en caso de éxito como de fallo.

        Args:
            selector (Union[str, Page.locator]): El **selector del campo de texto** a verificar.
                                                  Puede ser una cadena (CSS, XPath, etc.)
                                                  o un objeto `Locator` de Playwright preexistente.
            valor_numerico_esperado (int): El **valor numérico entero exacto** que se espera
                                           encontrar en el campo. Se convertirá a cadena para la
                                           comparación con el valor del campo HTML.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        campo contenga el `valor_numerico_esperado`. Si no lo contiene
                                        dentro de este plazo, la función devolverá `False`.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Returns:
            bool: `True` si el valor numérico (entero) del campo coincide con `valor_numerico_esperado`
                  dentro del tiempo especificado; `False` en caso contrario (timeout o aserción fallida).

        Raises:
            TypeError: Si `valor_numerico_esperado` no es un número entero.
            Error: Si ocurre un problema específico de Playwright que impida la verificación
                   (ej., selector inválido, elemento no es un campo de texto).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nVerificando que el campo '{selector}' contiene el valor numérico entero esperado: '{valor_numerico_esperado}'. Tiempo máximo de espera: {tiempo}s.")

        # --- Validación de entrada: Asegura que el valor esperado es un entero ---
        # Es crucial que el valor esperado sea un entero para la lógica de la función.
        if not isinstance(valor_numerico_esperado, int):
            error_msg = (
                f"\n❌ ERROR de tipo: 'valor_numerico_esperado' debe ser un número entero (int), "
                f"pero se recibió un tipo: {type(valor_numerico_esperado).__name__} con valor '{valor_numerico_esperado}'."
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_valor_int", directorio)
            raise TypeError(error_msg) # Se eleva un TypeError para un tipo de dato incorrecto.

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación del valor numérico ---
        # Registra el tiempo justo antes de iniciar la aserción del valor.
        start_time_numeric_check = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado del campo *antes* de la verificación.
            # Esto puede ser útil para ver el valor inicial si es diferente al esperado.
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_int", directorio)

            # Playwright espera a que el campo contenga el valor especificado (convertido a cadena).
            # El `timeout` se especifica en milisegundos.
            # Se usa `str(valor_numerico_esperado)` porque el valor en un campo de texto HTML
            # siempre se leerá como una cadena, incluso si representa un número.
            expect(locator).to_have_value(str(valor_numerico_esperado))
            
            # --- Medición de rendimiento: Fin de la verificación ---
            # Registra el tiempo una vez que la aserción del valor ha sido exitosa.
            end_time_numeric_check = time.time()
            # Calcula la duración total de la verificación del valor.
            # Esta métrica es importante para evaluar la **velocidad con la que los campos
            # numéricos se pueblan o actualizan** en la UI, lo cual puede depender de la carga
            # de datos, cálculos en el frontend o lógica de la aplicación que establece los valores.
            duration_numeric_check = end_time_numeric_check - start_time_numeric_check
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en verificar que el campo '{selector}' contiene el valor numérico '{valor_numerico_esperado}': {duration_numeric_check:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: El campo '{selector}' contiene el valor numérico entero esperado: '{valor_numerico_esperado}'.")
            # Toma una captura de pantalla al verificar que el campo tiene el valor esperado.
            self.tomar_captura(f"{nombre_base}_despues_verificar_valor_int", directorio)
            return True

        except TimeoutError as e:
            # Captura específica para cuando el campo no contiene el valor esperado dentro del tiempo.
            # Se intenta obtener el valor actual del campo para incluirlo en el mensaje de error.
            actual_value_str = "No se pudo obtener el valor o no es un campo de entrada"
            try:
                # Intenta obtener el valor actual como cadena
                actual_value_str = locator.input_value()
            except Exception:
                pass # Ignora si no se puede obtener el valor (ej., elemento no existe o no es input)

            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_numeric_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El campo '{selector}' no contiene el valor entero esperado '{valor_numerico_esperado}' "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s). "
                f"Valor actual en el campo: '{actual_value_str}'. Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' ya que la función devuelve False.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_int", directorio)
            return False

        except AssertionError as e:
            # Captura si la aserción de valor falla por alguna otra razón (menos común con to_have_value, pero posible).
            actual_value_str = "No se pudo obtener el valor o no es un campo de entrada"
            try:
                actual_value_str = locator.input_value()
                # Intenta convertir el valor actual a entero para una comparación más significativa en el log.
                actual_value_int = int(actual_value_str)
                comparison_msg = f"\n (Valor actual: {actual_value_int}, Esperado: {valor_numerico_esperado})"
            except ValueError: # Si el valor actual no se puede convertir a int.
                comparison_msg = f"\n (Valor actual no numérico: '{actual_value_str}', Esperado: {valor_numerico_esperado})"
            except Exception: # Si no se puede obtener el valor en absoluto.
                comparison_msg = f"\n (No se pudo obtener el valor actual, Esperado: {valor_numerico_esperado})"

            error_msg = (
                f"\n❌ FALLO (Aserción): El campo '{selector}' NO contiene el valor numérico entero esperado. "
                f"{comparison_msg}. Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usa 'warning' aquí también.
            # Toma una captura de pantalla en el momento del fallo de aserción.
            self.tomar_captura(f"{nombre_base}_fallo_verificar_valor_int", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, elemento no es un campo de entrada).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar el valor numérico entero del campo '{selector}'. "
                f"Esto indica un problema fundamental con el selector o el tipo de elemento.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_int", directorio)
            raise # Re-lanza la excepción porque esto es un fallo de ejecución, no una verificación de estado.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar el valor numérico entero del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_int", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre.
            # Aplica una espera fija después de la operación. Puede ser útil para observar
            # el estado del elemento o esperar efectos secundarios en la UI.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 22- Función para verificar el valor de un campo numérico (flotante) con medición de rendimiento
    def verificar_valor_campo_numerico_float(self, selector: Union[str, Page.locator], valor_numerico_esperado: float, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5, tolerancia: float = 1e-6) -> bool:
        """
        Verifica que el **valor de un campo de texto**, interpretado como un **número flotante**,
        coincida con el `valor_numerico_esperado` dentro de una `tolerancia` específica.
        Esta función espera hasta que el campo esté visible, obtiene su valor, y luego realiza
        la comparación. Registra el tiempo que tarda esta verificación como una métrica de rendimiento.
        Toma capturas de pantalla tanto en caso de éxito como de fallo.

        Args:
            selector (Union[str, Page.locator]): El **selector del campo de texto** a verificar.
                                                  Puede ser una cadena (CSS, XPath, etc.)
                                                  o un objeto `Locator` de Playwright preexistente.
            valor_numerico_esperado (float): El **valor numérico flotante exacto** que se espera
                                           encontrar en el campo.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        campo se haga visible y su valor sea obtenible.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).
            tolerancia (float): **Margen de error** aceptable para la comparación de números flotantes.
                                 Debido a la naturaleza de la representación de punto flotante,
                                 raramente se comparan flotantes para una igualdad exacta.
                                 Por defecto, `1e-6` (0.000001).

        Returns:
            bool: `True` si el valor numérico (flotante) del campo coincide con `valor_numerico_esperado`
                  dentro de la tolerancia y el tiempo especificado; `False` en caso contrario.

        Raises:
            TypeError: Si `valor_numerico_esperado` o `tolerancia` no son números flotantes.
            Error: Si ocurre un problema específico de Playwright que impida la verificación
                   (ej., selector inválido, elemento no es un campo de texto).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nVerificando que el campo '{selector}' contiene el valor numérico flotante esperado: '{valor_numerico_esperado}' con tolerancia {tolerancia}. Tiempo máximo de espera: {tiempo}s.")

        # --- Validación de entrada: Asegura que el valor esperado es un flotante y la tolerancia es un flotante ---
        if not isinstance(valor_numerico_esperado, float):
            error_msg = (
                f"\n❌ ERROR de tipo: 'valor_numerico_esperado' debe ser un número flotante (float), "
                f"pero se recibió un tipo: {type(valor_numerico_esperado).__name__} con valor '{valor_numerico_esperado}'."
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_valor_float", directorio)
            raise TypeError(error_msg) # Se eleva un TypeError para un tipo de dato incorrecto.
        
        if not isinstance(tolerancia, float) or tolerancia < 0:
            error_msg = (
                f"\n❌ ERROR de tipo: 'tolerancia' debe ser un número flotante (float) no negativo, "
                f"pero se recibió un tipo: {type(tolerancia).__name__} con valor '{tolerancia}'."
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_tolerancia_float", directorio)
            raise TypeError(error_msg)

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación del valor flotante ---
        # Registra el tiempo justo antes de iniciar la operación de verificación.
        start_time_float_check = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado del campo *antes* de la verificación.
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_float", directorio)

            # Primero, asegurar que el campo es visible y está presente en el DOM
            # Esto es necesario porque `input_value()` no tiene un mecanismo de espera.
            expect(locator).to_be_visible() 
            
            # Obtener el valor actual del campo como una cadena.
            actual_value_str = locator.input_value()

            # Intentar convertir la cadena a un número flotante.
            actual_value_float = float(actual_value_str)
            
            # Realizar la comparación de flotantes con la tolerancia.
            # `math.isclose` es la forma recomendada para comparar flotantes.
            if math.isclose(actual_value_float, valor_numerico_esperado, rel_tol=tolerancia, abs_tol=tolerancia):
                # --- Medición de rendimiento: Fin de la verificación (éxito) ---
                end_time_float_check = time.time()
                duration_float_check = end_time_float_check - start_time_float_check
                self.logger.info(f"PERFORMANCE: Tiempo que tardó en verificar que el campo '{selector}' contiene el valor flotante '{valor_numerico_esperado}': {duration_float_check:.4f} segundos.")

                self.logger.info(f"\n✔ ÉXITO: El campo '{selector}' contiene el valor numérico flotante esperado: '{valor_numerico_esperado}' (Actual: {actual_value_float}).")
                # Toma una captura de pantalla al verificar que el campo tiene el valor esperado.
                self.tomar_captura(f"{nombre_base}_despues_verificar_valor_float", directorio)
                return True
            else:
                # Si la comparación con tolerancia falla
                error_msg = (
                    f"\n❌ FALLO (Inexactitud): El campo '{selector}' NO contiene el valor numérico flotante esperado. "
                    f"Actual: {actual_value_float}, Esperado: {valor_numerico_esperado}, "
                    f"Diferencia: {abs(actual_value_float - valor_numerico_esperado):.10f} (Tolerancia: {tolerancia})."
                )
                self.logger.warning(error_msg)
                self.tomar_captura(f"{nombre_base}_fallo_inexactitud_float", directorio)
                return False

        except TimeoutError as e:
            # Captura si el campo no se hace visible o no se puede obtener su valor a tiempo.
            # Se intenta obtener el valor actual del campo si es posible.
            actual_value_str_on_timeout = "N/A"
            try:
                # Intenta obtener el valor actual como cadena justo antes de la excepción.
                actual_value_str_on_timeout = locator.input_value()
            except Exception:
                pass # Ignora si no se puede obtener.

            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_float_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El campo '{selector}' no se hizo visible o no se pudo obtener su valor "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s) para verificar el flotante '{valor_numerico_esperado}'. "
                f"Valor actual (si disponible): '{actual_value_str_on_timeout}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_float", directorio)
            return False

        except ValueError:
            # Captura si el valor obtenido del campo no es una cadena que pueda convertirse a float.
            error_msg = (
                f"\n❌ FALLO (Valor no numérico): El valor actual del campo '{selector}' ('{actual_value_str}') "
                f"no pudo ser convertido a flotante para comparación. Se esperaba '{valor_numerico_esperado}'."
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_valor_no_float", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, elemento no es un campo de entrada).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar el valor numérico flotante del campo '{selector}'. "
                f"Esto indica un problema fundamental con el selector o el tipo de elemento.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_float", directorio)
            raise # Re-lanza la excepción porque esto es un fallo de ejecución, no una verificación de estado.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar el valor numérico flotante del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_float", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre.
            # Aplica una espera fija después de la operación. Puede ser útil para observar
            # el estado del elemento o esperar efectos secundarios en la UI.
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    # 23- Función para verificar el texto 'alt' de una imagen con medición de rendimiento
    def verificar_alt_imagen(self, selector: Union[str, Page.locator], texto_alt_esperado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica que el **texto del atributo 'alt' de una imagen** coincida con el
        `texto_alt_esperado`. Esta función espera a que la imagen sea visible,
        obtiene su atributo 'alt', y luego realiza la comparación.
        Registra el tiempo que tarda esta verificación como una métrica de rendimiento.
        Toma capturas de pantalla tanto en caso de éxito como de fallo.

        Args:
            selector (Union[str, Page.locator]): El **selector de la imagen** a verificar.
                                                  Puede ser una cadena (CSS, XPath, etc.)
                                                  o un objeto `Locator` de Playwright preexistente.
            texto_alt_esperado (str): El **valor exacto del texto 'alt'** que se espera
                                      encontrar en la imagen.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que la
                                        imagen se haga visible y su atributo 'alt' sea obtenible.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (se ajustó de 0.5 para robustez).

        Returns:
            bool: `True` si el texto 'alt' de la imagen coincide con `texto_alt_esperado`
                  dentro del tiempo especificado; `False` en caso contrario (timeout o no coincidencia).

        Raises:
            Error: Si ocurre un problema específico de Playwright que impida la verificación
                   (ej., selector inválido, el elemento no es una imagen).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nVerificando el texto 'alt' para la imagen con selector: '{selector}'. Valor esperado: '{texto_alt_esperado}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación del texto 'alt' ---
        # Registra el tiempo justo antes de iniciar la operación de verificación.
        start_time_alt_check = time.time()

        try:
            # Resalta visualmente el elemento en el navegador. Útil para depuración.
            locator.highlight()
            # Toma una captura de pantalla del estado de la imagen *antes* de la verificación.
            self.tomar_captura(f"{nombre_base}_antes_verificar_alt_imagen", directorio)

            # Esperar a que la imagen sea visible y esté adjunta al DOM.
            # Esto es crucial antes de intentar obtener atributos, ya que asegura que el elemento está cargado.
            expect(locator).to_be_visible()
            self.logger.debug(f"\nLa imagen con selector '{selector}' es visible.")

            # Obtener el atributo 'alt' de la imagen.
            # `get_attribute` también tiene un `timeout` que esperará hasta que el atributo esté presente.
            alt_text_actual = locator.get_attribute("alt")

            # Validar que el atributo 'alt' no sea None y coincida con el texto esperado.
            # La comparación debe ser estricta para asegurar que el atributo existe y es correcto.
            if alt_text_actual == texto_alt_esperado:
                # --- Medición de rendimiento: Fin de la verificación (éxito) ---
                end_time_alt_check = time.time()
                duration_alt_check = end_time_alt_check - start_time_alt_check
                self.logger.info(f"PERFORMANCE: Tiempo que tardó en verificar el texto 'alt' de la imagen '{selector}': {duration_alt_check:.4f} segundos.")

                self.logger.info(f"\n✔ ÉXITO: El texto 'alt' de la imagen es '{alt_text_actual}' y coincide con el esperado ('{texto_alt_esperado}').")
                # Toma una captura de pantalla al verificar que el 'alt' de la imagen es el esperado.
                self.tomar_captura(f"{nombre_base}_alt_ok", directorio)
                return True
            else:
                # Si el texto 'alt' no coincide con el esperado
                error_msg = (
                    f"\n❌ FALLO (No Coincide): El texto 'alt' actual de la imagen '{selector}' es '{alt_text_actual}', "
                    f"pero se esperaba '{texto_alt_esperado}'."
                )
                self.logger.warning(error_msg) # Usa 'warning' ya que la función devuelve False.
                # Toma una captura de pantalla si el texto 'alt' no coincide.
                self.tomar_captura(f"{nombre_base}_alt_error", directorio)
                return False

        except TimeoutError as e:
            # Captura si la imagen no se hace visible o no se puede obtener su atributo 'alt' a tiempo.
            error_msg = (
                f"\n❌ FALLO (Timeout): La imagen con selector '{selector}' no se hizo visible "
                f"o no se pudo obtener su atributo 'alt' después de {tiempo} segundos para verificar el texto '{texto_alt_esperado}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla en el momento del fallo por timeout.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_alt_imagen", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, el elemento no es una imagen).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar el texto 'alt' de la imagen '{selector}'. "
                f"Esto indica un problema fundamental con el selector o el tipo de elemento.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Registra el error con la traza completa.
            # Toma una captura de pantalla para el error específico de Playwright.
            self.tomar_captura(f"{nombre_base}_error_playwright_alt_imagen", directorio)
            raise # Re-lanza la excepción porque esto es un fallo de ejecución, no una verificación de estado.

        except Exception as e:
            # Captura cualquier otra excepción inesperada que pueda ocurrir.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar el texto 'alt' de la imagen '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            # Toma una captura de pantalla para errores completamente inesperados.
            self.tomar_captura(f"{nombre_base}_error_inesperado_alt_imagen", directorio)
            raise # Re-lanza la excepción.

        finally:
            # El bloque `finally` se ejecuta siempre.
            # Aplica una espera fija después de la operación. Puede ser útil para observar
            # el estado del elemento o esperar efectos secundarios en la UI.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    # 24- Función para verificar que una imagen se cargue exitosamente (sin enlaces rotos) con pruebas de rendimiento.
    def verificar_carga_exitosa_imagen(self, selector: Union[str, Page.locator], nombre_base: str, directorio: str, tiempo_espera_red: Union[int, float] = 10.0, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica que una **imagen especificada por su selector** se cargue exitosamente,
        lo que implica que sea visible en el DOM y que su recurso se descargue con un
        código de estado HTTP exitoso (2xx). Integra mediciones de rendimiento para
        registrar el tiempo total de esta verificación.

        Args:
            selector (Union[str, Page.locator]): El **selector de la imagen** (e.g., 'img#logo', 'img[alt="banner"]').
                                                  Puede ser una cadena o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo_espera_red (Union[int, float]): **Tiempo máximo de espera** (en segundos) para
                                                  que la imagen sea visible y para que su respuesta
                                                  de red se complete. Por defecto, `10.0` segundos.
                                                  Este es el principal timeout de rendimiento.
            tiempo (Union[int, float]): **Tiempo de espera fijo** (en segundos) al final de la
                                        operación, útil para observar cambios. Por defecto, `1.0` segundo.

        Returns:
            bool: `True` si la imagen se carga exitosamente (visible y respuesta 2xx);
                  `False` en caso contrario (timeout, src vacío, o estado HTTP de error).

        Raises:
            Error: Si ocurre un problema específico de Playwright durante la interacción con el elemento
                   (ej., selector inválido, no es un elemento de imagen válido).
            Exception: Para cualquier otro error inesperado.
        """
        image_url = None
        self.logger.info(f"\nIniciando verificación de carga exitosa para la imagen con selector: '{selector}'. Tiempo de espera de red: {tiempo_espera_red}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la verificación de carga de imagen ---
        # Registra el tiempo justo antes de iniciar la cadena de verificaciones (visibilidad, src, respuesta de red).
        start_time_image_load_check = time.time()

        try:
            # 1. Resaltar el elemento (útil para depuración visual en el navegador)
            locator.highlight()
            self.logger.debug(f"\nElemento con selector '{selector}' resaltado.")
            self.tomar_captura(f"{nombre_base}_antes_verificar_carga_imagen", directorio) # Captura antes de iniciar la carga.

            # 2. Esperar a que la imagen sea visible en el DOM
            # Esto asegura que el elemento <img> está presente y renderizado.
            self.logger.debug(f"\nEsperando visibilidad de la imagen con selector '{selector}' (timeout: {tiempo_espera_red}s).")
            expect(locator).to_be_visible()
            self.logger.info(f"\nLa imagen con selector '{selector}' es visible en el DOM.")

            # 3. Obtener la URL de la imagen del atributo 'src'
            # Playwright esperará implícitamente a que el atributo 'src' esté presente.
            image_url = locator.get_attribute("src")
            if not image_url:
                error_msg = f"\n❌ FALLO: El atributo 'src' de la imagen con selector '{selector}' está vacío o no existe."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_src_vacio", directorio)
                return False

            self.logger.info(f"\nURL de la imagen a verificar: {image_url}")

            # 4. Monitorear la carga de la imagen en la red
            # Usamos page.wait_for_response para esperar la respuesta HTTP de la imagen específica.
            # Esto es más robusto que solo verificar la visibilidad, ya que asegura que el recurso
            # fue descargado correctamente de la red. Filtramos por la URL y el tipo de recurso 'image'.
            self.logger.debug(f"\nEsperando respuesta de red para la imagen con URL: {image_url} (timeout: {tiempo_espera_red}s).")
            response = self.page.wait_for_response(
                lambda resp: resp.url == image_url and resp.request.resource_type == "image",
                timeout=tiempo_espera_red * 1000 # Playwright espera milisegundos
            )

            # 5. Verificar el código de estado de la respuesta HTTP
            if 200 <= response.status <= 299:
                # --- Medición de rendimiento: Fin de la verificación (éxito) ---
                end_time_image_load_check = time.time()
                duration_image_load_check = end_time_image_load_check - start_time_image_load_check
                self.logger.info(f"PERFORMANCE: Tiempo total para verificar la carga exitosa de la imagen '{selector}' (URL: {image_url}): {duration_image_load_check:.4f} segundos.")

                self.logger.info(f"\n✔ ÉXITO: La imagen con URL '{image_url}' cargó exitosamente con estado HTTP {response.status}.")
                self.tomar_captura(f"{nombre_base}_carga_ok", directorio)
                return True
            else:
                # Si el estado HTTP no es un 2xx (indica un problema de carga)
                self.logger.error(f"\n❌ FALLO: La imagen con URL '{image_url}' cargó con un estado de error: {response.status}.")
                self.tomar_captura(f"{nombre_base}_carga_fallida_status_{response.status}", directorio)
                return False

        except TimeoutError as e:
            # Captura si el elemento no aparece o la respuesta de red no llega a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_image_load_check # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo verificar la carga de la imagen con selector '{selector}' "
                f"y URL '{image_url if image_url else 'N/A'}' después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_red}s).\n"
                f"Posibles causas: El elemento no apareció a tiempo o la respuesta de red no se completó.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True) # Usa 'warning' ya que la función devuelve False.
            self.tomar_captura(f"{nombre_base}_timeout_verificacion", directorio)
            return False

        except Error as e: # Captura errores específicos de Playwright (ej., selector inválido, no es un elemento de imagen)
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al verificar la carga de la imagen con selector '{selector}'.\n"
                f"Esto indica un problema fundamental con el selector o que el elemento no es una imagen válida.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            return False

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar la carga de la imagen con selector '{selector}' "
                f"y URL '{image_url if image_url else 'N/A'}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usa nivel crítico para errores graves.
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    # 25- Función para cargar archivo(s) con medición de rendimiento
    def cargar_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio: str, base_dir: str, file_names: Union[str, List[str]], tiempo: Union[int, float] = 0.5) -> bool:
        """
        Carga uno o varios archivos en un elemento de entrada de tipo 'file' en la página.
        Verifica que los archivos existan localmente antes de intentar cargarlos.
        Mide el rendimiento de la operación de carga de archivos.

        Args:
            selector (Union[str, Locator]): El **selector del elemento de entrada de archivo** (input[type="file"]).
                                            Puede ser una cadena (CSS, XPath, etc.) o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            base_dir (str): **Directorio base** donde se encuentran los archivos a cargar.
            file_names (Union[str, List[str]]): El **nombre o una lista de nombres de archivo(s)**
                                                  (solo el nombre del archivo, no la ruta completa)
                                                  que se desea cargar. Estos nombres se combinarán
                                                  con `base_dir` para obtener la ruta completa.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté visible y habilitado. También es el tiempo
                                        de espera fijo después de la operación exitosa.
                                        Por defecto, `5.0` segundos (ajustado para robustez).

        Returns:
            bool: `True` si el archivo(s) se carga(n) exitosamente; `False` en caso de fallo
                  (ej., archivo no encontrado, timeout, elemento no interactuable).

        Raises:
            FileNotFoundError: Si alguno de los archivos especificados no existe en el `base_dir`.
            Error: Si ocurre un problema específico de Playwright (ej., selector inválido,
                   elemento no es un input de tipo file, timeout de visibilidad/habilitación).
            Exception: Para cualquier otro error inesperado.
        """
        # Normalizar `file_names` a una lista para manejar consistentemente uno o varios archivos
        file_names_list = [file_names] if isinstance(file_names, str) else file_names

        self.logger.info(f"\nIntentando cargar archivo(s) '{file_names_list}' en el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Construir las rutas completas de los archivos y verificar su existencia localmente
        full_file_paths = []
        for name in file_names_list:
            full_path = os.path.join(base_dir, name)
            full_file_paths.append(full_path)
            self.logger.debug(f"\nConstruida ruta completa para archivo: '{full_path}'")

            if not os.path.exists(full_path):
                error_msg = f"\n❌ Error: El archivo no existe en la ruta especificada: '{full_path}'."
                self.logger.error(error_msg, exc_info=True)
                self.tomar_captura(f"{nombre_base}_archivo_no_encontrado", directorio)
                raise FileNotFoundError(error_msg) # Elevar un error específico si el archivo no se encuentra.

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de carga de archivos ---
        # Registra el tiempo justo antes de iniciar la interacción con el elemento de entrada de archivo.
        start_time_file_upload = time.time()

        try:
            # 1. Esperar a que el elemento de entrada de archivo esté visible y habilitado
            # Es fundamental asegurar que el elemento está listo para interactuar.
            self.logger.debug(f"\nEsperando que el selector '{selector}' esté visible y habilitado (timeout: {tiempo}s).")
            expect(locator).to_be_visible()
            expect(locator).to_be_enabled() # También se puede usar to_be_editable() si es un input
            self.logger.info(f"\nEl selector '{selector}' está visible y habilitado.")

            # 2. Opcional: Resaltar el elemento para depuración visual
            locator.highlight()
            self.logger.debug(f"\nElemento con selector '{selector}' resaltado.")
            self.tomar_captura(f"{nombre_base}_antes_cargar_archivos", directorio) # Captura antes de adjuntar los archivos.

            # 3. Usar set_input_files para adjuntar el archivo(s)
            # Playwright maneja la interacción con el diálogo de carga de archivos.
            # Se le pasa una lista de rutas completas de los archivos a adjuntar.
            self.logger.info(f"\nAdjuntando archivo(s) {file_names_list} al selector '{selector}'.")
            locator.set_input_files(full_file_paths)

            # --- Medición de rendimiento: Fin de la operación de carga de archivos ---
            # Registra el tiempo una vez que Playwright ha adjuntado los archivos.
            end_time_file_upload = time.time()
            duration_file_upload = end_time_file_upload - start_time_file_upload
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en cargar el archivo(s) '{file_names_list}' en el selector '{selector}': {duration_file_upload:.4f} segundos.")

            # Construir mensaje de éxito basado en si es uno o varios archivos
            if len(file_names_list) == 1:
                success_msg = f"\n✅ Archivo '{file_names_list[0]}' cargado exitosamente desde '{base_dir}' en el selector '{selector}'."
            else:
                success_msg = f"\n✅ Archivos {file_names_list} cargados exitosamente desde '{base_dir}' en el selector '{selector}'."
            self.logger.info(success_msg)
            
            self.tomar_captura(f"{nombre_base}_archivos_cargados", directorio)
            return True

        except TimeoutError as e:
            # Captura si el elemento no se hace visible o habilitado a tiempo.
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_file_upload # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible o habilitado "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s) para cargar el archivo(s) '{error_files_info}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Usa 'error' porque un timeout al cargar archivos es un fallo crítico.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_cargar_archivo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, el elemento no es un input[type="file"]).
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al cargar el archivo(s) '{error_files_info}' "
                f"en el selector '{selector}'. Esto puede deberse a un selector incorrecto o que el elemento "
                f"no es un input de tipo archivo válido.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_cargar_archivo", directorio)
            raise # Re-lanza la excepción porque es un fallo de ejecución.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al cargar el archivo(s) '{error_files_info}' "
                f"en el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_cargar_archivo", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
        
    # 26- Función para remover carga de archivo(s) con medición de rendimiento
    def remover_carga_de_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Remueve la carga de archivo(s) de un elemento de entrada de tipo 'file'
        estableciendo su valor a una lista vacía. Mide el rendimiento de esta operación.

        Args:
            selector (Union[str, Locator]): El **selector del elemento de entrada de archivo** (input[type="file"])
                                            del cual se removerá la carga.
                                            Puede ser una cadena (CSS, XPath, etc.) o un objeto `Locator` de Playwright.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que el
                                        elemento esté visible y habilitado antes de intentar
                                        remover la carga. También es el tiempo de espera fijo
                                        después de la operación exitosa.
                                        Por defecto, `5.0` segundos (ajustado para robustez).

        Returns:
            bool: `True` si la carga del archivo se remueve exitosamente; `False` en caso de fallo
                  (ej., timeout, elemento no interactuable).

        Raises:
            Error: Si ocurre un problema específico de Playwright (ej., selector inválido,
                   elemento no es un input de tipo file, timeout de visibilidad/habilitación).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIntentando remover la carga de archivo para el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # Asegura que 'selector' sea un objeto Locator de Playwright para un uso consistente.
        if isinstance(selector, str):
            locator = self.page.locator(selector)
        else:
            locator = selector

        # --- Medición de rendimiento: Inicio de la operación de remoción de archivos ---
        # Registra el tiempo justo antes de iniciar la interacción con el elemento.
        start_time_file_removal = time.time()

        try:
            # 1. Esperar a que el elemento de entrada de archivo esté visible y habilitado
            # Es fundamental asegurar que el elemento está listo para interactuar y aceptar la limpieza.
            self.logger.debug(f"\nEsperando que el selector '{selector}' esté visible y habilitado (timeout: {tiempo}s) para remover la carga.")
            expect(locator).to_be_visible()
            expect(locator).to_be_enabled() # O to_be_editable()
            self.logger.info(f"\nEl selector '{selector}' está visible y habilitado.")

            # 2. Resaltar el elemento para depuración visual
            locator.highlight()
            self.logger.debug(f"\nElemento con selector '{selector}' resaltado.")
            self.tomar_captura(f"{nombre_base}_antes_remover_carga", directorio) # Captura antes de remover.

            # 3. Usar set_input_files con una lista vacía para remover el archivo
            # Esto simula el usuario cancelando o limpiando la selección de archivos.
            self.logger.info(f"\nEstableciendo input files a vacío para el selector '{selector}'.")
            locator.set_input_files([])

            # --- Medición de rendimiento: Fin de la operación de remoción de archivos ---
            # Registra el tiempo una vez que Playwright ha limpiado el input de archivos.
            end_time_file_removal = time.time()
            duration_file_removal = end_time_file_removal - start_time_file_removal
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en remover la carga de archivo para el selector '{selector}': {duration_file_removal:.4f} segundos.")

            self.logger.info(f"\n✅ Carga de archivo removida exitosamente para el selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_remocion_completa", directorio)
            return True

        except TimeoutError as e:
            # Captura si el elemento no se hace visible o habilitado a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_file_removal # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible o habilitado "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s) para remover la carga de archivo. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Usa 'error' porque un timeout es un fallo crítico.
            self.tomar_captura(f"{nombre_base}_fallo_timeout_remocion_archivo", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, el elemento no es un input[type="file"]).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar remover la carga de archivo "
                f"para el selector '{selector}'. Esto puede deberse a un selector incorrecto o que el elemento "
                f"no es un input de tipo archivo válido.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_remocion_archivo", directorio)
            raise # Re-lanza la excepción porque es un fallo de ejecución.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar remover la carga de archivo "
                f"para el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_remocion_archivo", directorio)
            raise # Re-lanza la excepción.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
        
    # 27- Función para contar filas y columnas de una tabla con pruebas de rendimiento
    def obtener_dimensiones_tabla(self, selector: Locator, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> Tuple[int, int]:
        """
        Obtiene las dimensiones (número de filas y columnas) de una tabla HTML
        identificada por un Playwright Locator. Mide el tiempo que toma esta operación.

        Prioriza el conteo de filas de `tbody tr` y columnas de `th` (encabezados).
        Si no hay encabezados, intenta contar las celdas `td` de la primera fila de datos.

        Args:
            selector (Locator): El **Locator de Playwright** que representa el elemento
                                `<table>` (o un elemento padre que contenga la tabla).
                                Es crucial que sea un Locator, no una cadena, para aprovechar
                                sus funcionalidades de espera y contexto.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que la tabla
                                        y sus elementos internos (filas/columnas) estén presentes
                                        y visibles antes de intentar contarlos.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (ajustado de 1.0 para robustez).

        Returns:
            tuple[int, int]: Una tupla `(num_filas, num_columnas)` representando las dimensiones de la tabla.
                             Retorna `(-1, -1)` en caso de `TimeoutError` si la tabla no está lista.

        Raises:
            Error: Si ocurre un problema específico de Playwright al interactuar con el selector
                   (ej., el selector no apunta a una tabla o un elemento válido).
            Exception: Para cualquier otro error inesperado.
        """
        # Intentar obtener información útil del selector para los logs y nombres de captura.
        # Esto ayuda a identificar la tabla en los logs, especialmente si no tiene ID/NAME.
        selector_info = selector.get_attribute('id') or selector.get_attribute('name')
        if not selector_info:
            try:
                # Si no hay id/name, intentar obtener el HTML externo de la etiqueta principal
                selector_info = selector.evaluate("el => el.outerHTML.split('>')[0] + '>'")
            except Exception:
                selector_info = f"Tabla con selector genérico: {selector}" # Fallback si evaluate falla

        self.logger.info(f"\nObteniendo dimensiones de la tabla con selector: '{selector_info}'. Tiempo máximo de espera: {tiempo}s.")

        # --- Medición de rendimiento: Inicio de la operación de obtener dimensiones ---
        # Registra el tiempo justo antes de iniciar la interacción para obtener las dimensiones.
        start_time_get_dimensions = time.time()

        try:
            # 1. Asegurar que la tabla principal esté visible
            # Es crucial que la tabla esté cargada y visible para poder contar sus elementos.
            self.logger.debug(f"\nEsperando que la tabla con selector '{selector_info}' esté visible (timeout: {tiempo}s).")
            expect(selector).to_be_visible()
            
            # Resaltar el elemento de la tabla para depuración visual.
            selector.highlight()
            self.logger.debug(f"\nTabla con selector '{selector_info}' resaltada.")
            self.tomar_captura(f"{nombre_base}_antes_obtener_dimensiones", directorio) # Captura antes de contar.

            # 2. Contar el número de filas de datos
            # Se buscan filas `<tr>` dentro de un `<tbody>` para contar solo las filas de datos,
            # excluyendo potencialmente encabezados o pies de tabla.
            filas_datos = selector.locator("tbody tr")
            num_filas = filas_datos.count()
            self.logger.debug(f"\nFilas de datos encontradas (tbody tr): {num_filas}.")

            # 3. Contar el número de columnas
            num_columnas = 0
            # Intentar contar desde los encabezados de la tabla (th) primero.
            headers = selector.locator("th")
            if headers.count() > 0:
                num_columnas = headers.count()
                self.logger.debug(f"\nColumnas contadas desde encabezados (th): {num_columnas}.")
            else:
                # Si no hay thead/th, intentar contar td's de la primera fila de datos.
                # Esto es útil para tablas que no usan thead o que son simples.
                first_row_tds = selector.locator("tr").first.locator("td")
                if first_row_tds.count() > 0:
                    num_columnas = first_row_tds.count()
                    self.logger.debug(f"\nColumnas contadas desde celdas de la primera fila (td): {num_columnas}.")
                else:
                    self.logger.warning(f"\nADVERTENCIA: No se pudieron encontrar encabezados (th) ni celdas (td) en la primera fila "
                                        f"para la tabla con selector '{selector_info}'. Asumiendo 0 columnas.")
                    # En este caso, num_columnas seguirá siendo 0.

            # --- Medición de rendimiento: Fin de la operación de obtener dimensiones ---
            # Registra el tiempo una vez que se han contado las filas y columnas.
            end_time_get_dimensions = time.time()
            duration_get_dimensions = end_time_get_dimensions - start_time_get_dimensions
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en obtener las dimensiones de la tabla '{selector_info}': {duration_get_dimensions:.4f} segundos.")

            self.tomar_captura(f"{nombre_base}_dimensiones_obtenidas", directorio)
            self.logger.info(f"\n✅ ÉXITO: Dimensiones de la tabla '{selector_info}' obtenidas.")
            self.logger.info(f"--> Filas encontradas: {num_filas}")
            self.logger.info(f"--> Columnas encontradas: {num_columnas}")
            return (num_filas, num_columnas)

        except TimeoutError as e:
            # Captura si la tabla principal o sus elementos internos no se hacen visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_get_dimensions # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo obtener las dimensiones de la tabla con selector '{selector_info}' "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo}s).\n"
                f"La tabla o sus elementos internos no estuvieron disponibles a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True) # Usa 'warning' ya que devuelve un valor indicativo de fallo.
            self.tomar_captura(f"{nombre_base}_dimensiones_timeout", directorio)
            return (-1, -1) # Retorna valores indicativos de fallo.

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector de tabla inválido, problema al interactuar con el DOM).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar obtener las dimensiones de la tabla con selector '{selector_info}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada, elemento no es una tabla.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dimensiones_error_playwright", directorio)
            raise # Relanzar porque es un error de ejecución de Playwright, no un fallo de aserción.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al obtener las dimensiones de la tabla con selector '{selector_info}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Nivel crítico para errores muy graves.
            self.tomar_captura(f"{nombre_base}_dimensiones_error_inesperado", directorio)
            raise # Relanzar por ser un error inesperado.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
        
    # 28- Función para buscar datos parcial e imprimir la fila con pruebas de rendimiento
    def busqueda_coincidencia_e_imprimir_fila(self, table_selector: Locator, texto_buscado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Busca una **coincidencia parcial de texto** dentro de las filas de una tabla
        especificada por un Playwright Locator. Si encuentra el texto, resalta la fila
        y registra su contenido. Mide el rendimiento de esta operación de búsqueda.

        Args:
            table_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` (o un elemento padre que contenga la tabla).
                                      Es crucial que sea un Locator, no una cadena, para aprovechar
                                      sus funcionalidades de espera y contexto.
            texto_buscado (str): El **texto a buscar** dentro de las filas de la tabla.
                                 La búsqueda no es sensible a mayúsculas/minúsculas.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que la tabla
                                        esté visible antes de iniciar la búsqueda.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (ajustado de 1.0 para robustez).

        Returns:
            bool: `True` si se encuentra al menos una coincidencia parcial del `texto_buscado`
                  en alguna fila de la tabla; `False` en caso contrario o si ocurre un `TimeoutError`.

        Raises:
            Error: Si ocurre un problema específico de Playwright al interactuar con el selector
                   de la tabla (ej., el selector no apunta a una tabla válida).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIniciando búsqueda de coincidencia parcial para '{texto_buscado}' en la tabla con selector: '{table_selector}'. Tiempo máximo de espera: {tiempo}s.")
        encontrado = False

        # --- Medición de rendimiento: Inicio de la búsqueda en la tabla ---
        # Registra el tiempo justo antes de iniciar la interacción con la tabla para la búsqueda.
        start_time_table_search = time.time()

        try:
            # 1. Esperar a que la tabla esté visible
            # Esto es fundamental antes de intentar iterar sobre sus filas.
            self.logger.debug(f"\nEsperando que la tabla con selector '{table_selector}' esté visible (timeout: {tiempo}s).")
            expect(table_selector).to_be_visible()
            self.logger.info(f"\nTabla con selector '{table_selector}' está visible.")
            
            # Resaltar la tabla completa para depuración visual.
            table_selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_busqueda_coincidencia", directorio) # Captura antes de buscar.

            # 2. Obtener todas las filas de datos de la tabla
            # Se buscan filas `<tr>` dentro de un `<tbody>` para enfocar la búsqueda en los datos.
            filas = table_selector.locator("tbody tr")
            num_filas = filas.count()
            self.logger.debug(f"\nNúmero de filas de datos encontradas en la tabla: {num_filas}.")

            # 3. Iterar sobre cada fila para buscar la coincidencia
            for i in range(num_filas):
                fila = filas.nth(i) # Obtiene el Locator para la fila actual.
                fila_texto = fila.text_content() # Obtiene todo el texto visible de la fila.
                self.logger.debug(f"\nAnalizando fila {i+1}: '{fila_texto}'.")

                # Realizar la búsqueda de coincidencia parcial sin distinguir mayúsculas/minúsculas.
                if texto_buscado.lower() in fila_texto.lower():
                    self.logger.info(f"\n✅ ÉXITO: Texto '{texto_buscado}' encontrado (coincidencia parcial) en la fila {i+1}.")
                    self.logger.info(f"Contenido completo de la fila: '{fila_texto}'")
                    fila.highlight() # Resalta la fila donde se encontró la coincidencia.
                    self.tomar_captura(f"{nombre_base}_coincidencia_parcial_encontrada_fila_{i+1}", directorio)
                    encontrado = True
                    # Si solo se necesita encontrar la primera coincidencia y terminar, descomentar el 'break'
                    # break 
            
            if not encontrado:
                self.logger.info(f"\nℹ️ Texto '{texto_buscado}' (coincidencia parcial) NO encontrado en ninguna fila de la tabla.")
                self.tomar_captura(f"{nombre_base}_coincidencia_parcial_no_encontrada", directorio)

            # --- Medición de rendimiento: Fin de la búsqueda en la tabla ---
            # Registra el tiempo una vez que se ha completado la iteración sobre todas las filas (o hasta la primera coincidencia si se usa break).
            end_time_table_search = time.time()
            duration_table_search = end_time_table_search - start_time_table_search
            self.logger.info(f"PERFORMANCE: Tiempo que tardó la búsqueda de '{texto_buscado}' en la tabla '{table_selector}': {duration_table_search:.4f} segundos.")

            return encontrado

        except TimeoutError as e:
            # Captura si la tabla principal o sus filas no se hacen visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_table_search # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo encontrar la tabla con selector '{table_selector}' "
                f"o sus filas no estuvieron disponibles a tiempo ({duration_fail:.4f}s, timeout configurado: {tiempo}s) "
                f"durante la búsqueda de '{texto_buscado}'.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True) # Usa 'warning' ya que devuelve False.
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_timeout", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector de tabla inválido, problemas de interacción con el DOM).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al buscar coincidencia para '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada, o problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_error_playwright", directorio)
            raise # Relanzar porque es un error de ejecución de Playwright.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al buscar coincidencia para '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Nivel crítico para errores muy graves.
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_error_inesperado", directorio)
            raise # Relanzar por ser un error inesperado.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
        
    # 29- Función para buscar datos exacto e imprimir la fila con pruebas de rendimiento
    def busqueda_estricta_imprimir_fila(self, table_selector: Locator, texto_buscado: str, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Busca una **coincidencia exacta de texto** dentro de las celdas de una tabla
        especificada por un Playwright Locator. Si encuentra el texto, resalta la celda
        y la fila correspondiente, y registra el contenido completo de la fila.
        Mide el rendimiento de esta operación de búsqueda estricta.

        Args:
            table_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` (o un elemento padre que contenga la tabla).
                                      Es crucial que sea un Locator, no una cadena, para aprovechar
                                      sus funcionalidades de espera y contexto.
            texto_buscado (str): El **texto exacto a buscar** dentro de las celdas de la tabla.
                                 La búsqueda es sensible a mayúsculas/minúsculas y requiere una
                                 coincidencia exacta (después de eliminar espacios en blanco).
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): **Tiempo máximo de espera** (en segundos) para que la tabla
                                        esté visible antes de iniciar la búsqueda.
                                        También es el tiempo de espera fijo después de la operación.
                                        Por defecto, `5.0` segundos (ajustado de 1.0 para robustez).

        Returns:
            bool: `True` si se encuentra al menos una coincidencia exacta del `texto_buscado`
                  en alguna celda de la tabla; `False` en caso contrario o si ocurre un `TimeoutError`.

        Raises:
            Error: Si ocurre un problema específico de Playwright al interactuar con el selector
                   de la tabla (ej., el selector no apunta a una tabla válida).
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\nIniciando búsqueda estricta para '{texto_buscado}' en la tabla con selector: '{table_selector}'. Tiempo máximo de espera: {tiempo}s.")
        encontrado = False

        # --- Medición de rendimiento: Inicio de la búsqueda estricta en la tabla ---
        # Registra el tiempo justo antes de iniciar la interacción con la tabla para la búsqueda.
        start_time_strict_search = time.time()

        try:
            # 1. Esperar a que la tabla esté visible
            # Esto es fundamental antes de intentar iterar sobre sus filas y celdas.
            self.logger.debug(f"\nEsperando que la tabla con selector '{table_selector}' esté visible (timeout: {tiempo}s).")
            expect(table_selector).to_be_visible()
            self.logger.info(f"\nTabla con selector '{table_selector}' está visible.")
            
            # Resaltar la tabla completa para depuración visual.
            table_selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_busqueda_estricta", directorio) # Captura antes de buscar.

            # 2. Obtener todas las filas de datos de la tabla
            # Se buscan filas `<tr>` dentro de un `tbody` para enfocar la búsqueda en los datos.
            filas = table_selector.locator("tbody tr")
            num_filas = filas.count()
            self.logger.debug(f"\nNúmero de filas de datos encontradas en la tabla: {num_filas}.")

            # 3. Iterar sobre cada fila y cada celda para buscar la coincidencia exacta
            for i in range(num_filas):
                fila = filas.nth(i) # Obtiene el Locator para la fila actual.
                celdas = fila.locator("td") # Asumiendo celdas de datos son 'td'.
                num_celdas = celdas.count()
                fila_texto_completo = "" # Para reconstruir y loggear el contenido completo de la fila.
                self.logger.debug(f"\nAnalizando fila {i+1} para búsqueda estricta.")

                for j in range(num_celdas):
                    celda = celdas.nth(j) # Obtiene el Locator para la celda actual.
                    celda_texto = celda.text_content().strip() # Obtiene el texto de la celda y elimina espacios en blanco alrededor.
                    fila_texto_completo += celda_texto + " | " # Concatenar para imprimir la fila completa en el log.

                    # Realizar la búsqueda de coincidencia estricta.
                    if celda_texto == texto_buscado: # Coincidencia estricta
                        self.logger.info(f"\n✅ ÉXITO: Texto '{texto_buscado}' encontrado (coincidencia estricta) en la celda {j+1} de la fila {i+1}.")
                        self.logger.info(f"Contenido completo de la fila: '{fila_texto_completo.strip(' | ')}'")
                        celda.highlight() # Resaltar la celda donde se encontró la coincidencia.
                        fila.highlight() # También resaltar la fila para mejor visibilidad.
                        self.tomar_captura(f"{nombre_base}_coincidencia_estricta_encontrada_fila_{i+1}_celda_{j+1}", directorio)
                        encontrado = True
                        # Si solo se necesita encontrar la primera coincidencia y terminar, descomentar ambos 'break'.
                        # break # Rompe el bucle de celdas.
                
                if encontrado:
                    # break # Rompe el bucle de filas si se encontró una coincidencia y se desea parar.
                    pass # Si se desea seguir buscando en otras filas, manteniendo el 'encontrado' en True.

            if not encontrado:
                self.logger.info(f"\nℹ️ Texto '{texto_buscado}' (coincidencia estricta) NO encontrado en ninguna celda de la tabla.")
                self.tomar_captura(f"{nombre_base}_coincidencia_estricta_no_encontrada", directorio)

            # --- Medición de rendimiento: Fin de la búsqueda estricta en la tabla ---
            # Registra el tiempo una vez que se ha completado la iteración sobre todas las celdas/filas.
            end_time_strict_search = time.time()
            duration_strict_search = end_time_strict_search - start_time_strict_search
            self.logger.info(f"PERFORMANCE: Tiempo que tardó la búsqueda estricta de '{texto_buscado}' en la tabla '{table_selector}': {duration_strict_search:.4f} segundos.")

            return encontrado

        except TimeoutError as e:
            # Captura si la tabla principal o sus elementos internos (filas/celdas) no se hacen visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_strict_search # Mide desde el inicio de la operación.
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo encontrar la tabla con selector '{table_selector}' "
                f"o sus elementos internos no estuvieron disponibles a tiempo ({duration_fail:.4f}s, timeout configurado: {tiempo}s) "
                f"durante la búsqueda estricta de '{texto_buscado}'.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True) # Usa 'warning' ya que devuelve False.
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_timeout", directorio)
            return False

        except Error as e:
            # Captura errores específicos de Playwright (ej., selector de tabla inválido, problemas de interacción con el DOM).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al buscar estrictamente '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada, o problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_error_playwright", directorio)
            raise # Relanzar porque es un error de ejecución de Playwright.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al buscar estrictamente '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Nivel crítico para errores muy graves.
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_error_inesperado", directorio)
            raise # Relanzar por ser un error inesperado.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Aplica una espera fija al final de la operación, útil para observación.
            if tiempo > 0:
                self.esperar_fijo(tiempo)
        
    # 30- Función para validar que todos los valores en una columna específica de una tabla sean numéricos, con pruebas de rendimiento
    def verificar_precios_son_numeros(self, tabla_selector: Locator, columna_nombre: str, nombre_base: str, directorio: str, tiempo_espera_celda: Union[int, float] = 0.5, tiempo_general_timeout: Union[int, float] = 15.0) -> bool:
        """
        Verifica que todos los valores en una **columna específica** de una tabla HTML
        sean **numéricos válidos**. Esto es crucial para la integridad de los datos
        mostrados en la UI, especialmente para precios o cantidades.
        Mide el rendimiento de esta operación de validación.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` (o un elemento padre que contenga la tabla).
                                      Es crucial que sea un Locator para aprovechar sus
                                      funcionalidades de espera y contexto.
            columna_nombre (str): El **nombre exacto de la columna** (texto del encabezado `<th>`)
                                  cuyos valores se desean verificar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_celda (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que una celda de precio individual
                                                     sea visible. Por defecto, `5.0` segundos.
            tiempo_general_timeout (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que la tabla y su `<tbody>` estén
                                                        visibles y listos para la interacción.
                                                        Por defecto, `15.0` segundos.

        Returns:
            bool: `True` si todos los valores en la columna especificada son numéricos válidos;
                  `False` si se encuentra algún valor no numérico o si la columna no existe.

        Raises:
            AssertionError: Si la tabla o sus elementos clave no están disponibles a tiempo,
                            o si ocurre un error inesperado de Playwright o genérico.
        """
        self.logger.info(f"\n⚙️ Verificando que todos los precios en la columna '{columna_nombre}' de la tabla '{tabla_selector}' son números.")

        # --- Medición de rendimiento: Inicio de la validación de la tabla ---
        # Registra el tiempo justo antes de iniciar cualquier interacción con la tabla.
        start_time_validation = time.time()

        try:
            # 1. Asegurar que la tabla principal esté visible
            # Es el primer paso para garantizar que la tabla se ha cargado en el DOM.
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_general_timeout}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.debug(f"\nTabla resaltada para verificación: {tabla_selector}")

            # 2. Esperar a que el tbody exista y tenga contenido
            # Es crucial esperar por la sección de cuerpo de la tabla y al menos una fila,
            # ya que a menudo se cargan de forma asíncrona.
            tbody_locator = tabla_selector.locator("tbody")
            self.logger.debug(f"\nEsperando que el tbody de la tabla sea visible (timeout: {tiempo_general_timeout}s).")
            expect(tbody_locator).to_be_visible()
            self.logger.info("\n✅ El tbody de la tabla es visible.")
            
            self.logger.debug(f"\nEsperando que al menos la primera fila de datos sea visible (timeout: {tiempo_general_timeout}s).")
            expect(tbody_locator.locator("tr").first).to_be_visible()
            self.logger.info("\n✅ Al menos la primera fila de datos en la tabla es visible.")
            self.tomar_captura(f"{nombre_base}_tabla_visible_para_verificacion", directorio) # Captura el estado inicial.

            # 3. Encontrar el índice de la columna por su nombre
            # Primero, asegurar que los encabezados existan y sean visibles.
            headers = tabla_selector.locator("th")
            self.logger.debug(f"\nEsperando que los encabezados (th) de la tabla sean visibles (timeout: {tiempo_general_timeout}s).")
            expect(headers.first).to_be_visible()

            col_index = -1
            header_texts = []
            for i in range(headers.count()):
                header_text = headers.nth(i).text_content().strip()
                header_texts.append(header_text)
                if header_text == columna_nombre:
                    col_index = i
            
            self.logger.info(f"\n🔍 Cabeceras encontradas: {header_texts}")

            if col_index == -1:
                self.logger.error(f"\n❌ Error: No se encontró la columna '{columna_nombre}' en la tabla. Cabeceras disponibles: {header_texts}")
                self.tomar_captura(f"{nombre_base}_columna_no_encontrada", directorio)
                # No lanzamos una excepción aquí, ya que el retorno False es suficiente para indicar el fallo lógico.
                return False

            self.logger.info(f"\n🔍 Columna '{columna_nombre}' encontrada en el índice: {col_index}")

            # 4. Obtener todas las filas de la tabla (solo las de datos dentro de tbody)
            rows = tbody_locator.locator("tr")
            num_rows = rows.count()
            if num_rows == 0:
                self.logger.warning("\n⚠️ Advertencia: La tabla no contiene filas de datos para verificar.")
                self.tomar_captura(f"{nombre_base}_tabla_vacia_no_precios", directorio)
                return True # Considera esto un éxito si no hay datos que validar.

            self.logger.info(f"\n🔍 Se encontraron {num_rows} filas de datos para verificar precios.")

            all_prices_are_numbers = True
            for i in range(num_rows):
                row_locator = rows.nth(i)
                # Se busca la celda correspondiente al índice de la columna dentro de la fila actual.
                price_cell = row_locator.locator(f"td").nth(col_index)
                
                # Es crucial esperar a que la celda de precio sea visible si las filas se renderizan dinámicamente
                # o el contenido de las celdas aparece con un retardo.
                self.logger.debug(f"\n Esperando que la celda de precio en la fila {i+1} esté visible (timeout: {tiempo_espera_celda}s).")
                expect(price_cell).to_be_visible() # Convertir a milisegundos
                
                price_text = price_cell.text_content().strip() # Obtener texto y limpiar espacios.
                price_cell.highlight() # Resaltar la celda actual para depuración visual.

                self.logger.debug(f"\n Procesando fila {i+1}, texto de precio: '{price_text}'")

                try:
                    float(price_text) # Intentar convertir el texto a un número flotante.
                    self.logger.debug(f"\n ✅ '{price_text}' es un número válido.")
                except ValueError:
                    self.logger.error(f"\n ❌ Error: El valor '{price_text}' en la fila {i+1} de la columna '{columna_nombre}' no es un número válido.")
                    self.tomar_captura(f"{nombre_base}_precio_invalido_fila_{i+1}", directorio)
                    all_prices_are_numbers = False
                    # Continuamos el bucle para reportar todos los valores no numéricos, no solo el primero.

            # --- Medición de rendimiento: Fin de la validación ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo total de validación de precios en la columna '{columna_nombre}': {duration_validation:.4f} segundos.")

            if all_prices_are_numbers:
                self.logger.info(f"\n✅ Todos los precios en la columna '{columna_nombre}' son números válidos.")
                self.tomar_captura(f"{nombre_base}_precios_ok", directorio)
                return True
            else:
                self.logger.error(f"\n❌ Se encontraron precios no numéricos en la columna '{columna_nombre}'.")
                return False

        except TimeoutError as e:
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_validation
            error_msg = (
                f"\n❌ FALLO (Timeout): La tabla o sus elementos (tbody, filas, celdas) no se volvieron visibles a tiempo "
                f"después de {duration_fail:.4f} segundos (timeout general configurado: {tiempo_general_timeout}s, celda: {tiempo_espera_celda}s). "
                f"Error: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_timeout_verificacion_precios", directorio)
            # Elevar AssertionError para que la prueba falle claramente cuando la tabla no está lista.
            raise AssertionError(f"\nElementos de la tabla no disponibles a tiempo para verificación de precios: {tabla_selector}") from e
        
        except Error as e:
            # Captura errores específicos de Playwright (ej., selector inválido, DOM mal formado).
            error_msg = (
                f"\n❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al verificar los precios en la tabla '{tabla_selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Nivel crítico porque un error de Playwright es un problema fundamental.
            self.tomar_captura(f"{nombre_base}_playwright_error_verificacion_precios", directorio)
            raise AssertionError(f"\nError de Playwright al verificar precios en la tabla: {tabla_selector}") from e
        
        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al verificar los precios en la tabla '{tabla_selector}'. "
                f"Error: {type(e).__name__}: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_excepcion_inesperada", directorio)
            raise AssertionError(f"\nError inesperado al verificar precios en la tabla: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Se puede eliminar la espera fija si la prueba se basa puramente en el retorno de la función.
            # Sin embargo, se mantiene por si se desea una pausa visual al final de la ejecución.
            if tiempo_general_timeout > 0: # Usamos el tiempo_general_timeout para la espera final
                self.esperar_fijo(tiempo_general_timeout / 5.0) # Espera un tiempo más corto al final, por ejemplo.
        
    # 31- Función para extraer y retornar el valor de un elemento dado su Playwright Locator, con pruebas de rendimiento
    def obtener_valor_elemento(self, selector: Locator, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5) -> Optional[str]:
        """
        Extrae y retorna el valor de un elemento dado su Playwright Locator.
        Prioriza la extracción de valores de campos de formulario (`input_value`),
        luego intenta `text_content` y `inner_text` para otros tipos de elementos.
        Mide el rendimiento de la operación de extracción.

        Args:
            selector (Locator): El **Locator de Playwright** que representa el elemento
                                del cual se desea extraer el valor. Es crucial que sea
                                un Locator para aprovechar sus funcionalidades de espera y contexto.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el elemento sea visible y esté
                                                        habilitado antes de intentar extraer su valor.
                                                        Por defecto, `5.0` segundos.

        Returns:
            Optional[str]: El valor extraído del elemento como una cadena de texto (str).
                           Retorna `None` si no se pudo extraer ningún valor significativo
                           después de intentar todos los métodos, o si el elemento no tiene texto.

        Raises:
            AssertionError: Si el elemento no se vuelve visible/habilitado a tiempo,
                            o si ocurre un error inesperado de Playwright o genérico
                            que impida la extracción del valor.
        """
        self.logger.info(f"\n⚙️ Extrayendo valor del elemento con selector: '{selector}'. Tiempo máximo de espera: {tiempo_espera_elemento}s.")
        valor_extraido = None
        
        # --- Medición de rendimiento: Inicio de la extracción del valor ---
        # Registra el tiempo justo antes de iniciar la interacción con el elemento.
        start_time_extraction = time.time()

        try:
            # 1. Asegurar que el elemento esté visible y habilitado
            # Estas aserciones son cruciales para garantizar que el elemento está listo para interactuar.
            self.logger.debug(f"\nEsperando que el elemento '{selector}' sea visible (timeout: {tiempo_espera_elemento}s).")
            expect(selector).to_be_visible()
            
            self.logger.debug(f"\nEsperando que el elemento '{selector}' esté habilitado (timeout: {tiempo_espera_elemento}s).")
            expect(selector).to_be_enabled()

            # Resaltar el elemento para depuración visual y tomar una captura.
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_extraccion_valor", directorio)
            self.logger.debug(f"\nElemento '{selector}' es visible y habilitado.")

            # 2. Intentar extraer el valor usando diferentes métodos de Playwright
            # Priorizamos `input_value` para campos de formulario (<input>, <textarea>, <select>).
            try:
                valor_extraido = selector.input_value() # Un timeout corto para input_value
                self.logger.debug(f"\nValor extraído (input_value) de '{selector}': '{valor_extraido}'")
            except Error as e_input: # Capturamos el error si input_value no es aplicable (ej. no es un elemento de entrada)
                self.logger.debug(f"\ninput_value no aplicable o falló para '{selector}'. Intentando text_content/inner_text. Error: {e_input}")
                
                # Si input_value falla, intentamos con text_content o inner_text para otros elementos (p. ej. <div>, <span>, <p>)
                try:
                    valor_extraido = selector.text_content() # Un timeout corto para text_content
                    # Si text_content devuelve solo espacios en blanco o es vacío,
                    # intentamos inner_text, que a veces es más preciso para texto renderizado visiblemente.
                    if valor_extraido is not None and valor_extraido.strip() == "":
                        valor_extraido = selector.inner_text() # Un timeout corto para inner_text
                        self.logger.debug(f"\nValor extraído (inner_text) de '{selector}': '{valor_extraido}' (después de text_content vacío).")
                    else:
                        self.logger.debug(f"\nValor extraído (text_content) de '{selector}': '{valor_extraido}'")
                except Error as e_text_inner:
                    self.logger.warning(f"\nNo se pudo extraer input_value, text_content ni inner_text de '{selector}'. Detalles: {e_text_inner}")
                    valor_extraido = None # Asegurarse de que sea None si todos los intentos fallan

            # 3. Procesar el valor extraído y registrar el rendimiento
            valor_final = None
            if valor_extraido is not None:
                # Eliminar espacios en blanco al inicio y al final si el valor es una cadena.
                valor_final = valor_extraido.strip() if isinstance(valor_extraido, str) else valor_extraido
                self.logger.info(f"\n✅ Valor final obtenido del elemento '{selector}': '{valor_final}'")
                self.tomar_captura(f"{nombre_base}_valor_extraido_exito", directorio)
            else:
                self.logger.warning(f"\n❌ No se pudo extraer ningún valor significativo del elemento '{selector}'.")
                self.tomar_captura(f"{nombre_base}_fallo_extraccion_valor_no_encontrado", directorio)
            
            # --- Medición de rendimiento: Fin de la extracción del valor ---
            end_time_extraction = time.time()
            duration_extraction = end_time_extraction - start_time_extraction
            self.logger.info(f"PERFORMANCE: Tiempo total de extracción del valor del elemento '{selector}': {duration_extraction:.4f} segundos.")

            return valor_final

        except TimeoutError as e:
            # Captura si el elemento no se vuelve visible o habilitado a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_extraction
            mensaje_error = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no se volvió visible/habilitado a tiempo "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_elemento}s) "
                f"para extraer su valor. Detalles: {e}"
            )
            self.logger.error(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_extraccion_valor", directorio)
            # Elevar AssertionError para indicar un fallo de prueba claro.
            raise AssertionError(f"\nElemento no disponible para extracción de valor: {selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con el DOM.
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True) # Nivel crítico para errores de Playwright.
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_extraccion_valor", directorio)
            raise AssertionError(f"\nError de Playwright al extraer valor: {selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_extraccion_valor", directorio)
            raise AssertionError(f"\nError inesperado al extraer valor: {selector}") from e

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Puedes eliminar esta espera si no es necesaria para la observación.
            if tiempo_espera_elemento > 0:
                self.esperar_fijo(tiempo_espera_elemento / 5.0) # Una pequeña espera al final.
        
    # 32- Función para verificar que los encabezados de las columnas de una tabla sean correctos y estén presentes, con pruebas de rendimiento
    def verificar_encabezados_tabla(self, tabla_selector: Locator, encabezados_esperados: List[str], nombre_base: str, directorio: str, tiempo_espera_tabla: Union[int, float] = 1.0) -> bool:
        """
        Verifica que los encabezados (<th>) de las columnas de una tabla HTML
        sean correctos y estén presentes en el orden esperado.
        Mide el rendimiento de esta operación de verificación.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` (o un elemento padre que contenga la tabla).
                                      Es crucial que sea un Locator para aprovechar sus
                                      funcionalidades de espera y contexto.
            encabezados_esperados (List[str]): Una **lista de cadenas de texto** que representan
                                               los encabezados esperados, en el orden en que
                                               deben aparecer en la tabla.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_tabla (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que la tabla y su sección de encabezado
                                                     (`<thead>` y `<th>`) estén visibles y listos.
                                                     Por defecto, `10.0` segundos.

        Returns:
            bool: `True` si todos los encabezados de la tabla coinciden con los esperados
                  en cantidad y contenido; `False` en caso contrario o si la tabla/encabezados
                  no están disponibles a tiempo.

        Raises:
            AssertionError: Si la tabla o sus elementos de encabezado no están disponibles
                            a tiempo, o si ocurre un error inesperado de Playwright o genérico
                            que impida la verificación.
        """
        self.logger.info(f"\n⚙️ Verificando encabezados de la tabla con selector '{tabla_selector}'...")
        self.logger.info(f"\nEncabezados esperados: {encabezados_esperados}. Tiempo máximo de espera: {tiempo_espera_tabla}s.")
        
        # --- Medición de rendimiento: Inicio de la verificación de encabezados ---
        # Registra el tiempo justo antes de iniciar cualquier interacción con la tabla.
        start_time_header_verification = time.time()

        try:
            # 1. Verificar la presencia y visibilidad de la tabla misma
            # Esto es crucial para asegurar que la tabla se ha cargado en el DOM.
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_tabla}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.debug(f"\nTabla resaltada para verificación: {tabla_selector}")

            # 2. Verificar la presencia y visibilidad del elemento thead (cabecera de la tabla)
            thead_locator = tabla_selector.locator("thead")
            self.logger.debug(f"\nEsperando que el thead de la tabla con selector '{tabla_selector} thead' esté visible (timeout: {tiempo_espera_tabla}s).")
            expect(thead_locator).to_be_visible()
            self.logger.info("\n✅ El elemento '<thead>' de la tabla es visible.")
            
            # 3. Obtener los locators de los encabezados (<th>) dentro del thead
            encabezados_actuales_locators = thead_locator.locator("th")
            self.logger.debug(f"\nEsperando que al menos un '<th>' dentro del '<thead>' sea visible (timeout: {tiempo_espera_tabla}s).")
            expect(encabezados_actuales_locators.first).to_be_visible()
            
            # Resaltar todos los encabezados encontrados para depuración visual.
            for i in range(encabezados_actuales_locators.count()):
                encabezados_actuales_locators.nth(i).highlight()
            self.tomar_captura(f"{nombre_base}_encabezados_encontrados_y_resaltados", directorio)

            num_encabezados_actuales = encabezados_actuales_locators.count()
            num_encabezados_esperados = len(encabezados_esperados)

            # 4. Comparar la cantidad de encabezados
            if num_encabezados_actuales != num_encabezados_esperados:
                actual_texts = [h.text_content().strip() for h in encabezados_actuales_locators.all_js_handles()] # Obtener todos los textos para el log de error
                self.logger.error(f"\n❌ --> FALLO: El número de encabezados '<th>' encontrados ({num_encabezados_actuales}) "
                                  f"no coincide con el número de encabezados esperados ({num_encabezados_esperados}).\n"
                                  f"Actuales: {actual_texts}\nEsperados: {encabezados_esperados}")
                self.tomar_captura(f"{nombre_base}_cantidad_encabezados_incorrecta", directorio)
                return False

            # 5. Iterar y comparar el texto de cada encabezado
            todos_correctos = True
            for i in range(num_encabezados_esperados):
                encabezado_locator = encabezados_actuales_locators.nth(i)
                # Obtenemos el texto de la celda del encabezado y eliminamos espacios en blanco.
                texto_encabezado_actual = encabezado_locator.text_content().strip()
                encabezado_esperado = encabezados_esperados[i]

                if texto_encabezado_actual == encabezado_esperado:
                    self.logger.info(f"\n ✅ Encabezado {i+1}: '{texto_encabezado_actual}' coincide con el esperado '{encabezado_esperado}'.")
                    # encabezado_locator.highlight() # Opcional: resaltar el encabezado individual si es necesario para cada uno.
                else:
                    self.logger.error(f"\n ❌ FALLO: Encabezado {i+1} esperado era '{encabezado_esperado}', pero se encontró '{texto_encabezado_actual}'.")
                    encabezado_locator.highlight() # Resaltar el encabezado incorrecto.
                    self.tomar_captura(f"{nombre_base}_encabezado_incorrecto_{i+1}", directorio)
                    todos_correctos = False
                    # No es necesario un time.sleep() aquí si solo queremos el log y la captura.

            # --- Medición de rendimiento: Fin de la verificación de encabezados ---
            end_time_header_verification = time.time()
            duration_header_verification = end_time_header_verification - start_time_header_verification
            self.logger.info(f"PERFORMANCE: Tiempo total de verificación de encabezados de tabla '{tabla_selector}': {duration_header_verification:.4f} segundos.")

            if todos_correctos:
                self.logger.info("\n✅ ÉXITO: Todos los encabezados de columna son correctos y están en el orden esperado.")
                self.tomar_captura(f"{nombre_base}_encabezados_verificados_ok", directorio)
                return True
            else:
                self.logger.error("\n❌ FALLO: Uno o más encabezados de columna son incorrectos o no están en el orden esperado.")
                self.tomar_captura(f"{nombre_base}_encabezados_verificados_fallo", directorio)
                return False

        except TimeoutError as e:
            # Captura si la tabla, el thead o los th no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_header_verification
            error_msg = (
                f"\n❌ FALLO (Timeout): La tabla o sus encabezados con el selector '{tabla_selector}' no se volvieron visibles a tiempo "
                f"después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_tabla}s).\n"
                f"Posiblemente la tabla no estuvo disponible a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_timeout", directorio)
            # Elevar AssertionError para que la prueba falle claramente cuando la tabla no está lista.
            raise AssertionError(f"\nElementos de encabezado de tabla no disponibles a tiempo: {tabla_selector}") from e

        except Error as e: # Catch Playwright-specific errors
            # Captura errores de Playwright que impiden la interacción con el DOM.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar verificar la tabla o sus encabezados con el selector '{tabla_selector}'.\n"
                f"Posibles causas: Selector inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Nivel crítico para errores de Playwright.
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al verificar encabezados de tabla: {tabla_selector}") from e # Relanzar.

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar los encabezados de la tabla con el selector '{tabla_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar encabezados de tabla: {tabla_selector}") from e # Relanzar.

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Puedes eliminar esta espera si no es necesaria para la observación.
            if tiempo_espera_tabla > 0:
                self.esperar_fijo(tiempo_espera_tabla / 5.0) # Una pequeña espera al final, por ejemplo.
        
    # 33- Función para verificar los datos de las filas de una tabla, con pruebas de rendimiento integradas.
    def verificar_datos_filas_tabla(self, tabla_selector: Locator, datos_filas_esperados: List[Dict[str, Union[str, bool, int, float]]], nombre_base: str, directorio: str, tiempo_espera_general: Union[int, float] = 0.5) -> bool:
        """
        Verifica que los datos de las filas de una tabla HTML coincidan con los datos esperados.
        La función compara el número de filas, el texto de las celdas y el estado de los checkboxes
        en columnas específicas. Mide el rendimiento de todo el proceso de verificación.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` que contiene las filas a verificar.
            datos_filas_esperados (List[Dict[str, Union[str, bool, int, float]]]): Una lista
                                      de diccionarios, donde cada diccionario representa una fila
                                      esperada. Las claves del diccionario deben ser los nombres
                                      de los encabezados de las columnas y los valores, los
                                      datos esperados para esa columna en la fila.
                                      Ejemplo: `[{'ID': '123', 'Name': 'Product A', 'Price': '10.50', 'Select': True}]`.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_general (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que la tabla, sus encabezados y
                                                        las filas estén visibles y listos para
                                                        la interacción. Por defecto, `15.0` segundos.

        Returns:
            bool: `True` si todos los datos de las filas y los estados de los checkboxes
                  coinciden con los valores esperados; `False` en caso contrario o si
                  la tabla/datos no están disponibles a tiempo.

        Raises:
            AssertionError: Si la tabla o sus elementos clave (encabezados, filas) no están
                            disponibles a tiempo, o si ocurre un error inesperado de Playwright
                            o genérico que impida la verificación.
        """
        self.logger.info(f"\n--- Iniciando verificación de datos de las filas de la tabla con locator '{tabla_selector}' ---")
        self.logger.info(f"\nNúmero de filas esperadas: {len(datos_filas_esperados)}")
        self.tomar_captura(f"{nombre_base}_inicio_verificacion_datos_filas", directorio)

        # --- Medición de rendimiento: Inicio de la verificación de datos de filas ---
        # Registra el tiempo justo antes de iniciar cualquier interacción con la tabla.
        start_time_row_data_verification = time.time()

        try:
            # 1. Asegurarse de que la tabla esté visible y disponible
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_general}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.info("\n✅ Tabla visible. Procediendo a verificar los datos.")

            # 2. Obtener los encabezados para mapear los índices de las columnas
            header_locators = tabla_selector.locator("thead th")
            self.logger.debug(f"\nEsperando que los encabezados (th) de la tabla sean visibles (timeout: {tiempo_espera_general}s).")
            expect(header_locators.first).to_be_visible()
            headers = [h.text_content().strip() for h in header_locators.all()]
            
            if not headers:
                self.logger.error(f"\n❌ --> FALLO: No se encontraron encabezados en la tabla con locator '{tabla_selector}'. No se pueden verificar los datos de las filas.")
                self.tomar_captura(f"{nombre_base}_no_headers_para_datos_filas", directorio)
                return False
            self.logger.info(f"\n🔍 Encabezados de la tabla encontrados: {headers}")

            # 3. Obtener todas las filas del cuerpo de la tabla (excluyendo thead)
            tbody_locator = tabla_selector.locator("tbody")
            self.logger.debug(f"\nEsperando que el tbody de la tabla sea visible (timeout: {tiempo_espera_general}s).")
            expect(tbody_locator).to_be_visible()

            row_locators = tbody_locator.locator("tr")
            # Esperar a que al menos la primera fila de datos sea visible si se esperan filas.
            if len(datos_filas_esperados) > 0:
                self.logger.debug(f"\nEsperando que al menos la primera fila de datos sea visible (timeout: {tiempo_espera_general}s).")
                expect(row_locators.first).to_be_visible()

            num_filas_actuales = row_locators.count()
            num_filas_esperadas = len(datos_filas_esperados)

            # 4. Comparar el número total de filas
            if num_filas_actuales == 0 and num_filas_esperadas == 0:
                self.logger.info("\n✅ ÉXITO: No se esperaban filas y no se encontraron filas en la tabla. Verificación completada.")
                self.tomar_captura(f"{nombre_base}_no_rows_expected_and_found", directorio)
                # No necesitamos detener el tiempo si no se hizo nada realmente.
                return True
            
            if num_filas_actuales != num_filas_esperadas:
                self.logger.error(f"\n❌ --> FALLO: El número de filas encontradas ({num_filas_actuales}) "
                                  f"no coincide con el número de filas esperadas ({num_filas_esperadas}).")
                self.tomar_captura(f"{nombre_base}_cantidad_filas_incorrecta", directorio)
                return False
            self.logger.info(f"\n🔍 Número de filas actual y esperado coinciden: {num_filas_actuales} filas.")

            # --- Variable principal para el retorno ---
            todos_los_datos_correctos = True 

            # 5. Iterar sobre cada fila esperada y verificar sus datos
            for i in range(num_filas_esperadas):
                fila_actual_locator = row_locators.nth(i)
                datos_fila_esperada = datos_filas_esperados[i]
                self.logger.info(f"\n  Verificando Fila {i+1} (Datos esperados: {datos_fila_esperada})...")
                fila_actual_locator.highlight() # Resaltar la fila actual en la captura para debug.

                # Bandera para saber si la fila actual tiene algún fallo
                fila_actual_correcta = True 

                # Iterar sobre las columnas esperadas para esta fila
                for col_name, expected_value in datos_fila_esperada.items():
                    try:
                        # Encontrar el índice de la columna por su nombre
                        if col_name not in headers:
                            self.logger.error(f"\n  ❌ FALLO: Columna '{col_name}' esperada para la Fila {i+1} no encontrada en los encabezados de la tabla. Encabezados actuales: {headers}")
                            self.tomar_captura(f"{nombre_base}_fila_{i+1}_columna_{col_name}_no_encontrada", directorio)
                            todos_los_datos_correctos = False # Falla general
                            fila_actual_correcta = False # Falla en esta fila
                            continue # Pasa a la siguiente columna esperada o fila

                        col_index = headers.index(col_name)
                        
                        # Localizar la celda específica (td) dentro de la fila por su índice
                        celda_locator = fila_actual_locator.locator("td").nth(col_index)
                        
                        # Asegurarse de que la celda esté visible antes de interactuar.
                        expect(celda_locator).to_be_visible() # Timeout para celda individual

                        if col_name == "Select": # Lógica específica para el checkbox en la columna "Select"
                            checkbox_locator = celda_locator.locator("input[type='checkbox']")
                            if checkbox_locator.count() == 0: # Si no se encuentra el checkbox dentro de la celda
                                self.logger.error(f"\n  ❌ FALLO: Checkbox no encontrado en la columna '{col_name}' de la Fila {i+1}.")
                                celda_locator.highlight() # Resaltar la celda donde se esperaba el checkbox
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_no_checkbox", directorio)
                                todos_los_datos_correctos = False
                                fila_actual_correcta = False
                            elif isinstance(expected_value, bool): # Si se espera un estado específico (True/False)
                                if checkbox_locator.is_checked() != expected_value:
                                    self.logger.error(f"\n  ❌ FALLO: El checkbox de la Fila {i+1}, Columna '{col_name}' estaba "
                                                      f"{'marcado' if checkbox_locator.is_checked() else 'desmarcado'}, se esperaba {'marcado' if expected_value else 'desmarcado'}.")
                                    checkbox_locator.highlight() # Resaltar el checkbox incorrecto
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_estado_incorrecto", directorio)
                                    todos_los_datos_correctos = False
                                    fila_actual_correcta = False
                                else:
                                    self.logger.info(f"\n  ✅ Fila {i+1}, Columna '{col_name}': Checkbox presente y estado correcto ({'marcado' if expected_value else 'desmarcado'}).")
                            else: # Si se espera que el checkbox exista, pero no se especificó un estado booleano
                                self.logger.info(f"\n  ✅ Fila {i+1}, Columna '{col_name}': Checkbox presente (estado no verificado explícitamente).")
                        else: # Para otras columnas de texto (no checkbox)
                            actual_value = celda_locator.text_content().strip()
                            # Aseguramos que expected_value también sea una cadena para la comparación, eliminando espacios.
                            if actual_value != str(expected_value).strip(): 
                                self.logger.error(f"\n  ❌ FALLO: Fila {i+1}, Columna '{col_name}'. Se esperaba '{expected_value}', se encontró '{actual_value}'.")
                                celda_locator.highlight() # Resaltar la celda con el dato incorrecto
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_col_{col_name}_incorrecta", directorio)
                                todos_los_datos_correctos = False
                                fila_actual_correcta = False
                            else:
                                self.logger.info(f"\n  ✅ Fila {i+1}, Columna '{col_name}': '{actual_value}' coincide con lo esperado.")
                        
                    except TimeoutError as cell_timeout_e:
                        self.logger.error(f"\n  ❌ FALLO (Timeout): La celda de la Fila {i+1}, Columna '{col_name}' no se volvió visible a tiempo. Detalles: {cell_timeout_e}")
                        self.tomar_captura(f"{nombre_base}_fila_{i+1}_col_{col_name}_timeout", directorio)
                        todos_los_datos_correctos = False
                        fila_actual_correcta = False
                    except Error as col_playwright_e:
                        self.logger.error(f"\n  ❌ FALLO (Playwright): Error de Playwright al verificar la columna '{col_name}' de la Fila {i+1}. Detalles: {col_playwright_e}")
                        self.tomar_captura(f"{nombre_base}_fila_{i+1}_col_{col_name}_playwright_error", directorio)
                        todos_los_datos_correctos = False
                        fila_actual_correcta = False
                    except Exception as col_e:
                        self.logger.error(f"\n  ❌ ERROR INESPERADO al verificar la columna '{col_name}' de la Fila {i+1}: {col_e}", exc_info=True)
                        self.tomar_captura(f"{nombre_base}_fila_{i+1}_col_{col_name}_error_inesperado", directorio)
                        todos_los_datos_correctos = False
                        fila_actual_correcta = False
                        # Podrías decidir si quieres continuar con el resto de las columnas/filas
                        # o si este error debe detener la verificación.

                # Pausa solo si la fila actual tuvo algún fallo para que la captura sea más útil
                if not fila_actual_correcta:
                    self.esperar_fijo(1) # Pausa de 1 segundo para visualización si hay un fallo en la fila.

            # --- Medición de rendimiento: Fin de la verificación de datos de filas ---
            end_time_row_data_verification = time.time()
            duration_row_data_verification = end_time_row_data_verification - start_time_row_data_verification
            self.logger.info(f"PERFORMANCE: Tiempo total de verificación de datos de filas en la tabla '{tabla_selector}': {duration_row_data_verification:.4f} segundos.")

            # --- Retorno final basado en el estado acumulado ---
            if todos_los_datos_correctos:
                self.logger.info("\n✅ ÉXITO: Todos los datos de las filas y checkboxes son correctos y están presentes.")
                self.tomar_captura(f"{nombre_base}_datos_filas_verificados_ok", directorio)
                return True
            else:
                self.logger.error("\n❌ FALLO: Uno o más datos de las filas o checkboxes son incorrectos o faltan.")
                self.tomar_captura(f"{nombre_base}_datos_filas_verificados_fallo", directorio)
                return False

        except TimeoutError as e:
            # Captura si la tabla, el thead, el tbody o las filas no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_row_data_verification
            error_msg = (
                f"\n❌ FALLO (Timeout): La tabla, sus encabezados o sus filas con el locator '{tabla_selector}' no se volvieron visibles a tiempo "
                f"después de {duration_fail:.4f} segundos (timeout general configurado: {tiempo_espera_general}s).\n"
                f"Posiblemente la tabla no estuvo disponible a tiempo o tardó demasiado en cargar su contenido.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_timeout", directorio)
            # Elevar AssertionError para que la prueba falle claramente cuando la tabla no está lista.
            raise AssertionError(f"\nElementos de tabla no disponibles a tiempo para verificación de datos de filas: {tabla_selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con el DOM de la tabla.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar verificar las filas con el locator '{tabla_selector}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al verificar datos de filas de tabla: {tabla_selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada durante la verificación.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error desconocido al verificar los datos de las filas con el locator '{tabla_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar datos de filas de tabla: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre, haya o no una excepción.
            # Puedes eliminar esta espera si no es necesaria para la observación al final de la ejecución de la función.
            self.esperar_fijo(1) # Pequeña espera final para observación.
    
    # 34- Función para seleccionar y verificar el estado de checkboxes de filas aleatorias, con pruebas de rendimiento.
    def seleccionar_y_verificar_checkboxes_aleatorios(self, tabla_selector: Locator, num_checkboxes_a_interactuar: int, nombre_base: str, directorio: str, tiempo_espera_tabla: Union[int, float] = 1.0, pausa_interaccion: Union[int, float] = 0.5) -> bool:
        """
        Selecciona y verifica el estado de un número específico de checkboxes aleatorios
        dentro de una tabla. Mide el rendimiento de las operaciones de búsqueda e interacción.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` que contiene los checkboxes a interactuar.
            num_checkboxes_a_interactuar (int): El **número de checkboxes aleatorios** a
                                                seleccionar y verificar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_tabla (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que la tabla y sus checkboxes estén
                                                     visibles y listos. Por defecto, `10.0` segundos.
            pausa_interaccion (Union[int, float]): **Pausa opcional** (en segundos) después de
                                                   cada interacción con un checkbox para permitir
                                                   que el DOM se actualice visualmente. Por defecto, `0.5` segundos.

        Returns:
            bool: `True` si todos los checkboxes seleccionados aleatoriamente fueron
                  interactuados y verificados correctamente; `False` en caso contrario.

        Raises:
            AssertionError: Si la tabla o sus checkboxes no están disponibles a tiempo,
                            o si ocurre un error inesperado de Playwright o genérico
                            que impida la interacción.
        """
        self.logger.info(f"\n--- Iniciando selección y verificación de {num_checkboxes_a_interactuar} checkbox(es) aleatorio(s) en la tabla con locator '{tabla_selector}' ---")
        self.tomar_captura(f"{nombre_base}_inicio_seleccion_checkbox", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que la tabla esté visible
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_tabla}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.info("\n✅ Tabla visible. Procediendo a buscar checkboxes.")

            # --- Medición de rendimiento: Inicio del descubrimiento de checkboxes ---
            start_time_discovery = time.time()

            # 2. Obtener todos los locators de los checkboxes en las celdas del cuerpo de la tabla
            all_checkbox_locators = tabla_selector.locator("tbody tr td input[type='checkbox']")
            
            # Asegurarse de que al menos un checkbox sea visible si esperamos interactuar.
            if num_checkboxes_a_interactuar > 0:
                self.logger.debug(f"\nEsperando que al menos un checkbox en la tabla sea visible (timeout: {tiempo_espera_tabla}s).")
                expect(all_checkbox_locators.first).to_be_visible()

            num_checkboxes_disponibles = all_checkbox_locators.count()

            # --- Medición de rendimiento: Fin del descubrimiento de checkboxes ---
            end_time_discovery = time.time()
            duration_discovery = end_time_discovery - start_time_discovery
            self.logger.info(f"PERFORMANCE: Tiempo de descubrimiento de checkboxes disponibles: {duration_discovery:.4f} segundos. ({num_checkboxes_disponibles} encontrados)")

            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{tabla_selector.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados", directorio)
                return False
            
            if num_checkboxes_a_interactuar <= 0:
                self.logger.warning("\n⚠️ ADVERTENCIA: El número de checkboxes a interactuar es 0 o negativo. No se realizará ninguna acción.")
                return True

            if num_checkboxes_a_interactuar > num_checkboxes_disponibles:
                self.logger.error(f"\n❌ --> FALLO: Se solicitaron {num_checkboxes_a_interactuar} checkboxes para interactuar, pero solo hay {num_checkboxes_disponibles} disponibles.")
                self.tomar_captura(f"{nombre_base}_no_suficientes_checkboxes", directorio)
                return False

            self.logger.info(f"\nSe encontraron {num_checkboxes_disponibles} checkboxes. Seleccionando {num_checkboxes_a_interactuar} aleatoriamente...")

            # 3. Seleccionar N índices de checkboxes aleatorios y únicos
            random_indices = random.sample(range(num_checkboxes_disponibles), num_checkboxes_a_interactuar)
            
            todos_correctos = True
            interaction_times = [] # Lista para almacenar tiempos de interacción individuales

            # 4. Iterar sobre los checkboxes seleccionados aleatoriamente e interactuar con ellos
            for i, idx in enumerate(random_indices):
                checkbox_to_interact = all_checkbox_locators.nth(idx)
                
                # --- Medición de rendimiento: Inicio de interacción individual ---
                start_time_interaction = time.time()

                # Resaltar el checkbox actual para la captura/visualización
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_checkbox_{i+1}_aleatorio_idx_{idx}_resaltado", directorio)
                self.esperar_fijo(pausa_interaccion) # Pausa para ver el resaltado

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                product_id = "N/A" # Default en caso de error
                try:
                    # Encontrar la fila que contiene este checkbox para obtener información de contexto.
                    # Esto es un poco más complejo si el checkbox no está en la primera columna,
                    # pero si asumimos que está dentro de un 'td' de un 'tr' que representa una fila:
                    # Podemos buscar el ancestro 'tr' y luego el primer 'td' de ese 'tr'.
                    # Podría ser más robusto si el product ID estuviera en un atributo de datos,
                    # o si el checkbox tuviera un atributo id/name relacionado con el producto.
                    row_locator_for_id = checkbox_to_interact.locator("..").locator("..") # Sube dos niveles para llegar al 'tr'
                    # Asegurarse de que el 'td' existe en la primera posición.
                    if row_locator_for_id.locator("td").count() > 0:
                        product_id = row_locator_for_id.locator("td").nth(0).text_content().strip()
                    else:
                        self.logger.warning(f"No se pudo extraer el ID del producto para la fila del checkbox en el índice {idx}. La primera celda (td[0]) no fue encontrada o no tiene texto.")
                except Exception as id_e:
                    self.logger.warning(f"Error al intentar obtener el ID del producto para el checkbox en el índice {idx}: {id_e}")
                
                initial_state = checkbox_to_interact.is_checked()
                self.logger.info(f"\n  Checkbox del Producto ID: {product_id} (Fila índice: {idx}, Interacción {i+1}/{num_checkboxes_a_interactuar}): Estado inicial {'MARCADO' if initial_state else 'DESMARCADO'}.")

                # --- Lógica para asegurar que el click lo deje en estado 'seleccionado' (marcado) ---
                if initial_state: # Si ya está marcado, lo desmarcamos primero para asegurar la acción de marcar
                    self.logger.info(f"\n  El checkbox del Producto ID: {product_id} ya está MARCADO. Haciendo clic para desmarcar antes de seleccionar.")
                    checkbox_to_interact.uncheck()
                    self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice

                    if checkbox_to_interact.is_checked(): # Si después de uncheck sigue marcado, es un fallo
                        self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no se desmarcó correctamente para la interacción.")
                        checkbox_to_interact.highlight()
                        self.tomar_captura(f"{nombre_base}_fila_{idx+1}_no_se_desmarco", directorio)
                        todos_correctos = False
                        # No es necesario continuar con la verificación de 'check' si el 'uncheck' ya falló.
                        # Continua al siguiente checkbox aleatorio.
                        continue 
                
                # Ahora el checkbox debería estar DESMARCADO (o siempre lo estuvo si initial_state era False)
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para MARCARLO...")
                checkbox_to_interact.check() # Marca el checkbox
                self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if not final_state: # Si no está marcado (seleccionado) después del clic
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a MARCADO después del clic. Sigue DESMARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{idx+1}_no_se_marco", directorio)
                    todos_correctos = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está MARCADO (seleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{idx+1}_marcado_ok", directorio)
                
                # --- Medición de rendimiento: Fin de interacción individual ---
                end_time_interaction = time.time()
                duration_interaction = end_time_interaction - start_time_interaction
                interaction_times.append(duration_interaction)
                self.logger.info(f"PERFORMANCE: Tiempo de interacción para checkbox {i+1} (Producto ID: {product_id}): {duration_interaction:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación de selección y verificación de checkboxes: {duration_total_operation:.4f} segundos.")

            if interaction_times:
                avg_interaction_time = sum(interaction_times) / len(interaction_times)
                self.logger.info(f"PERFORMANCE: Tiempo promedio de interacción por checkbox: {avg_interaction_time:.4f} segundos.")

            if todos_correctos:
                self.logger.info(f"\n✅ ÉXITO: Todos los {num_checkboxes_a_interactuar} checkbox(es) aleatorio(s) fueron seleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_seleccionados_ok", directorio)
                return True
            else:
                self.logger.error(f"\n❌ FALLO: Uno o más checkbox(es) aleatorio(s) no pudieron ser seleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_seleccion", directorio)
                return False

        except TimeoutError as e:
            # Captura si la tabla o los checkboxes no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{tabla_selector}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_tabla}s).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_timeout", directorio)
            raise AssertionError(f"\nElementos de tabla/checkboxes no disponibles a tiempo para interacción: {tabla_selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con los checkboxes.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al seleccionar y verificar checkboxes en la tabla '{tabla_selector}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al interactuar con checkboxes: {tabla_selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al seleccionar y verificar checkboxes aleatorios.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al interactuar con checkboxes: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre.
            self.esperar_fijo(1) # Pequeña espera final para observación.
    
    # 35- Función para seleccionar y verificar el estado de checkboxes de filas CONSECUTIVAS, con pruebas de rendimiento.
    def seleccionar_y_verificar_checkboxes_consecutivos(self, tabla_selector: Locator, start_index: int, num_checkboxes_a_interactuar: int, nombre_base: str, directorio: str, tiempo_espera_tabla: Union[int, float] = 1.0, pausa_interaccion: Union[int, float] = 0.5) -> bool:
        """
        Selecciona y verifica el estado de un número específico de checkboxes en filas consecutivas
        dentro de una tabla, comenzando desde un índice dado. Mide el rendimiento de las
        operaciones de búsqueda e interacción.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` que contiene los checkboxes a interactuar.
            start_index (int): El **índice de la primera fila** (basado en 0) donde se encuentra
                                el primer checkbox consecutivo a interactuar.
            num_checkboxes_a_interactuar (int): El **número de checkboxes consecutivos** a
                                                seleccionar y verificar a partir de `start_index`.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_tabla (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que la tabla y sus checkboxes estén
                                                     visibles y listos. Por defecto, `10.0` segundos.
            pausa_interaccion (Union[int, float]): **Pausa opcional** (en segundos) después de
                                                   cada interacción con un checkbox para permitir
                                                   que el DOM se actualice visualmente. Por defecto, `0.5` segundos.

        Returns:
            bool: `True` si todos los checkboxes consecutivos fueron interactuados y
                  verificados correctamente; `False` en caso contrario.

        Raises:
            AssertionError: Si la tabla o sus checkboxes no están disponibles a tiempo,
                            o si ocurre un error inesperado de Playwright o genérico
                            que impida la interacción.
        """
        self.logger.info(f"\n--- Iniciando selección y verificación de {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) "
                         f"a partir del índice {start_index} en la tabla con locator '{tabla_selector}' ---")
        self.tomar_captura(f"{nombre_base}_inicio_seleccion_consecutiva_checkbox", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que la tabla esté visible
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_tabla}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.info("\n✅ Tabla visible. Procediendo a buscar checkboxes.")

            # --- Medición de rendimiento: Inicio del descubrimiento de checkboxes ---
            start_time_discovery = time.time()

            # 2. Obtener todos los locators de los checkboxes en las celdas del cuerpo de la tabla
            all_checkbox_locators = tabla_selector.locator("tbody tr td input[type='checkbox']")
            
            # Asegurarse de que al menos un checkbox sea visible si esperamos interactuar.
            if num_checkboxes_a_interactuar > 0:
                self.logger.debug(f"\nEsperando que al menos el primer checkbox en el rango deseado sea visible (timeout: {tiempo_espera_tabla}s).")
                # Intentamos esperar al primer checkbox de la secuencia.
                if num_checkboxes_a_interactuar > 0 and start_index < all_checkbox_locators.count():
                    expect(all_checkbox_locators.nth(start_index)).to_be_visible()
                elif num_checkboxes_a_interactuar > 0: # Si el start_index es inválido, pero aún se esperan interacciones
                    # Esto será capturado por las validaciones de rango más adelante.
                    pass 

            num_checkboxes_disponibles = all_checkbox_locators.count()

            # --- Medición de rendimiento: Fin del descubrimiento de checkboxes ---
            end_time_discovery = time.time()
            duration_discovery = end_time_discovery - start_time_discovery
            self.logger.info(f"PERFORMANCE: Tiempo de descubrimiento de checkboxes disponibles: {duration_discovery:.4f} segundos. ({num_checkboxes_disponibles} encontrados)")

            # 3. Validaciones de precondición
            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{tabla_selector.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados_consec", directorio)
                return False
            
            if num_checkboxes_a_interactuar <= 0:
                self.logger.warning("\n⚠️ ADVERTENCIA: El número de checkboxes a interactuar es 0 o negativo. No se realizará ninguna acción.")
                return True # Consideramos éxito si no hay nada que hacer

            if start_index < 0 or start_index >= num_checkboxes_disponibles:
                self.logger.error(f"\n❌ --> FALLO: El 'posición de inicio' ({start_index}) está fuera del rango válido de checkboxes disponibles (0 a {num_checkboxes_disponibles - 1}).")
                self.tomar_captura(f"{nombre_base}_start_index_invalido_consec", directorio)
                return False
            
            if (start_index + num_checkboxes_a_interactuar) > num_checkboxes_disponibles:
                self.logger.error(f"\n❌ --> FALLO: Se solicitaron {num_checkboxes_a_interactuar} checkboxes a partir del índice {start_index}, "
                                  f"pero solo hay {num_checkboxes_disponibles} disponibles. El rango excede los límites de la tabla.")
                self.tomar_captura(f"{nombre_base}_rango_excedido_consec", directorio)
                return False

            self.logger.info(f"\nInteractuando con {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) "
                             f"desde el índice {start_index} hasta el {start_index + num_checkboxes_a_interactuar - 1}...")
            
            todos_correctos = True
            interaction_times = [] # Lista para almacenar tiempos de interacción individuales

            # 4. Iterar sobre los checkboxes consecutivos e interactuar con ellos
            for i in range(num_checkboxes_a_interactuar):
                current_idx = start_index + i
                checkbox_to_interact = all_checkbox_locators.nth(current_idx)
                
                # --- Medición de rendimiento: Inicio de interacción individual ---
                start_time_interaction = time.time()

                # Resaltar el checkbox actual para la captura/visualización
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_checkbox_consecutivo_{i+1}_idx_{current_idx}_resaltado", directorio)
                self.esperar_fijo(pausa_interaccion) # Pausa para ver el resaltado

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                product_id = "N/A" # Default en caso de error
                try:
                    # Se asume que el checkbox está dentro de un 'td' y este 'td' está dentro de un 'tr'.
                    # Se suben dos niveles para llegar al 'tr' y luego se busca el primer 'td'.
                    row_locator_for_id = checkbox_to_interact.locator("..").locator("..") 
                    if row_locator_for_id.locator("td").count() > 0:
                        product_id = row_locator_for_id.locator("td").nth(0).text_content().strip()
                    else:
                        self.logger.warning(f"No se pudo extraer el ID del producto para la fila del checkbox en el índice {current_idx}. La primera celda (td[0]) no fue encontrada o no tiene texto.")
                except Exception as id_e:
                    self.logger.warning(f"Error al intentar obtener el ID del producto para el checkbox en el índice {current_idx}: {id_e}")
                
                initial_state = checkbox_to_interact.is_checked()
                self.logger.info(f"\n  Checkbox del Producto ID: {product_id} (Fila índice: {current_idx}, Interacción {i+1}/{num_checkboxes_a_interactuar}): Estado inicial {'MARCADO' if initial_state else 'DESMARCADO'}.")

                # --- Lógica para asegurar que el click lo deje en estado 'seleccionado' (marcado) ---
                if initial_state: # Si ya está marcado, lo desmarcamos primero para asegurar la acción de marcar
                    self.logger.info(f"\n  El checkbox del Producto ID: {product_id} ya está MARCADO. Haciendo clic para desmarcar antes de seleccionar.")
                    checkbox_to_interact.uncheck()
                    self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice

                    if checkbox_to_interact.is_checked(): # Si después de uncheck sigue marcado, es un fallo
                        self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no se desmarcó correctamente para la interacción.")
                        checkbox_to_interact.highlight()
                        self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_no_se_desmarco_consec", directorio)
                        todos_correctos = False
                        # No es necesario continuar con la verificación de 'check' si el 'uncheck' ya falló.
                        continue 
                
                # Ahora el checkbox debería estar DESMARCADO (o siempre lo estuvo si initial_state era False)
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para MARCARLO...")
                checkbox_to_interact.check() # Marca el checkbox
                self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if not final_state: # Si no está marcado (seleccionado) después del clic
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a MARCADO después del clic. Sigue DESMARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_no_se_marco_consec", directorio)
                    todos_correctos = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está MARCADO (seleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_marcado_ok_consec", directorio)
                
                # --- Medición de rendimiento: Fin de interacción individual ---
                end_time_interaction = time.time()
                duration_interaction = end_time_interaction - start_time_interaction
                interaction_times.append(duration_interaction)
                self.logger.info(f"PERFORMANCE: Tiempo de interacción para checkbox {i+1} (Producto ID: {product_id}): {duration_interaction:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación de selección y verificación de checkboxes consecutivos: {duration_total_operation:.4f} segundos.")

            if interaction_times:
                avg_interaction_time = sum(interaction_times) / len(interaction_times)
                self.logger.info(f"PERFORMANCE: Tiempo promedio de interacción por checkbox: {avg_interaction_time:.4f} segundos.")

            if todos_correctos:
                self.logger.info(f"\n✅ ÉXITO: Todos los {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) fueron seleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_seleccionados_ok_consec", directorio)
                return True
            else:
                self.logger.error(f"\n❌ FALLO: Uno o más checkbox(es) consecutivo(s) no pudieron ser seleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_seleccion_consec", directorio)
                return False

        except TimeoutError as e:
            # Captura si la tabla o los checkboxes no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{tabla_selector}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_tabla}s).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_timeout", directorio)
            raise AssertionError(f"\nElementos de tabla/checkboxes no disponibles a tiempo para interacción: {tabla_selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con los checkboxes.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al seleccionar y verificar checkboxes consecutivos en la tabla '{tabla_selector}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al interactuar con checkboxes: {tabla_selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al seleccionar y verificar checkboxes consecutivos.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al interactuar con checkboxes: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre.
            self.esperar_fijo(1) # Pequeña espera final para observación.
        
    # 36- Función para deseleccionar todos los checkboxes actualmente marcados y verificar su estado.
    def deseleccionar_y_verificar_checkbox_marcado(self, tabla_selector: Locator, nombre_base: str, directorio: str, tiempo_espera_tabla: Union[int, float] = 1.0, pausa_interaccion: Union[int, float] = 0.5) -> bool:
        """
        Deselecciona y verifica el estado de **todos** los checkboxes que se encuentren
        actualmente marcados dentro de una tabla específica. Mide el rendimiento de
        las operaciones de búsqueda y deselección.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` que contiene los checkboxes a interactuar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_tabla (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que la tabla y sus checkboxes estén
                                                     visibles y listos. Por defecto, `10.0` segundos.
            pausa_interaccion (Union[int, float]): **Pausa opcional** (en segundos) después de
                                                   cada deselección con un checkbox para permitir
                                                   que el DOM se actualice visualmente. Por defecto, `0.5` segundos.

        Returns:
            bool: `True` si todos los checkboxes que estaban marcados fueron deseleccionados
                  y verificados correctamente; `False` en caso contrario.

        Raises:
            AssertionError: Si la tabla o sus checkboxes no están disponibles a tiempo,
                            o si ocurre un error inesperado de Playwright o genérico
                            que impida la interacción.
        """
        self.logger.info(f"\n--- Iniciando deselección y verificación de TODOS los checkboxes marcados "
                         f"en la tabla con locator '{tabla_selector}' ---")
        self.tomar_captura(f"{nombre_base}_inicio_deseleccion_todos_marcados", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que la tabla esté visible
            self.logger.debug(f"\nEsperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_tabla}s).")
            expect(tabla_selector).to_be_visible()
            tabla_selector.highlight()
            self.logger.info("\n✅ Tabla visible. Procediendo a buscar checkboxes.")

            # --- Medición de rendimiento: Inicio del descubrimiento de checkboxes ---
            start_time_discovery = time.time()

            # 2. Obtener todos los locators de los checkboxes en las celdas de la tabla
            all_checkbox_locators = tabla_selector.locator("tbody tr td input[type='checkbox']")
            
            # Asegurarse de que al menos un checkbox sea visible si esperamos interactuar (si no hay ninguno, lo gestionamos)
            if all_checkbox_locators.count() > 0:
                self.logger.debug(f"\nEsperando que al menos un checkbox en la tabla sea visible (timeout: {tiempo_espera_tabla}s).")
                expect(all_checkbox_locators.first).to_be_visible()

            num_checkboxes_disponibles = all_checkbox_locators.count()

            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{tabla_selector.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados_todos", directorio)
                return False
            
            # 3. Recolectar todos los checkboxes que están actualmente marcados para deseleccionar
            checkboxes_to_deselect = []
            for i in range(num_checkboxes_disponibles):
                checkbox = all_checkbox_locators.nth(i)
                if checkbox.is_checked():
                    checkboxes_to_deselect.append({"locator": checkbox, "original_index": i})
            
            # --- Medición de rendimiento: Fin del descubrimiento de checkboxes ---
            end_time_discovery = time.time()
            duration_discovery = end_time_discovery - start_time_discovery
            self.logger.info(f"PERFORMANCE: Tiempo de descubrimiento de checkboxes y filtrado de marcados: {duration_discovery:.4f} segundos. ({len(checkboxes_to_deselect)} marcados encontrados de {num_checkboxes_disponibles} disponibles)")

            if not checkboxes_to_deselect:
                self.logger.warning("\n⚠️ ADVERTENCIA: No se encontró ningún checkbox actualmente MARCADO en la tabla para deseleccionar. La función finaliza sin acciones de deselección.")
                self.tomar_captura(f"{nombre_base}_no_marcados_para_deseleccionar", directorio)
                return True # Consideramos éxito si no hay nada que deseleccionar

            self.logger.info(f"\nSe encontraron {len(checkboxes_to_deselect)} checkbox(es) marcado(s) para deseleccionar. Iniciando el proceso...")

            todos_deseleccionados_correctamente = True
            interaction_times = [] # Lista para almacenar tiempos de interacción individuales

            # 4. Iterar sobre los checkboxes marcados y deseleccionarlos
            for i, checkbox_info in enumerate(checkboxes_to_deselect):
                checkbox_to_interact = checkbox_info["locator"]
                original_idx = checkbox_info["original_index"]
                
                # --- Medición de rendimiento: Inicio de interacción individual ---
                start_time_interaction = time.time()

                # Resaltar el checkbox actual
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_deseleccion_actual_{i+1}_idx_{original_idx}_resaltado", directorio)
                self.esperar_fijo(pausa_interaccion)

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                product_id = "N/A" # Default en caso de error
                try:
                    # Se asume que el checkbox está dentro de un 'td' y este 'td' dentro de un 'tr'.
                    # Se suben dos niveles para llegar al 'tr' y luego se busca el primer 'td'.
                    row_locator_for_id = checkbox_to_interact.locator("..").locator("..")
                    if row_locator_for_id.locator("td").count() > 0:
                        product_id = row_locator_for_id.locator("td").nth(0).text_content().strip()
                    else:
                        self.logger.warning(f"No se pudo extraer el ID del producto para la fila del checkbox en el índice {original_idx}. La primera celda (td[0]) no fue encontrada o no tiene texto.")
                except Exception as id_e:
                    self.logger.warning(f"Error al intentar obtener el ID del producto para el checkbox en el índice {original_idx}: {id_e}")
                
                self.logger.info(f"\n  Procesando checkbox del Producto ID: {product_id} (Fila índice: {original_idx}, Interacción {i+1}/{len(checkboxes_to_deselect)}). Estado inicial: MARCADO (esperado).")

                # --- Interacción: Clic para deseleccionar ---
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para DESMARCARLO...")
                # Usar .uncheck() es más directo para desmarcar que .click() si ya sabes el estado esperado.
                checkbox_to_interact.uncheck()
                self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if final_state: # Si sigue marcado después de .uncheck()
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a DESMARCADO después del clic. Sigue MARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{original_idx+1}_no_desmarcado", directorio)
                    todos_deseleccionados_correctamente = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está DESMARCADO (deseleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{original_idx+1}_desmarcado_ok", directorio)
                
                # --- Medición de rendimiento: Fin de interacción individual ---
                end_time_interaction = time.time()
                duration_interaction = end_time_interaction - start_time_interaction
                interaction_times.append(duration_interaction)
                self.logger.info(f"PERFORMANCE: Tiempo de deselección para checkbox {i+1} (Producto ID: {product_id}): {duration_interaction:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación de deselección y verificación de checkboxes: {duration_total_operation:.4f} segundos.")

            if interaction_times:
                avg_interaction_time = sum(interaction_times) / len(interaction_times)
                self.logger.info(f"PERFORMANCE: Tiempo promedio de deselección por checkbox: {avg_interaction_time:.4f} segundos.")


            if todos_deseleccionados_correctamente:
                self.logger.info(f"\n✅ ÉXITO: Todos los {len(checkboxes_to_deselect)} checkbox(es) marcados fueron deseleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_deseleccionados_ok", directorio)
                return True
            else:
                self.logger.error(f"\n❌ FALLO: Uno o más checkbox(es) marcados no pudieron ser deseleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_deseleccion_todos", directorio)
                return False

        except TimeoutError as e:
            # Captura si la tabla o los checkboxes no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{tabla_selector}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo después de {duration_fail:.4f} segundos (timeout configurado: {tiempo_espera_tabla}s).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_timeout", directorio)
            raise AssertionError(f"\nElementos de tabla/checkboxes no disponibles a tiempo para interacción: {tabla_selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con los checkboxes.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al deseleccionar y verificar todos los checkboxes marcados en la tabla '{tabla_selector}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al interactuar con checkboxes: {tabla_selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al deseleccionar y verificar todos los checkboxes marcados.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al interactuar con checkboxes: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre.
            self.esperar_fijo(1) # Pequeña espera final para observación.
    
    # 37- Función para buscar un 'texto_a_buscar' en las celdas de una tabla (tbody) y, si lo encuentra,
    # intenta marcar el checkbox asociado en la misma fila. Incluye pruebas de rendimiento.
    def seleccionar_checkbox_por_contenido_celda(self, tabla_selector: Locator, texto_a_buscar: str, nombre_base: str, directorio: str, case_sensitive: bool = False, tiempo_espera_tabla: Union[int, float] = 1.0, pausa_interaccion: Union[int, float] = 0.5) -> bool:
        """
        Busca un 'texto_a_buscar' en todas las celdas (<td>) del cuerpo de una tabla (<tbody>).
        Si encuentra el texto en una celda, intenta localizar y marcar el checkbox
        asociado en la misma fila. Mide el rendimiento de la búsqueda y la interacción.

        Args:
            tabla_selector (Locator): El **Locator de Playwright** que representa el elemento
                                      `<table>` que contiene las filas y checkboxes.
            texto_a_buscar (str): El **texto exacto o parcial** a buscar dentro de las celdas de la tabla.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            case_sensitive (bool): Si es `True`, la búsqueda de texto será **sensible a mayúsculas y minúsculas**.
                                   Por defecto, `False` (insensible).
            tiempo_espera_tabla (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que la tabla esté visible y cargada.
                                                     Por defecto, `10.0` segundos.
            pausa_interaccion (Union[int, float]): **Pausa opcional** (en segundos) después de
                                                   resaltar la fila y de marcar el checkbox,
                                                   para permitir la actualización visual. Por defecto, `0.5` segundos.

        Returns:
            bool: `True` si se encontró al menos una coincidencia y se pudo marcar un checkbox asociado;
                  `False` si no se encontraron coincidencias o si hubo errores críticos.

        Raises:
            AssertionError: Si la tabla no está disponible a tiempo, o si ocurre un error
                            inesperado de Playwright o genérico durante la interacción.
        """
        self.logger.info(f"\n--- Iniciando búsqueda de '{texto_a_buscar}' en la tabla '{tabla_selector}' para marcar checkboxes ---")
        self.tomar_captura(f"{nombre_base}_inicio_busqueda_celdas", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que la tabla está visible y cargada
            self.logger.debug(f"Esperando que la tabla con selector '{tabla_selector}' esté visible (timeout: {tiempo_espera_tabla}s).")
            # Convertir timeout de segundos a milisegundos para expect()
            expect(tabla_selector).to_be_visible() 
            tabla_selector.highlight()
            self.logger.info("\n✅ Tabla visible. Comenzando a iterar por filas y celdas.")

            # --- Medición de rendimiento: Inicio del escaneo de la tabla ---
            start_time_scan = time.time()

            # Obtener todas las filas del cuerpo de la tabla
            filas = tabla_selector.locator("tbody tr")
            num_filas = filas.count()

            if num_filas == 0:
                self.logger.error(f"\n❌ --> FALLO: No se encontraron filas en el 'tbody' de la tabla con locator '{tabla_selector}'.")
                self.tomar_captura(f"{nombre_base}_no_filas_encontradas", directorio)
                return False

            self.logger.info(f"\nSe encontraron {num_filas} filas en la tabla. Iniciando escaneo de celdas...")
            
            checkboxes_marcados_exitosamente = 0
            
            # Normalizar el texto de búsqueda si no es sensible a mayúsculas/minúsculas
            search_text_normalized = texto_a_buscar if case_sensitive else texto_a_buscar.lower()
            
            found_any_match = False # Bandera para saber si se encontró al menos una coincidencia
            interaction_times = [] # Para medir el tiempo de marcado de cada checkbox

            for i in range(num_filas):
                fila_actual = filas.nth(i)
                # Obtener todas las celdas (td) de la fila actual
                celdas = fila_actual.locator("td")
                num_celdas = celdas.count()

                if num_celdas == 0:
                    self.logger.warning(f"\n  ADVERTENCIA: La fila {i+1} no contiene celdas (td). Saltando.")
                    continue

                celda_encontrada_en_fila = False
                for j in range(num_celdas):
                    celda_actual = celdas.nth(j)
                    celda_texto = celda_actual.text_content().strip()
                    
                    # Normalizar el texto de la celda para la comparación
                    celda_texto_normalized = celda_texto if case_sensitive else celda_texto.lower()

                    if search_text_normalized in celda_texto_normalized:
                        self.logger.info(f"\n  ✅ Coincidencia encontrada en Fila {i+1}, Celda {j+1}: '{celda_texto}' contiene '{texto_a_buscar}'.")
                        celda_encontrada_en_fila = True
                        found_any_match = True
                        
                        # Buscar el checkbox dentro de la misma fila
                        checkbox_locator = fila_actual.locator("input[type='checkbox']")
                        
                        if checkbox_locator.count() > 0:
                            checkbox = checkbox_locator.first
                            checkbox.highlight()
                            self.tomar_captura(f"{nombre_base}_fila_{i+1}_coincidencia_resaltada", directorio)
                            self.esperar_fijo(pausa_interaccion)

                            # --- Medición de rendimiento: Inicio de interacción de checkbox ---
                            start_time_checkbox_interaction = time.time()

                            if not checkbox.is_checked():
                                self.logger.info(f"\n  --> Marcando checkbox en Fila {i+1} (texto '{celda_texto}')...")
                                checkbox.check()
                                self.esperar_fijo(pausa_interaccion) # Pausa para que el DOM se actualice
                                
                                if checkbox.is_checked():
                                    self.logger.info(f"\n  ✅ Checkbox en Fila {i+1} marcado correctamente.")
                                    checkboxes_marcados_exitosamente += 1
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_marcado", directorio)
                                else:
                                    self.logger.error(f"\n  ❌ FALLO: No se pudo marcar el checkbox en Fila {i+1} (texto '{celda_texto}').")
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_no_marcado", directorio)
                            else:
                                self.logger.warning(f"\n  ⚠️ Checkbox en Fila {i+1} (texto '{celda_texto}') ya estaba marcado. No se requiere acción.")
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_ya_marcado", directorio)
                            
                            # --- Medición de rendimiento: Fin de interacción de checkbox ---
                            end_time_checkbox_interaction = time.time()
                            duration_checkbox_interaction = end_time_checkbox_interaction - start_time_checkbox_interaction
                            interaction_times.append(duration_checkbox_interaction)
                            self.logger.info(f"PERFORMANCE: Tiempo de interacción con checkbox en Fila {i+1}: {duration_checkbox_interaction:.4f} segundos.")

                        else:
                            self.logger.warning(f"\n  ⚠️ ADVERTENCIA: No se encontró un checkbox en la Fila {i+1} a pesar de la coincidencia del texto.")
                        break # Salir del bucle de celdas una vez encontrada la coincidencia en la fila

                if not celda_encontrada_en_fila:
                    self.logger.debug(f"\n  No se encontró '{texto_a_buscar}' en la Fila {i+1}. Continuando con la siguiente fila.")

            # --- Medición de rendimiento: Fin del escaneo de la tabla ---
            end_time_scan = time.time()
            duration_scan = end_time_scan - start_time_scan
            self.logger.info(f"PERFORMANCE: Tiempo total de escaneo de {num_filas} filas en la tabla: {duration_scan:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (búsqueda y marcado): {duration_total_operation:.4f} segundos.")

            if interaction_times:
                avg_interaction_time = sum(interaction_times) / len(interaction_times)
                self.logger.info(f"PERFORMANCE: Tiempo promedio de marcado por checkbox: {avg_interaction_time:.4f} segundos.")


            if checkboxes_marcados_exitosamente > 0:
                self.logger.info(f"\n✅ ÉXITO: Se marcaron {checkboxes_marcados_exitosamente} checkbox(es) basados en la búsqueda de '{texto_a_buscar}'.")
                self.tomar_captura(f"{nombre_base}_busqueda_finalizada_exito", directorio)
                return True
            elif found_any_match and checkboxes_marcados_exitosamente == 0:
                 self.logger.warning(f"\n⚠️ ADVERTENCIA: Se encontraron coincidencias para '{texto_a_buscar}', pero no se pudo marcar ningún checkbox. Posiblemente ya estaban marcados o hubo un problema al interactuar.")
                 self.tomar_captura(f"{nombre_base}_busqueda_finalizada_coincidencia_sin_marcados", directorio)
                 return True # Consideramos éxito si se encontró la coincidencia, aunque no se marcaran nuevos.
            else:
                self.logger.warning(f"\n⚠️ ADVERTENCIA: No se encontraron coincidencias para '{texto_a_buscar}' en ninguna celda de la tabla.")
                self.tomar_captura(f"{nombre_base}_busqueda_finalizada_sin_coincidencias", directorio)
                return False # Falla si no se encuentra ninguna coincidencia.

        except TimeoutError as e:
            # Captura si la tabla no se vuelve visible a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): La tabla con el locator '{tabla_selector}' no estuvo visible a tiempo (timeout configurado: {tiempo_espera_tabla}s).\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_timeout_tabla", directorio)
            raise AssertionError(f"\nTabla no disponible a tiempo: {tabla_selector}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con la tabla o checkboxes.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error al interactuar con la tabla o los checkboxes.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright durante la búsqueda/marcado: {tabla_selector}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado durante la búsqueda y marcado de checkboxes.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado durante la búsqueda/marcado: {tabla_selector}") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(1) # Pequeña espera final para observación o para liberar recursos si es necesario.
        
    # 38- Función para verificar que la página inicial esperada esté seleccionada y resaltada en un componente de paginación.
    # Incluye pruebas de rendimiento.
    def verificar_pagina_inicial_seleccionada(self, selector_paginado: Locator, texto_pagina_inicial: str, nombre_base: str, directorio: str, clase_resaltado: str = "active", tiempo_espera_componente: Union[int, float] = 1.0) -> bool:
        """
        Verifica que la página inicial esperada esté seleccionada y correctamente resaltada
        dentro de un componente de paginación. Mide el rendimiento de la localización y verificación.

        Args:
            selector_paginado (Locator): El **Locator de Playwright** que representa el
                                         contenedor principal del componente de paginación.
                                         (e.g., un `<div>` o `<nav>` que encierra el paginador).
            texto_pagina_inicial (str): El **texto exacto** de la página que se espera que sea
                                        la página inicial seleccionada (ej. "1", "Inicio").
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            clase_resaltado (str): La **clase CSS** que indica que un elemento de paginación
                                   está activo/seleccionado (ej. "active", "selected", "current-page").
                                   Por defecto, "active".
            tiempo_espera_componente (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                         para que el componente de paginación y
                                                         el elemento de la página inicial estén visibles.
                                                         Por defecto, `10.0` segundos.

        Returns:
            bool: `True` si la página inicial esperada está visible y tiene la clase de resaltado;
                  `False` en caso contrario.

        Raises:
            AssertionError: Si el componente de paginación o el elemento de la página inicial
                            no están disponibles a tiempo, o si ocurre un error inesperado
                            de Playwright o genérico.
        """
        self.logger.info(f"\n--- Iniciando verificación del estado inicial de la paginación ---")
        self.logger.info(f"\nContenedor de paginación locator: '{selector_paginado}'")
        self.logger.info(f"P\nágina inicial esperada: '{texto_pagina_inicial}'")
        self.tomar_captura(f"{nombre_base}_inicio_verificacion_paginacion", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que el contenedor de paginación esté visible
            self.logger.debug(f"\nEsperando que el contenedor de paginación '{selector_paginado}' esté visible (timeout: {tiempo_espera_componente}s).")
            # Convertir tiempo_espera_componente de segundos a milisegundos para expect()
            expect(selector_paginado).to_be_visible()
            selector_paginado.highlight()
            self.logger.info("\n✅ Contenedor de paginación visible. Procediendo a verificar la página inicial.")

            # --- Medición de rendimiento: Inicio de localización de la página inicial ---
            start_time_locator_page = time.time()

            # 2. Intentar encontrar el elemento de la página inicial por su texto dentro del contenedor
            # Se usa text= para una coincidencia exacta del texto visible del número de página.
            # Es crucial que el selector apunte al elemento que realmente tiene el texto de la página (ej. un <a> o <span> dentro de un <li>).
            # Si el texto '1' aparece en otros lugares, puede ser necesario un selector más específico,
            # como `selector_paginado.locator(f"li > a:has-text('{texto_pagina_inicial}')")` o similar.
            pagina_inicial_locator = selector_paginado.locator(f"text='{texto_pagina_inicial}'").first

            # Esperar a que el elemento de la página inicial esté visible y sea interactuable
            self.logger.debug(f"\nEsperando que el elemento de la página inicial '{texto_pagina_inicial}' esté visible (timeout: {tiempo_espera_componente}s).")
            expect(pagina_inicial_locator).to_be_visible()
            self.logger.info(f"\n✅ Elemento para la página '{texto_pagina_inicial}' encontrado y visible.")

            # --- Medición de rendimiento: Fin de localización de la página inicial ---
            end_time_locator_page = time.time()
            duration_locator_page = end_time_locator_page - start_time_locator_page
            self.logger.info(f"PERFORMANCE: Tiempo de localización del elemento de la página inicial: {duration_locator_page:.4f} segundos.")

            # --- Medición de rendimiento: Inicio de verificación de estado ---
            start_time_verification = time.time()

            # 3. Verificar que la página inicial esperada esté seleccionada (marcada con la clase de resaltado)
            self.logger.info(f"\nVerificando si la página '{texto_pagina_inicial}' tiene la clase de resaltado esperada '{clase_resaltado}'...")
            pagina_inicial_locator.highlight() # Resaltar el elemento para la captura visual
            self.tomar_captura(f"{nombre_base}_pagina_inicial_encontrada_resaltada", directorio)

            # Obtener todas las clases del elemento y verificar si la clase de resaltado está presente
            current_classes_attribute = pagina_inicial_locator.get_attribute("class")
            
            # Un elemento puede no tener atributo 'class' o puede ser una cadena vacía
            if current_classes_attribute is not None:
                current_classes_list = current_classes_attribute.split()
            else:
                current_classes_list = [] # Si no hay atributo 'class', la lista está vacía

            if clase_resaltado in current_classes_list:
                self.logger.info(f"\n  ✅ ÉXITO: La página '{texto_pagina_inicial}' está seleccionada y resaltada con la clase '{clase_resaltado}'.")
                self.tomar_captura(f"{nombre_base}_pagina_inicial_seleccionada_ok", directorio)
                success = True
            else:
                self.logger.error(f"\n  ❌ FALLO: La página '{texto_pagina_inicial}' no tiene la clase de resaltado esperada '{clase_resaltado}'.")
                self.logger.info(f"\n  Clases actuales del elemento: '{current_classes_attribute}'")
                self.tomar_captura(f"{nombre_base}_pagina_inicial_no_resaltada", directorio)
                success = False
            
            # --- Medición de rendimiento: Fin de verificación de estado ---
            end_time_verification = time.time()
            duration_verification = end_time_verification - start_time_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de la clase de resaltado: {duration_verification:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de paginación inicial): {duration_total_operation:.4f} segundos.")

            return success

        except TimeoutError as e:
            # Captura si el contenedor de paginación o el elemento de la página inicial no se vuelven visibles a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): El contenedor de paginación '{selector_paginado}' "
                f"o la página inicial '{texto_pagina_inicial}' no estuvieron visibles a tiempo "
                f"(timeout configurado: {tiempo_espera_componente}s).\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_timeout_paginacion", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nComponente de paginación o página inicial no disponibles a tiempo: {selector_paginado}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con el DOM.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error al interactuar con el componente de paginación.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError de Playwright al verificar paginación: {selector_paginado}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la paginación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError inesperado al verificar paginación: {selector_paginado}") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.5) # Pequeña espera final para observación o para liberar recursos si es necesario.
        
    # 39- Función para navegar a un número de página específico en un componente de paginación y verificar su estado.
    # Incluye pruebas de rendimiento.
    def navegar_y_verificar_pagina(self, selector_paginado: Locator, numero_pagina_a_navegar: str, nombre_base: str, directorio: str, clase_resaltado: str = "active", tiempo_espera_componente: Union[int, float] = 1.0, pausa_post_clic: Union[int, float] = 0.5) -> bool:
        """
        Navega a un número de página específico en un componente de paginación haciendo clic en el enlace
        correspondiente y verifica que la página de destino esté seleccionada y resaltada.
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            selector_paginado (Locator): El **Locator de Playwright** que representa el
                                         contenedor principal del componente de paginación.
                                         (e.g., un `<div>` o `<nav>` que encierra el paginador).
            numero_pagina_a_navegar (str): El **número de la página** a la que se desea navegar (como cadena).
                                          Ej. "2", "5".
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            clase_resaltado (str): La **clase CSS** que indica que un elemento de paginación
                                   está activo/seleccionado (ej. "active", "selected", "current-page").
                                   Por defecto, "active".
            tiempo_espera_componente (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                         para que el componente de paginación y
                                                         los elementos de página estén visibles.
                                                         Por defecto, `10.0` segundos.
            pausa_post_clic (Union[int, float]): **Pausa opcional** (en segundos) después de
                                                  hacer clic en un número de página, para permitir
                                                  que la página cargue y los estilos se apliquen.
                                                  Por defecto, `0.5` segundos.

        Returns:
            bool: `True` si la navegación fue exitosa y la página de destino está resaltada;
                  `False` en caso contrario.

        Raises:
            AssertionError: Si el componente de paginación o el elemento de la página de destino
                            no están disponibles a tiempo, o si ocurre un error inesperado
                            de Playwright o genérico.
        """
        self.logger.info(f"\n--- Iniciando navegación a la página '{numero_pagina_a_navegar}' y verificación de resaltado ---")
        self.logger.info(f"\nContenedor de paginación locator: '{selector_paginado}'")
        self.tomar_captura(f"{nombre_base}_inicio_navegacion_pagina_{numero_pagina_a_navegar}", directorio)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que el contenedor de paginación está visible
            self.logger.debug(f"\nEsperando que el contenedor de paginación '{selector_paginado}' esté visible (timeout: {tiempo_espera_componente}s).")
            # Convertir tiempo_espera_componente de segundos a milisegundos para expect()
            expect(selector_paginado).to_be_visible()
            selector_paginado.highlight()
            self.logger.info("\n✅ Contenedor de paginación visible. Procediendo.")

            # --- Medición de rendimiento: Inicio detección de página actual y total ---
            start_time_detection = time.time()

            # Obtener la página actualmente seleccionada
            # Este locator debería apuntar al elemento que realmente tiene la clase 'active'
            # y cuyo texto es el número de página (ej. un <a> dentro de un <li>)
            pagina_actual_locator = selector_paginado.locator(f"a.{clase_resaltado}").first
            # O si el <li> es el que tiene la clase, y necesitas el texto del <a> dentro:
            # pagina_actual_locator = selector_paginado.locator(f"li.{clase_resaltado} a").first

            # Usar .is_visible() y .text_content() para obtener el texto de forma segura
            pagina_actual_texto = "Desconocida"
            if pagina_actual_locator.count() > 0 and pagina_actual_locator.is_visible():
                pagina_actual_texto = pagina_actual_locator.text_content().strip()
            self.logger.info(f"\n  Página actualmente seleccionada detectada: {pagina_actual_texto}")

            # Calcular el número total de páginas disponibles
            # Asumimos que los elementos de paginación son 'li' y que el último 'li' visible
            # que contenga un número representa la última página.
            todos_los_botones_pagina = selector_paginado.locator("li")
            num_botones = todos_los_botones_pagina.count()
            
            total_paginas = 0
            if num_botones > 0:
                for i in range(num_botones - 1, -1, -1): # Iterar al revés para encontrar el último número
                    btn = todos_los_botones_pagina.nth(i)
                    btn_text = btn.text_content().strip()
                    if btn_text.isdigit(): # Si el texto es un número válido
                        total_paginas = int(btn_text)
                        break
            
            self.logger.info(f"\n  Número total de páginas detectadas: {total_paginas}")
            
            # --- Medición de rendimiento: Fin detección de página actual y total ---
            end_time_detection = time.time()
            duration_detection = end_time_detection - start_time_detection
            self.logger.info(f"PERFORMANCE: Tiempo de detección de página actual y total: {duration_detection:.4f} segundos.")

            # 2. Validaciones previas a la navegación
            try:
                # Convertir a int para comparaciones numéricas
                num_pagina_int = int(numero_pagina_a_navegar)
                pagina_actual_int = int(pagina_actual_texto) if pagina_actual_texto.isdigit() else -1 # Usar -1 si es desconocido
            except ValueError:
                self.logger.error(f"\n❌ FALLO: El número de página a navegar '{numero_pagina_a_navegar}' no es un número válido.")
                self.tomar_captura(f"{nombre_base}_pagina_destino_invalida", directorio)
                return False

            # Condicional 1: Página de destino es mayor que el total de páginas
            if total_paginas > 0 and num_pagina_int > total_paginas:
                self.logger.warning(f"\n⚠️ ADVERTENCIA: La página de destino '{numero_pagina_a_navegar}' es mayor que el número total de páginas disponibles '{total_paginas}'.")
                self.tomar_captura(f"{nombre_base}_pagina_destino_fuera_rango", directorio)
                return False # Considerar como fallo si la página está fuera de rango

            # Condicional 2: La página de destino es la misma que la página actual
            if num_pagina_int == pagina_actual_int:
                self.logger.warning(f"\n⚠️ ADVERTENCIA: Ya estás en la página '{numero_pagina_a_navegar}'. No se requiere navegación.")
                # Opcional: Podrías verificar de nuevo que siga resaltada, pero si ya estaba, no es una "navegación".
                self.tomar_captura(f"{nombre_base}_pagina_destino_actual", directorio)
                return True # Considerar esto un éxito, ya que el estado es el esperado.

            # 3. Encontrar y hacer clic en el botón de la página deseada
            # Este selector busca un 'a' dentro de un 'li' que contenga el texto del número de página.
            # Ajusta esto si tu estructura HTML es diferente (ej. si el número está directamente en el 'li').
            pagina_destino_locator = selector_paginado.locator(
                f"li:has-text('{numero_pagina_a_navegar}') a" 
            ).first
            
            # --- Medición de rendimiento: Inicio de localización del botón de la página de destino ---
            start_time_locator_button = time.time()
            expect(pagina_destino_locator).to_be_visible()
            expect(pagina_destino_locator).to_be_enabled()
            self.logger.info(f"\n✅ Elemento de la página '{numero_pagina_a_navegar}' encontrado y habilitado para clic.")
            
            # --- Medición de rendimiento: Fin de localización del botón de la página de destino ---
            end_time_locator_button = time.time()
            duration_locator_button = end_time_locator_button - start_time_locator_button
            self.logger.info(f"PERFORMANCE: Tiempo de localización del botón de la página de destino: {duration_locator_button:.4f} segundos.")

            pagina_destino_locator.highlight()
            self.tomar_captura(f"{nombre_base}_pagina_a_navegar_encontrada", directorio)
            
            self.logger.info(f"\n  Haciendo clic en la página '{numero_pagina_a_navegar}'...")
            
            # --- Medición de rendimiento: Inicio de click y espera de carga ---
            start_time_click_and_wait = time.time()
            pagina_destino_locator.click()
            self.esperar_fijo(pausa_post_clic) # Pausa para permitir la carga de la página y la aplicación de estilos
            
            # --- Medición de rendimiento: Fin de click y espera de carga ---
            end_time_click_and_wait = time.time()
            duration_click_and_wait = end_time_click_and_wait - start_time_click_and_wait
            self.logger.info(f"PERFORMANCE: Tiempo de click y espera de carga para la página '{numero_pagina_a_navegar}': {duration_click_and_wait:.4f} segundos.")

            self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_clic", directorio)

            # 4. Verificar que la página de destino se resalte
            self.logger.info(f"\nVerificando si la página '{numero_pagina_a_navegar}' tiene la clase de resaltado '{clase_resaltado}'...")
            
            # Asegurarse de que el elemento de destino aún esté visible y, opcionalmente, que sus atributos se hayan actualizado.
            expect(pagina_destino_locator).to_be_visible()
            pagina_destino_locator.highlight() # Resaltar el elemento para la captura final

            # --- Medición de rendimiento: Inicio de verificación de estado final ---
            start_time_final_verification = time.time()

            current_classes_attribute = pagina_destino_locator.get_attribute("class")
            
            if current_classes_attribute is not None:
                current_classes_list = current_classes_attribute.split()
            else:
                current_classes_list = []

            if clase_resaltado in current_classes_list:
                self.logger.info(f"\n  ✅ ÉXITO: La página '{numero_pagina_a_navegar}' está seleccionada y resaltada con la clase '{clase_resaltado}'.")
                self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_seleccionada_ok", directorio)
                success = True
            else:
                self.logger.error(f"\n  ❌ FALLO: La página '{numero_pagina_a_navegar}' no tiene la clase de resaltado esperada '{clase_resaltado}'.")
                self.logger.info(f"\n  Clases actuales del elemento: '{current_classes_attribute}'")
                self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_no_resaltada", directorio)
                success = False

            # --- Medición de rendimiento: Fin de verificación de estado final ---
            end_time_final_verification = time.time()
            duration_final_verification = end_time_final_verification - start_time_final_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de la clase de resaltado final: {duration_final_verification:.4f} segundos.")

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (navegación y verificación de paginación): {duration_total_operation:.4f} segundos.")

            return success

        except TimeoutError as e:
            # Captura si el contenedor de paginación o el elemento de la página de destino no se vuelven visibles/interactuables a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Timeout): El contenedor de paginación '{selector_paginado}' "
                f"o la página '{numero_pagina_a_navegar}' no estuvieron visibles/interactuables a tiempo "
                f"(timeout configurado: {tiempo_espera_componente}s).\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_timeout_navegacion", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nComponente de paginación o página de destino no disponibles a tiempo: {selector_paginado} o página {numero_pagina_a_navegar}") from e

        except Error as e:
            # Captura errores específicos de Playwright durante la interacción con el DOM.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error al interactuar con el componente de paginación durante la navegación.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM, elemento no clickable.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError de Playwright al navegar/verificar paginación: {selector_paginado}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al navegar y verificar la paginación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError inesperado al navegar/verificar paginación: {selector_paginado}") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.5) # Pequeña espera final para observación o para liberar recursos si es necesario.
        
    # 40- Función para verificar una alerta simple utilizando page.expect_event().
    # Integra pruebas de rendimiento para medir la aparición y manejo de la alerta.
    def verificar_alerta_simple_con_expect_event(self, selector: Locator, mensaje_esperado: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_espera_alerta: Union[int, float] = 0.5) -> bool:
        """
        Verifica una alerta de tipo 'alert' que aparece después de hacer clic en un selector dado.
        Utiliza `page.expect_event("dialog")` de Playwright para esperar y capturar el diálogo.
        Comprueba el tipo de diálogo y su mensaje, y finalmente lo acepta.
        Integra mediciones de rendimiento para la aparición y manejo de la alerta.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara la alerta.
            mensaje_esperado (str): El **mensaje esperado** dentro del cuerpo de la alerta.
                                    Se verifica si este mensaje está contenido en el texto de la alerta.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_espera_alerta (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                      para que la alerta (diálogo) aparezca después
                                                      de hacer clic en el selector. Por defecto, `5.0` segundos.

        Returns:
            bool: `True` si la alerta apareció, es del tipo 'alert', contiene el mensaje esperado
                  y fue aceptada correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si la alerta no aparece,
                            si el tipo de diálogo es incorrecto, o si ocurre un error inesperado
                            de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de alerta con expect_event: {nombre_base} ---")
        self.logger.info(f"\nVerificando alerta al hacer clic en '{selector}'")
        self.logger.info(f"\n  --> Mensaje de alerta esperado: '{mensaje_esperado}'")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará la alerta
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_alerta", directorio)


            self.logger.debug(f"\n  --> Preparando expect_event para la alerta y haciendo clic (timeout de alerta: {tiempo_espera_alerta}s)...")
            
            # 2. Esperar el evento de diálogo (alerta) y hacer clic en el selector
            # Se usa `timeout` en `expect_event` para el tiempo máximo de aparición de la alerta.
            # Se usa `timeout` en `click` para el tiempo máximo de clic en el elemento.
            # Se recomienda que el timeout del `expect_event` sea al menos tan grande como el del `click`
            # para dar tiempo a que la alerta aparezca.
            # Playwright automáticamente acepta diálogos si no hay un handler. Aquí, lo manejamos explícitamente.
            with self.page.expect_event("dialog") as info_dialogo:
                # --- Medición de rendimiento: Inicio de click y espera de alerta ---
                start_time_alert_detection = time.time()
                self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}' para disparar la alerta...")
                selector.click()
            
            dialogo: Dialog = info_dialogo.value # Obtener el objeto Dialog de la alerta
            # --- Medición de rendimiento: Fin de click y espera de alerta ---
            end_time_alert_detection = time.time()
            duration_alert_detection = end_time_alert_detection - start_time_alert_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la alerta: {duration_alert_detection:.4f} segundos.")

            self.logger.info(f"\n  --> Alerta detectada. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}'")
            self.tomar_captura(f"{nombre_base}_alerta_detectada", directorio)

            # 3. Validar el tipo de diálogo
            if dialogo.type != "alert":
                dialogo.accept() # Aceptar para no bloquear si es un tipo inesperado
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'alert'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'alert'.")

            # 4. Validar el mensaje de la alerta
            # --- Medición de rendimiento: Inicio de verificación del mensaje ---
            start_time_message_verification = time.time()
            if mensaje_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_alerta_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje de alerta incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                dialogo.accept() # Aceptar para no bloquear antes de fallar
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            # --- Medición de rendimiento: Fin de verificación del mensaje ---
            end_time_message_verification = time.time()
            duration_message_verification = end_time_message_verification - start_time_message_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del mensaje de la alerta: {duration_message_verification:.4f} segundos.")


            # 5. Aceptar la alerta
            dialogo.accept()
            self.logger.info("\n  ✅  --> Alerta ACEPTADA correctamente.")

            # Opcional: Verificar el resultado en la página después de la interacción
            # Si tu aplicación cambia el estado del DOM (ej. un mensaje de éxito/error)
            # después de que la alerta es aceptada, puedes verificarlo aquí.
            # Por ejemplo: expect(self.page.locator("#status_message")).to_have_text("Operación completada");
            self.esperar_fijo(0.5) # Pequeña pausa para que el DOM se actualice si es necesario

            self.tomar_captura(f"{nombre_base}_alerta_exitosa", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: La alerta se mostró, mensaje verificado y aceptada correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de alerta): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo o si la alerta no aparece a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo "
                f"o la alerta no apareció/fue detectada a tiempo ({tiempo_espera_elemento}s para elemento, {tiempo_espera_alerta}s para alerta).\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_alerta_NO_aparece_timeout", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nTimeout al verificar alerta para selector '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError de Playwright al verificar alerta para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (tipo de diálogo, mensaje incorrecto).
            self.logger.critical(f"\n❌ FALLO (Validación de Alerta): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError inesperado al verificar alerta para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
    
    # 41- Función para verificar una alerta simple utilizando page.on("dialog") con page.once().
    # Integra pruebas de rendimiento para medir la aparición y manejo de la alerta a través de un listener.
    def verificar_alerta_simple_con_on(self, selector: Locator, mensaje_alerta_esperado: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_max_deteccion_alerta: Union[int, float] = 0.7) -> bool:
        """
        Verifica una alerta de tipo 'alert' que aparece después de hacer clic en un selector dado.
        Utiliza `page.once("dialog")` para registrar un manejador de eventos que captura
        y acepta la alerta cuando aparece. Mide el rendimiento de cada fase.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara la alerta.
            mensaje_alerta_esperado (str): El **mensaje esperado** dentro del cuerpo de la alerta.
                                           Se verifica si este mensaje está contenido en el texto de la alerta.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_max_deteccion_alerta (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                              después de hacer clic para que el listener
                                                              detecte y maneje la alerta. Debe ser mayor que
                                                              el tiempo de procesamiento esperado de la alerta.
                                                              Por defecto, `7.0` segundos.

        Returns:
            bool: `True` si la alerta apareció, es del tipo 'alert', contiene el mensaje esperado
                  y fue aceptada correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si la alerta no aparece,
                            si el tipo de diálogo es incorrecto, o si ocurre un error inesperado
                            de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de alerta con page.once('dialog'): {nombre_base} ---")
        self.logger.info(f"\nVerificando alerta simple al hacer clic en el botón '{selector}'")
        self.logger.info(f"\n  --> Mensaje de alerta esperado: '{mensaje_alerta_esperado}'")

        # Resetear el estado de las banderas para cada ejecución del test
        # Esto es crucial para evitar que valores de una ejecución anterior afecten la actual.
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará la alerta
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_alerta", directorio)

            # 2. Registrar el listener ANTES de la acción que dispara la alerta
            self.logger.debug("\n  --> Registrando listener para la alerta con page.once('dialog')...")
            # Usa page.once para que el listener se desregistre automáticamente después de detectar el primer diálogo.
            # El handler `_get_simple_alert_handler_for_on()` también acepta la alerta internamente.
            self.page.once("dialog", self._get_simple_alert_handler_for_on())

            # 3. Hacer clic en el botón que dispara la alerta
            self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
            # --- Medición de rendimiento: Inicio de click y espera de detección de alerta ---
            start_time_click_and_alert_detection = time.time()
            selector.click() # Reutilizar tiempo_espera_elemento para el click

            # 4. Esperar a que el listener haya detectado y manejado la alerta
            self.logger.debug(f"\n  --> Esperando a que la alerta sea detectada y manejada por el listener (timeout: {tiempo_max_deteccion_alerta}s)...")
            # Bucle de espera activa hasta que la bandera _alerta_detectada sea True
            # Se añade un timeout para el bucle, calculado a partir de tiempo_max_deteccion_alerta
            wait_end_time = time.time() + tiempo_max_deteccion_alerta
            while not self._alerta_detectada and time.time() < wait_end_time:
                time.sleep(0.1) # Pausa breve para evitar consumo excesivo de CPU

            # --- Medición de rendimiento: Fin de click y espera de detección de alerta ---
            end_time_click_and_alert_detection = time.time()
            duration_click_and_alert_detection = end_time_click_and_alert_detection - start_time_click_and_alert_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la alerta por el listener: {duration_click_and_alert_detection:.4f} segundos.")

            if not self._alerta_detectada:
                error_msg = f"\n❌ FALLO: La alerta no fue detectada por el listener después de {tiempo_max_deteccion_alerta} segundos."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_alerta_NO_detectada_timeout", directorio)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            self.tomar_captura(f"{nombre_base}_alerta_detectada_por_listener", directorio)
            self.logger.info(f"\n  ✅  Alerta detectada con éxito por el listener.")

            # 5. Validaciones después de que el listener ha actuado
            # --- Medición de rendimiento: Inicio de verificación de contenido de alerta ---
            start_time_alert_content_verification = time.time()
            if self._alerta_tipo_capturado != "alert":
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'alert'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'alert'.")

            if mensaje_alerta_esperado not in self._alerta_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_alerta_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje de alerta incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_alerta_esperado}'\n"
                    f"  --> Obtenido: '{self._alerta_mensaje_capturado}'"
                )
                self.logger.error(error_msg)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            # --- Medición de rendimiento: Fin de verificación de contenido de alerta ---
            end_time_alert_content_verification = time.time()
            duration_alert_content_verification = end_time_alert_content_verification - start_time_alert_content_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de tipo y mensaje de la alerta: {duration_alert_content_verification:.4f} segundos.")


            # La alerta ya fue aceptada por el handler `_get_simple_alert_handler_for_on()`.
            self.logger.info("\n  ✅  --> Alerta ACEPTADA (por el listener).")

            # Opcional: Verificar el resultado en la página después de la interacción
            # Si tu aplicación cambia el estado del DOM (ej. un mensaje de éxito/error)
            # después de que la alerta es aceptada, puedes verificarlo aquí.
            # Por ejemplo: expect(self.page.locator("#status_message")).to_have_text("Operación completada");
            self.esperar_fijo(0.5) # Pequeña pausa para que el DOM se actualice si es necesario

            self.tomar_captura(f"{nombre_base}_alerta_exitosa", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: La alerta se mostró, mensaje verificado y aceptada correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de alerta por listener): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo. La detección de alerta por timeout se maneja en el bucle.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo "
                f"antes de intentar hacer clic ({tiempo_espera_elemento}s).\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_elemento_NO_listo_timeout", directorio)
            raise AssertionError(f"\nTimeout al preparar el elemento disparador para '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al verificar alerta para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (alerta no detectada, tipo incorrecto, mensaje incorrecto).
            self.logger.critical(f"\n❌ FALLO (Validación de Alerta): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar alerta para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
        
    # 42- Función para verificar una alerta de confirmación utilizando page.expect_event().
    # Este método maneja el diálogo exclusivamente con expect_event e integra pruebas de rendimiento.
    def verificar_confirmacion_expect_event(self, selector: Locator, mensaje_esperado: str, accion_confirmacion: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_espera_confirmacion: Union[int, float] = 0.7) -> bool:
        """
        Verifica una alerta de tipo 'confirm' que aparece después de hacer clic en un selector dado.
        Utiliza `page.expect_event("dialog")` de Playwright para esperar y capturar el diálogo.
        Comprueba el tipo de diálogo y su mensaje, y finalmente realiza la acción solicitada (aceptar o cancelar).
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara la confirmación.
            mensaje_esperado (str): El **mensaje esperado** dentro del cuerpo de la confirmación.
                                    Se verifica si este mensaje está contenido en el texto de la confirmación.
            accion_confirmacion (str): La **acción a realizar** en la confirmación:
                                       'accept' para aceptar el diálogo o 'dismiss' para cancelarlo.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_espera_confirmacion (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                            para que la confirmación (diálogo) aparezca después
                                                            de hacer clic en el selector. Debe ser mayor que
                                                            el tiempo de procesamiento esperado. Por defecto, `7.0` segundos.

        Returns:
            bool: `True` si la confirmación apareció, es del tipo 'confirm', contiene el mensaje esperado
                  y fue manejada correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si la confirmación no aparece,
                            si el tipo de diálogo es incorrecto, si el mensaje no coincide, si la acción
                            de confirmación no es válida, o si ocurre un error inesperado de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de confirmación con expect_event: {nombre_base} ---")
        self.logger.info(f"\nVerificando confirmación al hacer clic en '{selector}' para '{accion_confirmacion}'")
        self.logger.info(f"\n  --> Mensaje de confirmación esperado: '{mensaje_esperado}'")

        # Validar la acción de confirmación antes de iniciar la operación
        if accion_confirmacion not in ['accept', 'dismiss']:
            error_msg = f"\n❌ FALLO: Acción de confirmación no válida: '{accion_confirmacion}'. Use 'accept' o 'dismiss'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_accion_invalida", directorio)
            raise AssertionError(error_msg)

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará la confirmación
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_confirmacion", directorio)

            # 2. Esperar el evento de diálogo (confirmación) y hacer clic en el selector
            self.logger.debug(f"\n  --> Preparando expect_event para la confirmación y haciendo clic (timeout de confirmación: {tiempo_espera_confirmacion}s)...")
            
            # Se usa `timeout` en `expect_event` para el tiempo máximo de aparición de la confirmación.
            # Se usa `timeout` en `click` para el tiempo máximo de clic en el elemento.
            # Es importante que el timeout de `expect_event` sea suficiente para que la confirmación aparezca.
            with self.page.expect_event("dialog", timeout=int(tiempo_espera_confirmacion * 1000)) as info_dialogo:
                # --- Medición de rendimiento: Inicio de click y espera de confirmación ---
                start_time_confirm_detection = time.time()
                self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}' para disparar la confirmación...")
                selector.click(timeout=int(tiempo_espera_elemento * 1000)) # Reutilizar tiempo_espera_elemento para el click
            
            dialogo: Dialog = info_dialogo.value # Obtener el objeto Dialog de la confirmación
            # --- Medición de rendimiento: Fin de click y espera de confirmación ---
            end_time_confirm_detection = time.time()
            duration_confirm_detection = end_time_confirm_detection - start_time_confirm_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la confirmación: {duration_confirm_detection:.4f} segundos.")

            self.logger.info(f"\n  --> Confirmación detectada. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}'")
            self.tomar_captura(f"{nombre_base}_confirmacion_detectada", directorio)

            # 3. Validar el tipo de diálogo
            if dialogo.type != "confirm":
                # Realizar la acción solicitada incluso si el tipo es incorrecto para no bloquear
                if accion_confirmacion == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'confirm'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'confirm'.")

            # 4. Validar el mensaje de la confirmación
            # --- Medición de rendimiento: Inicio de verificación del mensaje ---
            start_time_message_verification = time.time()
            if mensaje_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_confirmacion_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje de confirmación incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                # Realizar la acción solicitada para no bloquear antes de fallar
                if accion_confirmacion == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            # --- Medición de rendimiento: Fin de verificación del mensaje ---
            end_time_message_verification = time.time()
            duration_message_verification = end_time_message_verification - start_time_message_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del mensaje de la confirmación: {duration_message_verification:.4f} segundos.")

            # 5. Realizar la acción solicitada (Aceptar o Cancelar)
            # --- Medición de rendimiento: Inicio de la acción sobre la confirmación ---
            start_time_confirm_action = time.time()
            if accion_confirmacion == 'accept':
                dialogo.accept()
                self.logger.info("\n  ✅  --> Confirmación ACEPTADA.")
            elif accion_confirmacion == 'dismiss':
                dialogo.dismiss()
                self.logger.info("\n  ✅  --> Confirmación CANCELADA.")
            # --- Medición de rendimiento: Fin de la acción sobre la confirmación ---
            end_time_confirm_action = time.time()
            duration_confirm_action = end_time_confirm_action - start_time_confirm_action
            self.logger.info(f"PERFORMANCE: Tiempo de acción ('{accion_confirmacion}') sobre la confirmación: {duration_confirm_action:.4f} segundos.")


            # 6. Opcional: Verificar el resultado en la página después de la interacción
            # Es crucial para confirmar que la acción en el diálogo tuvo el efecto esperado en la UI.
            # Asumo un selector '#demo' y textos específicos, ajusta esto a tu aplicación real.
            # --- Medición de rendimiento: Inicio de verificación del resultado en la página ---
            start_time_post_action_verification = time.time()
            if accion_confirmacion == 'accept':
                # Esto es un ejemplo, ajusta el selector y el texto esperado
                expect(self.page.locator("#demo")).to_have_text("You pressed OK!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed OK!' verificado.")
            elif accion_confirmacion == 'dismiss':
                # Esto es un ejemplo, ajusta el selector y el texto esperado
                expect(self.page.locator("#demo")).to_have_text("You pressed Cancel!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed Cancel!' verificado.")
            
            # --- Medición de rendimiento: Fin de verificación del resultado en la página ---
            end_time_post_action_verification = time.time()
            duration_post_action_verification = end_time_post_action_verification - start_time_post_action_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del resultado en la página: {duration_post_action_verification:.4f} segundos.")


            self.tomar_captura(f"{nombre_base}_confirmacion_exitosa_{accion_confirmacion}", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: La confirmación se mostró, mensaje verificado y '{accion_confirmacion}' correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de confirmación): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo o si la confirmación no aparece a tiempo, o la verificación post-acción falla.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo, "
                f"la confirmación no apareció/fue detectada a tiempo ({tiempo_espera_elemento}s para elemento, {tiempo_espera_confirmacion}s para confirmación), "
                f"o la verificación del resultado en la página falló.\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_confirmacion_NO_aparece_timeout", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nTimeout al verificar confirmación para selector '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError de Playwright al verificar confirmación para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (tipo de diálogo, mensaje incorrecto, acción inválida).
            self.logger.critical(f"\n❌ FALLO (Validación de Confirmación): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar confirmación para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
        
    # 43- Función para verificar una alerta de confirmación utilizando page.on("dialog") con page.once().
    # Integra pruebas de rendimiento para medir la aparición y manejo de la confirmación a través de un listener.
    def verificar_confirmacion_on_dialog(self, selector: Locator, mensaje_esperado: str, accion_confirmacion: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_max_deteccion_confirmacion: Union[int, float] = 0.7) -> bool:
        """
        Verifica una confirmación de tipo 'confirm' que aparece después de hacer clic en un selector dado.
        Utiliza `page.once("dialog")` para registrar un manejador de eventos que captura
        la confirmación y realiza la acción solicitada (aceptar o cancelar).
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara la confirmación.
            mensaje_esperado (str): El **mensaje esperado** dentro del cuerpo de la confirmación.
                                    Se verifica si este mensaje está contenido en el texto de la confirmación.
            accion_confirmacion (str): La **acción a realizar** en la confirmación:
                                       'accept' para aceptar el diálogo o 'dismiss' para cancelarlo.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_max_deteccion_confirmacion (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                                  después de hacer clic para que el listener
                                                                  detecte y maneje la confirmación. Debe ser mayor que
                                                                  el tiempo de procesamiento esperado de la confirmación.
                                                                  Por defecto, `7.0` segundos.

        Returns:
            bool: `True` si la confirmación apareció, es del tipo 'confirm', contiene el mensaje esperado
                  y fue manejada correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si la confirmación no aparece,
                            si el tipo de diálogo es incorrecto, si el mensaje no coincide, si la acción
                            de confirmación no es válida, o si ocurre un error inesperado de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de confirmación con page.on('dialog'): {nombre_base} ---")
        self.logger.info(f"\nVerificando confirmación al hacer clic en '{selector}' para '{accion_confirmacion}'")
        self.logger.info(f"\n  --> Mensaje de confirmación esperado: '{mensaje_esperado}'")

        # Validar la acción de confirmación antes de iniciar la operación
        if accion_confirmacion not in ['accept', 'dismiss']:
            error_msg = f"\n❌ FALLO: Acción de confirmación no válida: '{accion_confirmacion}'. Use 'accept' o 'dismiss'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_accion_invalida", directorio)
            raise AssertionError(error_msg)

        # Resetear el estado de las banderas para cada ejecución del test
        # Esto es crucial para evitar que valores de una ejecución anterior afecten la actual.
        self._dialogo_detectado = False
        self._dialogo_mensaje_capturado = ""
        self._dialogo_tipo_capturado = ""

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará la confirmación
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_confirmacion", directorio)

            # 2. Registrar el listener ANTES de la acción que dispara la confirmación
            self.logger.debug("\n  --> Registrando listener para la confirmación con page.once('dialog')...")
            # Usa page.once para que el listener se desregistre automáticamente después de detectar el primer diálogo.
            # El handler `_get_confirmation_dialog_handler_for_on()` también acepta/cancela la confirmación internamente.
            self.page.once("dialog", self._get_confirmation_dialog_handler_for_on(accion_confirmacion))

            # 3. Hacer clic en el botón que dispara la confirmación
            self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
            # --- Medición de rendimiento: Inicio de click y espera de detección de confirmación ---
            start_time_click_and_confirm_detection = time.time()
            selector.click()

            # 4. Esperar a que el listener haya detectado y manejado la confirmación
            self.logger.debug(f"\n  --> Esperando a que la confirmación sea detectada y manejada por el listener (timeout: {tiempo_max_deteccion_confirmacion}s)...")
            # Bucle de espera activa hasta que la bandera _dialogo_detectado sea True
            # Se añade un timeout para el bucle, calculado a partir de tiempo_max_deteccion_confirmacion
            wait_end_time = time.time() + tiempo_max_deteccion_confirmacion
            while not self._dialogo_detectado and time.time() < wait_end_time:
                time.sleep(0.1) # Pausa breve para evitar consumo excesivo de CPU

            # --- Medición de rendimiento: Fin de click y espera de detección de confirmación ---
            end_time_click_and_confirm_detection = time.time()
            duration_click_and_confirm_detection = end_time_click_and_confirm_detection - start_time_click_and_confirm_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la confirmación por el listener: {duration_click_and_confirm_detection:.4f} segundos.")

            if not self._dialogo_detectado:
                error_msg = f"\n❌ FALLO: La confirmación no fue detectada por el listener después de {tiempo_max_deteccion_confirmacion} segundos."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_confirmacion_NO_detectada_timeout", directorio)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            self.tomar_captura(f"{nombre_base}_confirmacion_detectada_por_listener", directorio)
            self.logger.info(f"\n  ✅  Confirmación detectada con éxito por el listener.")

            # 5. Validaciones después de que el listener ha actuado
            # --- Medición de rendimiento: Inicio de verificación de contenido de confirmación ---
            start_time_dialog_content_verification = time.time()
            if self._dialogo_tipo_capturado != "confirm":
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{self._dialogo_tipo_capturado}'. Se esperaba 'confirm'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{self._dialogo_tipo_capturado}'. Se esperaba 'confirm'.")

            if mensaje_esperado not in self._dialogo_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_confirmacion_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje de confirmación incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_esperado}'\n"
                    f"  --> Obtenido: '{self._dialogo_mensaje_capturado}'"
                )
                self.logger.error(error_msg)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            # --- Medición de rendimiento: Fin de verificación de contenido de confirmación ---
            end_time_dialog_content_verification = time.time()
            duration_dialog_content_verification = end_time_dialog_content_verification - start_time_dialog_content_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de tipo y mensaje de la confirmación: {duration_dialog_content_verification:.4f} segundos.")

            # La confirmación ya fue aceptada/cancelada por el handler `_get_confirmation_dialog_handler_for_on()`.
            self.logger.info(f"\n  ✅  --> Confirmación manejada (acción '{accion_confirmacion}' por el listener).")

            # 6. Opcional: Verificar el resultado en la página después de la interacción
            # Es crucial para confirmar que la acción en el diálogo tuvo el efecto esperado en la UI.
            # --- Medición de rendimiento: Inicio de verificación del resultado en la página ---
            start_time_post_action_verification = time.time()
            if accion_confirmacion == 'accept':
                # Asumo un selector '#demo' y texto "You pressed OK!", ajusta esto a tu aplicación real
                expect(self.page.locator("#demo")).to_have_text("You pressed OK!")
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed OK!' verificado.")
            elif accion_confirmacion == 'dismiss':
                # Asumo un selector '#demo' y texto "You pressed Cancel!", ajusta esto a tu aplicación real
                expect(self.page.locator("#demo")).to_have_text("You pressed Cancel!")
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed Cancel!' verificado.")
            
            # --- Medición de rendimiento: Fin de verificación del resultado en la página ---
            end_time_post_action_verification = time.time()
            duration_post_action_verification = end_time_post_action_verification - start_time_post_action_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del resultado en la página: {duration_post_action_verification:.4f} segundos.")

            self.tomar_captura(f"{nombre_base}_confirmacion_exitosa_{accion_confirmacion}", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: La confirmación se mostró, mensaje verificado y '{accion_confirmacion}' correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de confirmación por listener): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo, si la confirmación no aparece a tiempo, o si la verificación post-acción falla.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo, "
                f"la confirmación no fue detectada por el listener después de {tiempo_max_deteccion_confirmacion} segundos, "
                f"o la verificación del resultado en la página falló.\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_confirmacion_NO_detectada_timeout", directorio)
            raise AssertionError(f"\nTimeout al verificar confirmación para selector '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al verificar confirmación para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (acción inválida, tipo de diálogo, mensaje incorrecto).
            self.logger.critical(f"\n❌ FALLO (Validación de Confirmación): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar confirmación para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
    
    # 44- Función para verificar_prompt_expect_event (Implementación para Prompt Alert con expect_event).
    # Integra pruebas de rendimiento para medir la aparición, interacción y manejo de un diálogo prompt.
    def verificar_prompt_expect_event(self, selector: Locator, mensaje_prompt_esperado: str, input_text: Optional[str], accion_prompt: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_espera_prompt: Union[int, float] = 0.7) -> bool:
        """
        Verifica un cuadro de diálogo 'prompt' que aparece después de hacer clic en un selector dado.
        Utiliza `page.expect_event("dialog")` de Playwright para esperar y capturar el diálogo.
        Comprueba el tipo de diálogo y su mensaje. Si la acción es 'accept', introduce el texto
        proporcionado; de lo contrario, cancela el prompt.
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara el diálogo prompt.
            mensaje_prompt_esperado (str): El **mensaje esperado** dentro del cuerpo del prompt.
                                           Se verifica si este mensaje está contenido en el texto del prompt.
            input_text (Optional[str]): El **texto a introducir** en el prompt si `accion_prompt` es 'accept'.
                                        Debe ser `None` si `accion_prompt` es 'dismiss'.
            accion_prompt (str): La **acción a realizar** en el prompt:
                                 'accept' para introducir texto y aceptar, o 'dismiss' para cancelar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_espera_prompt (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                     para que el prompt aparezca después de hacer clic en el selector.
                                                     Debe ser mayor que el tiempo de procesamiento esperado.
                                                     Por defecto, `7.0` segundos.

        Returns:
            bool: `True` si el prompt apareció, es del tipo 'prompt', contiene el mensaje esperado
                  y fue manejado correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si el prompt no aparece,
                            si el tipo de diálogo es incorrecto, si el mensaje no coincide,
                            si la acción del prompt no es válida, si `input_text` es incorrecto
                            para la acción, o si ocurre un error inesperado de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de prompt con expect_event: {nombre_base} ---")
        self.logger.info(f"\nVerificando prompt al hacer clic en '{selector}' para '{accion_prompt}'")
        self.logger.info(f"\n  --> Mensaje del prompt esperado: '{mensaje_prompt_esperado}'")
        if accion_prompt == 'accept':
            self.logger.info(f"\n  --> Texto a introducir: '{input_text}'")

        # Validar la acción y el input_text antes de iniciar la operación
        if accion_prompt not in ['accept', 'dismiss']:
            error_msg = f"\n❌ FALLO: Acción de prompt no válida: '{accion_prompt}'. Use 'accept' o 'dismiss'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_accion_invalida", directorio)
            raise AssertionError(error_msg)
        if accion_prompt == 'accept' and input_text is None:
            error_msg = "\n❌ FALLO: 'input_text' no puede ser None cuando 'accion_prompt' es 'accept'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_input_text_missing", directorio)
            raise AssertionError(error_msg)
        if accion_prompt == 'dismiss' and input_text is not None:
            self.logger.warning("\n⚠️ ADVERTENCIA: 'input_text' se ignora cuando 'accion_prompt' es 'dismiss'.")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará el prompt
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_prompt", directorio)

            # 2. Esperar el evento de diálogo (prompt) y hacer clic en el selector
            self.logger.debug(f"\n  --> Preparando expect_event para el prompt y haciendo clic (timeout de prompt: {tiempo_espera_prompt}s)...")
            
            # Se usa `timeout` en `expect_event` para el tiempo máximo de aparición del prompt.
            # Se usa `timeout` en `click` para el tiempo máximo de clic en el elemento.
            with self.page.expect_event("dialog") as info_dialogo:
                # --- Medición de rendimiento: Inicio de click y espera de prompt ---
                start_time_prompt_detection = time.time()
                self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}' para disparar el prompt...")
                selector.click()
            
            dialogo: Dialog = info_dialogo.value # Obtener el objeto Dialog del prompt
            # --- Medición de rendimiento: Fin de click y espera de prompt ---
            end_time_prompt_detection = time.time()
            duration_prompt_detection = end_time_prompt_detection - start_time_prompt_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección del prompt: {duration_prompt_detection:.4f} segundos.")

            self.logger.info(f"\n  --> Prompt detectado. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}', Valor por defecto: '{dialogo.default_value}'")
            self.tomar_captura(f"{nombre_base}_prompt_detectado", directorio)

            # 3. Validar el tipo de diálogo
            if dialogo.type != "prompt":
                # Si el tipo es inesperado, intenta cerrarlo para no bloquear el test antes de fallar.
                if accion_prompt == 'accept':
                    dialogo.accept(input_text if input_text is not None else "") # Aceptar con o sin texto
                else:
                    dialogo.dismiss()
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'prompt'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'prompt'.")

            # 4. Validar el mensaje del prompt
            # --- Medición de rendimiento: Inicio de verificación del mensaje ---
            start_time_message_verification = time.time()
            if mensaje_prompt_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_prompt_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje del prompt incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_prompt_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                # Intenta cerrar el diálogo antes de fallar
                if accion_prompt == 'accept':
                    dialogo.accept(input_text if input_text is not None else "")
                else:
                    dialogo.dismiss()
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            # --- Medición de rendimiento: Fin de verificación del mensaje ---
            end_time_message_verification = time.time()
            duration_message_verification = end_time_message_verification - start_time_message_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del mensaje del prompt: {duration_message_verification:.4f} segundos.")

            # 5. Realizar la acción solicitada (Introducir texto y Aceptar, o Cancelar)
            # --- Medición de rendimiento: Inicio de la acción sobre el prompt ---
            start_time_prompt_action = time.time()
            if accion_prompt == 'accept':
                # El método `accept()` para prompts puede tomar un argumento `promptText`
                dialogo.accept(input_text)
                self.logger.info(f"\n  ✅  --> Texto '{input_text}' introducido en el prompt y ACEPTADO.")
            elif accion_prompt == 'dismiss':
                dialogo.dismiss()
                self.logger.info("\n  ✅  --> Prompt CANCELADO.")
            # No se necesita 'else' aquí, ya se validó 'accion_prompt' al principio
            # --- Medición de rendimiento: Fin de la acción sobre el prompt ---
            end_time_prompt_action = time.time()
            duration_prompt_action = end_time_prompt_action - start_time_prompt_action
            self.logger.info(f"PERFORMANCE: Tiempo de acción ('{accion_prompt}') sobre el prompt: {duration_prompt_action:.4f} segundos.")


            # 6. Opcional: Verificar el resultado en la página después de la interacción
            # Es crucial para confirmar que la acción en el diálogo tuvo el efecto esperado en la UI.
            # Asumo un selector '#demo' y textos específicos, ajusta esto a tu aplicación real.
            # --- Medición de rendimiento: Inicio de verificación del resultado en la página ---
            start_time_post_action_verification = time.time()
            if accion_prompt == 'accept':
                # Ejemplo: Si el texto introducido se muestra en un elemento de la página
                expect(self.page.locator("#demo")).to_have_text(f"You entered: {input_text}")
                self.logger.info(f"\n  ✅  --> Resultado en página: 'You entered: {input_text}' verificado.")
            elif accion_prompt == 'dismiss':
                # Ejemplo: Si se muestra un mensaje de cancelación
                expect(self.page.locator("#demo")).to_have_text("You cancelled the prompt.")
                self.logger.info("\n  ✅  --> Resultado en página: 'You cancelled the prompt.' verificado.")
            
            # --- Medición de rendimiento: Fin de verificación del resultado en la página ---
            end_time_post_action_verification = time.time()
            duration_post_action_verification = end_time_post_action_verification - start_time_post_action_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación del resultado en la página: {duration_post_action_verification:.4f} segundos.")

            self.tomar_captura(f"{nombre_base}_prompt_exitosa_{accion_prompt}", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: El prompt se mostró, mensaje verificado, texto introducido y '{accion_prompt}' correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de prompt): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo o si el prompt no aparece a tiempo, o la verificación post-acción falla.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo, "
                f"el prompt no apareció/fue detectado a tiempo ({tiempo_espera_elemento}s para elemento, {tiempo_espera_prompt}s para prompt), "
                f"o la verificación del resultado en la página falló.\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_prompt_NO_aparece_timeout", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nTimeout al verificar prompt para selector '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nError de Playwright al verificar prompt para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (acción inválida, tipo de diálogo, mensaje incorrecto).
            self.logger.critical(f"\n❌ FALLO (Validación de Prompt): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar prompt para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
        
    # 45- Función para verificar una alerta de tipo 'prompt' utilizando page.on("dialog") con page.once().
    # Integra pruebas de rendimiento para medir la aparición, interacción y manejo de un diálogo prompt.
    def verificar_prompt_on_dialog(self, selector: Locator, mensaje_prompt_esperado: str, input_text: Optional[str], accion_prompt: str, nombre_base: str, directorio: str, tiempo_espera_elemento: Union[int, float] = 0.5, tiempo_max_deteccion_prompt: Union[int, float] = 0.7) -> bool:
        """
        Verifica un cuadro de diálogo 'prompt' que aparece después de hacer clic en un selector dado.
        Utiliza `page.once("dialog")` para registrar un manejador de eventos que captura el prompt,
        introduce el texto si es necesario y realiza la acción solicitada (aceptar o cancelar).
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón)
                                que, al ser clicado, dispara el diálogo prompt.
            mensaje_prompt_esperado (str): El **mensaje esperado** dentro del cuerpo del prompt.
                                           Se verifica si este mensaje está contenido en el texto del prompt.
            input_text (Optional[str]): El **texto a introducir** en el prompt si `accion_prompt` es 'accept'.
                                        Debe ser `None` si `accion_prompt` es 'dismiss'.
            accion_prompt (str): La **acción a realizar** en el prompt:
                                 'accept' para introducir texto y aceptar, o 'dismiss' para cancelar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_elemento (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                        para que el `selector` esté visible y habilitado
                                                        antes de intentar hacer clic. Por defecto, `5.0` segundos.
            tiempo_max_deteccion_prompt (Union[int, float]): **Tiempo máximo de espera** (en segundos)
                                                               para que el listener detecte y maneje el prompt
                                                               después de hacer clic en el selector.
                                                               Por defecto, `7.0` segundos.

        Returns:
            bool: `True` si el prompt apareció, es del tipo 'prompt', contiene el mensaje esperado
                  y fue manejado correctamente; `False` en caso contrario o si ocurre un Timeout.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, si el prompt no aparece,
                            si el tipo de diálogo es incorrecto, si el mensaje no coincide,
                            si la acción del prompt no es válida, si `input_text` es incorrecto
                            para la acción, o si ocurre un error inesperado de Playwright o genérico.
        """
        self.logger.info(f"\n--- Ejecutando verificación de prompt con page.on('dialog'): {nombre_base} ---")
        self.logger.info(f"\nVerificando prompt al hacer clic en '{selector}' para '{accion_prompt}'")
        self.logger.info(f"\n  --> Mensaje del prompt esperado: '{mensaje_prompt_esperado}'")
        if accion_prompt == 'accept':
            self.logger.info(f"\n  --> Texto a introducir: '{input_text}'")

        # Validar la acción y el input_text antes de iniciar la operación
        if accion_prompt not in ['accept', 'dismiss']:
            error_msg = f"\n❌ FALLO: Acción de prompt no válida: '{accion_prompt}'. Use 'accept' o 'dismiss'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_accion_invalida", directorio)
            raise AssertionError(error_msg)
        if accion_prompt == 'accept' and input_text is None:
            error_msg = "\n❌ FALLO: 'input_text' no puede ser None cuando 'accion_prompt' es 'accept'."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_input_text_missing", directorio)
            raise AssertionError(error_msg)
        if accion_prompt == 'dismiss' and input_text is not None:
            self.logger.warning("\n⚠️ ADVERTENCIA: 'input_text' se ignora cuando 'accion_prompt' es 'dismiss'.")

        # Resetear el estado de las banderas para cada ejecución del test.
        # Esto es crucial para evitar que valores de una ejecución anterior afecten la actual.
        self._dialogo_detectado = False
        self._dialogo_mensaje_capturado = ""
        self._dialogo_tipo_capturado = ""
        self._dialogo_input_capturado = "" # Resetear también el input capturado del handler

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar visibilidad y habilitación del selector que disparará el prompt
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}' (timeout: {tiempo_espera_elemento}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad y habilitación del elemento ---
            start_time_element_ready = time.time()
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic
            # --- Medición de rendimiento: Fin de visibilidad y habilitación del elemento ---
            end_time_element_ready = time.time()
            duration_element_ready = end_time_element_ready - start_time_element_ready
            self.logger.info(f"PERFORMANCE: Tiempo para que el elemento disparador esté listo: {duration_element_ready:.4f} segundos.")
            
            self.tomar_captura(f"{nombre_base}_elemento_listo_para_prompt", directorio)

            # 2. Registrar el listener ANTES de la acción que dispara el prompt
            self.logger.debug("\n  --> Registrando listener para el prompt con page.once('dialog')...")
            # Usa page.once para que el listener se desregistre automáticamente después de detectar el primer diálogo.
            # El handler `_get_prompt_dialog_handler_for_on()` también acepta/cancela la confirmación internamente.
            self.page.once("dialog", self._get_prompt_dialog_handler_for_on(input_text, accion_prompt))

            # 3. Hacer clic en el botón que dispara el prompt
            self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
            # --- Medición de rendimiento: Inicio de click y espera de detección del prompt ---
            start_time_click_and_prompt_detection = time.time()
            selector.click()

            # 4. Esperar a que el listener haya detectado y manejado el prompt
            self.logger.debug(f"\n  --> Esperando a que el prompt sea detectado y manejado por el listener (timeout: {tiempo_max_deteccion_prompt}s)...")
            # Bucle de espera activa hasta que la bandera _dialogo_detectado sea True
            # Se añade un timeout para el bucle, calculado a partir de tiempo_max_deteccion_prompt
            wait_end_time = time.time() + tiempo_max_deteccion_prompt
            while not self._dialogo_detectado and time.time() < wait_end_time:
                time.sleep(0.1) # Pausa breve para evitar consumo excesivo de CPU

            # --- Medición de rendimiento: Fin de click y espera de detección del prompt ---
            end_time_click_and_prompt_detection = time.time()
            duration_click_and_prompt_detection = end_time_click_and_prompt_detection - start_time_click_and_prompt_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección del prompt por el listener: {duration_click_and_prompt_detection:.4f} segundos.")

            if not self._dialogo_detectado:
                error_msg = f"\n❌ FALLO: El prompt no fue detectado por el listener después de {tiempo_max_deteccion_prompt} segundos."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_prompt_NO_detectada_timeout", directorio)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            self.tomar_captura(f"{nombre_base}_prompt_detectado_por_listener", directorio)
            self.logger.info(f"\n  ✅  Prompt detectado con éxito por el listener.")

            # 5. Validaciones después de que el listener ha actuado
            # --- Medición de rendimiento: Inicio de verificación de contenido del prompt ---
            start_time_dialog_content_verification = time.time()
            if self._dialogo_tipo_capturado != "prompt":
                self.logger.error(f"\n⚠️ Tipo de diálogo inesperado: '{self._dialogo_tipo_capturado}'. Se esperaba 'prompt'.")
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(f"\nTipo de diálogo inesperado: '{self._dialogo_tipo_capturado}'. Se esperaba 'prompt'.")

            if mensaje_prompt_esperado not in self._dialogo_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_prompt_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Mensaje del prompt incorrecto.\n"
                    f"  --> Esperado (contiene): '{mensaje_prompt_esperado}'\n"
                    f"  --> Obtenido: '{self._dialogo_mensaje_capturado}'"
                )
                self.logger.error(error_msg)
                # Re-lanzar como AssertionError para un fallo claro de la prueba
                raise AssertionError(error_msg)
            
            # Verificar que el texto introducido (si es el caso) se ha guardado correctamente
            if accion_prompt == 'accept' and self._dialogo_input_capturado != input_text:
                self.tomar_captura(f"{nombre_base}_prompt_input_incorrecto", directorio)
                error_msg = (
                    f"\n❌ FALLO: Texto introducido en el prompt incorrecto.\n"
                    f"  --> Esperado: '{input_text}'\n"
                    f"  --> Obtenido (capturado): '{self._dialogo_input_capturado}'"
                )
                self.logger.error(error_msg)
                raise AssertionError(error_msg)

            # --- Medición de rendimiento: Fin de verificación de contenido del prompt ---
            end_time_dialog_content_verification = time.time()
            duration_dialog_content_verification = end_time_dialog_content_verification - start_time_dialog_content_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de tipo, mensaje y texto introducido del prompt: {duration_dialog_content_verification:.4f} segundos.")

            # El prompt ya fue aceptado/cancelado por el handler `_get_prompt_dialog_handler_for_on()`.
            self.logger.info(f"\n  ✅  --> Prompt manejado (acción '{accion_prompt}' por el listener).")

            # Nota: La verificación del resultado en la página se considera una aserción separada
            # que debe realizarse después de esta función para desacoplar responsabilidades.
            # Sin embargo, a modo de ejemplo, se puede mantener aquí si la aplicación es simple.
            # Aquí se eliminará para reflejar un mejor desacoplamiento de esta función.

            self.tomar_captura(f"{nombre_base}_prompt_exitosa_{accion_prompt}", directorio)
            self.logger.info(f"\n✅  --> ÉXITO: El prompt se mostró, mensaje y texto verificado, y acción '{accion_prompt}' completada correctamente.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificación de prompt por listener): {duration_total_operation:.4f} segundos.")

            return True

        except TimeoutError as e:
            # Captura si el selector no está listo, si el prompt no aparece a tiempo, o si la verificación post-acción falla.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): El elemento '{selector}' no estuvo listo, "
                f"el prompt no apareció/fue detectado por el listener después de {tiempo_max_deteccion_prompt} segundos.\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_prompt_NO_detectada_timeout", directorio)
            raise AssertionError(f"\nTimeout al verificar prompt para selector '{selector}'") from e

        except Error as e:
            # Captura errores específicos de Playwright (ej. click fallido, problemas con el diálogo).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise AssertionError(f"\nError de Playwright al verificar prompt para selector '{selector}'") from e

        except AssertionError as e:
            # Captura las AssertionError lanzadas internamente por la función (acción inválida, tipo de diálogo, mensaje incorrecto).
            self.logger.critical(f"\n❌ FALLO (Validación de Prompt): {e}", exc_info=True)
            # La captura ya se tomó en la lógica interna donde se lanzó el AssertionError
            raise # Re-lanzar la excepción original para que el framework la maneje

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise AssertionError(f"\nError inesperado al verificar prompt para selector '{selector}'") from e

        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
        
    # 46- Función para esperar por una nueva pestaña/página (popup) que se haya abierto
    # y cambia el foco de la instancia 'page' actual a esa nueva pestaña.
    # Integra mediciones de rendimiento para la apertura y carga de la nueva página.
    def abrir_y_cambiar_a_nueva_pestana(self, selector_boton_apertura: Locator, nombre_base: str, directorio: str, tiempo_espera_max_total: Union[int, float] = 1.5) -> Optional[Page]:
        """
        Espera por la apertura de una nueva pestaña/página (popup) después de hacer clic
        en un elemento dado, cambia el foco de la instancia 'page' actual a esa nueva pestaña,
        y mide el rendimiento de este proceso.

        Args:
            selector_boton_apertura (Locator): El **Locator de Playwright** del botón o elemento
                                               que, al ser clicado, dispara la apertura de una nueva pestaña/ventana.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_espera_max_total (Union[int, float]): **Tiempo máximo total de espera** (en segundos)
                                                         para todo el proceso: desde el clic hasta que la nueva
                                                         página esté completamente cargada y lista. Por defecto, `15.0` segundos.

        Returns:
            Optional[Page]: El objeto `Page` de la nueva pestaña/ventana si se abrió y cargó correctamente.
                            Retorna `None` si ocurre un `TimeoutError` o un fallo durante el proceso.

        Raises:
            AssertionError: Si el elemento disparador no está disponible, o si ocurre un error inesperado
                            durante la interacción o la espera.
        """
        self.logger.info(f"\n🔄 Preparando para hacer clic en '{selector_boton_apertura}' y esperar nueva pestaña/popup. Esperando hasta {tiempo_espera_max_total} segundos...")

        nueva_pagina = None
        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar que el botón es visible y habilitado antes de hacer clic
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector_boton_apertura}'...")
            # Aquí puedes reutilizar un tiempo de espera más corto para la validación inicial del elemento si lo deseas,
            # o usar el tiempo_espera_max_total. Para simplicidad, se usará el total aquí.
            expect(selector_boton_apertura).to_be_visible()
            expect(selector_boton_apertura).to_be_enabled()
            selector_boton_apertura.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual

            # 2. Usar page.context.expect_event("page") para esperar la nueva página
            # y realizar la acción de click DENTRO de este contexto.
            # Esto asegura que la página capturada es la que se abre DESPUÉS del click.
            self.logger.debug(f"\n  --> Configurando listener para nueva página y haciendo clic en '{selector_boton_apertura}'...")
            # El timeout de expect_event cubre el tiempo desde el clic hasta que Playwright detecta la nueva página.
            with self.page.context.expect_event("page") as event_info:
                # --- Medición de rendimiento: Inicio de click y detección de nueva página ---
                start_time_click_and_new_page_detection = time.time()
                # Realizar el clic en el botón que abre la nueva pestaña
                self.hacer_click_en_elemento(selector_boton_apertura, f"{nombre_base}_click_para_nueva_pestana", directorio, tiempo_espera_max_total)
            
            # Obtener el objeto 'Page' de la nueva pestaña
            nueva_pagina = event_info.value 
            # --- Medición de rendimiento: Fin de click y detección de nueva página ---
            end_time_click_and_new_page_detection = time.time()
            duration_click_and_new_page_detection = end_time_click_and_new_page_detection - start_time_click_and_new_page_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la nueva página: {duration_click_and_new_page_detection:.4f} segundos.")
            
            # 3. Esperar a que la nueva página cargue completamente el DOM y los recursos (load state)
            self.logger.debug(f"\n  --> Esperando que la nueva página cargue completamente (Load state, timeout: {tiempo_espera_max_total}s)...")
            # --- Medición de rendimiento: Inicio de carga de nueva página ---
            start_time_new_page_load = time.time()
            nueva_pagina.wait_for_load_state("load")
            # --- Medición de rendimiento: Fin de carga de nueva página ---
            end_time_new_page_load = time.time()
            duration_new_page_load = end_time_new_page_load - start_time_new_page_load
            self.logger.info(f"PERFORMANCE: Tiempo de carga (load state) de la nueva página: {duration_new_page_load:.4f} segundos.")

            # 4. Esperar a que un elemento clave de la nueva página sea visible (ej. body o un elemento específico)
            # Esto es más relevante para el rendimiento percibido por el usuario.
            self.logger.debug(f"\n  --> Esperando que el 'body' de la nueva página sea visible (timeout: {tiempo_espera_max_total}s)...")
            # --- Medición de rendimiento: Inicio de visibilidad de contenido de nueva página ---
            start_time_new_page_content_visible = time.time()
            expect(nueva_pagina.locator("body")).to_be_visible()
            # --- Medición de rendimiento: Fin de visibilidad de contenido de nueva página ---
            end_time_new_page_content_visible = time.time()
            duration_new_page_content_visible = end_time_new_page_content_visible - start_time_new_page_content_visible
            self.logger.info(f"PERFORMANCE: Tiempo hasta que el 'body' de la nueva página fue visible: {duration_new_page_content_visible:.4f} segundos.")

            self.logger.info(f"\n✅ Nueva pestaña abierta y detectada: URL = {nueva_pagina.url}, Título = {nueva_pagina.title}")
            
            # 5. Actualizar self.page para que las subsiguientes operaciones usen la nueva página
            self.page = nueva_pagina 
            self.tomar_captura(f"{nombre_base}_nueva_pestana_abierta_y_cargada", directorio)
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (apertura y cambio a nueva pestaña): {duration_total_operation:.4f} segundos.")

            return nueva_pagina

        except TimeoutError as e:
            # Captura si la nueva página no se abre o no carga a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_total_operation
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido): No se detectó ninguna nueva pestaña/página después de {tiempo_espera_max_total} segundos "
                f"al intentar hacer clic en el botón de apertura ('{selector_boton_apertura}'), o la nueva página no cargó completamente/no mostró su contenido.\n"
                f"La operación duró {duration_fail:.4f} segundos antes del fallo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_no_se_detecto_popup_timeout", directorio)
            # Re-lanzar como AssertionError para que el framework de pruebas registre un fallo.
            raise AssertionError(f"\nTimeout al abrir o cargar nueva pestaña para selector '{selector_boton_apertura}'") from e
            # Retornar None si prefieres manejar el error en el nivel superior y no lanzar.
            # return None 
        except Error as e:
            # Captura errores específicos de Playwright (ej. clic fallido).
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la nueva pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_abrir_pestana", directorio)
            raise AssertionError(f"\nError de Playwright al abrir y cambiar a nueva pestaña para selector '{selector_boton_apertura}'") from e
        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar abrir y cambiar a la nueva pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_abrir_pestana", directorio)
            raise AssertionError(f"\nError inesperado al abrir y cambiar a nueva pestaña para selector '{selector_boton_apertura}'") from e
        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.

    # 47- Función que cierra la pestaña actual y, si hay otras pestañas abiertas en el mismo contexto,
    # cambia el foco de la instancia 'page' a la primera pestaña disponible.
    # Integra mediciones de rendimiento para el cierre y el cambio de foco.
    def cerrar_pestana_actual(self, nombre_base: str, directorio: str, tiempo_post_cierre: Union[int, float] = 1.0) -> None:
        """
        Cierra la pestaña Playwright actualmente activa (`self.page`).
        Si quedan otras pestañas abiertas en el mismo contexto del navegador,
        cambia el foco (`self.page`) a la primera pestaña disponible.
        Mide el rendimiento de las operaciones de cierre y cambio de foco.

        Args:
            nombre_base (str): Nombre base utilizado para la **captura de pantalla**
                               tomada antes de cerrar la pestaña.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            tiempo_post_cierre (Union[int, float]): **Tiempo de espera** (en segundos) después de
                                                    cerrar la pestaña, antes de intentar cambiar el foco.
                                                    Por defecto, `1.0` segundos.

        Raises:
            AssertionError: Si ocurre un error inesperado durante el cierre o el cambio de foco.
                            Se lanza para asegurar que el test falle si la operación no es exitosa.
        """
        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        # Guardar la URL actual antes de cerrarla (para logging)
        current_page_url = "N/A (Página ya cerrada o no accesible)"
        try:
            current_page_url = self.page.url
            self.logger.info(f"\n🚪 Cerrando la pestaña actual: URL = {current_page_url}")
        except Exception as e:
            self.logger.warning(f"\nNo se pudo obtener la URL de la página actual antes de intentar cerrarla: {e}")


        try:
            # ¡IMPORTANTE! Tomar la captura *antes* de cerrar la página.
            self.tomar_captura(f"{nombre_base}_antes_de_cerrar", directorio) 
            
            self.logger.debug(f"\n  --> Iniciando cierre de la página: {current_page_url}")
            # --- Medición de rendimiento: Inicio del cierre de la pestaña ---
            start_time_close_page = time.time()
            self.page.close()
            # --- Medición de rendimiento: Fin del cierre de la pestaña ---
            end_time_close_page = time.time()
            duration_close_page = end_time_close_page - start_time_close_page
            self.logger.info(f"PERFORMANCE: Tiempo de cierre de la pestaña: {duration_close_page:.4f} segundos.")
            
            self.logger.info(f"\n✅ Pestaña con URL '{current_page_url}' cerrada exitosamente.")
            
            # Pequeña espera después de cerrar la pestaña para asegurar que el DOM se libere
            self.esperar_fijo(tiempo_post_cierre) 

            # Verificar si hay otras páginas abiertas en el contexto y cambiar el foco
            self.logger.debug("\n  --> Verificando otras pestañas en el contexto para cambiar el foco...")
            # --- Medición de rendimiento: Inicio del cambio de foco ---
            start_time_switch_focus = time.time()
            if self.page.context.pages:
                # Playwright mantiene automáticamente la lista de páginas abiertas.
                # Al cerrar una página, si era la única, la lista se vacía.
                # Si hay más, la primera página en la lista es generalmente la que queda activa o la primera en crearse.
                self.page = self.page.context.pages[0] # Cambia el foco a la primera página disponible
                # --- Medición de rendimiento: Fin del cambio de foco ---
                end_time_switch_focus = time.time()
                duration_switch_focus = end_time_switch_focus - start_time_switch_focus
                self.logger.info(f"PERFORMANCE: Tiempo de cambio de foco a la nueva pestaña activa: {duration_switch_focus:.4f} segundos.")

                self.logger.info(f"\n🔄 Foco cambiado automáticamente a la primera pestaña disponible: URL = {self.page.url}")
                # Opcional: Podrías tomar otra captura aquí si quieres mostrar el estado de la nueva pestaña activa.
                # self.tomar_captura(f"{nombre_base}_foco_cambiado", directorio)
            else:
                self.logger.warning("\n⚠️ No hay más pestañas abiertas en el contexto del navegador. La instancia 'self.page' ahora es None.")
                self.page = None # No hay página activa en este contexto

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (cierre de pestaña y cambio de foco): {duration_total_operation:.4f} segundos.")

        except Error as e:
            # Captura errores específicos de Playwright, como si la página ya está cerrada o el contexto se cerró.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar cerrar la pestaña o cambiar de foco.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            # No se toma captura aquí porque la página podría estar inactiva/cerrada.
            raise AssertionError(f"\nError de Playwright al cerrar pestaña actual: {e}") from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar cerrar la pestaña actual o cambiar el foco.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            # No se toma captura aquí porque la página podría estar inactiva/cerrada.
            raise AssertionError(f"\nError inesperado al cerrar pestaña actual: {e}") from e
        finally:
            # Este bloque se ejecuta siempre, independientemente del resultado.
            self.esperar_fijo(0.2) # Pequeña espera final para observación o para liberar recursos.
        
    # 48- Función para hacer clic en un selector y esperar que se abran nuevas ventanas/pestañas.
    # Retorna una lista de objetos Page para las nuevas ventanas.
    # Integra pruebas de rendimiento para la detección y carga de múltiples páginas.
    def hacer_clic_y_abrir_nueva_ventana(self, selector: Locator, nombre_base: str, directorio: str, nombre_paso: str = "", tiempo_espera_max_total: Union[int, float] = 3.0) -> List[Page]:
        """
        Hace clic en un selector y espera que se abran una o más nuevas ventanas/pestañas (popups).
        Captura las nuevas páginas utilizando un listener global (`context.on("page")`),
        espera a que cada una cargue completamente y mide el rendimiento del proceso.

        Args:
            selector (Locator): El **Locator de Playwright** del elemento (ej. botón, enlace)
                                que, al ser clicado, dispara la apertura de nuevas ventanas/pestañas.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str): Una descripción opcional del paso que se está ejecutando.
            tiempo_espera_max_total (Union[int, float]): **Tiempo máximo total de espera** (en segundos)
                                                         para todo el proceso: desde el clic hasta que
                                                         al menos una nueva página se detecta y se cargan
                                                         todas las detectadas. Por defecto, `30.0` segundos.

        Returns:
            List[Page]: Una lista de objetos `Page` que representan las nuevas ventanas/pestañas abiertas
                        y cargadas correctamente. Retorna una lista vacía si ocurre un `TimeoutError`
                        o si no se detectan nuevas páginas.

        Raises:
            AssertionError: Si ocurre un error inesperado durante el clic, la detección o la carga de las páginas.
        """
        self.logger.info(f"\n--- {nombre_paso}: Haciendo clic en '{selector}' para abrir nuevas ventanas/pestañas ---")
        self.tomar_captura(f"{nombre_base}_antes_clic_nueva_ventana", directorio)

        # Limpiar la lista de páginas detectadas antes de cada interacción.
        # Esto es crucial para asegurar que solo se capturan las páginas de la ejecución actual.
        self._all_new_pages_opened_by_click = []

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Validar que el elemento es visible y habilitado antes de hacer clic
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del elemento '{selector}'...")
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            selector.highlight()
            self.esperar_fijo(0.2) # Pequeña pausa visual antes del clic

            # 2. Hacer clic en el elemento que debería abrir la(s) nueva(s) ventana(s)
            self.logger.debug(f"\n  --> Realizando clic en '{selector}'...")
            # --- Medición de rendimiento: Inicio del clic ---
            start_time_click = time.time()
            selector.click()
            # --- Medición de rendimiento: Fin del clic ---
            end_time_click = time.time()
            duration_click = end_time_click - start_time_click
            self.logger.info(f"PERFORMANCE: Tiempo de la acción de clic: {duration_click:.4f} segundos.")
            
            # 3. Esperar a que al menos una nueva página sea detectada por el listener
            # Usamos un bucle de espera activa con un timeout para dar tiempo a que los popups aparezcan
            self.logger.debug(f"\n  --> Esperando detección de nueva(s) ventana(s) por el listener (timeout: {tiempo_espera_max_total}s)...")
            # --- Medición de rendimiento: Inicio de la espera de detección de páginas ---
            start_time_page_detection = time.time()
            wait_for_detection_end_time = time.time() + tiempo_espera_max_total
            while not self._all_new_pages_opened_by_click and time.time() < wait_for_detection_end_time:
                time.sleep(0.1) # Pausa breve para evitar consumo excesivo de CPU

            if not self._all_new_pages_opened_by_click:
                raise TimeoutError(f"\nNo se detectó ninguna nueva ventana/pestaña después de hacer clic en '{selector}' dentro del tiempo de espera de {tiempo_espera_max_total} segundos.")
            
            # --- Medición de rendimiento: Fin de la espera de detección de páginas ---
            end_time_page_detection = time.time()
            duration_page_detection = end_time_page_detection - start_time_page_detection
            self.logger.info(f"PERFORMANCE: Tiempo desde el clic hasta la detección de la primera nueva página: {duration_page_detection:.4f} segundos.")

            # 4. Esperar a que cada una de las nuevas páginas cargue completamente
            self.logger.debug(f"\n  --> Esperando la carga completa de {len(self._all_new_pages_opened_by_click)} nueva(s) página(s)...")
            loaded_pages = []
            for i, new_page in enumerate(self._all_new_pages_opened_by_click):
                try:
                    self.logger.debug(f"\n    --> Cargando página {i+1}/{len(self._all_new_pages_opened_by_click)}: URL inicial = {new_page.url}")
                    # --- Medición de rendimiento: Inicio de carga de página individual ---
                    start_time_single_page_load = time.time()
                    
                    # Esperar los estados de carga con el timeout global
                    new_page.wait_for_load_state("load", timeout=tiempo_espera_max_total * 1000)
                    new_page.wait_for_load_state("domcontentloaded", timeout=tiempo_espera_max_total * 1000)
                    new_page.wait_for_load_state("networkidle", timeout=tiempo_espera_max_total * 1000)
                    
                    # Opcional: Esperar a que el 'body' o un elemento clave sea visible para garantizar renderizado.
                    # expect(new_page.locator("body")).to_be_visible(timeout=tiempo_espera_max_total * 1000)

                    # --- Medición de rendimiento: Fin de carga de página individual ---
                    end_time_single_page_load = time.time()
                    duration_single_page_load = end_time_single_page_load - start_time_single_page_load
                    self.logger.info(f"PERFORMANCE: Tiempo de carga completa para página {i+1} (URL: {new_page.url}): {duration_single_page_load:.4f} segundos.")
                    
                    self.logger.info(f"\n  ✅ Nueva página cargada exitosamente: URL = {new_page.url}, Título = {new_page.title}")
                    self.tomar_captura(f"{nombre_base}_pagina_abierta_{i+1}", directorio, page_to_capture=new_page)
                    loaded_pages.append(new_page)

                except TimeoutError as te:
                    self.logger.error(f"\n  ❌ FALLO: Tiempo de espera excedido al cargar la página {i+1} (URL: {new_page.url}). Detalles: {te}")
                    self.tomar_captura(f"{nombre_base}_pagina_no_cargada_{i+1}", directorio, page_to_capture=new_page)
                except Error as pe:
                    self.logger.error(f"\n  ❌ FALLO: Error de Playwright al interactuar con la página {i+1} (URL: {new_page.url}). Detalles: {pe}")
                    self.tomar_captura(f"{nombre_base}_pagina_error_playwright_{i+1}", directorio, page_to_capture=new_page)
                except Exception as ex:
                    self.logger.error(f"\n  ❌ FALLO: Error inesperado al cargar la página {i+1} (URL: {new_page.url}). Detalles: {ex}")
                    self.tomar_captura(f"{nombre_base}_pagina_error_inesperado_{i+1}", directorio, page_to_capture=new_page)

            if not loaded_pages:
                self.logger.error(f"\n ❌ FALLO: Ninguna de las nuevas ventanas/pestañas se cargó correctamente.")
                self.tomar_captura(f"{nombre_base}_ninguna_ventana_cargada", directorio)
                # Re-lanzar un AssertionError si no se pudo cargar ninguna página
                raise AssertionError("\nNinguna de las nuevas ventanas/pestañas se cargó correctamente.")

            self.tomar_captura(f"{nombre_base}_despues_clic_nueva_ventana_final", directorio)
            self.logger.info(f"\n✅ Se han detectado y cargado {len(loaded_pages)} de {len(self._all_new_pages_opened_by_click)} nueva(s) ventana(s) con éxito.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer clic y abrir/cargar nuevas ventanas): {duration_total_operation:.4f} segundos.")

            return loaded_pages

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Tiempo de espera excedido) - {nombre_paso}: No se detectó ninguna nueva ventana "
                f"después de hacer clic en '{selector}' dentro del tiempo de espera de {tiempo_espera_max_total} segundos, "
                f"o el elemento no estuvo visible/habilitado a tiempo.\nDetalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_no_nueva_ventana_timeout", directorio)
            # Re-lanzar como AssertionError para que el test falle correctamente.
            raise AssertionError(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n❌ FALLO (Playwright) - {nombre_paso}: Error de Playwright al hacer clic o al detectar/interactuar con las nuevas ventanas.\nDetalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_abrir_ventanas", directorio)
            raise AssertionError(error_msg) from e

        except AssertionError as e:
            # Captura las aserciones lanzadas internamente (ej. ninguna página cargada correctamente)
            self.logger.critical(f"\n❌ FALLO (Validación) - {nombre_paso}: {e}", exc_info=True)
            raise # Re-lanzar la excepción original para que el test falle

        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al intentar abrir nuevas ventanas.\nDetalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_abrir_nueva_ventana", directorio)
            raise AssertionError(error_msg) from e

        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.

    # 49- Función para cambiar el foco del navegador a una ventana/pestaña específica,
    # ya sea por su índice (int) o por una parte de su URL o título (str).
    # Integra mediciones de rendimiento para la búsqueda y el cambio de foco.
    def cambiar_foco_entre_ventanas(self, opcion_ventana: Union[int, str], nombre_base: str, directorio: str, nombre_paso: str = "") -> Page:
        """
        Cambia el foco de la instancia 'self.page' a una ventana/pestaña específica
        dentro del mismo contexto del navegador. La ventana objetivo puede ser identificada
        por su índice numérico o por una subcadena presente en su URL o título.
        Mide el rendimiento de la operación de cambio de foco.

        Args:
            opcion_ventana (Union[int, str]): El **criterio para seleccionar la ventana/pestaña objetivo**:
                                              - Si es `int`: el índice (0-basado) de la pestaña en la lista de páginas abiertas.
                                              - Si es `str`: una subcadena que debe coincidir con parte de la URL o el título de la pestaña.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str): Una descripción opcional del paso que se está ejecutando para los logs.

        Returns:
            Page: El objeto `Page` de la ventana/pestaña a la que se ha cambiado el foco exitosamente.

        Raises:
            IndexError: Si se proporciona un índice fuera de rango.
            ValueError: Si no se encuentra ninguna pestaña que coincida con la subcadena.
            TypeError: Si el tipo de `opcion_ventana` no es `int` ni `str`.
            AssertionError: Si ocurre un error inesperado durante el proceso de cambio de foco.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando cambiar el foco a la ventana/pestaña: '{opcion_ventana}' ---")
        
        target_page_to_focus: Optional[Page] = None
        
        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Obtener todas las páginas actuales en el contexto del navegador
            self.logger.debug("\n  --> Recuperando todas las páginas en el contexto del navegador...")
            # --- Medición de rendimiento: Inicio de recuperación de páginas ---
            start_time_get_pages = time.time()
            all_pages_in_context = self.page.context.pages
            # --- Medición de rendimiento: Fin de recuperación de páginas ---
            end_time_get_pages = time.time()
            duration_get_pages = end_time_get_pages - start_time_get_pages
            self.logger.info(f"PERFORMANCE: Tiempo de recuperación de todas las páginas en el contexto: {duration_get_pages:.4f} segundos.")

            self.logger.info(f"\n  Ventanas/pestañas abiertas actualmente: {len(all_pages_in_context)}")
            for i, p in enumerate(all_pages_in_context):
                try:
                    self.logger.info(f"\n    [{i}] URL: {p.url} | Título: {p.title()}")
                except Exception as e:
                    self.logger.warning(f"\n    [{i}] No se pudo obtener URL/Título: {e}")

            # 2. Buscar la página objetivo basada en la opción_ventana
            self.logger.debug(f"\n  --> Buscando la página objetivo '{opcion_ventana}'...")
            # --- Medición de rendimiento: Inicio de búsqueda de página objetivo ---
            start_time_find_target_page = time.time()

            if isinstance(opcion_ventana, int):
                if 0 <= opcion_ventana < len(all_pages_in_context):
                    target_page_to_focus = all_pages_in_context[opcion_ventana]
                    self.logger.info(f"  --> Seleccionada por índice: {opcion_ventana}")
                else:
                    error_msg = f"\n❌ FALLO: El índice '{opcion_ventana}' está fuera del rango de pestañas abiertas (0-{len(all_pages_in_context)-1})."
                    self.logger.error(error_msg)
                    self.tomar_captura(f"{nombre_base}_error_indice_invalido", directorio)
                    raise IndexError(error_msg)
            elif isinstance(opcion_ventana, str):
                # Intentar encontrar por URL o título
                found_match = False
                for p in all_pages_in_context:
                    try:
                        if opcion_ventana in p.url or opcion_ventana in p.title():
                            target_page_to_focus = p
                            found_match = True
                            self.logger.info(f"\n  --> Seleccionada por coincidencia de URL/Título: '{opcion_ventana}' (URL: {p.url}, Título: {p.title()})")
                            break
                    except Error as e:
                        # La página podría haberse cerrado justo en el momento de acceder a URL/title
                        self.logger.warning(f"\n  --> Error de Playwright al acceder a URL/título de una página durante la búsqueda: {e}")
                
                if not found_match:
                    error_msg = f"\n❌ FALLO: No se encontró ninguna pestaña con la URL o título que contenga '{opcion_ventana}'."
                    self.logger.error(error_msg)
                    self.tomar_captura(f"{nombre_base}_error_no_coincidencia_foco", directorio)
                    raise ValueError(error_msg)
            else:
                error_msg = f"\n❌ FALLO: El tipo de 'opcion_ventana' no es válido. Debe ser int o str (tipo recibido: {type(opcion_ventana).__name__})."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_error_tipo_opcion_foco", directorio)
                raise TypeError(error_msg)
            
            # --- Medición de rendimiento: Fin de búsqueda de página objetivo ---
            end_time_find_target_page = time.time()
            duration_find_target_page = end_time_find_target_page - start_time_find_target_page
            self.logger.info(f"PERFORMANCE: Tiempo de búsqueda de la página objetivo: {duration_find_target_page:.4f} segundos.")

            # 3. Cambiar el foco si la página objetivo no es la actual
            if target_page_to_focus == self.page:
                self.logger.info(f"\n✅ El foco ya está en la ventana seleccionada (URL: {self.page.url}). No es necesario cambiar.")
            else:
                self.logger.debug(f"\n  --> Cambiando el foco de '{self.page.url}' a '{target_page_to_focus.url}'...")
                # --- Medición de rendimiento: Inicio del cambio de foco ---
                start_time_switch_focus = time.time()
                self.page = target_page_to_focus
                # --- Medición de rendimiento: Fin del cambio de foco ---
                end_time_switch_focus = time.time()
                duration_switch_focus = end_time_switch_focus - start_time_switch_focus
                self.logger.info(f"PERFORMANCE: Tiempo de asignación del foco (self.page = ...): {duration_switch_focus:.4f} segundos.")
                
                self.logger.info(f"\n✅ Foco cambiado exitosamente a la ventana/pestaña seleccionada.")
            
            # 4. Reportar el estado final y tomar captura
            self.logger.info(f"\n  URL de la pestaña actual: {self.page.url}")
            self.logger.info(f"\n  Título de la pestaña actual: {self.page.title()}")
            self.tomar_captura(f"{nombre_base}_foco_cambiado", directorio)

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (cambio de foco entre ventanas): {duration_total_operation:.4f} segundos.")
            
            return self.page # Retorna la página a la que se cambió el foco

        except (IndexError, ValueError, TypeError) as e:
            # Captura errores de validación de entrada o de búsqueda de la página
            self.logger.critical(f"\n❌ FALLO (Validación) - {nombre_paso}: {e}", exc_info=True)
            # La captura ya se tomó en los bloques if/elif donde se lanzó el error
            raise # Re-lanzar la excepción original para que el test falle

        except Error as e:
            # Captura errores específicos de Playwright
            error_msg = (
                f"\n❌ FALLO (Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar cambiar el foco de ventana.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_cambiar_foco", directorio)
            raise AssertionError(error_msg) from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al intentar cambiar el foco de ventana.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_cambiar_foco", directorio)
            raise AssertionError(error_msg) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.

    # 50- Función que cierra un objeto 'Page' específico.
    # Si la página cerrada era la página activa (self.page), intenta cambiar el foco
    # a la primera página disponible en el mismo contexto del navegador.
    # Integra mediciones de rendimiento para el cierre de la pestaña y el posible cambio de foco.
    def cerrar_pestana_especifica(self, page_to_close: Page, nombre_base: str, directorio: str, nombre_paso: str = "") -> None:
        """
        Cierra un objeto `Page` específico proporcionado.
        Si la página que se va a cerrar es la actualmente activa (`self.page`),
        la función intentará cambiar el foco a la primera página disponible
        en el contexto del navegador. Mide el rendimiento de estas operaciones.

        Args:
            page_to_close (Page): El objeto `Page` específico que se desea cerrar.
            nombre_base (str): Nombre base utilizado para las **capturas de pantalla**
                               tomadas durante la ejecución de la función.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str): Una descripción opcional del paso que se está ejecutando para los logs.

        Raises:
            AssertionError: Si ocurre un error de Playwright o un error inesperado
                            durante el cierre de la pestaña o el cambio de foco.
        """
        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            closed_url = "N/A (Página no válida o ya cerrada)"
            try:
                # Intenta obtener la URL para el log, pero maneja el error si la página ya está cerrada
                if page_to_close and not page_to_close.is_closed():
                    closed_url = page_to_close.url
                self.logger.info(f"\n--- {nombre_paso}: Intentando cerrar la pestaña con URL: {closed_url} ---")
            except Error as e:
                self.logger.warning(f"\nNo se pudo obtener la URL de la página a cerrar. Podría estar inactiva: {e}")

            if not page_to_close or page_to_close.is_closed():
                self.logger.info(f"\n ℹ️ La pestaña (URL: {closed_url}) ya estaba cerrada o no es un objeto Page válido. No se requiere acción.")
                # --- Medición de rendimiento: Fin total de la función (sin acción real) ---
                end_time_total_operation = time.time()
                duration_total_operation = end_time_total_operation - start_time_total_operation
                self.logger.info(f"PERFORMANCE: Tiempo total de la operación (pestaña ya cerrada): {duration_total_operation:.4f} segundos.")
                return # Salir si la página ya está cerrada o no es válida

            # 1. Determinar si la página a cerrar es la página actual (self.page)
            # --- Medición de rendimiento: Inicio de la detección de página actual ---
            start_time_is_current_page_check = time.time()
            is_current_page = (self.page == page_to_close)
            # --- Medición de rendimiento: Fin de la detección de página actual ---
            end_time_is_current_page_check = time.time()
            duration_is_current_page_check = end_time_is_current_page_check - start_time_is_current_page_check
            self.logger.info(f"PERFORMANCE: Tiempo de verificación si es la página actual: {duration_is_current_page_check:.4f} segundos.")

            self.logger.debug(f"\n  --> Tomando captura antes de cerrar la pestaña: {closed_url}")
            self.tomar_captura(f"{nombre_base}_antes_de_cerrar_especifica", directorio, page_to_capture=page_to_close)
            
            # 2. Cerrar la pestaña específica
            self.logger.debug(f"\n  --> Procediendo a cerrar la pestaña: {closed_url}")
            # --- Medición de rendimiento: Inicio del cierre de la pestaña ---
            start_time_close_page = time.time()
            page_to_close.close()
            # --- Medición de rendimiento: Fin del cierre de la pestaña ---
            end_time_close_page = time.time()
            duration_close_page = end_time_close_page - start_time_close_page
            self.logger.info(f"PERFORMANCE: Tiempo de cierre de la pestaña '{closed_url}': {duration_close_page:.4f} segundos.")
            
            self.logger.info(f"\n✅ Pestaña '{closed_url}' cerrada exitosamente.")
            # No se toma una captura después de cerrar la página porque ya no es accesible.

            # 3. Si la página cerrada era la página actual (self.page), cambiar el foco
            if is_current_page:
                self.logger.info("\n  --> Detectado: La pestaña cerrada era la pestaña activa.")
                # --- Medición de rendimiento: Inicio del cambio de foco ---
                start_time_switch_focus = time.time()
                # Buscar la primera página disponible en el contexto
                if self.page.context.pages:
                    self.page = self.page.context.pages[0]
                    # --- Medición de rendimiento: Fin del cambio de foco ---
                    end_time_switch_focus = time.time()
                    duration_switch_focus = end_time_switch_focus - start_time_switch_focus
                    self.logger.info(f"PERFORMANCE: Tiempo de cambio de foco a la nueva pestaña activa: {duration_switch_focus:.4f} segundos.")

                    self.logger.info(f"\n🔄 Foco cambiado automáticamente a la primera pestaña disponible: URL = {self.page.url}")
                    self.tomar_captura(f"{nombre_base}_foco_cambiado_despues_cerrar", directorio, page_to_capture=self.page)
                else:
                    self.logger.warning("\n⚠️ No hay más pestañas abiertas en el contexto del navegador. La instancia 'self.page' ahora es None.")
                    self.page = None # No hay página activa en este contexto
                    self.logger.info("PERFORMANCE: No se realizó cambio de foco, no hay más páginas en el contexto.")
            else:
                self.logger.info("\n  --> La pestaña cerrada no era la pestaña activa. El foco actual permanece sin cambios.")
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (cierre de pestaña específica y gestión de foco): {duration_total_operation:.4f} segundos.")

        except Error as e: # Captura errores específicos de Playwright
            # Esto puede ocurrir si la página ya se cerró por alguna razón externa, o si hubo un problema con el contexto.
            if "Target page, context or browser has been closed" in str(e) or "Page closed" in str(e):
                self.logger.warning(f"\n⚠️ Advertencia (Playwright): La pestaña ya estaba cerrada o el contexto ya no es válido durante la operación. Detalles: {e}")
                # En este caso, no necesitamos relanzar una excepción, ya que el objetivo (cerrar la página)
                # se cumple implícitamente o la página ya estaba en el estado deseado.
                # Asegúrate de que el estado de self.page es consistente si se cerró la activa
                if self.page and self.page.is_closed():
                    self.logger.warning("\n  --> La página activa se ha cerrado. Intentando reasignar el foco.")
                    if self.page.context.pages:
                        self.page = self.page.context.pages[0]
                        self.logger.info(f"\n  --> Foco reasignado a: {self.page.url}")
                    else:
                        self.page = None
                        self.logger.warning("\n  --> No hay más páginas en el contexto. self.page es None.")
            else:
                error_msg = (
                    f"\n❌ FALLO (Playwright Error) - {nombre_paso}: Ocurrió un error de Playwright al intentar cerrar la pestaña.\n"
                    f"Detalles: {e}"
                )
                self.logger.critical(error_msg, exc_info=True)
                self.tomar_captura(f"{nombre_base}_error_cerrar_pestana_playwright", directorio)
                raise AssertionError(error_msg) from e
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error al intentar cerrar la pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_cerrar_pestana", directorio)
            raise AssertionError(error_msg) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
            
    #51- Función para realizar una operación de "Drag and Drop" de un elemento a otro.
    def realizar_drag_and_drop(self, elemento_origen: Locator, elemento_destino: Locator, nombre_base: str, directorio: str, nombre_paso: str = "", tiempo_espera_manual: float = 0.5, timeout_ms: int = 15000) -> None:
        """
        Realiza una operación de "Drag and Drop" de un elemento de origen a un elemento de destino.
        Intenta primero con el método estándar de Playwright (`locator.drag_to()`).
        Si el método estándar falla (ej. por `TimeoutError` u otro `Playwright Error`),
        recurre a un método manual que simula las acciones de ratón (`hover`, `mouse.down`, `mouse.up`).
        Integra pruebas de rendimiento para ambos enfoques.

        Args:
            elemento_origen (Locator): El **Locator** del elemento que se desea arrastrar.
            elemento_destino (Locator): El **Locator** del área o elemento donde se desea soltar el elemento arrastrado.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            tiempo_espera_manual (float, opcional): Tiempo en segundos para las pausas entre las acciones
                                                   del ratón en el método manual (aplicado con `esperar_fijo`). Por defecto `0.5` segundos.
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la operación de Drag and Drop
                                        (tanto para `drag_to` como para las validaciones iniciales y pasos manuales).
                                        Por defecto `15000`ms (15 segundos).

        Raises:
            AssertionError: Si la operación de Drag and Drop (estándar o manual) falla,
                            o si los elementos no están listos para la interacción.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando realizar 'Drag and Drop' de '{elemento_origen}' a '{elemento_destino}' ---")
        
        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Pre-validación: Verificar que ambos elementos estén visibles y habilitados antes de interactuar.
            self.logger.info(f"\n🔍 Validando que el elemento de origen '{elemento_origen}' esté habilitado y listo para interactuar...")
            # --- Medición de rendimiento: Inicio pre-validación ---
            start_time_pre_validation = time.time()
            expect(elemento_origen).to_be_enabled()
            expect(elemento_destino).to_be_enabled()
            # --- Medición de rendimiento: Fin pre-validación ---
            end_time_pre_validation = time.time()
            duration_pre_validation = end_time_pre_validation - start_time_pre_validation
            self.logger.info(f"PERFORMANCE: Tiempo de pre-validación de elementos: {duration_pre_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ Ambos elementos están habilitados y listos para 'Drag and Drop'.")
            self.tomar_captura(f"{nombre_base}_antes_drag_and_drop", directorio)

            # 2. Intento 1: Usar el método .drag_to() del Locator (recomendado por Playwright)
            self.logger.info(f"\n🔄 Intentando 'Drag and Drop' con el método estándar de Playwright (locator.drag_to())...")
            # --- Medición de rendimiento: Inicio drag_to ---
            start_time_drag_to = time.time()
            try:
                elemento_origen.drag_to(elemento_destino)
                # --- Medición de rendimiento: Fin drag_to ---
                end_time_drag_to = time.time()
                duration_drag_to = end_time_drag_to - start_time_drag_to
                self.logger.info(f"PERFORMANCE: Tiempo del método estándar 'drag_to': {duration_drag_to:.4f} segundos.")

                self.logger.info(f"\n✅ 'Drag and Drop' realizado exitosamente con el método estándar.")
                self.tomar_captura(f"{nombre_base}_drag_and_drop_exitoso_estandar", directorio)
                
                # --- Medición de rendimiento: Fin total de la función ---
                end_time_total_operation = time.time()
                duration_total_operation = end_time_total_operation - start_time_total_operation
                self.logger.info(f"PERFORMANCE: Tiempo total de la operación (estándar D&D): {duration_total_operation:.4f} segundos.")
                return # Si funciona, salimos de la función

            except (Error, TimeoutError) as e:
                # Captura errores específicos de Playwright (incluyendo TimeoutError de drag_to)
                self.logger.warning(f"\n⚠️ Advertencia: El método directo 'locator.drag_to()' falló con error de Playwright: {type(e).__name__}: {e}")
                self.logger.info("\n🔄 Recurriendo a 'Drag and Drop' con método manual de Playwright (mouse.hover, mouse.down, mouse.up)...")
                self.tomar_captura(f"{nombre_base}_fallo_directo_intentando_manual", directorio)
                
                # Registrar el rendimiento del intento fallido de drag_to
                end_time_drag_to = time.time() # Registrar el tiempo que tomó fallar
                duration_drag_to = end_time_drag_to - start_time_drag_to
                self.logger.info(f"PERFORMANCE: Tiempo del método estándar 'drag_to' (fallido): {duration_drag_to:.4f} segundos.")

                # 3. Intento 2 (Fallback): Usar el método manual
                self._realizar_drag_and_drop_manual(elemento_origen, elemento_destino, nombre_base, directorio, nombre_paso, tiempo_pausa_mouse=tiempo_espera_manual, timeout_ms=timeout_ms)
                self.logger.info(f"\n✅ 'Drag and Drop' realizado exitosamente con el método manual.")
                self.tomar_captura(f"{nombre_base}_drag_and_drop_exitoso_manual", directorio)

        except (Error, TimeoutError) as e: # Captura errores de Playwright que puedan ocurrir fuera del drag_to o en la pre-validación
            error_msg = (
                f"\n❌ FALLO (Playwright Error) - {nombre_paso}: Ocurrió un error de Playwright al realizar 'Drag and Drop'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_drag_and_drop", directorio)
            raise AssertionError(error_msg) from e
        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al intentar realizar 'Drag and Drop'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_drag_and_drop", directorio)
            raise AssertionError(error_msg) from e
        finally:
            # --- Medición de rendimiento: Fin total de la función (si no se salió antes) ---
            if 'start_time_total_operation' in locals() and 'end_time_total_operation' not in locals():
                end_time_total_operation = time.time()
                duration_total_operation = end_time_total_operation - start_time_total_operation
                self.logger.info(f"PERFORMANCE: Tiempo total de la operación (fallback manual D&D): {duration_total_operation:.4f} segundos.")
            
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
        
    # 52- Función para mover sliders de rango (con dos pulgares)
    # Integra pruebas de rendimiento para cada fase del movimiento de los pulgares.
    def mover_slider_rango(self, pulgar_izquierdo_locator: Locator, pulgar_derecho_locator: Locator, barra_slider_locator: Locator,
                            porcentaje_destino_izquierdo: float, porcentaje_destino_derecho: float,
                            nombre_base: str, directorio: str, nombre_paso: str = "",
                            tolerancia_pixeles: int = 3, timeout_ms: int = 15000) -> None:
        """
        Mueve los dos "pulgares" (handles) de un slider de rango horizontal a porcentajes de destino específicos.
        Utiliza las acciones de ratón de Playwright para simular el arrastre.
        Integra mediciones de rendimiento detalladas para cada paso del movimiento.

        Args:
            pulgar_izquierdo_locator (Locator): El **Locator** del pulgar izquierdo (mínimo) del slider.
            pulgar_derecho_locator (Locator): El **Locator** del pulgar derecho (máximo) del slider.
            barra_slider_locator (Locator): El **Locator** de la barra principal del slider (donde se mueven los pulgares).
            porcentaje_destino_izquierdo (float): El porcentaje de la barra (0.0 a 1.0) al que se moverá el pulgar izquierdo.
            porcentaje_destino_derecho (float): El porcentaje de la barra (0.0 a 1.0) al que se moverá el pulgar derecho.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str, opcional): Descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            tolerancia_pixeles (int, opcional): Margen de error en píxeles para considerar que un pulgar
                                                ya está en su posición deseada. Por defecto `3` píxeles.
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la visibilidad/habilitación
                                        de los elementos. Por defecto `15000`ms (15 segundos).

        Raises:
            ValueError: Si los porcentajes de destino son inválidos o el izquierdo es mayor que el derecho.
            RuntimeError: Si no se puede obtener el bounding box de los elementos.
            AssertionError: Si ocurre un error de Playwright o un error inesperado durante la interacción.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando mover el slider de rango. Pulgar Izquierdo a {porcentaje_destino_izquierdo*100:.0f}%, Pulgar Derecho a {porcentaje_destino_derecho*100:.0f}% ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        # 1. Validaciones iniciales de porcentajes
        if not (0.0 <= porcentaje_destino_izquierdo <= 1.0) or not (0.0 <= porcentaje_destino_derecho <= 1.0):
            error_msg = "\n❌ Los porcentajes de destino para ambos pulgares deben ser valores flotantes entre 0.0 (0%) y 1.0 (100%)."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_porcentajes", directorio)
            raise ValueError(error_msg)
        
        # Validación de negocio: el porcentaje izquierdo no puede ser mayor que el derecho
        if porcentaje_destino_izquierdo > porcentaje_destino_derecho:
            error_msg = "\n❌ El porcentaje del pulgar izquierdo no puede ser mayor que el del pulgar derecho."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_orden_porcentajes", directorio)
            raise ValueError(error_msg)
        
        elementos_a_validar: Dict[str, Locator] = {
            "pulgar izquierdo": pulgar_izquierdo_locator,
            "pulgar derecho": pulgar_derecho_locator,
            "barra del slider": barra_slider_locator
        }

        try:
            # 2. Pre-validación: Verificar visibilidad y habilitación de todos los elementos
            self.logger.info("\n🔍 Validando visibilidad y habilitación de los elementos del slider...")
            # --- Medición de rendimiento: Inicio pre-validación ---
            start_time_pre_validation = time.time()
            for nombre_elemento, localizador_elemento in elementos_a_validar.items():
                expect(localizador_elemento).to_be_visible()
                expect(localizador_elemento).to_be_enabled()
                localizador_elemento.highlight() # Para visualización durante la ejecución
                self.esperar_fijo(0.1) # Pequeña pausa para que se vea el highlight
            
            # --- Medición de rendimiento: Fin pre-validación ---
            end_time_pre_validation = time.time()
            duration_pre_validation = end_time_pre_validation - start_time_pre_validation
            self.logger.info(f"PERFORMANCE: Tiempo de pre-validación de elementos del slider: {duration_pre_validation:.4f} segundos.")
            self.logger.info("\n✅ Todos los elementos del slider están visibles y habilitados.")
            self.tomar_captura(f"{nombre_base}_slider_elementos_listos", directorio)

            # 3. Obtener el bounding box de la barra del slider (esencial para el cálculo de posiciones)
            self.logger.debug("\n  --> Obteniendo bounding box de la barra del slider...")
            # --- Medición de rendimiento: Inicio obtener bounding box ---
            start_time_get_bounding_box = time.time()
            caja_barra = barra_slider_locator.bounding_box()
            if not caja_barra:
                raise RuntimeError(f"\n❌ No se pudo obtener el bounding box de la barra del slider '{barra_slider_locator}'.")
            # --- Medición de rendimiento: Fin obtener bounding box ---
            end_time_get_bounding_box = time.time()
            duration_get_bounding_box = end_time_get_bounding_box - start_time_get_bounding_box
            self.logger.info(f"PERFORMANCE: Tiempo de obtención de bounding box de la barra: {duration_get_bounding_box:.4f} segundos.")

            inicio_x_barra = caja_barra['x']
            ancho_barra = caja_barra['width']
            posicion_y_barra = caja_barra['y'] + (caja_barra['height'] / 2) # Y central de la barra para movimientos

            # --- 4. Mover Pulgar Izquierdo (Mínimo) ---
            self.logger.info(f"\n🔄 Moviendo pulgar izquierdo a {porcentaje_destino_izquierdo*100:.0f}%...")
            # --- Medición de rendimiento: Inicio movimiento pulgar izquierdo ---
            start_time_move_left_thumb = time.time()

            caja_pulgar_izquierdo = pulgar_izquierdo_locator.bounding_box()
            if not caja_pulgar_izquierdo:
                raise RuntimeError(f"\n❌ No se pudo obtener el bounding box del pulgar izquierdo '{pulgar_izquierdo_locator}'.")

            posicion_x_destino_izquierdo = inicio_x_barra + (ancho_barra * porcentaje_destino_izquierdo)
            # Usar la Y central de la barra para movimientos, para mantener una línea recta si el pulgar no es perfectamente redondo
            posicion_y_movimiento_izquierdo = posicion_y_barra 

            # Calcular la posición X central actual del pulgar izquierdo para iniciar el arrastre
            posicion_x_actual_izquierdo_centro = caja_pulgar_izquierdo['x'] + (caja_pulgar_izquierdo['width'] / 2)

            # Verificar si el pulgar izquierdo ya está en la posición deseada dentro de la tolerancia
            if abs(posicion_x_actual_izquierdo_centro - posicion_x_destino_izquierdo) < tolerancia_pixeles:
                self.logger.info(f"\n  > Pulgar izquierdo ya se encuentra en la posición deseada ({porcentaje_destino_izquierdo*100:.0f}%). No se requiere movimiento.")
            else:
                self.logger.info(f"\n  > Iniciando arrastre de pulgar izquierdo de X={posicion_x_actual_izquierdo_centro:.0f} a X={posicion_x_destino_izquierdo:.0f}...")
                
                # Acciones del ratón para el arrastre
                self.logger.debug("\n    -> mouse.move al origen")
                self.page.mouse.move(posicion_x_actual_izquierdo_centro, posicion_y_movimiento_izquierdo) # Mover al centro del pulgar actual
                self.esperar_fijo(0.1) # Pequeña pausa
                
                self.logger.debug("\n    -> mouse.down")
                self.page.mouse.down() # Presionar el botón del ratón
                self.esperar_fijo(0.2) # Pausa para simular la interacción humana
                
                self.logger.debug("\n    -> mouse.move al destino (arrastrando)")
                self.page.mouse.move(posicion_x_destino_izquierdo, posicion_y_movimiento_izquierdo, steps=10) # Arrastrar suavemente
                self.esperar_fijo(0.2) # Pausa para simular la interacción humana
                
                self.logger.debug("\n    -> mouse.up")
                self.page.mouse.up() # Soltar el botón del ratón
                self.logger.info(f"\n  > Pulgar izquierdo movido a X={posicion_x_destino_izquierdo:.0f}.")
            
            # --- Medición de rendimiento: Fin movimiento pulgar izquierdo ---
            end_time_move_left_thumb = time.time()
            duration_move_left_thumb = end_time_move_left_thumb - start_time_move_left_thumb
            self.logger.info(f"PERFORMANCE: Tiempo de movimiento de pulgar izquierdo: {duration_move_left_thumb:.4f} segundos.")
            self.tomar_captura(f"{nombre_base}_slider_izquierdo_movido", directorio)
            self.esperar_fijo(0.5) # Pausa adicional después de procesar el primer pulgar para estabilización

            # --- 5. Mover Pulgar Derecho (Máximo) ---
            self.logger.info(f"\n🔄 Moviendo pulgar derecho a {porcentaje_destino_derecho*100:.0f}%...")
            # --- Medición de rendimiento: Inicio movimiento pulgar derecho ---
            start_time_move_right_thumb = time.time()

            caja_pulgar_derecho = pulgar_derecho_locator.bounding_box()
            if not caja_pulgar_derecho:
                raise RuntimeError(f"\n❌ No se pudo obtener el bounding box del pulgar derecho '{pulgar_derecho_locator}'.")

            posicion_x_destino_derecho = inicio_x_barra + (ancho_barra * porcentaje_destino_derecho)
            # Usar la Y central de la barra para movimientos
            posicion_y_movimiento_derecho = posicion_y_barra 

            # Calcular la posición X central actual del pulgar derecho para iniciar el arrastre
            posicion_x_actual_derecho_centro = caja_pulgar_derecho['x'] + (caja_pulgar_derecho['width'] / 2)

            # Verificar si el pulgar derecho ya está en la posición deseada dentro de la tolerancia
            if abs(posicion_x_actual_derecho_centro - posicion_x_destino_derecho) < tolerancia_pixeles:
                self.logger.info(f"\n  > Pulgar derecho ya se encuentra en la posición deseada ({porcentaje_destino_derecho*100:.0f}%). No se requiere movimiento.")
            else:
                self.logger.info(f"\n  > Iniciando arrastre de pulgar derecho de X={posicion_x_actual_derecho_centro:.0f} a X={posicion_x_destino_derecho:.0f}...")
                
                # Acciones del ratón para el arrastre
                self.logger.debug("\n    -> mouse.move al origen")
                self.page.mouse.move(posicion_x_actual_derecho_centro, posicion_y_movimiento_derecho) # Mover al centro del pulgar actual
                self.esperar_fijo(0.1) # Pequeña pausa
                
                self.logger.debug("\n    -> mouse.down")
                self.page.mouse.down() # Presionar el botón del ratón
                self.esperar_fijo(0.2) # Pausa para simular la interacción humana
                
                self.logger.debug("\n    -> mouse.move al destino (arrastrando)")
                self.page.mouse.move(posicion_x_destino_derecho, posicion_y_movimiento_derecho, steps=10) # Arrastrar suavemente
                self.esperar_fijo(0.2) # Pausa para simular la interacción humana
                
                self.logger.debug("    -> mouse.up")
                self.page.mouse.up() # Soltar el botón del ratón
                self.logger.info(f"\n  > Pulgar derecho movido a X={posicion_x_destino_derecho:.0f}.")
            
            # --- Medición de rendimiento: Fin movimiento pulgar derecho ---
            end_time_move_right_thumb = time.time()
            duration_move_right_thumb = end_time_move_right_thumb - start_time_move_right_thumb
            self.logger.info(f"PERFORMANCE: Tiempo de movimiento de pulgar derecho: {duration_move_right_thumb:.4f} segundos.")

            self.logger.info(f"\n✅ Slider de rango procesado exitosamente. Izquierdo a {porcentaje_destino_izquierdo*100:.0f}%, Derecho a {porcentaje_destino_derecho*100:.0f}%.")
            self.tomar_captura(f"{nombre_base}_slider_rango_procesado_{int(porcentaje_destino_izquierdo*100)}_{int(porcentaje_destino_derecho*100)}pc_final", directorio)

            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (mover slider de rango): {duration_total_operation:.4f} segundos.")

        except (ValueError, RuntimeError) as e:
            # Captura errores de validación de entrada o de obtención de bounding box
            self.logger.critical(f"\n❌ FALLO (Validación/Configuración) - {nombre_paso}: {e}", exc_info=True)
            # La captura ya se tomó en los bloques if/elif donde se lanzó el error de validación
            raise AssertionError(f"\nError de validación/configuración en mover_slider_rango: {e}") from e

        except (Error, TimeoutError) as e:
            # Captura errores específicos de Playwright, incluyendo TimeoutError de expect()
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar mover el slider de rango.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_slider_rango", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada
            mensaje_error = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al intentar mover el slider de rango.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_slider_rango", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
    
    # 53- Función para seleccionar una opción en un ComboBox (elemento <select>) por su atributo 'value'.
    # Integra pruebas de rendimiento para las fases de validación, selección y verificación.
    def seleccionar_opcion_por_valor(self, combobox_locator: Locator, valor_a_seleccionar: str, nombre_base: str, directorio: str, nombre_paso: str = "", timeout_ms: int = 15000) -> None:
        """
        Selecciona una opción dentro de un elemento ComboBox (`<select>`) utilizando su atributo 'value'.
        La función valida la visibilidad y habilitación del ComboBox, realiza la selección y
        verifica que la opción haya sido aplicada correctamente.
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            combobox_locator (Locator): El **Locator** del elemento `<select>` (ComboBox).
            valor_a_seleccionar (str): El **valor del atributo 'value'** de la opción `<option>` que se desea seleccionar.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la visibilidad,
                                        habilitación y verificación de la selección. Por defecto `15000`ms (15 segundos).

        Raises:
            AssertionError: Si el ComboBox no es visible/habilitado, la opción no se puede seleccionar,
                            la selección no se verifica correctamente o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Iniciando selección de '{valor_a_seleccionar}' en ComboBox por valor: '{combobox_locator}' ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n🔍 Esperando que el ComboBox '{combobox_locator}' sea visible y habilitado...")
            # --- Medición de rendimiento: Inicio validación/espera ---
            start_time_validation = time.time()
            expect(combobox_locator).to_be_visible()
            combobox_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_locator).to_be_enabled()
            # --- Medición de rendimiento: Fin validación/espera ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo de validación de visibilidad y habilitación: {duration_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox '{combobox_locator}' es visible y habilitado.")
            
            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_combo", directorio)

            # 3. Seleccionar la opción por su valor
            self.logger.info(f"\n🔄 Seleccionando opción '{valor_a_seleccionar}' en '{combobox_locator}'...")
            # --- Medición de rendimiento: Inicio selección ---
            start_time_selection = time.time()
            combobox_locator.select_option(value=valor_a_seleccionar, timeout=timeout_ms) # Asegúrate de pasar el 'value=' explícitamente si es necesario
            # --- Medición de rendimiento: Fin selección ---
            end_time_selection = time.time()
            duration_selection = end_time_selection - start_time_selection
            self.logger.info(f"PERFORMANCE: Tiempo de selección de la opción: {duration_selection:.4f} segundos.")
            
            self.logger.info(f"\n✅ Opción '{valor_a_seleccionar}' seleccionada exitosamente en '{combobox_locator}'.")

            # 4. Verificar que la opción fue seleccionada correctamente
            self.logger.info(f"\n🔍 Verificando que ComboBox '{combobox_locator}' tenga el valor '{valor_a_seleccionar}'...")
            # --- Medición de rendimiento: Inicio verificación ---
            start_time_verification = time.time()
            expect(combobox_locator).to_have_value(valor_a_seleccionar, timeout=timeout_ms)
            # --- Medición de rendimiento: Fin verificación ---
            end_time_verification = time.time()
            duration_verification = end_time_verification - start_time_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de la selección: {duration_verification:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox '{combobox_locator}' verificado con valor '{valor_a_seleccionar}'.")

            # 5. Tomar captura después de la selección exitosa
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_combo_exito", directorio)
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (seleccionar ComboBox): {duration_total_operation:.4f} segundos.")

        except TimeoutError as e:
            # Captura TimeoutError específicamente para mensajes más claros
            mensaje_error = (
                f"\n❌ FALLO (Timeout) - {nombre_paso}: El ComboBox '{combobox_locator}' "
                f"no se volvió visible/habilitado o la opción '{valor_a_seleccionar}' no se pudo seleccionar/verificar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_combo", directorio)
            raise AssertionError(mensaje_error) from e

        except Error as e:
            # Captura otros errores de Playwright
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar seleccionar la opción '{valor_a_seleccionar}' en '{combobox_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select>, opción no existe, o ComboBox no interactuable.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_combo", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            # Captura cualquier otra excepción inesperada
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado) - {nombre_paso}: Ocurrió un error desconocido al manejar el ComboBox '{combobox_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_combo", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
        
    # 54- Función para seleccionar una opción en un ComboBox (elemento <select>) por su texto visible (label).
    # Integra pruebas de rendimiento para las fases de validación, selección y verificación.
    def seleccionar_opcion_por_label(self, combobox_locator: Locator, label_a_seleccionar: str, nombre_base: str, directorio: str, value_esperado: Optional[str] = None, nombre_paso: str = "", timeout_ms: int = 15000) -> None:
        """
        Selecciona una opción dentro de un elemento ComboBox (`<select>`) utilizando su texto visible (label).
        La función valida la visibilidad y habilitación del ComboBox, realiza la selección y
        verifica que la opción haya sido aplicada correctamente, ya sea por su 'value' esperado
        o asumiendo que el 'value' es igual al 'label'.
        Integra mediciones de rendimiento para cada fase de la operación.

        Args:
            combobox_locator (Locator): El **Locator** del elemento `<select>` (ComboBox).
            label_a_seleccionar (str): El **texto visible (label)** de la opción `<option>` que se desea seleccionar.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            value_esperado (str, opcional): El **valor del atributo 'value'** que se espera que tenga el ComboBox
                                            después de seleccionar la opción por su label. Si no se proporciona,
                                            se asume que `value_esperado` es igual a `label_a_seleccionar`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la visibilidad,
                                        habilitación y verificación de la selección. Por defecto `15000`ms (15 segundos).

        Raises:
            AssertionError: Si el ComboBox no es visible/habilitado, la opción no se puede seleccionar,
                            la selección no se verifica correctamente o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Iniciando selección de '{label_a_seleccionar}' en ComboBox por label: '{combobox_locator}' ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n🔍 Esperando que el ComboBox '{combobox_locator}' sea visible y habilitado...")
            # --- Medición de rendimiento: Inicio validación/espera ---
            start_time_validation = time.time()
            expect(combobox_locator).to_be_visible()
            combobox_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_locator).to_be_enabled()
            # --- Medición de rendimiento: Fin validación/espera ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo de validación de visibilidad y habilitación: {duration_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox '{combobox_locator}' es visible y habilitado.")
            
            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_combo_label", directorio)

            # 3. Seleccionar la opción por su texto visible (label)
            self.logger.info(f"\n🔄 Seleccionando opción con texto '{label_a_seleccionar}' en '{combobox_locator}'...")
            # --- Medición de rendimiento: Inicio selección ---
            start_time_selection = time.time()
            # El método select_option() espera automáticamente a que el elemento
            # sea visible, habilitado y con la opción disponible.
            combobox_locator.select_option(label=label_a_seleccionar) # Usa 'label=' para claridad
            # --- Medición de rendimiento: Fin selección ---
            end_time_selection = time.time()
            duration_selection = end_time_selection - start_time_selection
            self.logger.info(f"PERFORMANCE: Tiempo de selección de la opción por label: {duration_selection:.4f} segundos.")
            
            self.logger.info(f"\n✅ Opción '{label_a_seleccionar}' seleccionada exitosamente en '{combobox_locator}' por label.")

            # 4. Verificar que la opción fue seleccionada correctamente
            # Usamos to_have_value() para asegurar que el valor del select cambió al esperado.
            # Esto es más robusto que to_have_text() para <select>, ya que el texto visible puede variar
            # o incluir espacios, mientras que el 'value' es el dato real subyacente.
            valor_para_comparar_verificacion = value_esperado if value_esperado is not None else label_a_seleccionar
            
            self.logger.info(f"\n🔍 Verificando que ComboBox '{combobox_locator}' tenga el valor esperado '{valor_para_comparar_verificacion}'...")
            # --- Medición de rendimiento: Inicio verificación ---
            start_time_verification = time.time()
            expect(combobox_locator).to_have_value(valor_para_comparar_verificacion)
            # --- Medición de rendimiento: Fin verificación ---
            end_time_verification = time.time()
            duration_verification = end_time_verification - start_time_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de la selección: {duration_verification:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox '{combobox_locator}' verificado con valor seleccionado '{valor_para_comparar_verificacion}'.")

            # 5. Tomar captura después de la selección exitosa
            # Asegura que la captura refleje el estado final y el valor seleccionado
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_combo_label_exito", directorio)
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (seleccionar ComboBox por label): {duration_total_operation:.4f} segundos.")

        except TimeoutError as e:
            mensaje_error = (
                f"\n❌ FALLO (Timeout) - {nombre_paso}: El ComboBox '{combobox_locator}' "
                f"no se volvió visible/habilitado o la opción con label '{label_a_seleccionar}' no se pudo seleccionar/verificar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_combo_label", directorio)
            raise AssertionError(mensaje_error) from e

        except Error as e:
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error al intentar seleccionar la opción con label '{label_a_seleccionar}' en '{combobox_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select>, opción con ese label no existe, o ComboBox no interactuable.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_combo_label", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado) - {nombre_paso}: Ocurrió un error desconocido al manejar el ComboBox '{combobox_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_combo_label", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
    
    # 55- Función para presionar la tecla TAB en el teclado
    # Integra pruebas de rendimiento para medir el tiempo de ejecución de la acción.
    def Tab_Press(self, tiempo_espera_post_tab: float = 0.5, nombre_paso: str = "") -> None:
        """
        Simula la acción de presionar la tecla 'TAB' en el teclado.
        Esta función es útil para navegar entre elementos interactivos (inputs, botones, enlaces)
        en una página web, moviendo el foco al siguiente elemento tabulable.
        Integra mediciones de rendimiento para la operación.

        Args:
            tiempo_espera_post_tab (float, opcional): Tiempo en segundos para esperar *después* de presionar 'TAB'.
                                                      Útil para dar tiempo a que la UI procese el cambio de foco
                                                      o se carguen elementos dinámicamente. Por defecto `0.5` segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Raises:
            Exception: Si ocurre algún error inesperado durante la simulación de la tecla TAB.
        """
        self.logger.info(f"\n--- {nombre_paso}: Presionando la tecla TAB y esperando {tiempo_espera_post_tab} segundos ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # --- Medición de rendimiento: Inicio de la acción 'keyboard.press' ---
            start_time_press_action = time.time()
            self.page.keyboard.press("Tab")
            # --- Medición de rendimiento: Fin de la acción 'keyboard.press' ---
            end_time_press_action = time.time()
            duration_press_action = end_time_press_action - start_time_press_action
            self.logger.info(f"PERFORMANCE: Tiempo de la acción 'keyboard.press(\"Tab\")': {duration_press_action:.4f} segundos.")
            
            self.logger.info("\nTecla TAB presionada exitosamente.")

            # Espera fija después de presionar TAB (configuracion por parametro)
            if tiempo_espera_post_tab > 0:
                self.esperar_fijo(tiempo_espera_post_tab)

        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al presionar la tecla TAB.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            # En este caso, una captura de pantalla podría no ser tan útil,
            # ya que es una acción de teclado global, pero se podría añadir
            # si el contexto lo amerita (e.g., para ver el estado del foco).
            # self.tomar_captura(f"error_tab_press", "directorio_errores") # Descomentar si se desea una captura
            raise AssertionError(f"\nError al presionar la tecla TAB: {e}") from e
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (Tab_Press): {duration_total_operation:.4f} segundos.")
        
    # 56- Función optimizada para seleccionar múltiples opciones en un ComboBox múltiple.
    # Integra pruebas de rendimiento utilizando mediciones de tiempo para cada fase clave.
    def seleccionar_multiples_opciones_combo(self, combobox_multiple_locator: Locator, valores_a_seleccionar: List[str], nombre_base: str, directorio: str, nombre_paso: str = "", timeout_ms: int = 15000) -> None:
        """
        Selecciona múltiples opciones en un ComboBox (`<select multiple>`) por sus valores o labels.
        La función valida la visibilidad y habilitación del ComboBox, realiza la selección de
        todas las opciones especificadas y verifica que todas ellas hayan sido aplicadas correctamente.
        Integra mediciones de rendimiento detalladas para cada fase de la operación.

        Args:
            combobox_multiple_locator (Locator): El **Locator** del elemento `<select multiple>` (ComboBox múltiple).
            valores_a_seleccionar (List[str]): Una **lista de cadenas** que representan los 'value' o 'label'
                                              de las opciones que se desean seleccionar.
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la visibilidad,
                                        habilitación y verificación de la selección. Por defecto `15000`ms (15 segundos).

        Raises:
            AssertionError: Si el ComboBox no es visible/habilitado, las opciones no se pueden seleccionar,
                            la verificación de las selecciones falla o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Iniciando selección de múltiples opciones {valores_a_seleccionar} en ComboBox: '{combobox_multiple_locator}' ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n🔍 Esperando que el ComboBox múltiple '{combobox_multiple_locator}' sea visible y habilitado...")
            # --- Medición de rendimiento: Inicio validación/espera ---
            start_time_validation = time.time()
            expect(combobox_multiple_locator).to_be_visible()
            combobox_multiple_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_multiple_locator).to_be_enabled()
            # --- Medición de rendimiento: Fin validación/espera ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo de validación de visibilidad y habilitación: {duration_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox múltiple '{combobox_multiple_locator}' es visible y habilitado.")
            
            # Opcional: Verificar que sea un select múltiple.
            # Esta aserción es útil para fallar temprano si el locator no apunta al tipo de elemento correcto.
            self.logger.debug(f"\nVerificando que '{combobox_multiple_locator}' sea un <select multiple>...")
            expect(combobox_multiple_locator).to_have_attribute("multiple") # El atributo 'multiple' existe
            self.logger.debug("\n  > ComboBox verificado como select múltiple.")

            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_multi_combo", directorio)

            # 3. Seleccionar las opciones
            self.logger.info(f"\n🔄 Seleccionando opciones '{valores_a_seleccionar}' en '{combobox_multiple_locator}'...")
            # --- Medición de rendimiento: Inicio selección de múltiples opciones ---
            start_time_selection = time.time()
            # Playwright's select_option() para listas maneja tanto valores como labels.
            # Pasando una lista de strings seleccionará las opciones correspondientes.
            combobox_multiple_locator.select_option(valores_a_seleccionar)
            # --- Medición de rendimiento: Fin selección de múltiples opciones ---
            end_time_selection = time.time()
            duration_selection = end_time_selection - start_time_selection
            self.logger.info(f"PERFORMANCE: Tiempo de selección de las múltiples opciones: {duration_selection:.4f} segundos.")
            
            self.logger.info(f"\n✅ Opciones '{valores_a_seleccionar}' seleccionadas exitosamente en '{combobox_multiple_locator}'.")

            # 4. Verificar que las opciones fueron seleccionadas correctamente
            self.logger.info(f"\n🔍 Verificando que ComboBox múltiple '{combobox_multiple_locator}' tenga los valores seleccionados: {valores_a_seleccionar}...")
            # --- Medición de rendimiento: Inicio verificación de selecciones ---
            start_time_verification = time.time()
            # to_have_values() es la aserción correcta para verificar múltiples selecciones por su 'value'.
            expect(combobox_multiple_locator).to_have_values(valores_a_seleccionar)
            # --- Medición de rendimiento: Fin verificación de selecciones ---
            end_time_verification = time.time()
            duration_verification = end_time_verification - start_time_verification
            self.logger.info(f"PERFORMANCE: Tiempo de verificación de las selecciones: {duration_verification:.4f} segundos.")
            
            self.logger.info(f"\n✅ ComboBox múltiple '{combobox_multiple_locator}' verificado con valores seleccionados: {valores_a_seleccionar}.")

            # 5. Tomar captura después de la selección exitosa
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_multi_combo_exito", directorio)
            
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (seleccionar ComboBox múltiple): {duration_total_operation:.4f} segundos.")

        except TimeoutError as e:
            mensaje_error = (
                f"\n❌ FALLO (Timeout) - {nombre_paso}: El ComboBox múltiple '{combobox_multiple_locator}' "
                f"no se volvió visible/habilitado o las opciones '{valores_a_seleccionar}' no se pudieron seleccionar/verificar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_multi_combo", directorio)
            raise AssertionError(mensaje_error) from e

        except Error as e:
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error al intentar seleccionar las opciones '{valores_a_seleccionar}' en '{combobox_multiple_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select multiple>, alguna opción no existe o el ComboBox no es interactuable.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_multi_combo", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado) - {nombre_paso}: Ocurrió un error desconocido al manejar el ComboBox múltiple '{combobox_multiple_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_multi_combo", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
        
    # 57- Función que obtiene y imprime los valores y el texto de todas las opciones en un dropdown list.
    # Integra pruebas de rendimiento para medir el tiempo de extracción de datos del dropdown.
    def obtener_valores_dropdown(self, selector_dropdown: Locator, nombre_base: str, directorio: str, nombre_paso: str = "", timeout_ms: int = 15000) -> Optional[List[Dict[str, str]]]:
        """
        Obtiene los atributos 'value' y el texto visible de todas las opciones (`<option>`)
        dentro de un elemento dropdown (`<select>`).
        La función valida la visibilidad y habilitación del dropdown antes de extraer los datos.
        Integra mediciones de rendimiento para cada fase clave de la extracción.

        Args:
            selector_dropdown (Locator): El **Locator** del elemento `<select>` (dropdown list).
            nombre_base (str): Nombre base para las **capturas de pantalla** tomadas durante la ejecución.
            directorio (str): **Ruta del directorio** donde se guardarán las capturas de pantalla.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la visibilidad
                                        y habilitación del dropdown. Por defecto `15000`ms (15 segundos).

        Returns:
            Optional[List[Dict[str, str]]]: Una lista de diccionarios, donde cada diccionario contiene
                                           'value' y 'text' de una opción. Retorna `None` si no se
                                           encuentran opciones.

        Raises:
            AssertionError: Si el dropdown no es visible/habilitado, o si ocurre un error inesperado
                            durante la extracción de los datos.
        """
        self.logger.info(f"\n--- {nombre_paso}: Extrayendo valores del dropdown '{selector_dropdown}' ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        valores_opciones: List[Dict[str, str]] = []

        try:
            # 1. Asegurar que el dropdown es visible y habilitado
            self.logger.info(f"\n🔍 Esperando que el dropdown '{selector_dropdown}' sea visible y habilitado...")
            # --- Medición de rendimiento: Inicio validación/espera ---
            start_time_validation = time.time()
            expect(selector_dropdown).to_be_visible()
            selector_dropdown.highlight() # Para visualización durante la ejecución
            expect(selector_dropdown).to_be_enabled()
            # --- Medición de rendimiento: Fin validación/espera ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo de validación de visibilidad y habilitación del dropdown: {duration_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ Dropdown '{selector_dropdown}' es visible y habilitado.")
            self.tomar_captura(f"{nombre_base}_dropdown_antes_extraccion", directorio)

            # 2. Obtener todos los locators de las opciones dentro del dropdown
            self.logger.info(f"\n🔄 Obteniendo locators de todas las opciones dentro de '{selector_dropdown}'...")
            # --- Medición de rendimiento: Inicio obtención de option locators ---
            start_time_get_options = time.time()
            option_locators = selector_dropdown.locator("option").all()
            # --- Medición de rendimiento: Fin obtención de option locators ---
            end_time_get_options = time.time()
            duration_get_options = end_time_get_options - start_time_get_options
            self.logger.info(f"PERFORMANCE: Tiempo de obtención de todos los option locators: {duration_get_options:.4f} segundos.")

            if not option_locators:
                self.logger.warning(f"\n⚠️ No se encontraron opciones dentro del dropdown '{selector_dropdown}'.")
                self.tomar_captura(f"{nombre_base}_dropdown_sin_opciones", directorio)
                return None

            self.logger.info(f"\n Encontradas {len(option_locators)} opciones para '{selector_dropdown}':")

            # 3. Iterar sobre cada opción y extraer su 'value' y 'text_content'
            self.logger.info("\n📊 Extrayendo valores y textos de cada opción...")
            # --- Medición de rendimiento: Inicio iteración y extracción ---
            start_time_extract_loop = time.time()
            for i, option_locator in enumerate(option_locators):
                value = option_locator.get_attribute("value")
                text = option_locator.text_content()

                # Limpieza de espacios en blanco
                clean_value = value.strip() if value is not None else "" # Manejo de None para get_attribute
                clean_text = text.strip() if text is not None else "" # Manejo de None para text_content

                valores_opciones.append({'value': clean_value, 'text': clean_text})
                self.logger.info(f"  Opción {i+1}: Value='{clean_value}', Text='{clean_text}'")
            # --- Medición de rendimiento: Fin iteración y extracción ---
            end_time_extract_loop = time.time()
            duration_extract_loop = end_time_extract_loop - start_time_extract_loop
            self.logger.info(f"PERFORMANCE: Tiempo de iteración y extracción de {len(option_locators)} opciones: {duration_extract_loop:.4f} segundos.")


            self.logger.info(f"\n✅ Valores obtenidos exitosamente del dropdown '{selector_dropdown}'.")
            self.tomar_captura(f"{nombre_base}_dropdown_valores_extraidos", directorio)
            return valores_opciones

        except TimeoutError as e:
            mensaje_error = (
                f"\n❌ FALLO (Timeout) - {nombre_paso}: El dropdown '{selector_dropdown}' "
                f"no se volvió visible/habilitado o sus opciones no cargaron a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_timeout", directorio)
            raise AssertionError(mensaje_error) from e

        except Error as e:
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar obtener los valores del dropdown '{selector_dropdown}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_playwright_error", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado) - {nombre_paso}: Ocurrió un error desconocido al intentar obtener los valores del dropdown '{selector_dropdown}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_inesperado", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (obtener valores dropdown): {duration_total_operation:.4f} segundos.")
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
        
    # 58- Función que obtiene y compara los valores y el texto de todas las opciones en un dropdown list.
    # Integra pruebas de rendimiento para medir el tiempo de extracción y comparación de datos.
    def obtener_y_comparar_valores_dropdown(self, dropdown_locator: Locator, nombre_base: str, directorio: str, expected_options: Optional[List[Union[str, Dict[str, str]]]] = None, compare_by_text: bool = True, compare_by_value: bool = False, nombre_paso: str = "", timeout_ms: int = 15000) -> Optional[List[Dict[str, str]]]:
        """
        Obtiene los atributos 'value' y el texto visible de todas las opciones (`<option>`)
        dentro de un elemento dropdown (`<select>`). Opcionalmente, compara las opciones obtenidas
        con una lista de opciones esperadas.
        La función valida la visibilidad y habilitación del dropdown y mide el rendimiento
        de la extracción y, si aplica, de la comparación.

        Args:
            dropdown_locator (Locator): El **Locator** de Playwright para el elemento `<select>` del dropdown.
            nombre_base (str): Nombre base para las **capturas de pantalla**.
            directorio (str): Directorio donde se guardarán las **capturas de pantalla**.
            expected_options (List[Union[str, Dict[str, str]]], optional):
                Lista de opciones esperadas para la comparación. Puede ser:
                - `List[str]`: Si solo se desea comparar por el texto visible de las opciones.
                - `List[Dict[str, str]]`: Si se desea comparar por 'value' y 'text'.
                  Ej: `[{'value': 'usa', 'text': 'Estados Unidos'}]`.
                Por defecto es `None` (no se realiza comparación).
            compare_by_text (bool): Si es `True`, compara el texto visible de las opciones.
                                  Usado si `expected_options` es `List[str]` o `List[Dict]`.
            compare_by_value (bool): Si es `True`, compara el atributo 'value' de las opciones.
                                   Usado si `expected_options` es `List[Dict]`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs y nombres de capturas. Por defecto "".
            timeout_ms (int): Tiempo máximo de espera en milisegundos para la visibilidad,
                              habilitación y la obtención de opciones. Por defecto `15000`ms (15 segundos).

        Returns:
            Optional[List[Dict[str, str]]]: Una lista de diccionarios con las opciones reales extraídas
                                           ({'value': '...', 'text': '...'}).
                                           Retorna `None` si no se encuentran opciones o si ocurre un error.

        Raises:
            AssertionError: Si el dropdown no es visible/habilitado, las opciones no se cargan,
                            si no se encuentran opciones cuando se esperaban,
                            o si la comparación de opciones falla.
        """
        self.logger.info(f"\n--- {nombre_paso}: Extrayendo y comparando valores del dropdown '{dropdown_locator}' ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        valores_opciones_reales: List[Dict[str, str]] = []

        try:
            # 1. Asegurar que el dropdown es visible y habilitado
            self.logger.info(f"\n🔍 Esperando que el dropdown '{dropdown_locator}' sea visible y habilitado...")
            # --- Medición de rendimiento: Inicio validación/espera ---
            start_time_validation = time.time()
            expect(dropdown_locator).to_be_visible()
            dropdown_locator.highlight() # Para visualización durante la ejecución
            expect(dropdown_locator).to_be_enabled()
            # --- Medición de rendimiento: Fin validación/espera ---
            end_time_validation = time.time()
            duration_validation = end_time_validation - start_time_validation
            self.logger.info(f"PERFORMANCE: Tiempo de validación de visibilidad y habilitación del dropdown: {duration_validation:.4f} segundos.")
            
            self.logger.info(f"\n✅ Dropdown '{dropdown_locator}' es visible y habilitado.")
            self.tomar_captura(f"{nombre_base}_dropdown_antes_extraccion_y_comparacion", directorio)

            # 2. Obtener todos los locators de las opciones dentro del dropdown
            self.logger.info(f"\n🔄 Obteniendo locators de todas las opciones dentro de '{dropdown_locator}'...")
            # --- Medición de rendimiento: Inicio obtención de option locators ---
            start_time_get_options = time.time()
            option_locators = dropdown_locator.locator("option").all()
            # --- Medición de rendimiento: Fin obtención de option locators ---
            end_time_get_options = time.time()
            duration_get_options = end_time_get_options - start_time_get_options
            self.logger.info(f"PERFORMANCE: Tiempo de obtención de todos los option locators: {duration_get_options:.4f} segundos.")

            if not option_locators:
                self.logger.warning(f"\n⚠️ No se encontraron opciones dentro del dropdown '{dropdown_locator}'.")
                self.tomar_captura(f"{nombre_base}_dropdown_sin_opciones", directorio)
                # Si se esperaban opciones y no hay ninguna, esto es un fallo de aserción.
                if expected_options:
                    raise AssertionError(f"\n❌ FALLO: No se encontraron opciones en el dropdown '{dropdown_locator}', pero se esperaban {len(expected_options)}.")
                return None

            self.logger.info(f"\n Encontradas {len(option_locators)} opciones reales para '{dropdown_locator}':")

            # 3. Iterar sobre cada opción y extraer su 'value' y 'text_content'
            self.logger.info("\n📊 Extrayendo valores y textos de cada opción...")
            # --- Medición de rendimiento: Inicio iteración y extracción ---
            start_time_extract_loop = time.time()
            for i, option_locator in enumerate(option_locators):
                value = option_locator.get_attribute("value")
                text = option_locator.text_content()

                # Limpieza de espacios en blanco
                # Asegura que value y text no sean None antes de strip().
                clean_value = value.strip() if value is not None else ""
                clean_text = text.strip() if text is not None else ""

                valores_opciones_reales.append({'value': clean_value, 'text': clean_text})
                self.logger.info(f"\n  Opción Real {i+1}: Value='{clean_value}', Text='{clean_text}'")
            # --- Medición de rendimiento: Fin iteración y extracción ---
            end_time_extract_loop = time.time()
            duration_extract_loop = end_time_extract_loop - start_time_extract_loop
            self.logger.info(f"PERFORMANCE: Tiempo de iteración y extracción de {len(option_locators)} opciones: {duration_extract_loop:.4f} segundos.")

            self.logger.info(f"\n✅ Valores obtenidos exitosamente del dropdown '{dropdown_locator}'.")
            self.tomar_captura(f"{nombre_base}_dropdown_valores_extraidos", directorio)

            # 4. Comparar con las opciones esperadas (si se proporcionan)
            if expected_options is not None:
                self.logger.info("\n--- Realizando comparación de opciones ---")
                # --- Medición de rendimiento: Inicio de la fase de comparación ---
                start_time_comparison = time.time()
                try:
                    expected_set = set()
                    real_set = set()

                    # Preparar los conjuntos para la comparación (normalizando a minúsculas y sin espacios extra)
                    for opt in expected_options:
                        if isinstance(opt, str):
                            if compare_by_text:
                                expected_set.add(opt.strip().lower())
                            else:
                                self.logger.warning(f"\n⚠️ Advertencia: Opciones esperadas en formato `str` pero `compare_by_text` es `False`. Ignorando '{opt}'.")
                        elif isinstance(opt, dict):
                            if compare_by_text and 'text' in opt and opt['text'] is not None:
                                expected_set.add(opt['text'].strip().lower())
                            if compare_by_value and 'value' in opt and opt['value'] is not None:
                                expected_set.add(opt['value'].strip().lower())
                            if not (compare_by_text or compare_by_value):
                                self.logger.warning(f"\n⚠️ Advertencia: `compare_by_text` y `compare_by_value` son `False`. Ninguna comparación se realizará para la opción esperada: {opt}.")
                        else:
                            self.logger.warning(f"\n⚠️ Advertencia: Formato de opción esperada no reconocido: '{opt}'. Ignorando.")

                    # Construir el conjunto de opciones reales para comparación
                    for opt_real in valores_opciones_reales:
                        if compare_by_text and 'text' in opt_real and opt_real['text'] is not None:
                            real_set.add(opt_real['text'].strip().lower())
                        if compare_by_value and 'value' in opt_real and opt_real['value'] is not None:
                            real_set.add(opt_real['value'].strip().lower())

                    # Comprobar si los conjuntos son idénticos
                    if expected_set == real_set:
                        self.logger.info("\n✅ ÉXITO: Las opciones del dropdown coinciden con las opciones esperadas.")
                        self.tomar_captura(f"{nombre_base}_dropdown_comparacion_exitosa", directorio)
                    else:
                        missing_in_real = list(expected_set - real_set)
                        missing_in_expected = list(real_set - expected_set)
                        error_msg = f"\n❌ FALLO: Las opciones del dropdown NO coinciden con las esperadas.\n"
                        if missing_in_real:
                            error_msg += f"  - Opciones esperadas no encontradas en el dropdown: {missing_in_real}\n"
                        if missing_in_expected:
                            error_msg += f"  - Opciones encontradas en el dropdown que no estaban esperadas: {missing_in_expected}\n"
                        self.logger.error(error_msg)
                        self.tomar_captura(f"{nombre_base}_dropdown_comparacion_fallida", directorio)
                        raise AssertionError(f"\nComparación de opciones del dropdown fallida para '{dropdown_locator}'. {error_msg.strip()}")

                except Exception as e:
                    self.logger.critical(f"\n❌ FALLO: Ocurrió un error durante la comparación de opciones: {e}", exc_info=True)
                    self.tomar_captura(f"{nombre_base}_dropdown_error_comparacion", directorio)
                    raise AssertionError(f"\nError al comparar opciones del dropdown '{dropdown_locator}': {e}") from e
                # --- Medición de rendimiento: Fin de la fase de comparación ---
                end_time_comparison = time.time()
                duration_comparison = end_time_comparison - start_time_comparison
                self.logger.info(f"PERFORMANCE: Tiempo de la fase de comparación: {duration_comparison:.4f} segundos.")

            return valores_opciones_reales

        except TimeoutError as e:
            mensaje_error = (
                f"\n❌ FALLO (Timeout) - {nombre_paso}: El dropdown '{dropdown_locator}' "
                f"no se volvió visible/habilitado o sus opciones no cargaron a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_timeout", directorio)
            raise AssertionError(mensaje_error) from e

        except Error as e:
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar obtener los valores del dropdown '{dropdown_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_playwright_error", directorio)
            raise AssertionError(mensaje_error) from e

        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado) - {nombre_paso}: Ocurrió un error desconocido al intentar obtener los valores del dropdown '{dropdown_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_inesperado", directorio)
            raise AssertionError(mensaje_error) from e
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (obtener y comparar valores dropdown): {duration_total_operation:.4f} segundos.")
            self.esperar_fijo(0.2) # Pequeña espera final para observación o liberar recursos.
    
    # 59- Función que detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo Excel.
    # Integra pruebas de rendimiento para medir el tiempo de lectura del archivo Excel.
    def num_Filas_excel(self, archivo_excel_path: str, hoja: str, has_header: bool = False, nombre_paso: str = "") -> int:
        """
        Detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo Excel.
        Opcionalmente, descuenta una fila para el encabezado si 'has_header' es True.
        Esta función mide el tiempo que tarda en cargar el archivo Excel y obtener el número de filas,
        lo cual es útil para pruebas de rendimiento en escenarios de procesamiento de datos.

        Args:
            archivo_excel_path (str): La **ruta completa al archivo Excel** (`.xlsx` o `.xlsm`).
            hoja (str): El **nombre de la hoja/pestaña** dentro del archivo Excel de la cual se desean contar las filas.
            has_header (bool, opcional): Si es `True`, se descuenta una fila del total
                                         para considerar que la primera fila es un encabezado.
                                         Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            int: El **número de filas de datos** en la hoja especificada.
                 Retorna `0` si el archivo no se encuentra, la hoja no existe, o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener el número de filas para la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        num_physical_rows = 0
        num_data_rows = 0

        try:
            self.logger.info(f"\n⏳ Cargando el libro de trabajo Excel: '{archivo_excel_path}'...")
            workbook = openpyxl.load_workbook(archivo_excel_path) # Carga el libro de trabajo Excel
            self.logger.info(f"\n✅ Libro de trabajo cargado. Seleccionando la hoja '{hoja}'...")
            sheet = workbook[hoja] # Selecciona la hoja específica del libro
            
            # Obtiene el número total de filas con contenido.
            # openpyxl.worksheet.max_row devuelve el índice de la última fila no vacía.
            num_physical_rows = sheet.max_row 

            if has_header and num_physical_rows > 0:
                # Si tiene encabezado y hay al menos una fila (el encabezado)
                num_data_rows = num_physical_rows - 1 # Resta 1 para no contar el encabezado
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en la hoja '{hoja}'.")
                return num_data_rows
            else:
                # Para hojas sin encabezado, o si num_physical_rows es 0 (hoja vacía),
                # el número de filas de datos es igual al número de filas físicas.
                num_data_rows = num_physical_rows
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas ocupadas en la hoja '{hoja}'.")
                return num_data_rows

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo Excel no se encontró en la ruta: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return 0
        except KeyError:
            error_msg = f"\n❌ FALLO (Hoja no encontrada): La hoja '{hoja}' no se encontró en el archivo Excel: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return 0
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el número de filas del Excel.\n"
                f"Archivo: '{archivo_excel_path}', Hoja: '{hoja}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return 0
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (num_Filas_excel): {duration_total_operation:.4f} segundos.")
            # Es importante cerrar el workbook si se ha abierto explícitamente y no con 'with open()'
            # Sin embargo, openpyxl.load_workbook no requiere un cierre explícito en la mayoría de los casos
            # ya que maneja el archivo internamente. Aun así, se puede añadir un log de depuración.
            self.logger.debug("\nFinalizada la operación de lectura de Excel.")

    # 60- Función que obtiene el valor de una celda específica de una hoja Excel,
    # ajustando la fila si se indica que hay un encabezado.
    # Integra pruebas de rendimiento para medir el tiempo de lectura de la celda.
    def dato_Columna_excel(self, archivo_excel_path: str, hoja: str, numero_fila_logica: int, nombre_o_indice_columna: Union[str, int], has_header_excel: bool = False, nombre_paso: str = "") -> Union[str, int, float, None]:
        """
        Obtiene el valor de una celda específica de una hoja de un archivo Excel.
        Ajusta el número de fila si se indica que la hoja tiene un encabezado.
        Permite especificar la columna por su nombre (si hay encabezado) o por su índice numérico.
        Esta función mide el tiempo que tarda en cargar el archivo, ubicar la columna/fila,
        y extraer el dato, lo cual es útil para identificar cuellos de botella en la lectura de datos.

        Args:
            archivo_excel_path (str): La **ruta completa al archivo Excel** (`.xlsx` o `.xlsm`).
            hoja (str): El **nombre de la hoja/pestaña** dentro del archivo Excel.
            numero_fila_logica (int): El **número de fila lógica** (basado en 1) de la celda a leer.
                                     Si `has_header_excel` es `True`, esta es la fila de datos
                                     (e.g., `1` para la primera fila después del encabezado).
            nombre_o_indice_columna (Union[str, int]): El **nombre del encabezado de la columna** (string)
                                                       o el **índice numérico de la columna** (entero, basado en 1).
            has_header_excel (bool, opcional): Si es `True`, indica que la hoja tiene un encabezado en la primera fila.
                                             Esto ajusta el cálculo de la fila física y permite la búsqueda por nombre de columna.
                                             Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            Union[str, int, float, None]: El valor de la celda. El tipo del valor se conserva,
                                          pero si se convierte a `str` para consumo general.
                                          Retorna `None` si el archivo no se encuentra, la hoja/columna no existe,
                                          la fila/columna está fuera de rango, o si ocurre un error.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener dato de la celda (Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna}) de la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header_excel}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        cell_value: Any = None # Inicializamos el valor de la celda

        try:
            # --- Medición de rendimiento: Carga del Workbook y selección de hoja ---
            start_time_load_workbook = time.time()
            self.logger.info(f"\n⏳ Cargando el libro de trabajo Excel: '{archivo_excel_path}'...")
            workbook = openpyxl.load_workbook(archivo_excel_path)
            self.logger.info(f"\n✅ Libro de trabajo cargado. Seleccionando la hoja '{hoja}'...")
            sheet = workbook[hoja]
            end_time_load_workbook = time.time()
            duration_load_workbook = end_time_load_workbook - start_time_load_workbook
            self.logger.info(f"PERFORMANCE: Tiempo de carga del workbook y selección de hoja: {duration_load_workbook:.4f} segundos.")

            # 1. Determinar el índice físico de la columna
            col_index: int = -1
            if isinstance(nombre_o_indice_columna, str):
                # --- Medición de rendimiento: Búsqueda de columna por nombre ---
                start_time_find_column = time.time()
                self.logger.info(f"\n🔎 Buscando columna por nombre: '{nombre_o_indice_columna}' en el encabezado de la hoja '{hoja}'...")
                header_found = False
                # sheet[1] se refiere a la primera fila física del Excel
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value is not None and str(cell.value).strip().lower() == nombre_o_indice_columna.strip().lower():
                        col_index = col_idx
                        header_found = True
                        break
                end_time_find_column = time.time()
                duration_find_column = end_time_find_column - start_time_find_column
                self.logger.info(f"PERFORMANCE: Tiempo de búsqueda de columna por nombre: {duration_find_column:.4f} segundos.")

                if not header_found:
                    self.logger.error(f"\n❌ Error: La columna '{nombre_o_indice_columna}' no fue encontrada en el encabezado de la hoja '{hoja}'.")
                    return None
            elif isinstance(nombre_o_indice_columna, int):
                col_index = nombre_o_indice_columna
            else:
                self.logger.error(f"\n❌ Error: El parámetro 'nombre_o_indice_columna' debe ser un string (nombre) o un entero (índice). Se recibió: '{nombre_o_indice_columna}' ({type(nombre_o_indice_columna).__name__}).")
                return None

            # Validar que el índice de columna sea válido
            if not (1 <= col_index <= sheet.max_column):
                self.logger.error(f"\n❌ Error: Índice de columna '{col_index}' fuera de rango para la hoja '{hoja}' (máximo: {sheet.max_column}).")
                return None

            # 2. Determinar el índice físico de la fila
            # 'numero_fila_logica' es la fila de datos que el usuario piensa (1 para la primera fila de datos).
            # Si hay encabezado, la primera fila de datos (lógica 1) está en la fila física 2.
            # Por lo tanto, sumamos 1 si hay encabezado.
            actual_fila_fisica = numero_fila_logica + 1 if has_header_excel else numero_fila_logica

            # Validar que la fila física sea válida
            if not (1 <= actual_fila_fisica <= sheet.max_row):
                self.logger.warning(f"\n⚠️ Advertencia: La fila física {actual_fila_fisica} (lógica: {numero_fila_logica}) está fuera del rango de filas de la hoja '{hoja}' (máximo: {sheet.max_row}). Retornando None.")
                return None
            
            self.logger.info(f"\n🔎 Intentando obtener el dato de la celda (Fila lógica: {numero_fila_logica}, Fila física: {actual_fila_fisica}, Columna: {nombre_o_indice_columna}) de la hoja '{hoja}'.")
            
            # --- Medición de rendimiento: Lectura de la celda ---
            start_time_read_cell = time.time()
            cell_value = sheet.cell(row=actual_fila_fisica, column=col_index).value
            end_time_read_cell = time.time()
            duration_read_cell = end_time_read_cell - start_time_read_cell
            self.logger.info(f"PERFORMANCE: Tiempo de lectura de la celda: {duration_read_cell:.4f} segundos.")
            
            # Convertir a string para asegurar que 'rellenar_campo_de_texto' u otras funciones siempre reciban un str
            if cell_value is not None:
                valor_retorno = str(cell_value)
                self.logger.info(f"\n✅ Dato obtenido de (Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna}) en '{hoja}': '{valor_retorno}'.")
                return valor_retorno
            else:
                self.logger.warning(f"\n⚠️ La celda en Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna} en '{hoja}' está vacía. Retornando None.")
                return None

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo Excel no se encontró en la ruta: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return None
        except KeyError:
            error_msg = f"\n❌ FALLO (Hoja no encontrada): La hoja '{hoja}' no se encontró en el archivo Excel: '{archivo_excel_path}'."
            self.logger.critical(error_msg)
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el dato del Excel.\n"
                f"Archivo: '{archivo_excel_path}', Hoja: '{hoja}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (dato_Columna_excel): {duration_total_operation:.4f} segundos.")
            # Aunque openpyxl maneja la liberación de recursos, un log final es útil.
            self.logger.debug("\nFinalizada la operación de lectura de dato de Excel.")
    
    # 61- Función que detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo CSV.
    # Integra pruebas de rendimiento para medir el tiempo de lectura del archivo CSV.
    def num_Filas_csv(self, archivo_csv_path: str, delimiter: str = ',', has_header: bool = False, nombre_paso: str = "") -> int:
        """
        Detecta y devuelve el número total de filas de datos en un archivo CSV.
        Opcionalmente, descuenta una fila para el encabezado si 'has_header' es True.
        Esta función mide el tiempo que tarda en abrir el archivo CSV, leer todas sus filas
        y realizar el conteo, lo cual es útil para evaluar el rendimiento en escenarios
        de procesamiento de grandes volúmenes de datos CSV.

        Args:
            archivo_csv_path (str): La **ruta completa al archivo CSV**.
            delimiter (str, opcional): El **carácter utilizado como separador** de datos en el CSV
                                      (e.g., ',', ';', '\t'). Por defecto es `,`.
            has_header (bool, opcional): Si es `True`, se descuenta una fila del total
                                         para considerar que la primera fila es un encabezado.
                                         Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            int: El **número de filas de datos** en el archivo CSV.
                 Retorna `0` si el archivo no se encuentra, ocurre un error de formato CSV,
                 o si hay un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener el número de filas para el archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        
        row_count = 0 # Inicializamos el contador de filas

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo CSV: '{archivo_csv_path}'...")
            with open(archivo_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
                # Crea un objeto reader para iterar sobre las líneas del CSV, usando el delimitador especificado.
                # 'newline=''' es crucial para evitar problemas con saltos de línea en diferentes SO.
                # 'encoding='utf-8'' es una buena práctica para manejar caracteres especiales.
                csv_reader = csv.reader(csvfile, delimiter=delimiter)
                
                # Cuenta todas las filas en el CSV. sum(1 for row in csv_reader) es una forma eficiente.
                row_count = sum(1 for row in csv_reader)

            self.logger.info(f"\n✅ Lectura de archivo CSV completada. Filas totales encontradas: {row_count}.")

            if has_header and row_count > 0:
                # Si tiene encabezado y el archivo no está vacío (es decir, hay al menos el encabezado)
                num_data_rows = row_count - 1 # Resta 1 para no contar el encabezado, obteniendo solo las filas de datos
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en el archivo CSV '{archivo_csv_path}'.")
                return num_data_rows
            else:
                # Si no tiene encabezado o el archivo está vacío (row_count es 0 o 1 si solo es un encabezado sin datos)
                num_data_rows = row_count
                self.logger.info(f"\n✅ Se encontraron {num_data_rows} filas ocupadas en el archivo CSV '{archivo_csv_path}'.")
                return num_data_rows

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo CSV no se encontró en la ruta: '{archivo_csv_path}'."
            self.logger.critical(error_msg)
            return 0
        except csv.Error as e:
            error_msg = f"\n❌ FALLO (Error de formato CSV): Ocurrió un error al procesar el archivo CSV '{archivo_csv_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace para errores de CSV
            return 0
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el número de filas del CSV.\n"
                f"Archivo: '{archivo_csv_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace para errores inesperados
            return 0
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (num_Filas_csv): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nFinalizada la operación de lectura de CSV.")

    # 62- Función que obtiene el valor de una "celda" específica de un archivo CSV,
    # ajustando la fila si se indica que hay un encabezado y recibiendo el delimitador.
    # Integra pruebas de rendimiento para medir el tiempo de lectura de la celda.
    def dato_Columna_csv(self, archivo_csv_path: str, fila_logica: int, columna_logica: int, delimiter: str = ',', has_header: bool = False, nombre_paso: str = "") -> Optional[str]:
        """
        Obtiene el valor de una "celda" específica de un archivo CSV, ajustando el índice de la fila
        si se indica que la primera fila es un encabezado. Permite especificar el delimitador del CSV.
        Esta función mide el tiempo que tarda en cargar el archivo CSV, leer todas sus filas
        y extraer el dato de la celda solicitada, lo cual es crucial para evaluar el rendimiento
        en escenarios de automatización basados en datos de archivos CSV.

        Args:
            archivo_csv_path (str): La **ruta completa al archivo CSV**.
            fila_logica (int): El **número de fila lógico** (basado en 1) de la celda a leer.
                               Si `has_header` es `True`, esta es la fila de datos
                               (e.g., `1` para la primera fila después del encabezado).
            columna_logica (int): El **número de columna lógico** (basado en 1) de la celda a leer.
            delimiter (str, opcional): El **carácter utilizado como separador** de datos en el CSV
                                      (e.g., ',', ';', '\t'). Por defecto es `,`.
            has_header (bool, opcional): Si es `True`, indica que la primera fila del CSV es un encabezado.
                                         Esto ajusta el cálculo de la fila física. Por defecto es `False`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs. Por defecto "".

        Returns:
            Optional[str]: El **valor de la celda como string**. Retorna `None` si el archivo no se encuentra,
                           los índices de fila/columna están fuera de rango, hay un error de formato CSV,
                           o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando obtener dato de la celda (Fila lógica: {fila_logica}, Columna lógica: {columna_logica}) del archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}). ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()
        cell_value: Optional[str] = None # Inicializamos el valor de la celda

        try:
            # Convierte el número de fila lógica (1-basada) a un índice 0-basado para Python
            # Si hay encabezado, la primera fila de datos (lógica 1) está en el índice físico 1 (0-basado).
            # Por lo tanto, si has_header, fila_logica 1 -> índice 1. Sin has_header, fila_logica 1 -> índice 0.
            actual_fila_0_indexed = fila_logica - 1
            if has_header:
                actual_fila_0_indexed += 1 # Ajusta si hay encabezado para saltar la fila 0

            # Convierte el número de columna lógica (1-basada) a un índice 0-basado para Python
            actual_col_0_indexed = columna_logica - 1

            self.logger.info(f"\n🔎 Calculando índices físicos: Fila física (0-indexed): {actual_fila_0_indexed}, Columna física (0-indexed): {actual_col_0_indexed}.")

            # --- Medición de rendimiento: Carga del archivo CSV y lectura de todas las filas ---
            start_time_load_csv = time.time()
            self.logger.info(f"\n⏳ Abriendo y leyendo todas las filas del archivo CSV: '{archivo_csv_path}'...")
            with open(archivo_csv_path, 'r', newline='', encoding='utf-8') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=delimiter)
                rows = list(csv_reader) # Lee todas las filas del CSV en una lista de listas (cada sublista es una fila)
            end_time_load_csv = time.time()
            duration_load_csv = end_time_load_csv - start_time_load_csv
            self.logger.info(f"PERFORMANCE: Tiempo de carga del archivo CSV y lectura de todas las filas: {duration_load_csv:.4f} segundos.")
            
            self.logger.info(f"\n✅ Archivo CSV leído. Total de filas físicas encontradas: {len(rows)}.")

            # Validación de límites para la fila
            if actual_fila_0_indexed < 0 or actual_fila_0_indexed >= len(rows):
                self.logger.error(f"\n❌ Error: La fila lógica {fila_logica} (física 0-indexed: {actual_fila_0_indexed}) está fuera de los límites del archivo CSV '{archivo_csv_path}'. Total filas físicas: {len(rows)}.")
                return None

            # Validación de límites para la columna en la fila específica
            if actual_col_0_indexed < 0 or actual_col_0_indexed >= len(rows[actual_fila_0_indexed]):
                self.logger.error(f"\n❌ Error: La columna lógica {columna_logica} (física 0-indexed: {actual_col_0_indexed}) está fuera de los límites de la fila física {actual_fila_0_indexed} del archivo CSV '{archivo_csv_path}'. Total columnas en esa fila: {len(rows[actual_fila_0_indexed])}.")
                return None

            # Obtiene el valor de la celda especificada
            cell_value = rows[actual_fila_0_indexed][actual_col_0_indexed]
            
            self.logger.info(f"\n✅ Dato obtenido de (Fila lógica: {fila_logica}, Columna lógica: {columna_logica}) en '{archivo_csv_path}': '{cell_value}'.")
            return cell_value
        
        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo CSV no se encontró en la ruta: '{archivo_csv_path}'."
            self.logger.critical(error_msg)
            return None
        except ValueError:
            # Esto ocurriría si fila_logica o columna_logica no fueran enteros,
            # pero los type hints ya lo previenen. Se mantiene por robustez.
            error_msg = f"\n❌ FALLO (Valor inválido): Los parámetros 'fila_logica' y 'columna_logica' deben ser números enteros. Se recibieron: fila='{fila_logica}', columna='{columna_logica}'."
            self.logger.critical(error_msg)
            return None
        except csv.Error as e:
            error_msg = f"\n❌ FALLO (Error de formato CSV): Ocurrió un error al procesar el archivo CSV '{archivo_csv_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el dato de la columna del CSV.\n"
                f"Archivo: '{archivo_csv_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace
            return None
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (dato_Columna_csv): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nFinalizada la operación de lectura de dato de CSV.")
    
    # 63- Función que lee y parsea un archivo JSON, devolviendo su contenido.
    # Integra pruebas de rendimiento para medir el tiempo que tarda en leer y parsear el JSON.
    def leer_json(self, json_file_path: str, nombre_paso: str = "") -> Union[Dict, List, None]:
        """
        Lee y parsea un archivo JSON, devolviendo su contenido como un diccionario o lista de Python.
        Esta función mide el tiempo que tarda en abrir, leer y parsear el archivo JSON,
        lo cual es útil para evaluar el rendimiento en escenarios de automatización impulsados por datos.

        Args:
            json_file_path (str): La **ruta completa al archivo JSON**.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[Dict, List, None]: El contenido del archivo JSON como un **diccionario** o una **lista**,
                                     o **None** si el archivo no se encuentra, el formato JSON es inválido,
                                     o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo JSON: '{json_file_path}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        data_content: Union[Dict, List, None] = None # Inicializamos a None

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo JSON: '{json_file_path}'...")
            with open(json_file_path, 'r', encoding='utf-8') as file:
                # 'encoding='utf-8'' es una buena práctica para manejar caracteres especiales.
                data_content = json.load(file) # Carga (parsea) el contenido del archivo JSON
            
            self.logger.info(f"\n✅ Archivo JSON '{json_file_path}' leído y parseado exitosamente.")
            return data_content

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo JSON no se encontró en la ruta: '{json_file_path}'."
            self.logger.critical(error_msg)
            return None
        except json.JSONDecodeError as e:
            error_msg = f"\n❌ FALLO (Error de formato JSON): Error al decodificar JSON desde '{json_file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores de decodificación JSON
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error inesperado al leer el archivo JSON.\n"
                f"Archivo: '{json_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores inesperados
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_json): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo JSON finalizada.")
        
    # 64- Función que lee el contenido completo de un archivo de texto plano.
    # Si se proporciona un delimitador, divide el contenido del archivo por el delimitador.
    # Integra pruebas de rendimiento para medir el tiempo de lectura y procesamiento del archivo.
    def leer_texto(self, file_path: str, delimiter: Optional[str] = None, nombre_paso: str = "") -> Union[str, List[str], None]:
        """
        Lee el contenido completo de un archivo de texto plano.
        Si se proporciona un delimitador, divide el contenido del archivo por el delimitador
        y lo devuelve como una lista de cadenas.
        Esta función mide el tiempo que tarda en abrir, leer y procesar el archivo de texto,
        lo cual es útil para evaluar el rendimiento en operaciones de E/S de archivos.

        Args:
            file_path (str): La **ruta completa al archivo de texto**.
            delimiter (str, opcional): Si se proporciona, el contenido del archivo se dividirá por este delimitador
                                        y se devolverá como una lista de cadenas. Si es `None`, se devuelve el contenido
                                        completo como una sola cadena. Por defecto es `None`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[str, List[str], None]: El contenido del archivo como una **cadena** (si `delimiter` es `None`)
                                         o una **lista de cadenas** (si se usa `delimiter`).
                                         Retorna `None` si el archivo no se encuentra, hay un error de E/S,
                                         o si ocurre un error inesperado.
        """
        delimiter_log_info = f"'{delimiter}'" if delimiter is not None else "Ninguno"
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo de texto: '{file_path}' (Delimitador: {delimiter_log_info}). ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        content: Optional[str] = None # Inicializamos content

        try:
            self.logger.info(f"\n⏳ Abriendo y leyendo el archivo de texto: '{file_path}'...")
            with open(file_path, 'r', encoding='utf-8') as file:
                # 'encoding='utf-8'' es crucial para manejar correctamente una amplia gama de caracteres.
                content = file.read() # Lee todo el contenido del archivo
            
            self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído exitosamente.")

            if delimiter is not None:
                # --- Medición de rendimiento: División del contenido (si aplica) ---
                start_time_split = time.time()
                self.logger.info(f"\n🔎 Dividiendo el contenido por el delimitador: '{delimiter}'...")
                result = content.split(delimiter) # Divide el contenido por el delimitador y lo retorna como lista
                end_time_split = time.time()
                duration_split = end_time_split - start_time_split
                self.logger.info(f"PERFORMANCE: Tiempo de división del contenido: {duration_split:.4f} segundos.")
                self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído y dividido exitosamente. Se encontraron {len(result)} segmentos.")
                return result
            else:
                self.logger.info(f"\n✅ Archivo de texto '{file_path}' leído completamente como una sola cadena.")
                return content
            
        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo de texto no se encontró en la ruta: '{file_path}'."
            self.logger.critical(error_msg)
            return None
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al leer el archivo de texto '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el archivo de texto.\n"
                f"Archivo: '{file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_texto): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo de texto finalizada.")

    # 65- Función que escribe contenido en un archivo de texto plano.
    # Si el contenido es una lista de cadenas y se proporciona un delimitador,
    # las cadenas se unirán con el delimitador antes de escribirlas.
    # Integra pruebas de rendimiento para medir el tiempo de preparación y escritura del archivo.
    def escribir_texto(self, file_path: str, content: Union[str, List[str]], append: bool = False, delimiter: Optional[str] = None, nombre_paso: str = "") -> bool:
        """
        Escribe contenido en un archivo de texto plano. Si el contenido es una lista de cadenas
        y se proporciona un delimitador, las cadenas se unirán con el delimitador antes de escribirlas.
        Esta función mide el tiempo de preparación del contenido y la escritura en el archivo,
        lo cual es útil para evaluar el rendimiento de las operaciones de E/S.

        Args:
            file_path (str): La **ruta completa al archivo de texto**.
            content (Union[str, List[str]]): La cadena o lista de cadenas a escribir.
            append (bool, opcional): Si es `True`, el contenido se añadirá al final del archivo.
                                     Si es `False` (por defecto), el archivo se sobrescribirá si existe.
            delimiter (str, opcional): Si se proporciona y `content` es una lista de cadenas, las cadenas
                                       se unirán con este delimitador antes de la escritura. Si es `None`,
                                       las cadenas de una lista se escribirán directamente sin separación explícita.
                                       Por defecto es `None`.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            bool: `True` si la escritura fue exitosa, `False` en caso de error.
        """
        mode = 'a' if append else 'w' # Determina el modo de apertura: 'a' para añadir, 'w' para sobrescribir
        action = "añadir a" if append else "escribir en" # Descripción de la acción para el log
        
        delimiter_log_info = f"'{delimiter}'" if delimiter is not None else "Ninguno"
        self.logger.info(f"\n--- {nombre_paso}: Intentando {action} el archivo de texto: '{file_path}' (Delimitador de escritura: {delimiter_log_info}). ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        text_to_write: str = "" # Variable para almacenar el contenido final a escribir

        try:
            # Lógica para procesar el contenido antes de la escritura
            if isinstance(content, list):
                # --- Medición de rendimiento: Unión de la lista con el delimitador ---
                start_time_join = time.time()
                
                if delimiter is not None:
                    text_to_write = delimiter.join(content)
                    self.logger.info(f"\n🔎 El contenido de la lista será unido con el delimitador '{delimiter}' antes de escribir.")
                else:
                    text_to_write = "".join(content)
                    self.logger.warning("\n⚠️ Se proporcionó una lista para escribir_texto sin delimitador. Las cadenas se concatenarán sin separación explícita, lo que puede no ser el comportamiento deseado.")
                
                end_time_join = time.time()
                duration_join = end_time_join - start_time_join
                self.logger.info(f"PERFORMANCE: Tiempo de preparación del contenido (join): {duration_join:.4f} segundos.")

            elif isinstance(content, str):
                text_to_write = content # Si el contenido ya es una cadena, lo asigna tal cual
                self.logger.info("\n🔎 El contenido es una cadena, se escribirá directamente.")
            else:
                error_msg = f"\n❌ FALLO (Tipo de dato inválido): El tipo de contenido proporcionado no es válido. Se esperaba str o List[str], se recibió: {type(content)}."
                self.logger.critical(error_msg)
                return False

            # Asegurarse de que el directorio del archivo exista antes de intentar escribir
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True) # `exist_ok=True` evita errores si el directorio ya existe
                self.logger.info(f"\n☑️ Directorio creado para el archivo de texto: {directory}")

            # --- Medición de rendimiento: Escritura en el archivo ---
            self.logger.info(f"\n✍️ Escribiendo contenido en el archivo: '{file_path}'...")
            with open(file_path, mode, encoding='utf-8') as file:
                # `encoding='utf-8'` es crucial para manejar correctamente una amplia gama de caracteres
                file.write(text_to_write)
            
            self.logger.info(f"\n✅ Contenido {action} exitosamente en '{file_path}'.")
            return True
        
        except IOError as e:
            error_msg = f"\n❌ FALLO (Error de E/S): Ocurrió un error de entrada/salida al {action} el archivo de texto '{file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return False
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al {action} el archivo de texto.\n"
                f"Archivo: '{file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo
            return False
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (escribir_texto): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de escritura de archivo de texto finalizada.")
    
    # 66- Función para leer archivos XML.
    # Integra pruebas de rendimiento para medir el tiempo que tarda en leer y parsear el XML.
    def leer_xml(self, xml_file_path: str, nombre_paso: str = "") -> Union[ET.Element, None]:
        """
        Lee y parsea un archivo XML, devolviendo su elemento raíz como un objeto Element.
        Esta función mide el tiempo que tarda en abrir, leer y parsear el archivo XML,
        lo cual es útil para evaluar el rendimiento en escenarios donde se procesan archivos XML.

        Args:
            xml_file_path (str): La **ruta completa al archivo XML**.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Union[ET.Element, None]: El **elemento raíz del XML** como un objeto `xml.etree.ElementTree.Element`,
                                     o **None** si el archivo no se encuentra, el formato XML es inválido,
                                     o si ocurre un error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando leer el archivo XML: '{xml_file_path}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        root_element: Optional[ET.Element] = None # Inicializamos el elemento raíz

        try:
            self.logger.info(f"\n⏳ Abriendo y parseando el archivo XML: '{xml_file_path}'...")
            # ET.parse() se encarga de abrir y parsear el archivo.
            # No es necesario especificar la codificación en la mayoría de los casos ya que
            # ET lo detecta automáticamente si el XML tiene una declaración de codificación (e.g., <?xml version="1.0" encoding="UTF-8"?>).
            tree = ET.parse(xml_file_path)
            
            # Obtiene el elemento raíz del XML
            root_element = tree.getroot()
            
            self.logger.info(f"\n✅ Archivo XML '{xml_file_path}' leído y parseado exitosamente. Elemento raíz: '{root_element.tag}'.")
            return root_element

        except FileNotFoundError:
            error_msg = f"\n❌ FALLO (Archivo no encontrado): El archivo XML no se encontró en la ruta: '{xml_file_path}'."
            self.logger.critical(error_msg)
            return None
        except ET.ParseError as e:
            error_msg = f"\n❌ FALLO (Error de formato XML): Ocurrió un error al parsear el archivo XML '{xml_file_path}'.\nDetalles: {e}"
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores de parseo XML
            return None
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al leer el archivo XML.\n"
                f"Archivo: '{xml_file_path}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Incluye el stack trace completo para errores inesperados
            return None
        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (leer_xml): {duration_total_operation:.4f} segundos.")
            self.logger.debug("\nOperación de lectura de archivo XML finalizada.")
    
    # 67- Función que realiza una acción de click derecho (context click) sobre un elemento.
    # Integra pruebas de rendimiento utilizando Playwright y captura métricas de tiempo.
    # Esta versión asume que la clase base de errores de Playwright se importa como 'Error' sin alias.
    def hacer_click_derecho_en_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo_espera_post_click: Union[int, float] = 0.5, nombre_paso: str = ""):
        """
        Realiza una acción de click derecho (context click) sobre un elemento en la página.
        Esta función mide el tiempo de localización del elemento y el tiempo que tarda el click,
        proporcionando métricas de rendimiento clave para tus interacciones con Playwright.
        También toma capturas de pantalla antes y después de la acción para depuración y evidencia.

        Args:
            selector (Union[str, Locator]): El selector del elemento (puede ser un string CSS/XPath/texto,
                                            o un objeto Locator de Playwright ya existente).
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_espera_post_click (Union[int, float], opcional): Tiempo en segundos de espera explícita
                                                                    después de realizar el click derecho.
                                                                    Útil para permitir que el menú contextual
                                                                    aparezca o que la página reaccione. Por defecto es 0.5 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Raises:
            TimeoutError: Si el elemento no se encuentra o no es interactuable dentro del tiempo de espera de Playwright.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando hacer click derecho sobre el elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector
            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización del elemento '{selector}': {duration_locator:.4f} segundos.")

            # Resaltar el elemento antes de la interacción (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes del click derecho
            self.tomar_captura(f"{nombre_base}_antes_click_derecho", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del click derecho: '{nombre_base}_antes_click_derecho.png'")

            # --- Medición de rendimiento: Tiempo de ejecución del click derecho ---
            start_time_click = time.time()
            # El atributo 'button="right"' es clave para el click derecho (context click)
            # Playwright espera implícitamente que el elemento esté visible y habilitado.
            locator.click(button="right") 
            end_time_click = time.time()
            duration_click = end_time_click - start_time_click
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución del click derecho en '{selector}': {duration_click:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Click derecho realizado exitosamente en el elemento con selector '{selector}'.")
            
            # Tomar captura de pantalla después del click derecho
            self.tomar_captura(f"{nombre_base}_despues_click_derecho", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del click derecho: '{nombre_base}_despues_click_derecho.png'")

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El tiempo de espera se agotó al hacer click derecho en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo ({e.message if hasattr(e, 'message') else str(e)}).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_click_derecho", directorio)
            # Re-lanzamos la excepción TimeoutError que ya es específica de Playwright
            raise 

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al hacer click derecho en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_click_derecho", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al intentar hacer click derecho en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_click_derecho", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer_click_derecho_en_elemento): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la interacción, si se especificó
            # Nota: el parámetro de entrada 'tiempo' se ha renombrado a 'tiempo_espera_post_click' para mayor claridad.
            if tiempo_espera_post_click > 0:
                self.logger.info(f"\n⏳ Esperando {tiempo_espera_post_click} segundos después del click derecho.")
                self.esperar_fijo(tiempo_espera_post_click) # Asegúrate de que esta función exista
    
    # 68- Función que realiza una acción de 'mouse down' (presionar el botón del ratón) sobre un elemento.
    # Esta versión utiliza page.mouse.down() para una simulación más precisa de solo presionar.
    def hacer_mouse_down_en_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo_espera_post_accion: Union[int, float] = 0.5, nombre_paso: str = ""):
        """
        Realiza una acción de 'mouse down' (presionar el botón izquierdo del ratón) sobre el centro de un elemento.
        Esta función solo simula la acción de presionar el botón, sin la liberación ('mouse up').
        Mide el tiempo de localización del elemento y el tiempo que tarda la acción de 'mouse down',
        proporcionando métricas de rendimiento clave para tus interacciones con Playwright.
        También toma capturas de pantalla antes y después de la acción para depuración y evidencia.

        Args:
            selector (Union[str, Locator]): El selector del elemento (puede ser un string CSS/XPath/texto,
                                            o un objeto Locator de Playwright ya existente).
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_espera_post_accion (Union[int, float], opcional): Tiempo en segundos de espera explícita
                                                                    después de realizar la acción de 'mouse down'.
                                                                    Útil para permitir que la página reaccione
                                                                    a la presión del botón. Por defecto es 0.5 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Raises:
            TimeoutError: Si el elemento no se encuentra o no es visible/habilitado dentro del tiempo de espera de Playwright.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando hacer 'mouse down' sobre el elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator
        element_bounding_box: Optional[Dict[str, Any]] = None # Para almacenar las coordenadas del elemento

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector

            # Asegurarse de que el elemento esté visible y obtener su bounding box
            # Playwright ya espera visibilidad/habilitación con locator.wait_for() o actionability checks.
            # Pero para obtener el bounding_box, el elemento debe estar en el DOM y visible.
            element_bounding_box = locator.bounding_box()

            if not element_bounding_box:
                raise Error(f"\nNo se pudo obtener el bounding box del elemento '{selector}'. Es posible que no sea visible o no esté adjunto al DOM.")
            
            # Calcular el centro del elemento
            center_x = element_bounding_box['x'] + element_bounding_box['width'] / 2
            center_y = element_bounding_box['y'] + element_bounding_box['height'] / 2

            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización y obtención de coordenadas para '{selector}': {duration_locator:.4f} segundos. Coordenadas: ({center_x:.2f}, {center_y:.2f})")

            # Resaltar el elemento antes de la interacción (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes de la acción
            self.tomar_captura(f"{nombre_base}_antes_mouse_down", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del 'mouse down': '{nombre_base}_antes_mouse_down.png'")

            # --- Medición de rendimiento: Tiempo de ejecución de la acción de 'mouse down' ---
            start_time_action = time.time()
            # Realiza la acción de 'mouse down' puro en las coordenadas del centro del elemento.
            self.page.mouse.down(button="left", x=center_x, y=center_y) 
            end_time_action = time.time()
            duration_action = end_time_action - start_time_action
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución de la acción 'mouse down' en '{selector}': {duration_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Acción de 'mouse down' realizada exitosamente en el elemento con selector '{selector}'.")
            
            # Tomar captura de pantalla después de la acción
            self.tomar_captura(f"{nombre_base}_despues_mouse_down", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del 'mouse down': '{nombre_base}_despues_mouse_down.png'")

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El tiempo de espera se agotó al hacer 'mouse down' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo para obtener coordenadas ({e.message if hasattr(e, 'message') else str(e)}).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_mouse_down", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al hacer 'mouse down' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_mouse_down", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al intentar hacer 'mouse down' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_mouse_down", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer_mouse_down_en_elemento): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la interacción, si se especificó
            if tiempo_espera_post_accion > 0:
                self.logger.info(f"\n⏳ Esperando {tiempo_espera_post_accion} segundos después de la acción de 'mouse down'.")
                self.esperar_fijo(tiempo_espera_post_accion) # Asegúrate de que esta función exista
    
    # 69- Función que realiza una acción de 'mouse up' (soltar el botón del ratón) sobre un elemento.
    # Esta versión utiliza page.mouse.up() para una simulación precisa de solo soltar el botón.
    def hacer_mouse_up_de_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo_espera_post_accion: Union[int, float] = 0.5, nombre_paso: str = ""):
        """
        Realiza una acción de 'mouse up' (soltar el botón izquierdo del ratón) sobre el centro de un elemento.
        Esta función solo simula la acción de liberar el botón, típicamente usada después de un 'mouse down'
        en escenarios de arrastrar y soltar, o interacciones complejas.
        Mide el tiempo de localización del elemento y el tiempo que tarda la acción de 'mouse up',
        proporcionando métricas de rendimiento clave para tus interacciones con Playwright.
        También toma capturas de pantalla antes y después de la acción para depuración y evidencia.

        Args:
            selector (Union[str, Locator]): El selector del elemento (puede ser un string CSS/XPath/texto,
                                            o un objeto Locator de Playwright ya existente).
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_espera_post_accion (Union[int, float], opcional): Tiempo en segundos de espera explícita
                                                                    después de realizar la acción de 'mouse up'.
                                                                    Útil para permitir que la página reaccione
                                                                    a la liberación del botón. Por defecto es 0.5 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Raises:
            TimeoutError: Si el elemento no se encuentra o no es visible/habilitado dentro del tiempo de espera de Playwright.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando hacer 'mouse up' sobre el elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator
        element_bounding_box: Optional[Dict[str, Any]] = None # Para almacenar las coordenadas del elemento

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector

            # Asegurarse de que el elemento esté visible y obtener su bounding box
            # locator.bounding_box() puede esperar la visibilidad del elemento.
            element_bounding_box = locator.bounding_box()

            if not element_bounding_box:
                raise Error(f"\nNo se pudo obtener el bounding box del elemento '{selector}'. Es posible que no sea visible o no esté adjunto al DOM.")
            
            # Calcular el centro del elemento
            center_x = element_bounding_box['x'] + element_bounding_box['width'] / 2
            center_y = element_bounding_box['y'] + element_bounding_box['height'] / 2

            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización y obtención de coordenadas para '{selector}': {duration_locator:.4f} segundos. Coordenadas: ({center_x:.2f}, {center_y:.2f})")

            # Resaltar el elemento antes de la interacción (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes de la acción
            self.tomar_captura(f"{nombre_base}_antes_mouse_up", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del 'mouse up': '{nombre_base}_antes_mouse_up.png'")

            # --- Medición de rendimiento: Tiempo de ejecución de la acción de 'mouse up' ---
            start_time_action = time.time()
            # Realiza la acción de 'mouse up' puro en las coordenadas del centro del elemento.
            self.page.mouse.up(button="left", x=center_x, y=center_y) 
            end_time_action = time.time()
            duration_action = end_time_action - start_time_action
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución de la acción 'mouse up' en '{selector}': {duration_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: Acción de 'mouse up' realizada exitosamente en el elemento con selector '{selector}'.")
            
            # Tomar captura de pantalla después de la acción
            self.tomar_captura(f"{nombre_base}_despues_mouse_up", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del 'mouse up': '{nombre_base}_despues_mouse_up.png'")

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El tiempo de espera se agotó al hacer 'mouse up' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo para obtener coordenadas ({e.message if hasattr(e, 'message') else str(e)}).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_mouse_up", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al hacer 'mouse up' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_mouse_up", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al intentar hacer 'mouse up' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_mouse_up", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer_mouse_up_de_elemento): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la interacción, si se especificó
            if tiempo_espera_post_accion > 0:
                self.logger.info(f"\n⏳ Esperando {tiempo_espera_post_accion} segundos después de la acción de 'mouse up'.")
                self.esperar_fijo(tiempo_espera_post_accion) # Asegúrate de que esta función exista
    
    # 70- Función que realiza una acción de 'focus' (enfocar) sobre un elemento.
    # Integra pruebas de rendimiento utilizando Playwright y captura métricas de tiempo.
    def hacer_focus_en_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo_espera_post_accion: Union[int, float] = 0.5, nombre_paso: str = ""):
        """
        Realiza una acción de 'focus' (establecer el foco) sobre un elemento especificado.
        Esta función es útil para simular la interacción del usuario al tabular o hacer clic
        en un campo de entrada, botón, etc., y es fundamental para las pruebas de accesibilidad
        y el control de flujo en formularios.
        Mide el tiempo de localización del elemento y el tiempo que tarda la acción de 'focus',
        proporcionando métricas de rendimiento clave.
        También toma capturas de pantalla antes y después de la acción para depuración y evidencia.

        Args:
            selector (Union[str, Locator]): El selector del elemento (puede ser un string CSS/XPath/texto,
                                            o un objeto Locator de Playwright ya existente).
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_espera_post_accion (Union[int, float], opcional): Tiempo en segundos de espera explícita
                                                                    después de realizar la acción de 'focus'.
                                                                    Útil para permitir que la página reaccione
                                                                    o se carguen elementos dependientes. Por defecto es 0.5 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Raises:
            TimeoutError: Si el elemento no se encuentra o no es interactuable dentro del tiempo de espera de Playwright.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando hacer 'focus' sobre el elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector
            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización del elemento '{selector}': {duration_locator:.4f} segundos.")

            # Resaltar el elemento antes de la interacción (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes de la acción
            self.tomar_captura(f"{nombre_base}_antes_focus", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del 'focus': '{nombre_base}_antes_focus.png'")

            # --- Medición de rendimiento: Tiempo de ejecución de la acción de 'focus' ---
            start_time_action = time.time()
            # El método focus() de Playwright establece el foco en el elemento.
            # Playwright espera implícitamente que el elemento esté visible y habilitado antes de enfocarlo.
            locator.focus() # Eliminado 'timeout' del focus() para usar el de Playwright por defecto o global.
                            # Si se necesita un timeout específico para el focus, se puede volver a añadir: timeout=tiempo_espera_max_para_focus * 1000
            end_time_action = time.time()
            duration_action = end_time_action - start_time_action
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución de la acción 'focus' en '{selector}': {duration_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: 'Focus' realizado exitosamente en el elemento con selector '{selector}'.")
            
            # Tomar captura de pantalla después de la acción
            self.tomar_captura(f"{nombre_base}_despues_focus", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del 'focus': '{nombre_base}_despues_focus.png'")

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El tiempo de espera se agotó al hacer 'focus' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo ({e.message if hasattr(e, 'message') else str(e)}).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_focus", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al hacer 'focus' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_focus", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al intentar hacer 'focus' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_focus", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer_focus_en_elemento): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la interacción, si se especificó
            # Nota: el parámetro de entrada original 'tiempo' se ha renombrado a 'tiempo_espera_post_accion' para mayor claridad.
            if tiempo_espera_post_accion > 0:
                self.logger.info(f"\n⏳ Esperando {tiempo_espera_post_accion} segundos después de la acción de 'focus'.")
                self.esperar_fijo(tiempo_espera_post_accion) # Asegúrate de que esta función exista
    
    # 71- Función que realiza una acción de 'blur' (desenfocar) sobre un elemento.
    # Integra pruebas de rendimiento utilizando Playwright y captura métricas de tiempo.
    def hacer_blur_en_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, tiempo_espera_post_accion: Union[int, float] = 0.5, nombre_paso: str = ""):
        """
        Realiza una acción de 'blur' (quitar el foco) sobre un elemento que actualmente lo tiene.
        Esta función simula que el usuario ha movido el foco de un elemento (por ejemplo, al hacer
        clic fuera de un campo de texto o al presionar Tab para salir de él). Es útil para probar
        validaciones 'on blur' o la finalización de la edición.
        Mide el tiempo de localización del elemento y el tiempo que tarda la acción de 'blur',
        proporcionando métricas de rendimiento clave.
        También toma capturas de pantalla antes y después de la acción para depuración y evidencia.

        Args:
            selector (Union[str, Locator]): El selector del elemento (puede ser un string CSS/XPath/texto,
                                            o un objeto Locator de Playwright ya existente).
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_espera_post_accion (Union[int, float], opcional): Tiempo en segundos de espera explícita
                                                                    después de realizar la acción de 'blur'.
                                                                    Útil para permitir que la página reaccione
                                                                    a la pérdida del foco. Por defecto es 0.5 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Raises:
            TimeoutError: Si el elemento no se encuentra o no es interactuable (o enfocable/desenfocable)
                          dentro del tiempo de espera de Playwright.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando hacer 'blur' sobre el elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector
            
            # Opcional: Podrías querer resaltar el elemento ANTES de desenfocarlo
            # Es útil para ver cuál elemento se va a desenfocar.
            # locator.highlight() 

            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización del elemento '{selector}': {duration_locator:.4f} segundos.")

            # Tomar captura de pantalla antes de la acción
            self.tomar_captura(f"{nombre_base}_antes_blur", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del 'blur': '{nombre_base}_antes_blur.png'")

            # --- Medición de rendimiento: Tiempo de ejecución de la acción de 'blur' ---
            start_time_action = time.time()
            # El método blur() de Playwright quita el foco del elemento.
            # Playwright espera implícitamente que el elemento esté en el DOM y enfocado para poder desenfocarlo.
            locator.blur() # Eliminado 'timeout' del blur() para usar el de Playwright por defecto o global.
                           # Si se necesita un timeout específico para el blur, se puede volver a añadir: timeout=tiempo_espera_max_para_blur * 1000
            end_time_action = time.time()
            duration_action = end_time_action - start_time_action
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución de la acción 'blur' en '{selector}': {duration_action:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: 'Blur' realizado exitosamente en el elemento con selector '{selector}'.")
            
            # Tomar captura de pantalla después de la acción
            self.tomar_captura(f"{nombre_base}_despues_blur", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del 'blur': '{nombre_base}_despues_blur.png'")

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El tiempo de espera se agotó al hacer 'blur' en '{selector}'.\n"
                f"Posibles causas: El elemento no estaba presente, visible o no era el elemento enfocado a tiempo ({e.message if hasattr(e, 'message') else str(e)}).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_blur", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al hacer 'blur' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_blur", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al intentar hacer 'blur' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_blur", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (hacer_blur_en_elemento): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la interacción, si se especificó
            # Nota: el parámetro de entrada original 'tiempo' se ha renombrado a 'tiempo_espera_post_accion' para mayor claridad.
            if tiempo_espera_post_accion > 0:
                self.logger.info(f"\n⏳ Esperando {tiempo_espera_post_accion} segundos después de la acción de 'blur'.")
                self.esperar_fijo(tiempo_espera_post_accion) # Asegúrate de que esta función exista
    
    # 72- Función que verifica el estado de un checkbox (marcado/desmarcado) o el valor de una opción seleccionada en un select.
    # Integra pruebas de rendimiento utilizando Playwright y captura métricas de tiempo.
    def verificar_estado_checkbox_o_select(self, selector: Union[str, Locator], estado_esperado: Union[bool, str], nombre_base: str, directorio: str, tiempo_max_espera_verificacion: Union[int, float] = 0.5, nombre_paso: str = "") -> bool:
        """
        Verifica el estado de un checkbox (marcado/desmarcado) o el valor de una opción seleccionada en un select.
        Esta función utiliza las aserciones de Playwright (`expect`) para manejar las esperas y la validación
        de manera eficiente y robusta. Proporciona métricas de rendimiento para la localización
        y la verificación del estado.

        Args:
            selector (Union[str, Locator]): El selector del checkbox o del elemento <select> (por ejemplo, CSS, XPath).
                                            Puede ser un string o un objeto Locator de Playwright ya existente.
            estado_esperado (Union[bool, str]):
                - Para checkbox: True si se espera que esté marcado, False si se espera que esté desmarcado.
                - Para select: El valor (value) de la opción que se espera que esté seleccionada.
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_max_espera_verificacion (Union[int, float], opcional): Tiempo máximo en segundos que Playwright
                                                                           esperará a que el elemento cumpla la condición.
                                                                           Por defecto es 5.0 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            bool: True si la verificación es exitosa (el estado actual coincide con el esperado), False en caso contrario.

        Raises:
            ValueError: Si el 'estado_esperado' no es un tipo válido (bool para checkbox, str para select).
            PlaywrightError (a través de TimeoutError o Error): Si ocurre un problema grave de Playwright
                                                                que impide la verificación.
        """
        self.logger.info(f"\n--- {nombre_paso}: Verificando estado para el selector: '{selector}'. Estado esperado: '{estado_esperado}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator
        tipo_elemento: str = "elemento" # Valor por defecto para los mensajes de error
        valor_actual_str: str = "N/A" # Valor por defecto para los mensajes de error
        mensaje_fallo_esperado: str = "" # Mensaje por defecto para fallos de aserción

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                # Usar locator().first para manejar casos donde el selector podría devolver múltiples elementos
                # pero solo nos interesa el primero. Si el selector ya es preciso, no hay problema.
                locator = self.page.locator(selector) 
            else: # Asume que si no es str, ya es un Locator
                locator = selector
            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización del elemento '{selector}': {duration_locator:.4f} segundos.")
            
            # Resaltar el elemento antes de la interacción (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes de la verificación
            self.tomar_captura(f"{nombre_base}_antes_verificar_estado", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes de verificar estado: '{nombre_base}_antes_verificar_estado.png'")

            # --- Lógica de Verificación y Medición de Aserción ---
            start_time_assertion = time.time()
            if isinstance(estado_esperado, bool): # Verificación para Checkbox
                tipo_elemento = "checkbox"
                if estado_esperado:
                    expect(locator).to_be_checked()
                else:
                    expect(locator).not_to_be_checked()
                
                valor_actual_str = str(locator.is_checked())
                mensaje_fallo_esperado = f"se esperaba {'marcado' if estado_esperado else 'desmarcado'} pero está '{valor_actual_str}'."
            
            elif isinstance(estado_esperado, str): # Verificación para Select (option)
                tipo_elemento = "select/option"
                expect(locator).to_have_value(estado_esperado)
                
                valor_actual_str = locator.input_value() # Obtiene el 'value' de la opción seleccionada
                mensaje_fallo_esperado = f"se esperaba la opción con valor '{estado_esperado}' pero la actual es '{valor_actual_str}'."
            
            else:
                raise ValueError(f"\nEl 'estado_esperado' debe ser un booleano para checkbox o un string para select. Tipo proporcionado: {type(estado_esperado).__name__}")

            end_time_assertion = time.time()
            duration_assertion = end_time_assertion - start_time_assertion
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución de la verificación (aserción) para '{selector}': {duration_assertion:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: El {tipo_elemento} '{selector}' tiene el estado esperado '{estado_esperado}'.")
            self.tomar_captura(f"{nombre_base}_despues_verificar_estado", directorio)
            return True

        except TimeoutError as e:
            # En caso de Timeout, intentamos obtener el valor actual para el mensaje de error.
            # Se usa un try-except interno para evitar fallos si el locator ya no es válido después del timeout.
            try:
                if tipo_elemento == "checkbox":
                    valor_actual_str = str(locator.is_checked())
                elif tipo_elemento == "select/option":
                    valor_actual_str = locator.input_value()
            except Exception:
                valor_actual_str = "No disponible (error al obtener el valor actual)"

            error_msg = (
                f"\n❌ FALLO (Timeout): El {tipo_elemento} '{selector}' "
                f"no cumplió el estado esperado '{estado_esperado}' después de {tiempo_max_espera_verificacion} segundos. "
                f"Estado actual: '{valor_actual_str}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_estado", directorio)
            return False

        except AssertionError as e:
            # En caso de AssertionError (falla de expect sin timeout), el valor ya se obtiene arriba.
            error_msg = (
                f"\n❌ FALLO (Aserción): El {tipo_elemento} '{selector}' "
                f"NO cumple el estado esperado. {mensaje_fallo_esperado} "
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_verificar_estado", directorio)
            return False

        except ValueError as e:
            error_msg = (
                f"\n❌ ERROR (Valor Inválido): Se proporcionó un 'estado_esperado' no válido para el selector '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Incluir exc_info para ValueError también
            self.tomar_captura(f"{nombre_base}_error_valor_invalido_verificar_estado", directorio)
            raise # Re-lanzamos el ValueError ya que es un error de uso de la función.

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            error_msg = (
                f"\n❌ FALLO (Playwright): Ocurrió un problema de Playwright al verificar el estado del elemento '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_estado", directorio)
            raise # Re-lanza la excepción original de Playwright

        except Exception as e: # Captura cualquier otro error inesperado
            error_msg = (
                f"\n❌ FALLO (Inesperado): Se produjo un error desconocido al verificar el estado del elemento '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_estado", directorio)
            raise # Re-lanza la excepción

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (verificar_estado_checkbox_o_select): {duration_total_operation:.4f} segundos.")
            
            # Espera fija después de la verificación, si se especificó.
            # El parámetro original 'tiempo' se renombró a 'tiempo_max_espera_verificacion' para el timeout de expect.
            # Si aún se desea una espera fija *adicional* al final, se usa el parámetro 'tiempo_espera_post_accion'
            # que definí para otras funciones. Para esta función, la espera principal es el timeout de expect.
            # Si el 'tiempo' original se refería a una espera fija *después* de todo, lo mantengo así.
            # Asumiendo que el 'tiempo' original era el timeout para la verificación.
            # Si necesitas una espera adicional después, se puede añadir un nuevo parámetro.
            pass # No hay una espera fija aquí por defecto, ya que el timeout de expect() maneja la espera.
                 # Si 'tiempo' original era para una pausa, el parámetro ha sido absorbido por el timeout de expect.
                 # Si se desea una pausa *adicional* al final, se debería añadir un nuevo parámetro.

    # 73- Función para extrae y retorna el valor textual de un elemento dado su selector.
    # Esta función ahora prioriza la extracción de 'value' de campos de formulario (input_value),
    # y luego el contenido de texto visible (inner_text) o todo el texto (text_content).
    # Integra pruebas de rendimiento y captura métricas de tiempo.
    def obtener_valor_de_elemento(self, selector: Union[str, Locator], nombre_base: str, directorio: str, 
                                 tiempo_max_espera_visibilidad: Union[int, float] = 5.0, nombre_paso: str = "") -> Optional[str]:
        """
        Extrae y retorna el valor textual (contenido o atributo 'value') de un elemento de la página.
        La función intenta obtener el valor de diferentes maneras:
        1.  Usa `locator.input_value()` para elementos de formulario como `<input>`, `<textarea>` o `<select>`.
        2.  Si `input_value()` no es aplicable o falla, intenta `locator.inner_text()` para obtener el texto
            visible renderizado dentro del elemento.
        3.  Si `inner_text()` no es apropiado (ej., texto oculto), intenta `locator.text_content()` para todo el texto.
        
        Playwright espera implícitamente que el elemento sea visible antes de intentar la extracción,
        lo cual es configurado por 'tiempo_max_espera_visibilidad'.

        Args:
            selector (Union[str, Locator]): El selector del elemento (CSS, XPath, texto, etc.).
                                            Puede ser un string o un objeto Locator de Playwright ya existente.
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            tiempo_max_espera_visibilidad (Union[int, float], opcional): Tiempo máximo en segundos que Playwright
                                                                        esperará a que el elemento sea visible
                                                                        antes de intentar extraer su valor.
                                                                        Por defecto es 5.0 segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para el registro (logs).
                                         Por defecto es una cadena vacía "".

        Returns:
            Optional[str]: El valor extraído del elemento como string, o None si no se pudo extraer ningún valor.

        Raises:
            AssertionError: Si el elemento no se vuelve visible dentro del tiempo de espera.
            Error: Para otros errores específicos de Playwright durante la interacción.
            Exception: Para cualquier otro error inesperado.
        """
        self.logger.info(f"\n--- {nombre_paso}: Extrayendo valor del elemento con selector: '{selector}'. ---")

        # --- Medición de rendimiento: Inicio de la operación total de la función ---
        start_time_total_operation = time.time()
        
        locator: Locator = None # Inicializamos el locator
        valor_extraido: Optional[str] = None # Para almacenar el valor extraído

        try:
            # Asegurarse de que el directorio de capturas de pantalla exista
            if not os.path.exists(directorio):
                os.makedirs(directorio, exist_ok=True)
                self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

            # --- Medición de rendimiento: Tiempo de localización del elemento y espera de visibilidad ---
            start_time_locator = time.time()
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else: # Asume que si no es str, ya es un Locator
                locator = selector
            
            # Esperar a que el elemento sea visible antes de intentar extraer su valor
            expect(locator).to_be_visible()
            end_time_locator = time.time()
            duration_locator = end_time_locator - start_time_locator
            self.logger.info(f"PERFORMANCE: Tiempo de localización y espera de visibilidad para '{selector}': {duration_locator:.4f} segundos.")
            
            # Resaltar el elemento (útil para la depuración visual)
            # locator.highlight() 

            # Tomar captura de pantalla antes de la extracción
            self.tomar_captura(f"{nombre_base}_antes_extraccion_valor", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes de la extracción de valor: '{nombre_base}_antes_extraccion_valor.png'")

            # --- Medición de rendimiento: Tiempo de extracción del valor ---
            start_time_extraction = time.time()
            # Priorizamos input_value() para campos de formulario (incluyendo <select>, <input>, <textarea>)
            # input_value() extrae el valor del atributo 'value' o el contenido de <textarea>.
            try:
                valor_extraido = locator.input_value()
                self.logger.debug(f"\nValor extraído (input_value) de '{selector}': '{valor_extraido}'")
            except Error as e: # Captura si no es un elemento de entrada o si falla la operación
                self.logger.debug(f"\ninput_value no aplicable o falló para '{selector}' (Detalles: {e.message if hasattr(e, 'message') else str(e)}). Intentando text_content/inner_text.")
                
                # Si falla input_value, intentamos con inner_text o text_content para otros elementos
                # inner_text() es a menudo preferible ya que devuelve el texto visible y renderizado.
                try:
                    valor_extraido = locator.inner_text()
                    self.logger.debug(f"\nValor extraído (inner_text) de '{selector}': '{valor_extraido}'")
                except Error as e_inner:
                    self.logger.debug(f"\ninner_text falló para '{selector}' (Detalles: {e_inner.message if hasattr(e_inner, 'message') else str(e_inner)}). Intentando text_content.")
                    try:
                        valor_extraido = locator.text_content()
                        self.logger.debug(f"\nValor extraído (text_content) de '{selector}': '{valor_extraido}'")
                    except Error as e_text:
                        self.logger.warning(f"\nNo se pudo extraer input_value, inner_text ni text_content de '{selector}' (Detalles: {e_text.message if hasattr(e_text, 'message') else str(e_text)}).")
                        valor_extraido = None # Asegurarse de que sea None si todo falla

            end_time_extraction = time.time()
            duration_extraction = end_time_extraction - start_time_extraction
            self.logger.info(f"PERFORMANCE: Tiempo de extracción del valor para '{selector}': {duration_extraction:.4f} segundos.")

            if valor_extraido is not None:
                # Stripping whitespace for cleaner results if it's a string
                valor_final = valor_extraido.strip() if isinstance(valor_extraido, str) else valor_extraido
                self.logger.info(f"\n✅ Valor final obtenido del elemento '{selector}': '{valor_final}'")
                self.tomar_captura(f"{nombre_base}_valor_extraido_exito", directorio)
                return valor_final
            else:
                self.logger.warning(f"\n❌ No se pudo extraer ningún valor significativo del elemento '{selector}'.")
                self.tomar_captura(f"{nombre_base}_fallo_extraccion_valor_no_encontrado", directorio)
                return None

        except TimeoutError as e:
            mensaje_error = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' "
                f"no se volvió visible a tiempo ({tiempo_max_espera_visibilidad}s) para extraer su valor. Detalles: {e}"
            )
            self.logger.error(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_extraccion_valor", directorio)
            # Elevar una excepción clara para que el flujo de la prueba se detenga si el elemento no está disponible
            raise AssertionError(f"\nElemento no disponible para extracción de valor: {selector}. Error: {e.message if hasattr(e, 'message') else str(e)}") from e

        except Error as e: # Captura errores específicos de Playwright (directamente 'Error' sin alias)
            mensaje_error = (
                f"\n❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar extraer el valor de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_extraccion_valor", directorio)
            raise AssertionError(f"\nError de Playwright al extraer valor: {selector}. Error: {e.message if hasattr(e, 'message') else str(e)}") from e

        except Exception as e: # Captura cualquier otro error inesperado
            mensaje_error = (
                f"\n❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar extraer el valor de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_extraccion_valor", directorio)
            raise AssertionError(f"\nError inesperado al extraer valor: {selector}. Error: {e}") from e

        finally:
            # --- Medición de rendimiento: Fin de la operación total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación (obtener_valor_de_elemento): {duration_total_operation:.4f} segundos.")
            
            # El parámetro 'tiempo' original en tu función no tenía un uso claro aquí,
            # ya que las operaciones de extracción tienen sus propios timeouts o son sincrónicas.
            # Lo he renombrado a 'tiempo_max_espera_visibilidad' y se usa en expect().to_be_visible().
            # No se añade una espera fija aquí por defecto. Si se necesita una pausa
            # adicional después de la extracción, se debería añadir un nuevo parámetro.
            pass
        
    # 74- Función para presionar la tecla TAB y verificar que el foco cambie al elemento esperado.
    # Combina la acción de TAB con una validación directa del foco, aceptando tanto selectores como objetos Locator.
    def presionar_Tab_y_verificar_foco(self, selector_o_locator_esperado: Union[str, Locator], nombre_base: str, direccion: str, tiempo_espera_post_tab: float = 0.5, nombre_paso: str = "") -> None:
        """
        Simula la acción de presionar la tecla 'TAB' y verifica que el foco del navegador
        se mueva al elemento especificado.

        Esta función es flexible, ya que puede recibir el selector del elemento
        como una cadena de texto (str) o como un objeto Locator de Playwright.

        Args:
            selector_o_locator_esperado (str | Locator): El selector del elemento (str)
                                                        o el objeto Locator (Locator) que se espera
                                                        que reciba el foco después de presionar 'TAB'.
                                                        Ejemplo: '#input-contrasena' o self.page.locator('#input-contrasena').
            nombre_base_captura (str): Nombre base para las capturas de pantalla en caso de error.
            direccion_captura (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo_espera_post_tab (float, opcional): Tiempo en segundos para esperar después de presionar 'TAB'.
                                                    Por defecto `0.5` segundos.
            nombre_paso (str, opcional): Descripción del paso para los logs. Por defecto "".

        Raises:
            AssertionError: Si el foco no se encuentra en el elemento esperado.
            Exception: Si ocurre un error inesperado.
        """
        # 1. Convertir el selector o el Locator a un objeto Locator.
        if isinstance(selector_o_locator_esperado, str):
            localizador = self.page.locator(selector_o_locator_esperado)
        else:
            localizador = selector_o_locator_esperado

        paso_descripcion = nombre_paso if nombre_paso else f"Verificando el cambio de foco a '{localizador}' después de presionar TAB."
        self.logger.info(f"\n--- {paso_descripcion} ---")

        try:
            # 2. Presionar la tecla TAB utilizando la función existente.
            self.Tab_Press(tiempo_espera_post_tab=tiempo_espera_post_tab, nombre_paso="Presionando TAB para cambiar de foco")

            self.logger.info(f"\nVerificando que el foco se haya movido al elemento: '{localizador}'...")
            
            # 3. La verificación clave con `expect`
            #     Es fundamental usar `expect()` para que `to_be_focused()` funcione.
            expect(localizador).to_be_focused()
            
            self.logger.info(f"\n✅ ÉXITO - El foco se encuentra en el elemento esperado: '{localizador}'.")

        except AssertionError as ae:
            mensaje_error = f"\n❌ FALLO de Verificación - {paso_descripcion}\n{ae}"
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_foco_fallido", direccion)
            raise ae
        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Inesperado) - {paso_descripcion}: Ocurrió un error inesperado durante la verificación.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(mensaje_error, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", direccion)
            raise AssertionError(f"\nError inesperado al verificar el foco: {e}") from e
                
    # 75- Función para presionar la combinación de teclas SHIFT + TAB en el teclado
    # Integra pruebas de rendimiento para medir el tiempo de ejecución de la acción.
    def presionar_shift_tab(self, tiempo_espera_post_shift_tab: float = 0.5, nombre_paso: str = "") -> None:
        """
        Simula la acción de presionar la combinación de teclas 'Shift + Tab' en el teclado.
        Esta función es útil para navegar *hacia atrás* entre elementos interactivos (inputs,
        botones, enlaces) en una página web, moviendo el foco al elemento tabulable anterior.
        Integra mediciones de rendimiento para la operación.

        Args:
            tiempo_espera_post_shift_tab (float, opcional): Tiempo en segundos para esperar *después*
                                                            de presionar 'Shift + Tab'. Útil para dar
                                                            tiempo a que la UI procese el cambio de foco
                                                            o se carguen elementos dinámicamente. Por defecto
                                                            `0.5` segundos.
            nombre_paso (str, opcional): Una descripción del paso que se está ejecutando para los logs.
                                        Por defecto "".

        Raises:
            Exception: Si ocurre algún error inesperado durante la simulación de la combinación de teclas.
        """
        self.logger.info(f"\n--- {nombre_paso}: Presionando la combinación de teclas SHIFT + TAB y esperando {tiempo_espera_post_shift_tab} segundos ---")

        # --- Medición de rendimiento: Inicio total de la función ---
        start_time_total_operation = time.time()

        try:
            # --- Medición de rendimiento: Inicio de la acción 'keyboard.press' ---
            start_time_press_action = time.time()
            self.page.keyboard.press("Shift+Tab")
            # --- Medición de rendimiento: Fin de la acción 'keyboard.press' ---
            end_time_press_action = time.time()
            duration_press_action = end_time_press_action - start_time_press_action
            self.logger.info(f"\nPERFORMANCE: Tiempo de la acción 'keyboard.press(\"Shift+Tab\")': {duration_press_action:.4f} segundos.")
            
            self.logger.info("\nCombinación de teclas SHIFT + TAB presionada exitosamente.")

            # Espera fija después de presionar SHIFT + TAB (configuracion por parametro)
            if tiempo_espera_post_shift_tab > 0:
                self.esperar_fijo(tiempo_espera_post_shift_tab)

        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al presionar la combinación de teclas SHIFT + TAB.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            # Se lanza la excepción para que el framework de pruebas maneje el fallo.
            raise AssertionError(f"\nError al presionar la combinación de teclas SHIFT + TAB: {e}") from e
        finally:
            # --- Medición de rendimiento: Fin total de la función ---
            end_time_total_operation = time.time()
            duration_total_operation = end_time_total_operation - start_time_total_operation
            self.logger.info(f"\nPERFORMANCE: Tiempo total de la operación (Shift_Tab_Press): {duration_total_operation:.4f} segundos.")
    
    # 76- Función para presionar la combinación de teclas SHIFT + TAB y verificar que el foco cambie al elemento esperado.
    # Combina la acción de SHIFT + TAB con una validación directa del foco, aceptando
    # selectores o objetos Locator.
    def presionar_Shift_Tab_y_verificar_foco(self, selector_o_locator_esperado: Union[str, Locator], nombre_base: str, direccion: str, tiempo_espera_post_shift_tab: float = 0.5, nombre_paso: str = "") -> None:
        """
        Simula la acción de presionar la combinación de teclas 'Shift + Tab' y verifica
        que el foco del navegador se mueva al elemento especificado.

        Esta función es flexible, ya que puede recibir el selector del elemento
        como una cadena de texto (str) o como un objeto Locator de Playwright.

        Args:
            selector_o_locator_esperado (str | Locator): El selector del elemento (str) o
                                                          el objeto Locator (Locator) que se espera
                                                          que reciba el foco después de presionar
                                                          'Shift + Tab'.
                                                          Ejemplo: '#input-contrasena' o self.page.locator('#input-contrasena').
            nombre_base (str): Nombre base para las capturas de pantalla en caso de error.
            direccion (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo_espera_post_shift_tab (float, opcional): Tiempo en segundos para esperar después de presionar
                                                             'Shift + Tab'. Por defecto `0.5` segundos.
            nombre_paso (str, opcional): Descripción del paso para los logs. Por defecto "".

        Raises:
            AssertionError: Si el foco no se encuentra en el elemento esperado.
            Exception: Si ocurre un error inesperado.
        """
        # 1. Convertir el selector o el Locator a un objeto Locator.
        if isinstance(selector_o_locator_esperado, str):
            localizador = self.page.locator(selector_o_locator_esperado)
        else:
            localizador = selector_o_locator_esperado

        paso_descripcion = nombre_paso if nombre_paso else f"Verificando el cambio de foco a '{localizador}' después de presionar SHIFT + TAB."
        # self.logger.info(f"\n--- {paso_descripcion} ---")

        try:
            # 2. Presionar la combinación de teclas SHIFT + TAB utilizando la función existente.
            self.presionar_shift_tab(tiempo_espera_post_shift_tab=tiempo_espera_post_shift_tab, nombre_paso="Presionando SHIFT + TAB para cambiar de foco")

            # self.logger.info(f"\nVerificando que el foco se haya movido al elemento: '{localizador}'...")
            
            # 3. La verificación clave con `expect` de Playwright.
            #    `to_be_focused()` verifica si el elemento actualmente tiene el foco.
            expect(localizador).to_be_focused()
            
            self.logger.info(f"\n✅ ÉXITO - El foco se encuentra en el elemento esperado: '{localizador}'.")

        except AssertionError as ae:
            mensaje_error = f"\n❌ FALLO de Verificación - {paso_descripcion}\n{ae}"
            # self.logger.error(mensaje_error)
            # self.tomar_captura(f"{nombre_base}_foco_fallido", direccion)
            raise ae
        except Exception as e:
            mensaje_error = (
                f"\n❌ FALLO (Inesperado) - {paso_descripcion}: Ocurrió un error inesperado durante la verificación.\n"
                f"Detalles: {e}"
            )
            # self.logger.critical(mensaje_error, exc_info=True)
            # self.tomar_captura(f"{nombre_base}_error_inesperado", direccion)
            raise AssertionError(f"\nError inesperado al verificar el foco: {e}") from e
    
    # 77- Función para descargar un archivo al hacer clic en un selector específico.
    def descargar_archivo(self, selector: Union[str, Locator], nombre_base: str, directorio_capturas: str, directorio_descargas: str, tiempo: Union[int, float] = 30.0) -> str:
        """
        Descarga un archivo al hacer clic en un selector específico.
        
        Esta función espera que se inicie una descarga, la guarda en un directorio local
        especificado y mide el tiempo de la operación. Es ideal para elementos como enlaces
        o botones que inician la descarga de un archivo.

        Args:
            selector (Union[str, Locator]): El selector del elemento (enlace, botón, etc.) 
                                            que desencadena la descarga.
            nombre_base (str): Nombre base para las capturas de pantalla tomadas durante la ejecución.
            directorio_capturas (str): Ruta del directorio donde se guardarán las capturas de pantalla.
            directorio_descargas (str): El directorio de destino donde se guardará el archivo descargado.
            tiempo (Union[int, float], opcional): Tiempo máximo de espera (en segundos) para que la 
                                                descarga se complete. Por defecto, 30.0 segundos.
                                                
        Returns:
            str: La ruta completa del archivo descargado si la operación es exitosa; 
                `None` en caso de cualquier fallo.

        Raises:
            TimeoutError: Si la descarga no se inicia o no se completa dentro del tiempo especificado.
            Error: Si ocurre un problema específico de Playwright, como un selector no válido.
            Exception: Para cualquier otro error inesperado.
        """
        # 1. Asegurar que el selector sea un objeto Locator para un uso uniforme.
        locator = self.page.locator(selector) if isinstance(selector, str) else selector
        self.logger.info(f"\nIntentando descargar archivo desde el selector: '{selector}'. Tiempo máximo de espera: {tiempo}s.")

        # 2. Configurar la escucha de la descarga ANTES de la acción que la desencadena.
        #    La declaración `with` asegura que la escucha se active antes de hacer clic.
        start_time_download = time.time()
        try:
            with self.page.expect_download() as download_info:
                # 3. Realizar la acción que inicia la descarga (ej. hacer clic en un enlace).
                self.logger.info(f"\nRealizando la acción de clic en el selector '{selector}' para iniciar la descarga.")
                locator.click()

            # 4. Obtener el objeto de descarga y la ruta temporal del archivo.
            download = download_info.value
            path_temp = download.path()
            file_name = download.suggested_filename

            # 5. Definir la ruta de destino y mover el archivo descargado.
            #    Es crucial mover el archivo desde su ruta temporal antes de que Playwright
            #    limpie la sesión.
            ruta_completa_del_archivo = os.path.join(directorio_descargas, file_name)
            os.makedirs(directorio_descargas, exist_ok=True) # Crea el directorio si no existe.
            download.save_as(ruta_completa_del_archivo)
            self.logger.info(f"\nArchivo guardado exitosamente: '{ruta_completa_del_archivo}'.")

            # 6. Medición de rendimiento y registro de éxito.
            end_time_download = time.time()
            duration_download = end_time_download - start_time_download
            self.logger.info(f"PERFORMANCE: Tiempo que tardó en descargar el archivo '{file_name}': {duration_download:.4f} segundos.")
            self.logger.info(f"\n✅ Archivo descargado exitosamente y guardado en '{ruta_completa_del_archivo}'.")
            self.tomar_captura(f"{nombre_base}_archivo_descargado", directorio_capturas)
            return ruta_completa_del_archivo

        except TimeoutError as e:
            # Manejo de error: la descarga no se inició o no se completó a tiempo.
            end_time_fail = time.time()
            duration_fail = end_time_fail - start_time_download
            error_msg = (
                f"\n❌ FALLO (Timeout): El elemento '{selector}' no estuvo visible/habilitado o "
                f"la descarga no se inició/completó después de {duration_fail:.4f} segundos.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_descargar_archivo", directorio_capturas)
            return None

        except Error as e:
            # Manejo de error: problemas con el selector o interacción de Playwright.
            error_msg = (
                f"\n❌ FALLO (Playwright): Error de Playwright al intentar descargar "
                f"el archivo desde el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_descarga", directorio_capturas)
            raise # Re-lanzar la excepción para que el test falle.

        except Exception as e:
            # Manejo de cualquier otro error inesperado.
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar descargar "
                f"el archivo desde el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_descarga", directorio_capturas)
            raise # Re-lanzar la excepción.
        
        """finally:
            # Este bloque siempre se ejecuta, útil para limpiezas o esperas finales.
            # Aquí se usa para una espera fija al final de la operación, si se configuró.
            if tiempo > 0:
                self.esperar_fijo(tiempo)"""
    
    # --- Manejadores y funciones para Alertas y Confirmaciones ---

    # Handler para alertas simples (usado con page.once).
    # Este handler captura información de la alerta y la acepta. Integra medición de rendimiento.
    def _get_simple_alert_handler_for_on(self):
        """
        Retorna una función handler (callback) diseñada para ser usada con `page.on('dialog', handler)`.
        
        Este handler:
        - Marca una bandera interna (`_alerta_detectada`) a True.
        - Captura el mensaje y el tipo del diálogo (`dialog.message`, `dialog.type`).
        - Registra información sobre la alerta detectada.
        - Mide el tiempo que tarda la lógica interna del handler en ejecutarse.
        - Acepta automáticamente el diálogo (`dialog.accept()`).
        - Registra la acción de aceptar el diálogo.

        Esta función no toma parámetros de selector o capturas de pantalla directas porque
        es un callback de evento que Playwright invoca.

        Returns:
            callable: Una función que toma un objeto `Dialog` como argumento y maneja la alerta.
        """
        # Se reinician las banderas para cada nueva creación del handler, útil si se usa page.once repetidamente
        self._alerta_detectada = False 
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""

        def handler(dialog: Dialog):
            """
            Función callback interna que se ejecuta cuando Playwright detecta un diálogo (alerta, confirmación, etc.).
            """
            # --- Medición de rendimiento: Inicio de la ejecución del handler ---
            start_time_handler_execution = time.time()
            self.logger.info(f"\n--- [LISTENER START] Procesando diálogo tipo: '{dialog.type}'. ---")

            try:
                self._alerta_detectada = True
                self._alerta_mensaje_capturado = dialog.message
                self._alerta_tipo_capturado = dialog.type
                
                self.logger.info(f"\n--> [LISTENER ON - Simple Alert] Alerta detectada: Tipo='{dialog.type}', Mensaje='{dialog.message}'")
                
                # Aceptamos el diálogo. Esto simula hacer clic en "Aceptar" o "OK".
                # Para un 'prompt', puedes pasar un texto: dialog.accept("texto de respuesta")
                dialog.accept() 
                self.logger.info("\n--> [LISTENER ON - Simple Alert] Alerta ACEPTADA.")

            except Exception as e:
                # Captura cualquier error que ocurra dentro del handler.
                # Es crucial aquí no re-lanzar, ya que podría romper el listener de Playwright.
                self.logger.error(f"\n❌ ERROR en el handler de alerta para '{dialog.type}' (Mensaje: '{dialog.message}'). Detalles: {e}", exc_info=True)
            finally:
                # --- Medición de rendimiento: Fin de la ejecución del handler ---
                end_time_handler_execution = time.time()
                duration_handler_execution = end_time_handler_execution - start_time_handler_execution
                self.logger.info(f"PERFORMANCE: Tiempo de ejecución del handler de alerta: {duration_handler_execution:.4f} segundos.")
                self.logger.info("\n--- [LISTENER END] Diálogo procesado. ---")

        return handler

    # Handler para diálogos de confirmación (usado con page.once).
    # Este handler captura información del diálogo, realiza una acción configurable (aceptar/descartar),
    # y registra métricas de rendimiento.
    def _get_confirmation_dialog_handler_for_on(self, accion: str):
        """
        Retorna una función handler (callback) diseñada para ser usada con `page.on('dialog', handler)`.
        Este handler está específicamente diseñado para diálogos de tipo 'confirm' o 'prompt',
        permitiendo decidir dinámicamente si se acepta o se descarta el diálogo.

        Este handler:
        - Marca una bandera interna (`_alerta_detectada`) a True.
        - Captura el mensaje y el tipo del diálogo (`dialog.message`, `dialog.type`).
        - Registra información sobre el diálogo detectado.
        - Mide el tiempo que tarda la lógica interna del handler en ejecutarse.
        - Realiza la acción especificada ('accept' o 'dismiss') en el diálogo.
        - Registra la acción tomada.
        - Por defecto, si la acción no es 'accept' ni 'dismiss', acepta el diálogo y emite una advertencia.

        Args:
            accion (str): La acción a realizar en el diálogo. Puede ser 'accept' para aceptar
                          o 'dismiss' para cancelar/descartar.

        Returns:
            callable: Una función que toma un objeto `Dialog` como argumento y maneja el diálogo.
        """
        # Se reinician las banderas para cada nueva creación del handler, útil si se usa page.once repetidamente
        self._alerta_detectada = False 
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""

        def handler(dialog: Dialog):
            """
            Función callback interna que se ejecuta cuando Playwright detecta un diálogo
            (especialmente 'confirm' o 'prompt').
            """
            # --- Medición de rendimiento: Inicio de la ejecución del handler ---
            start_time_handler_execution = time.time()
            self.logger.info(f"\n--- [LISTENER START] Procesando diálogo de confirmación tipo: '{dialog.type}'. ---")

            try:
                self._alerta_detectada = True
                self._alerta_mensaje_capturado = dialog.message
                self._alerta_tipo_capturado = dialog.type
                
                self.logger.info(f"\n--> [LISTENER ON - Dinámico] Diálogo detectado: Tipo='{dialog.type}', Mensaje='{dialog.message}'")
                
                if accion == 'accept':
                    # Acepta el diálogo (equivalente a hacer clic en "OK" o "Aceptar").
                    # Para un prompt, puedes pasar un valor: dialog.accept("mi respuesta")
                    dialog.accept()
                    self.logger.info("\n--> [LISTENER ON - Dinámico] Diálogo ACEPTADO.")
                elif accion == 'dismiss':
                    # Descarta/cancela el diálogo (equivalente a hacer clic en "Cancelar").
                    dialog.dismiss()
                    self.logger.info("\n--> [LISTENER ON - Dinámico] Diálogo CANCELADO/DESCARTADO.")
                else:
                    # En caso de acción no reconocida, se registra una advertencia y se acepta por defecto.
                    self.logger.warning(f"\n--> [LISTENER ON - Dinámico] Acción desconocida '{accion}' para el diálogo '{dialog.type}'. Aceptando por defecto.")
                    dialog.accept()
                    self.logger.info("\n--> [LISTENER ON - Dinámico] Diálogo ACEPTADO por defecto debido a acción inválida.")

            except Exception as e:
                # Captura cualquier error que ocurra dentro del handler.
                # Es crucial aquí no re-lanzar, ya que podría romper el listener de Playwright.
                self.logger.error(f"\n❌ ERROR en el handler de diálogo para '{dialog.type}' (Mensaje: '{dialog.message}', Acción: '{accion}'). Detalles: {e}", exc_info=True)
            finally:
                # --- Medición de rendimiento: Fin de la ejecución del handler ---
                end_time_handler_execution = time.time()
                duration_handler_execution = end_time_handler_execution - start_time_handler_execution
                self.logger.info(f"PERFORMANCE: Tiempo de ejecución del handler de diálogo de confirmación: {duration_handler_execution:.4f} segundos.")
                self.logger.info("\n--- [LISTENER END] Diálogo procesado. ---")

        return handler
    
    # Handler para diálogos de pregunta (prompt) (usado con page.once).
    # Este handler captura información del diálogo prompt, introduce un texto opcional,
    # realiza una acción configurable (aceptar/descartar), y registra métricas de rendimiento.
    def _get_prompt_dialog_handler_for_on(self, input_text: str = "", accion: str = "accept"):
        """
        Retorna una función handler (callback) diseñada para ser usada con `page.on('dialog', handler)`.
        Este handler está específicamente diseñado para diálogos de tipo 'prompt', permitiendo
        introducir texto y decidir dinámicamente si se acepta o se descarta el diálogo.

        Este handler:
        - Marca una bandera interna (`_alerta_detectada`) a True.
        - Captura el mensaje, el tipo del diálogo y el texto de entrada (`dialog.message`, `dialog.type`, `input_text`).
        - Registra información sobre el diálogo detectado.
        - Mide el tiempo que tarda la lógica interna del handler en ejecutarse.
        - Realiza la acción especificada ('accept' o 'dismiss') en el diálogo.
        - Si la acción es 'accept' y el tipo de diálogo es 'prompt', introduce el `input_text`.
        - Registra la acción tomada.
        - Por defecto, si la acción no es 'accept' ni 'dismiss', descarta el diálogo y emite una advertencia.

        Args:
            input_text (str, opcional): El texto a introducir en el campo de entrada del prompt si se acepta.
                                        Por defecto es una cadena vacía "".
            accion (str, opcional): La acción a realizar en el diálogo. Puede ser 'accept' para aceptar
                                    o 'dismiss' para cancelar/descartar. Por defecto es 'accept'.

        Returns:
            callable: Una función que toma un objeto `Dialog` como argumento y maneja el diálogo.
        """
        # Se reinician las banderas para cada nueva creación del handler
        self._alerta_detectada = False 
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""
        self._alerta_input_capturado = ""

        def handler(dialog: Dialog):
            """
            Función callback interna que se ejecuta cuando Playwright detecta un diálogo
            (especialmente de tipo 'prompt').
            """
            # --- Medición de rendimiento: Inicio de la ejecución del handler ---
            start_time_handler_execution = time.time()
            self.logger.info(f"\n--- [LISTENER START] Procesando diálogo de prompt tipo: '{dialog.type}'. ---")

            try:
                self._alerta_detectada = True
                self._alerta_mensaje_capturado = dialog.message
                self._alerta_tipo_capturado = dialog.type
                self._alerta_input_capturado = input_text # Almacena el texto que se intentó introducir

                self.logger.info(f"\n--> [LISTENER ON - Prompt Dinámico] Diálogo detectado: Tipo='{dialog.type}', Mensaje='{dialog.message}'.")
                
                if accion == 'accept':
                    if dialog.type == "prompt":
                        # Acepta el prompt e introduce el texto proporcionado.
                        dialog.accept(input_text)
                        self.logger.info(f"\n--> [LISTENER ON - Prompt Dinámico] Texto '{input_text}' introducido y prompt ACEPTADO.")
                    else:
                        # Si no es un prompt pero se especificó 'accept', lo acepta sin texto.
                        self.logger.warning(f"\n--> [LISTENER ON - Prompt Dinámico] Se solicitó 'accept' con texto para un diálogo no-prompt ('{dialog.type}'). Aceptando sin texto.")
                        dialog.accept()
                        self.logger.info("\n--> [LISTENER ON - Prompt Dinámico] Diálogo ACEPTADO (sin texto, no es prompt).")
                elif accion == 'dismiss':
                    # Descarta/cancela el diálogo. El texto de input_text se ignora.
                    dialog.dismiss()
                    self.logger.info("\n--> [LISTENER ON - Prompt Dinámico] Diálogo CANCELADO/DESCARTADO.")
                else:
                    # En caso de acción no reconocida, se registra una advertencia y se descarta por defecto.
                    # Se elige 'dismiss' como valor por defecto más seguro para evitar que el prompt
                    # se quede abierto y bloquee la ejecución si la acción es inválida.
                    self.logger.warning(f"\n--> [LISTENER ON - Prompt Dinámico] Acción desconocida '{accion}' para el diálogo '{dialog.type}'. Descartando por defecto.")
                    dialog.dismiss()
                    self.logger.info("\n--> [LISTENER ON - Prompt Dinámico] Diálogo DESCARTADO por defecto debido a acción inválida.")

            except Exception as e:
                # Captura cualquier error que ocurra dentro del handler.
                # Es crucial aquí no re-lanzar, ya que podría romper el listener de Playwright.
                self.logger.error(f"\n❌ ERROR en el handler de prompt para '{dialog.type}' (Mensaje: '{dialog.message}', Acción: '{accion}', Texto: '{input_text}'). Detalles: {e}", exc_info=True)
            finally:
                # --- Medición de rendimiento: Fin de la ejecución del handler ---
                end_time_handler_execution = time.time()
                duration_handler_execution = end_time_handler_execution - start_time_handler_execution
                self.logger.info(f"PERFORMANCE: Tiempo de ejecución del handler de diálogo de prompt: {duration_handler_execution:.4f} segundos.")
                self.logger.info("\n--- [LISTENER END] Diálogo procesado. ---")

        return handler

    # Handler de eventos para cuando se abre una nueva página (popup/nueva pestaña).
    # Este handler se encarga de detectar y registrar información sobre nuevas páginas,
    # y también mide el tiempo de procesamiento interno.
    def _on_new_page(self, page: Page):
        """
        Manejador de eventos (callback) para detectar nuevas páginas o ventanas emergentes (popups)
        que se abren, por ejemplo, al hacer clic en un enlace con `target="_blank"`.
        
        Este handler:
        - Marca una bandera interna (`_popup_detectado`) a True.
        - Almacena la referencia al objeto `Page` de la nueva ventana.
        - Captura la URL y el título de la nueva página.
        - Añade la nueva página a una lista de todas las páginas detectadas.
        - Registra información sobre la nueva página detectada.
        - Mide el tiempo que tarda la lógica interna del handler en ejecutarse.

        Args:
            page (Page): El objeto `Page` de Playwright que representa la nueva ventana/pestaña abierta.
                         Este es proporcionado automáticamente por Playwright cuando se dispara el evento.
        """
        # --- Medición de rendimiento: Inicio de la ejecución del handler ---
        start_time_handler_execution = time.time()
        self.logger.info("\n--- [LISTENER START] Procesando evento de nueva página. ---")

        try:
            self._popup_detectado = True
            self._popup_page = page
            self._popup_url_capturado = page.url
            # El title() puede requerir una pequeña espera si la página no ha cargado lo suficiente.
            # Sin embargo, para un handler que debe ser rápido, se asume que estará disponible.
            # Si el título no se obtiene inmediatamente, podría ser None o vacío.
            self._popup_title_capturado = page.title() 
            self._all_new_pages_opened_by_click.append(page) # Añadir la nueva página a la lista

            self.logger.info(f"\n🌐 Nueva página (popup/pestaña) detectada. URL: '{page.url}', Título: '{page.title()}'")
            # Opcional: Si solo te interesa la primera popup o una específica, podrías manejarlo aquí.
            # Por ahora, solo la añadimos a la lista para seguimiento.

        except Exception as e:
            # Es crucial capturar excepciones en handlers para evitar que Playwright deshabilite el listener.
            self.logger.error(f"\n❌ ERROR en el handler de nueva página. Detalles: {e}", exc_info=True)
        finally:
            # --- Medición de rendimiento: Fin de la ejecución del handler ---
            end_time_handler_execution = time.time()
            duration_handler_execution = end_time_handler_execution - start_time_handler_execution
            self.logger.info(f"PERFORMANCE: Tiempo de ejecución del handler de nueva página: {duration_handler_execution:.4f} segundos.")
            self.logger.info("\n--- [LISTENER END] Evento de nueva página procesado. ---")
        
    # 78 (Número consecutivo si es parte de una serie)- Función privada para realizar Drag and Drop manual.
    # Utiliza las acciones de ratón de bajo nivel de Playwright para simular arrastrar y soltar.
    # Se usa como método de fallback si el drag_and_drop() automático no funciona.
    # Integra mediciones de rendimiento detalladas.
    def _realizar_drag_and_drop_manual(self, elemento_origen: Locator, elemento_destino: Locator, 
                                      nombre_base: str, directorio: str, nombre_paso: str, 
                                      tiempo_pausa_ms: Union[int, float] = 1000, timeout_locators_ms: int = 5000) -> None:
        """
        Realiza una operación de "Drag and Drop" (arrastrar y soltar) utilizando acciones de ratón
        de bajo nivel de Playwright. Este método es útil como alternativa cuando el método
        `locator.drag_and_drop()` no produce el comportamiento deseado o es insuficiente.
        
        Mide el tiempo de cada paso clave (hover, click, drag, drop) para proporcionar
        métricas de rendimiento detalladas de esta operación manual.

        Args:
            elemento_origen (Locator): El Locator del elemento que se desea arrastrar.
            elemento_destino (Locator): El Locator del elemento donde se desea soltar el origen.
            nombre_base (str): Nombre base para las capturas de pantalla, asegurando un nombre único.
            directorio (str): Directorio donde se guardarán las capturas de pantalla. El directorio
                              se creará si no existe.
            nombre_paso (str): Una descripción del paso que se está ejecutando para el registro (logs).
            tiempo_pausa_ms (Union[int, float], opcional): Tiempo de pausa en milisegundos después de
                                                            presionar el ratón y después de arrastrarlo
                                                            sobre el destino. Por defecto es 1000ms (1 segundo).
                                                            Esto simula un arrastre más "humano".
            timeout_locators_ms (int, opcional): Tiempo máximo en milisegundos que Playwright esperará
                                                a que los localizadores sean visibles/interactuables
                                                durante las operaciones de `hover`. Por defecto es 5000ms.

        Raises:
            Error: Si ocurre un error específico de Playwright durante las operaciones del ratón.
            Exception: Para cualquier otro error inesperado durante la ejecución.
        """
        self.logger.info(f"\n--- {nombre_paso}: Intentando 'Drag and Drop' manualmente de '{elemento_origen}' a '{elemento_destino}'. ---")

        # Asegurarse de que el directorio de capturas de pantalla exista
        if not os.path.exists(directorio):
            os.makedirs(directorio, exist_ok=True)
            self.logger.info(f"\n☑️ Directorio de capturas de pantalla creado: {directorio}")

        # --- Medición de rendimiento: Inicio de la operación total de Drag and Drop manual ---
        start_time_total_drag_drop = time.time()
        
        try:
            self.tomar_captura(f"{nombre_base}_antes_drag_drop_manual", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada antes del D&D manual: '{nombre_base}_antes_drag_drop_manual.png'")

            # 1. Mover el ratón sobre el elemento de origen
            start_time_hover_origin = time.time()
            self.logger.info(f"\n🖱️ Moviendo ratón sobre elemento de origen: '{elemento_origen}'...")
            elemento_origen.hover()
            end_time_hover_origin = time.time()
            duration_hover_origin = end_time_hover_origin - start_time_hover_origin
            self.logger.info(f"PERFORMANCE: Tiempo de 'hover' en origen: {duration_hover_origin:.4f} segundos.")

            # 2. Presionar el botón izquierdo del ratón (iniciar arrastre)
            start_time_mouse_down = time.time()
            self.logger.info("\n⬇️ Presionando botón izquierdo del ratón para iniciar arrastre...")
            self.page.mouse.down()
            end_time_mouse_down = time.time()
            duration_mouse_down = end_time_mouse_down - start_time_mouse_down
            self.logger.info(f"PERFORMANCE: Tiempo de 'mouse.down': {duration_mouse_down:.4f} segundos.")

            # Pausa para simular arrastre humano
            if tiempo_pausa_ms > 0:
                self.logger.info(f"\n⏳ Pausa durante arrastre (simulación): {tiempo_pausa_ms} ms...")
                self.page.wait_for_timeout()

            # 3. Mover el ratón sobre el elemento de destino
            start_time_hover_destination = time.time()
            self.logger.info(f"\n➡️ Moviendo ratón sobre elemento de destino: '{elemento_destino}'...")
            elemento_destino.hover(timeout=timeout_locators_ms)
            end_time_hover_destination = time.time()
            duration_hover_destination = end_time_hover_destination - start_time_hover_destination
            self.logger.info(f"PERFORMANCE: Tiempo de 'hover' en destino: {duration_hover_destination:.4f} segundos.")

            # Pausa adicional antes de soltar, si se desea un comportamiento más humano
            if tiempo_pausa_ms > 0:
                self.logger.info(f"\n⏳ Pausa antes de soltar (simulación): {tiempo_pausa_ms} ms...")
                self.page.wait_for_timeout()

            # 4. Soltar el botón izquierdo del ratón (finalizar arrastre)
            start_time_mouse_up = time.time()
            self.logger.info("\n⬆️ Soltando botón izquierdo del ratón para finalizar arrastre...")
            self.page.mouse.up()
            end_time_mouse_up = time.time()
            duration_mouse_up = end_time_mouse_up - start_time_mouse_up
            self.logger.info(f"PERFORMANCE: Tiempo de 'mouse.up': {duration_mouse_up:.4f} segundos.")

            self.logger.info(f"\n✔ ÉXITO: 'Drag and Drop' manual realizado exitosamente de '{elemento_origen}' a '{elemento_destino}'.")
            self.tomar_captura(f"{nombre_base}_despues_drag_drop_manual", directorio)
            self.logger.info(f"\n📸 Captura de pantalla tomada después del D&D manual: '{nombre_base}_despues_drag_drop_manual.png'")

        except Error as e:
            error_msg = (
                f"\n❌ FALLO (Playwright Error - Manual) - {nombre_paso}: Ocurrió un error de Playwright al intentar realizar 'Drag and Drop' manualmente.\n"
                f"Asegúrate de que los elementos sean visibles e interactuables. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_manual_drag_and_drop_playwright", directorio)
            raise # Re-lanza la excepción original de Playwright.
        
        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado - Manual) - {nombre_paso}: Ocurrió un error inesperado al intentar realizar 'Drag and Drop' manualmente.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Uso critical para errores inesperados graves.
            self.tomar_captura(f"{nombre_base}_error_inesperado_manual_drag_and_drop", directorio)
            raise # Re-lanza la excepción.
        
        finally:
            # --- Medición de rendimiento: Fin de la operación total de Drag and Drop manual ---
            end_time_total_drag_drop = time.time()
            duration_total_drag_drop = end_time_total_drag_drop - start_time_total_drag_drop
            self.logger.info(f"PERFORMANCE: Tiempo total de la operación 'Drag and Drop' manual: {duration_total_drag_drop:.4f} segundos.")